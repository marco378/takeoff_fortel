#!/usr/bin/env python3
"""Self-contained CI tests (NO client drawings — those are gitignored). Exit non-zero on failure."""
import os, sys, shutil
from pathlib import Path
from reportlab.pdfgen import canvas
from geometry import measure_regions
from scale import detect_scale_bar
from pricing import slab_rate, price_project

# Every training derivative produced by CI is explicitly labelled and isolated from the
# repository/live volume, so cleanup never has to infer test pollution from client-like names.
os.environ.setdefault("LEARNING_ENVIRONMENT", "test")
os.environ.setdefault("TRAINING_LOG_FILE", f"/tmp/fortel_ci_training_{os.getpid()}.jsonl")
os.environ.setdefault("LEARNED_PATTERNS_FILE", f"/tmp/fortel_ci_patterns_{os.getpid()}.json")

P = []
def ck(n, c, g=""):
    P.append(bool(c)); print(f"  [{'PASS' if c else 'FAIL'}] {n} {g}")

class _FixtureNotPresent(Exception):
    pass

def _require_fixture(path, reason):
    if not Path(path).exists():
        raise _FixtureNotPresent(reason)

print("geometry")
K = 0.1
a, _ = measure_regions([[(0,0),(2000,0),(2000,1300),(0,1300)]], K,
                       holes={0: [[(200,200),(600,200),(600,500),(200,500)], [(1400,800),(1700,800),(1700,1100),(1400,1100)]]})
ck("voids 23,900", a == 23900)
a, f = measure_regions([[(0,0),(1000,1000),(1000,0),(0,1000)]], K); ck("self-intersect repaired+flagged", a == 5000 and f)
a, f = measure_regions([[(0,0),(1000,0),(1000,1000),(0,1000)], [(500,500),(1500,500),(1500,1500),(500,1500)]], K)
ck("overlap unioned 17,500", a == 17500 and f)
a, _ = measure_regions([[(0,0),(1,1)]], K); ck("degenerate <3 -> 0", a == 0.0)
try: measure_regions([[(0,0),(1,0),(1,1)]], None); ck("missing scale raises", False)
except ValueError: ck("missing scale raises", True)

print("scale")
c = canvas.Canvas("/tmp/_sb.pdf", pagesize=(1400,2200)); c.rect(200,1000,1000,800); c.line(100,150,600,150); c.drawString(250,160,"0          50 m"); c.save()
k, info = detect_scale_bar("/tmp/_sb.pdf"); ck("scale-bar k=0.1", k == 0.1, info)

# Rotated Office GA bars expose intermediate ticks as fused 1m/5m/10m tokens. The terminal
# value must win over the tick nearest the raw bottom-right corner.
_ticks = canvas.Canvas("/tmp/_sb_fused_ticks.pdf", pagesize=(700, 500))
_ticks.line(100, 80, 383.44, 80)
_ticks.drawString(125, 90, "1m")
_ticks.drawString(240, 90, "5m")
_ticks.drawString(375, 90, "10m")
_ticks.save()
_tick_k, _tick_info = detect_scale_bar("/tmp/_sb_fused_ticks.pdf")
ck("scale bar uses terminal 10m tick, not intermediate 1m",
   _tick_k is not None and abs(_tick_k - 10 / 283.44) < 1e-5 and "10 m" in _tick_info,
   (_tick_k, _tick_info))

print("fast refusal for obvious multi-page reports")
_das_pdf = "/tmp/_synthetic_heavy_das_report.pdf"
_das_canvas = canvas.Canvas(_das_pdf, pagesize=(1400, 900))
for _das_page in range(12):
    _das_canvas.setFont("Helvetica-Bold", 22)
    _das_canvas.drawString(80, 820, "Design and Access Statement")
    _das_canvas.setFont("Helvetica", 8)
    for _das_row in range(80):
        _das_canvas.drawString(80, 790 - _das_row * 9,
                               f"Planning report narrative row {_das_row} page {_das_page + 1}")
    _das_canvas.showPage()
_das_canvas.save()
from takeoff_pipeline import takeoff as _pipeline_takeoff_fast_report
_das_result = _pipeline_takeoff_fast_report(_das_pdf, auto_extract_spec=False,
                                             send_approval=False)
ck("obvious DAS report refuses before page ranking/raster measurement",
   _das_result.get("measurement_state") == "UNMEASURED" and
   _das_result.get("method") == "fast-refuse" and
   _das_result.get("area_m2") is None and
   any("no page was rasterised" in flag for flag in _das_result.get("flags", [])),
   _das_result)

_spec_pdf = "/tmp/Project_Technical_Specification.pdf"
_spec_canvas = canvas.Canvas(_spec_pdf, pagesize=(900, 700))
for _spec_page in range(6):
    _spec_canvas.drawString(60, 650, "DEVELOPMENT TECHNICAL SPECIFICATION")
    _spec_canvas.drawString(60, 620, f"Written requirements page {_spec_page + 1}")
    _spec_canvas.showPage()
_spec_canvas.save()
_spec_result = _pipeline_takeoff_fast_report(
    _spec_pdf, auto_extract_spec=False, send_approval=False)
ck("obvious specification refuses before page ranking/raster measurement",
   _spec_result.get("measurement_state") == "UNMEASURED" and
   _spec_result.get("method") == "fast-refuse" and
   _spec_result.get("area_m2") is None and
   any("filename 'Specification'" in flag and "first-page text 'TECHNICAL SPECIFICATION'" in flag
       for flag in _spec_result.get("flags", [])),
   _spec_result)

_schedule_pdf = "/tmp/Tender_Query_Schedule.pdf"
_schedule_canvas = canvas.Canvas(_schedule_pdf, pagesize=(900, 700))
_schedule_canvas.drawString(60, 650, "Tender Query Schedule")
_schedule_canvas.save()
_schedule_result = _pipeline_takeoff_fast_report(
    _schedule_pdf, auto_extract_spec=False, send_approval=False)
ck("single-page schedule is refused from strong filename evidence",
   _schedule_result.get("measurement_state") == "UNMEASURED" and
   _schedule_result.get("method") == "fast-refuse" and
   _schedule_result.get("area_m2") is None,
   _schedule_result)

_incidental_pdf = "/tmp/External_Works_Layout.pdf"
_incidental_canvas = canvas.Canvas(_incidental_pdf, pagesize=(900, 700))
_incidental_canvas.drawString(60, 650, "External Works Layout")
_incidental_canvas.drawString(60, 620, "Refer to door schedule for ancillary information")
_incidental_canvas.save()
import fitz as _fitz_fast_refusal
with _fitz_fast_refusal.open(_incidental_pdf) as _incidental_doc:
    _incidental_reason = __import__("takeoff_pipeline")._fast_report_refusal(
        _incidental_pdf, _incidental_doc)
ck("incidental schedule note on a one-sheet drawing is not fast-refused",
   _incidental_reason is None, _incidental_reason)

_boundary_pdf = "/tmp/Standalone_Boundary_Treatment_Plan.pdf"
_boundary_canvas = canvas.Canvas(_boundary_pdf, pagesize=(900, 700))
_boundary_canvas.drawString(60, 650, "BOUNDARY TREATMENT PLAN")
_boundary_canvas.drawString(60, 620, "Paladin fencing and timber acoustic fence")
_boundary_canvas.save()
_boundary_result = _pipeline_takeoff_fast_report(
    _boundary_pdf, auto_extract_spec=False, send_approval=False)
ck("dedicated boundary-treatment drawing refuses before raster measurement",
   _boundary_result.get("measurement_state") == "UNMEASURED" and
   _boundary_result.get("method") == "fast-refuse" and
   _boundary_result.get("area_m2") is None and
   any("boundary-treatment/fencing" in flag and "no page was rasterised" in flag
       for flag in _boundary_result.get("flags", [])),
   _boundary_result)

_combined_boundary_pdf = "/tmp/Hub_Hard_Landscaping_and_Boundary_Treatment_Plan.pdf"
_combined_boundary_canvas = canvas.Canvas(_combined_boundary_pdf, pagesize=(900, 700))
_combined_boundary_canvas.drawString(60, 650, "HARD LANDSCAPING AND BOUNDARY TREATMENT PLAN")
_combined_boundary_canvas.save()
with _fitz_fast_refusal.open(_combined_boundary_pdf) as _combined_boundary_doc:
    _combined_boundary_reason = __import__("takeoff_pipeline")._fast_report_refusal(
        _combined_boundary_pdf, _combined_boundary_doc)
ck("combined hard-landscaping boundary plan is not fast-refused",
   _combined_boundary_reason is None, _combined_boundary_reason)

_fake_large_container = Path("/tmp/_synthetic_large_tender.zip")
_fake_large_container.write_bytes(b"PK\x03\x04" + b"not-a-pdf" * 1024)
_container_result = _pipeline_takeoff_fast_report(
    str(_fake_large_container), auto_extract_spec=False, send_approval=False)
ck("non-PDF container refuses from magic bytes before MuPDF format probing",
   _container_result.get("measurement_state") == "REJECTED" and
   _container_result.get("area_m2") is None and
   any("%PDF header absent" in flag for flag in _container_result.get("flags", [])),
   _container_result)

print("Office GA assisted-trace vector candidates")
from office_candidates import detect_office_candidates as _office_candidates
_office_pdf = "/tmp/_office_candidates.pdf"
_office_canvas = canvas.Canvas(_office_pdf, pagesize=(900, 500))
for _index, _level in enumerate((0, 1, 2)):
    _x = 60 + _index * 280
    _office_canvas.rect(_x, 220, 100, 100, stroke=1, fill=0)
    _office_canvas.drawString(_x, 190, f"Office Plan Level {_level:02d}")
_office_canvas.line(60, 120, 160, 120)  # open line: must not become a candidate
_office_canvas.rect(800, 100, 5, 5, stroke=1, fill=0)  # symbol: below 20 m2
_office_canvas.save()
_office_detected = _office_candidates(_office_pdf, scale_k=0.1, scale_verified=True)
_office_by_level = {candidate["level"]: candidate
                    for candidate in _office_detected["candidate_polygons"]}
ck("closed vector plates become one assisted candidate per labelled level",
   set(_office_by_level) == {0, 1, 2} and len(_office_by_level) == 3 and
   all(candidate["outline_status"] == "resolved"
       for candidate in _office_detected["candidate_polygons"]),
   _office_detected)
ck("Level 00 maps to ground and upper levels stay separate",
   _office_by_level[0]["category"] == "ground_floor" and
   all(_office_by_level[level]["category"] == "upper_floor" for level in (1, 2)))
ck("candidate records are geometry-only tracing aids",
   all("area_m2" not in candidate and
       candidate["coordinate_space"] == "rotated_pdf_points" and
       candidate["confidence"] in {"low", "medium"} and
       candidate.get("confidence_reasons")
       for candidate in _office_detected["candidate_polygons"]))
from office_candidates import _dedupe_iou as _office_dedupe_iou
from shapely.geometry import box as _office_box
_iou_records = _office_dedupe_iou([
    {"geometry": _office_box(0, 0, 100, 100), "source": "office-vector-closed-loop"},
    {"geometry": _office_box(1, 1, 101, 101), "source": "office-vector-white-fill-loop"},
    # Same area but disjoint: an area-only deduper would incorrectly remove this core.
    {"geometry": _office_box(200, 0, 300, 100), "source": "office-vector-closed-loop"},
])
ck("office loop dedupe uses IoU and preserves equal-area disjoint regions",
   len(_iou_records) == 2)

_office_unresolved_pdf = "/tmp/_office_candidates_unresolved.pdf"
_office_unresolved_canvas = canvas.Canvas(_office_unresolved_pdf, pagesize=(900, 500))
_office_unresolved_canvas.drawString(60, 190, "Office Plan Level 00")
_office_unresolved_canvas.drawString(340, 190, "Office Plan Level 01")
_office_unresolved_canvas.save()
_office_unresolved = _office_candidates(
    _office_unresolved_pdf, scale_k=0.1, scale_verified=True)
ck("detected level without a defensible outline is reported, never dropped",
   [candidate["level"] for candidate in _office_unresolved["candidate_polygons"]] == [0, 1] and
   all(candidate["outline_status"] == "unresolved"
       for candidate in _office_unresolved["candidate_polygons"]) and
   any("level detected but outline not resolved — trace manually" in flag
       for flag in _office_unresolved["flags"]),
   _office_unresolved)

_steelwork_title_pdf = "/tmp/_office_first_floor_steelwork_title.pdf"
_steelwork_title_canvas = canvas.Canvas(_steelwork_title_pdf, pagesize=(900, 500))
_steelwork_title_canvas.drawString(250, 60, "First Floor Steelwork Layout")
_steelwork_title_canvas.drawString(650, 450, "Metal Deck Notes")
_steelwork_title_canvas.setStrokeColorRGB(0, 0.59, 0)
for _deck_row in range(30):
    for _deck_col in range(40):
        _hx = 220 + _deck_col * 3
        _hy = 210 + _deck_row * 3
        _steelwork_title_canvas.line(_hx, _hy, _hx + 1, _hy + 1)
_steelwork_title_canvas.save()
_steelwork_title_result = _office_candidates(
    _steelwork_title_pdf, scale_k=0.1, scale_verified=False)
ck("First Floor Steelwork Layout creates a real quantity-free proposed polygon",
   len(_steelwork_title_result["candidate_polygons"]) == 1 and
   _steelwork_title_result["candidate_polygons"][0]["level"] == 1 and
   _steelwork_title_result["candidate_polygons"][0]["category"] == "upper_floor" and
   _steelwork_title_result["candidate_polygons"][0]["outline_status"] == "proposed" and
   len(_steelwork_title_result["candidate_polygons"][0]["polygon_pts"]) >= 4 and
   "Metal Deck" in (_steelwork_title_result["candidate_polygons"][0].get("basis") or "") and
   "area_m2" not in _steelwork_title_result["candidate_polygons"][0],
   _steelwork_title_result)
_steelwork_auto = _steelwork_title_result.get("auto_measurement") or {}
ck("corroborated repeated deck hatch measures only at MEASURED_UNVERIFIED",
   _steelwork_auto.get("area_m2", 0) > 0 and
   _steelwork_auto.get("measurement_state") == "MEASURED_UNVERIFIED" and
   _steelwork_auto.get("needs_assessor") is True and
   _steelwork_auto.get("perimeter_lm", 0) > 0 and
   len(_steelwork_auto.get("zones") or []) == 1 and
   _steelwork_auto["zones"][0]["category"] == "upper_floor",
   _steelwork_auto)
_steelwork_no_scale = _office_candidates(
    _steelwork_title_pdf, scale_k=None, scale_verified=False)
ck("deck-hatch polygon stays a quantity-free proposal when scale is absent",
   _steelwork_no_scale["candidate_polygons"][0]["outline_status"] == "proposed" and
   "auto_measurement" not in _steelwork_no_scale and
   "area_m2" not in _steelwork_no_scale["candidate_polygons"][0],
   _steelwork_no_scale)

_office_no_sibling_pdf = "/tmp/_office_candidates_sibling_prefill.pdf"
_office_no_sibling_canvas = canvas.Canvas(_office_no_sibling_pdf, pagesize=(900, 500))
_office_no_sibling_canvas.rect(60, 220, 100, 100, stroke=1, fill=0)
_office_no_sibling_canvas.drawString(60, 190, "Office Plan Level 01")
_office_no_sibling_canvas.drawString(340, 190, "Office Plan Level 02")
_office_no_sibling_canvas.save()
_office_no_sibling = _office_candidates(
    _office_no_sibling_pdf, scale_k=0.1, scale_verified=True)
_office_no_sibling_by_level = {
    candidate["level"]: candidate
    for candidate in _office_no_sibling["candidate_polygons"]
}
ck("missing local Office loop gets geometry-only sibling prefill, never a measured area",
   _office_no_sibling_by_level[1]["outline_status"] == "resolved" and
   _office_no_sibling_by_level[2]["outline_status"] == "prefill" and
   _office_no_sibling_by_level[2]["regions"] and
   "area_m2" not in _office_no_sibling_by_level[2] and
   "no level-local dark closed plate loop" in
       " ".join(_office_no_sibling_by_level[2]["confidence_reasons"]) and
   "sibling-level-prefill" in _office_no_sibling_by_level[2]["source"] and
   any("LOW-CONFIDENCE PREFILL" in flag
       for flag in _office_no_sibling_by_level[2]["flags"]),
   _office_no_sibling)

print("accuracy scorecard harness")
from accuracy_report import (
    discover_pairs as _accuracy_discover_pairs,
    normalise_drawing_name as _accuracy_normalise_name,
    pair_marked_and_raw as _accuracy_pair,
    score_drawing as _accuracy_score_drawing,
    scorecard as _accuracy_scorecard,
    signed_delta_pct as _accuracy_delta,
    within_tolerance as _accuracy_within,
)
_accuracy_marked_record = {
    "path": Path("/tmp/Markup Project-Alpha_marked.pdf"),
    "truth": {"area_m2": 100.0, "zones": []},
}
_accuracy_raw_record = {"path": Path("/tmp/raw/Project Alpha.pdf")}
_accuracy_other_raw = {"path": Path("/tmp/raw/Project Beta.pdf")}
_accuracy_pairs, _accuracy_unused = _accuracy_pair(
    [_accuracy_marked_record], [_accuracy_raw_record, _accuracy_other_raw])
ck("accuracy pairing normalises marked/raw prefixes and suffixes without cross-pairing",
   _accuracy_normalise_name(_accuracy_marked_record["path"]) == "projectalpha" and
   _accuracy_normalise_name("MarkupProject Alpha.pdf") == "projectalpha" and
   len(_accuracy_pairs) == 1 and
   _accuracy_pairs[0]["raw_path"] == _accuracy_raw_record["path"] and
   len(_accuracy_unused) == 1 and "Project Beta.pdf" in _accuracy_unused[0],
   {"pairs": _accuracy_pairs, "unused": _accuracy_unused})
ck("accuracy tolerance maths is signed and inclusive at exactly 5%",
   _accuracy_delta(95, 100) == -5 and
   _accuracy_within(105, 100, 5) and
   not _accuracy_within(105.01, 100, 5))

_accuracy_not_measured = _accuracy_score_drawing(
    marked_path=Path("/tmp/marked.pdf"),
    raw_label="temporary stripped copy",
    raw_mode="derived-stripped",
    pairing="test",
    truth={"area_m2": 100.0,
           "zones": [{"category": "external_yard", "area_m2": 100.0}]},
    pipeline_run={
        "ok": True, "elapsed_s": 0.1,
        "payload": {
            "area_m2": None, "measurement_state": "UNMEASURED",
            "flags": ["REFUSED: no reliable slab region"], "zones": [],
        },
    },
    tolerance_pct=5,
)
ck("accuracy NOT MEASURED is an explicit client miss, never a silent pass",
   _accuracy_not_measured["verdict"] == "NOT MEASURED" and
   _accuracy_not_measured["failure_mode"] == "not measured" and
   _accuracy_not_measured["measured_total_m2"] is None,
   _accuracy_not_measured)

_accuracy_summary = _accuracy_scorecard([
    {"verdict": "PASS", "failure_mode": None},
    {"verdict": "FAIL", "failure_mode": "over-measured"},
    {"verdict": "FAIL", "failure_mode": "under-measured"},
    {"verdict": "FAIL", "failure_mode": "zone mis-split"},
    {"verdict": "NOT MEASURED", "failure_mode": "not measured"},
], 5)
ck("accuracy scorecard arithmetic counts passes and each failure mode",
   _accuracy_summary["passed"] == 1 and
   _accuracy_summary["drawings"] == 5 and
   _accuracy_summary["accuracy_pct"] == 20.0 and
   _accuracy_summary["summary_line"] == "1 of 5 within 5% (20.0%)" and
   all(_accuracy_summary["breakdown"][mode] == 1 for mode in (
       "not measured", "over-measured", "under-measured", "zone mis-split")),
   _accuracy_summary)

try:
    _require_fixture("drawings/castle_donington",
                     "accuracy harness Castle Donington pairing integration")
    _accuracy_castle_pairs = _accuracy_discover_pairs(
        ["drawings/castle_donington"])["pairs"]
    ck("accuracy harness pairs all 8 Castle truths without measuring a marked answer",
       len(_accuracy_castle_pairs) == 8 and
       sum(pair["raw_mode"] == "existing" for pair in _accuracy_castle_pairs) == 4 and
       sum(pair["raw_mode"] == "derived-stripped" for pair in _accuracy_castle_pairs) == 4 and
       all(pair["raw_path"] is None or "_stripped" in str(pair["raw_path"])
           for pair in _accuracy_castle_pairs),
       [(pair["marked_path"].name, pair["raw_mode"], str(pair["raw_path"]))
        for pair in _accuracy_castle_pairs])
except _FixtureNotPresent as _e:
    print(f"  [SKIP] {_e} — fixture not present")

print("pricing")
r, _ = slab_rate({"depth_mm":190,"conc_rate":128,"mesh":"A252","layers":1,"steel_rate_t":850,"margin":0.11})
ck("yard rate 44.89", r == 44.89)
tot, rows = price_project([{"name":"Yard","area_m2":26080,"depth_mm":190,"conc_rate":128,"mesh":"A252","layers":1,"steel_rate_t":850,"margin":0.11}])
ck("yard slab line GBP1,170,731.20", rows[0][5] == 1170731.20)
ck("unknown mesh handled", slab_rate({"depth_mm":150,"conc_rate":128,"mesh":"A999","layers":1,"steel_rate_t":850,"margin":0.11})[0] is None)

print("guards (95,463 m² incident)")
from scale import scale_consensus
from sanity import plausible
ck("mixed-scale dimensions flagged (no auto-pick)", scale_consensus([(257.2,710),(166,420),(50,75),(35,80)])[0] is None)
ck("consistent dimensions accepted", abs(scale_consensus([(100,1000),(50,500)])[0] - 0.1) < 1e-6)
ck("impossible area blocked", len(plausible(95463, site_m2=34329)) >= 1)
ck("correct area passes", plausible(26080, site_m2=34329) == [])

print("Fortel scale verification (from the call)")
from scale import calibrate_verified, verify_against_feature, title_block_k, ratio_from_k
geom = 2235703  # real yard polygon area in pt²
k_v, _ = calibrate_verified(title_denominator=500, bay_width_pt=2.5/0.108)  # verify vs 2.5 m bay
ck("parking-bay verify flips wrong title scale to truth", abs(geom*k_v*k_v - 26080) < 50)
ck("title-only scale flagged as a lie", len(verify_against_feature(title_block_k(500), 2.5/0.108, 2.5)) >= 1)
ck("the same internal scale k is presented to assessors as the conventional 1:250 ratio",
   ratio_from_k(title_block_k(250)) == 250)

print("drawing selection (from the call)")
from router import drawing_priority
ck("construction/kerbing drawing beats site plan",
   drawing_priority("RIBVE-XX-DR-CE-0750 construction kerbing") > drawing_priority("Proposed Site Plan"))
ck("engineer external-works beats architect hard-landscaping",
   drawing_priority("External Construction Thickness Layout", source="engineer")
   > drawing_priority("Unit 1 Hard Landscaping", source="architect"))

print("unmarked pipeline (legend-anchored colour segmentation)")
try:
    import numpy as _np
    from takeoff_unmarked import (segment_hatch, find_concrete_swatch_rgb,
                                   _choose_vector_swatch, _choose_raster_swatch,
                                   _is_plausible_surface_tint,
                                   _swatch_body_agrees)
    _im = _np.full((200, 300, 3), 255, _np.uint8); _im[50:150, 60:210] = (216, 216, 216)  # 100x150 grey
    _comp = segment_hatch(_im, (216, 216, 216))
    ck("segment grey hatch ~15,000 px", _comp is not None and abs(int(_comp.sum()) - 15000) < 900)
    ck("segment ignores white background", int(_comp.sum()) < 200 * 300 * 0.4)
    _px = int(_comp.sum()); _area = _px * (1 / 2.0) ** 2 * 0.1 * 0.1   # S=2 (1px=0.5pt), k=0.1 m/pt
    ck("unmarked area math (px->m2)", abs(_area - _px * 0.0025) < 1e-6)
    ck("white-segmentation blowup blocked by plausibility", len(plausible(279905)) >= 1)
    _w = segment_hatch(_im, (255, 0, 0))   # colour not present
    ck("absent hatch colour -> no region", _w is None or int(_w.sum()) == 0)

    print("client legend-swatch regressions (rotation, bright tints, pattern overlays)")
    import fitz as _fitz_swatch
    _swatch_pdf = "/tmp/_ci_rotated_bright_swatch.pdf"
    _sd = _fitz_swatch.open()
    _sp = _sd.new_page(width=400, height=600)
    # Text and chip are authored in unrotated PDF coordinates; get_pixmap() applies /Rotate.
    # The regression is the same coordinate-space mismatch as the real 270-degree sheet.
    _sp.insert_text((100, 450), "Concrete Service Yard", fontsize=12, rotate=90)
    _sp.draw_rect(_fitz_swatch.Rect(92, 300, 108, 340),
                  color=None, fill=(1.0, 180 / 255, 1.0))
    _sp.set_rotation(270)
    _sd.save(_swatch_pdf)
    _sd.close()
    with _fitz_swatch.open(_swatch_pdf) as _rd:
        _rp = _rd[0]
        _rpx = _rp.get_pixmap(matrix=_fitz_swatch.Matrix(2, 2), alpha=False)
        _rim = _np.frombuffer(_rpx.samples, _np.uint8).reshape(
            _rpx.height, _rpx.width, _rpx.n)[..., :3]
        _raw_bbox = _rp.search_for("Concrete Service Yard")[0]
        _rendered_bbox = _raw_bbox * _rp.rotation_matrix
        ck("rotated label bbox transformed into rendered-page coordinates",
           _rendered_bbox != _raw_bbox and _rendered_bbox.y1 <= _rp.rect.height)
    _rrgb, _rlabel = find_concrete_swatch_rgb(_swatch_pdf, im=_rim, S=2.0)
    ck("rotation-corrected raster swatch reads saturated bright tint",
       _rrgb == (255, 180, 255) and "concrete service yard" in _rlabel,
       f"rgb={_rrgb} label={_rlabel}")
    ck("surface-tint plausibility accepts bright hue and 239 tint, rejects ink/paper",
       _is_plausible_surface_tint((255, 180, 255)) and
       _is_plausible_surface_tint((239, 239, 240)) and
       not _is_plausible_surface_tint((0, 0, 0)) and
       not _is_plausible_surface_tint((255, 255, 255)))

    _vr = _fitz_swatch.Rect(10, 10, 40, 24)
    _overlay_candidates = [
        (2.0, (0, 0, 0), _fitz_swatch.Rect(_vr)),
        (2.1, (239, 239, 240), _fitz_swatch.Rect(_vr)),
    ]
    ck("co-located black pattern overlay yields to its non-black base tint",
       _choose_vector_swatch(_overlay_candidates) == (239, 239, 240))
    _two_row_patch = _np.full((20, 120, 3), 255, _np.uint8)
    _two_row_patch[:, 85:105] = (239, 239, 240)  # target chip nearest its label
    _two_row_patch[:, 0:50] = (156, 192, 207)    # larger neighbouring legend chip
    ck("wide raster window chooses nearest legend chip, not a larger neighbouring row",
       _choose_raster_swatch(_two_row_patch) == (239, 239, 240))
    ck("legend/body agreement is a strict gate, not an informational warning",
       _swatch_body_agrees((239, 239, 240), (238, 238, 240)) and
       not _swatch_body_agrees((239, 239, 240), (214, 214, 214)))
    _rgb_lock = _np.full((80, 120, 3), 255, _np.uint8)
    _rgb_lock[10:70, 10:55] = (239, 239, 240)
    _rgb_lock[10:70, 65:110] = (239, 220, 240)  # same R/B; G is outside the lock
    _strict_rgb = segment_hatch(_rgb_lock, (239, 239, 240), tol=5,
                                exclude_border=False, full_rgb=True)
    ck("swatch lock constrains every RGB channel, including near-grey tints",
       _strict_rgb is not None and int(_strict_rgb.sum()) < 60 * 55)

    print("team feedback fixes (DEMO4)")
    from takeoff_unmarked import drawing_style
    # (a) drawing-style guard: solid fill = colour-coded; thin lines = line/hatch (don't guess on engineer sheets)
    _solid = _np.full((300, 300, 3), 255, _np.uint8); _solid[40:260, 40:260] = (120, 170, 90)
    ck("colour-coded sheet detected", drawing_style(_solid)[0] == "colour-coded")
    _lines = _np.full((300, 300, 3), 255, _np.uint8)
    for _i in range(0, 300, 12):
        _lines[:, _i] = (80, 80, 80)
    ck("line/hatch sheet detected", drawing_style(_lines)[0] == "line/hatch")
    # (b) dock-bay/void fix: a large interior void is kept as a DEDUCTION, not filled (team: D77 dock bays)
    _v = _np.full((400, 400, 3), 255, _np.uint8); _v[40:360, 40:360] = (214, 214, 214); _v[150:250, 150:250] = 255
    _kept = segment_hatch(_v, (214, 214, 214), k=0.05, S=2.0, max_void_m2=1.0)   # void=6.25 m² > 1 -> kept out
    _fill = segment_hatch(_v, (214, 214, 214), k=0.05, S=2.0, max_void_m2=999)   # huge thresh -> filled
    ck("large interior void kept as deduction", int(_kept.sum()) < int(_fill.sum()))
    ck("void filled only when below threshold", int(_fill.sum()) - int(_kept.sum()) > 8000)

    print("polygon contour (fan/spoke regression)")
    import math as _math
    from takeoff_unmarked import _hatch_contour
    # Non-convex yard: rectangle with a deep notch cut from the top edge (loading dock).
    # The old angular-sort-from-centroid tracer produced spokes radiating across the slab
    # (lines from a corner) because rays from the centroid cross the boundary >2 times.
    # cv2.findContours walks the perimeter in order, so the outline must be clean.
    _cmp = _np.zeros((700, 1000), bool)
    _cmp[120:560, 160:840] = True
    _cmp[120:340, 480:680] = False          # deep top-edge notch -> strongly non-star-shaped
    _poly = _hatch_contour(_cmp, S=2.0, max_pts=80)
    ck("hatch contour returned", _poly is not None and len(_poly) >= 4)
    _xs = [q[0] for q in _poly]; _ys = [q[1] for q in _poly]
    # Bounding box must match the slab extent in PDF pt (mask px / S): x 80..420, y 60..280.
    ck("contour bbox matches slab extent",
       abs(min(_xs)-80) < 3 and abs(max(_xs)-420) < 3 and
       abs(min(_ys)-60) < 3 and abs(max(_ys)-280) < 3)
    # Perimeter sanity: true outer+notch boundary ~1440 pt. The fan bug inflated this to
    # ~2450+ pt (spokes shooting across the shape). Require it within ~25% of truth.
    _seg = [_math.hypot(_xs[i]-_xs[i-1], _ys[i]-_ys[i-1]) for i in range(1, len(_xs))]
    _seg.append(_math.hypot(_xs[0]-_xs[-1], _ys[0]-_ys[-1]))
    ck("contour perimeter not inflated by spokes", 1100 < sum(_seg) < 1800)
    # Spoke signature: count radius oscillations (near->far->near) about the centroid.
    # A clean traced outline has very few; the fan pattern had ~23/78.
    _cx = sum(_xs)/len(_xs); _cy = sum(_ys)/len(_ys)
    _rad = [_math.hypot(x-_cx, y-_cy) for x, y in zip(_xs, _ys)]
    _osc = sum(1 for i in range(1, len(_rad)-1)
               if (_rad[i] > _rad[i-1]) != (_rad[i+1] > _rad[i]))
    ck("no fan/spoke oscillation", _osc <= 6)
    # Plain convex rectangle -> exactly its 4 corners; degenerate masks -> None.
    _rect = _np.zeros((600, 800), bool); _rect[150:450, 200:650] = True
    ck("rectangle -> 4 corners", len(_hatch_contour(_rect, S=2.0)) == 4)
    ck("empty mask -> None", _hatch_contour(_np.zeros((40, 40), bool), S=2.0) is None)
    _tiny = _np.zeros((40, 40), bool); _tiny[10:12, 10:12] = True
    ck("sub-pixel blob -> None", _hatch_contour(_tiny, S=2.0) is None)
except ImportError as _e:
    print(f"  [SKIP] takeoff_unmarked tests — missing dependency: {_e}")

print("scale_for verification logic (scale bar vs title block)")
try:
    from takeoff_unmarked import (scale_for as _scale_for, SCALE_BAR_AGREE_TOL as _TOL,
                                  boundary_precision_risk as _boundary_precision_risk)
    import scale as _SC
    from sanity import (measurement_state as _precision_measurement_state,
                        MEASURED_UNVERIFIED as _PRECISION_UNVERIFIED)
    # PT_PER_M = 0.0254/72; k for 1:500 = 500 * PT_PER_M ≈ 0.176389 m/pt
    _PT_PER_M = 0.0254 / 72
    _k500 = 500 * _PT_PER_M   # ≈ 0.176389 m/pt

    # --- CASE 1: scale bar AGREES with title block (bar within ±3%) -> verified=True ---
    # Bar: 88 m / 500 pt = 0.176 m/pt; diff vs k500 ≈ 0.22% < 3%
    _c1 = canvas.Canvas("/tmp/_sf_agree.pdf", pagesize=(1400, 2200))
    _c1.drawString(100, 2100, "Drawing Scale 1:500")   # title-block text
    _c1.drawString(200, 120, "0          88 m")        # scale-bar label (88 m over 500 pt bar)
    _c1.line(100, 110, 600, 110)                        # 500 pt horizontal bar
    _c1.save()
    _k1, _v1, _n1, _src1 = _scale_for("/tmp/_sf_agree.pdf")
    ck("bar agrees with title -> verified=True",  _v1 is True, f"k={_k1:.5f} note={_n1[:60]}")
    ck("agree: bar in scale_sources",             "scale_bar" in _src1)
    ck("agree: title_block in scale_sources",     "title_block" in _src1)
    ck("agree: returned k close to bar",          _k1 is not None and abs(_k1 - 88/500) < 1e-6)

    # --- CASE 2: scale bar DISAGREES with title block (>3%) -> verified=False. Bar: 150 m / 500 pt
    # = 0.30 m/pt (implies ~1:850, an individually PLAUSIBLE drawing ratio) vs title k500 ≈ 70% off.
    # Both sources are plausible on their own -> this is the MIXED/DISAGREE branch (CLAUDE.md
    # invariant 3: disagreement -> refuse, don't auto-pick). Neither is silently adopted; the
    # title-block k is used for display and the assessor must set the scale explicitly.
    _c2 = canvas.Canvas("/tmp/_sf_disagree.pdf", pagesize=(1400, 2200))
    _c2.drawString(100, 2100, "Drawing Scale 1:500")
    _c2.drawString(200, 120, "0         150 m")
    _c2.line(100, 110, 600, 110)
    _c2.save()
    _k2, _v2, _n2, _src2 = _scale_for("/tmp/_sf_disagree.pdf")
    ck("bar disagrees with title -> verified=False", _v2 is False, f"k={_k2:.5f} note={_n2[:60]}")
    ck("disagree: note flags MIXED/DISAGREE",        "MIXED/DISAGREE" in _n2)
    ck("disagree: NOT auto-picked to bar k (title k used instead)",
       _k2 is not None and abs(_k2 - _k500) < 1e-6)

    # --- CASE 3: no scale bar, title block only -> verified=False ---
    _c3 = canvas.Canvas("/tmp/_sf_titleonly.pdf", pagesize=(1400, 2200))
    _c3.drawString(100, 2100, "Drawing Scale 1:500")   # title-block only, no bar line or label
    _c3.save()
    _k3, _v3, _n3, _src3 = _scale_for("/tmp/_sf_titleonly.pdf")
    ck("title-only -> verified=False",            _v3 is False, f"note={_n3[:60]}")
    ck("title-only: title_block in scale_sources", "title_block" in _src3)
    ck("title-only: no scale_bar in scale_sources", "scale_bar" not in _src3)
    ck("title-only: k close to k500",            _k3 is not None and abs(_k3 - _k500) < 1e-5)

    # --- CASE 4: bar DISAGREES with title AND the bar-implied ratio is IMPLAUSIBLE (false
    # scale-bar anchor, e.g. an unrelated dimension callout mis-paired to a nearby short line
    # fragment) -> reject the bar entirely, fall back to title-block k, still UNVERIFIED.
    # Reproduces the real corpus incident: Proposed_Gatehouse's "7016 m / 34 pt" bar candidate
    # implies k=205.868 m/pt (~1:583,563) which is nowhere near a real drawing scale.
    # Bar: 7016 m / 34 pt = 206.35 m/pt -> implied ~1:584,000, way outside 1:20-1:5000.
    _c4 = canvas.Canvas("/tmp/_sf_implausible.pdf", pagesize=(1400, 2200))
    _c4.drawString(100, 2100, "Drawing Scale 1:1250")
    _c4.drawString(200, 120, "0          7016 m")
    _c4.line(100, 110, 134, 110)                        # 34 pt bar
    _c4.save()
    _k4, _v4, _n4, _src4 = _scale_for("/tmp/_sf_implausible.pdf")
    _k1250 = 1250 * _PT_PER_M
    ck("implausible bar -> verified=False",        _v4 is False, f"k={_k4:.5f} note={_n4[:70]}")
    ck("implausible bar -> note says rejected",     "rejected as implausible" in _n4)
    ck("implausible bar -> falls back to title k",  _k4 is not None and abs(_k4 - _k1250) < 1e-5)
    ck("implausible bar -> sources still recorded", "scale_bar" in _src4 and "title_block" in _src4)

    # --- CASE 5: bar DISAGREES with title but BOTH are individually plausible drawing ratios
    # (e.g. a genuine 1:2500 site-location viewport vs a stale 1:1500 title block) -> MIXED/
    # DISAGREE. Must NOT auto-pick either side; verified stays False; title k shown for display.
    # Reproduces the real corpus incident: Site_Location_Plan's "100 m / 113 pt" bar (k=0.882,
    # ~1:2500 — a perfectly plausible ratio) disagreeing with the sheet's stated title 1:1500.
    _c5 = canvas.Canvas("/tmp/_sf_mixed.pdf", pagesize=(1400, 2200))
    _c5.drawString(100, 2100, "Drawing Scale 1:1500")
    _c5.drawString(200, 120, "0          100 m")
    _c5.line(100, 110, 213, 110)                        # 113 pt bar -> k=0.885 (~1:2504, plausible)
    _c5.save()
    _k5, _v5, _n5, _src5 = _scale_for("/tmp/_sf_mixed.pdf")
    _k1500 = 1500 * _PT_PER_M
    ck("mixed/disagree -> verified=False",          _v5 is False, f"k={_k5:.5f} note={_n5[:70]}")
    ck("mixed/disagree -> note says MIXED/DISAGREE", "MIXED/DISAGREE" in _n5)
    ck("mixed/disagree -> returns title k (no auto-pick of bar)",
       _k5 is not None and abs(_k5 - _k1500) < 1e-5)
    ck("mixed/disagree -> sources still recorded",  "scale_bar" in _src5 and "title_block" in _src5)

    # --- CASE 6: multiple viewport scales.  A bar agreeing with one printed denominator
    # does not prove that denominator belongs to the slab viewport being segmented.  Inderjit
    # clarified on 7 Aug that each layout can carry its own scale; without a spatial
    # bar/title/region association the sheet must remain assessor-gated.
    _c6 = canvas.Canvas("/tmp/_sf_multiple_viewports.pdf", pagesize=(1400, 2200))
    _c6.drawString(100, 2100, "GROUND FLOOR LAYOUT  Scale 1:500")
    _c6.drawString(800, 2100, "SECTION DETAIL  Scale 1:100")
    _c6.drawString(200, 120, "0          88 m")
    _c6.line(100, 110, 600, 110)
    _c6.save()
    _k6, _v6, _n6, _src6 = _scale_for("/tmp/_sf_multiple_viewports.pdf")
    ck("multiple viewport scales cannot be globally VERIFIED without spatial association",
       _v6 is False and "MULTIPLE VIEWPORT SCALES" in _n6 and
       _src6.get("title_block_candidates") == [500, 100],
       {"verified":_v6, "note":_n6, "sources":_src6})

    _c7 = canvas.Canvas("/tmp/_sf_a0_nts.pdf", pagesize=(1400, 2200))
    _c7.drawString(100, 2100, "SHEET SIZE A0   SCALE AS INDICATED   NTS")
    _c7.save()
    _k7, _v7, _n7, _src7 = _scale_for("/tmp/_sf_a0_nts.pdf")
    ck("A0 is sheet size and NTS never becomes a numeric scale",
       _k7 is None and _v7 is False and _src7 == {} and "no scale" in _n7,
       {"k":_k7, "verified":_v7, "note":_n7, "sources":_src7})

    _risk_1500 = _boundary_precision_risk({"title_block":{"denom":1500}})
    _risk_2000 = _boundary_precision_risk({"title_block":{"denom":2000}})
    ck("1:1500 and 1:2000 sheets carry a visible boundary-click precision risk",
       "1:1500" in _risk_1500 and "no numeric adjustment" in _risk_1500 and
       "1:2000" in _risk_2000)
    ck("ordinary 1:1000 sheet does not gain the large-denominator precision flag",
       _boundary_precision_risk({"title_block":{"denom":1000}}) is None)
    ck("large-denominator risk caps an otherwise verified number at assessor-gated state",
       _precision_measurement_state(
           1000, scale_verified=True,
           confidence="low" if _risk_1500 else None)[0] == _PRECISION_UNVERIFIED)

except ImportError as _e:
    print(f"  [SKIP] scale_for tests — missing dependency: {_e}")

print("defaults (Fortel build-up assumptions)")
from defaults import spec_with_defaults, assumption_note, flag_assumed
_s, _assumed = spec_with_defaults()
ck("default spec depth 190mm", _s["depth_mm"] == 190)
ck("default spec mesh A252",   _s["mesh"] == "A252")
ck("default assumed=True",     _assumed is True)
ck("assumption note contains 190mm", "190 mm" in assumption_note(_s))
ck("flags empty when fully specified",
   flag_assumed({"depth_mm":200,"mesh":"A393","layers":1,"conc_mix":"C32/40"}, False) == [])
ck("flags non-empty when assumed", len(flag_assumed(_s, True)) >= 1)
_s2, _a2 = spec_with_defaults({"depth_mm": 175, "mesh": "A193", "layers": 1, "conc_mix": "C32/40"})
ck("full engineer spec -> assumed=False", _a2 is False)
ck("engineer depth 175 overrides default 190", _s2["depth_mm"] == 175)

print("spec extractor (construction-detail PDF text parsing)")
from spec_extractor import describe_spec, extract_spec, extract_spec_from_text
_e1 = extract_spec_from_text("175 mm thick with A193 mesh, C32/40 concrete")
ck("depth 175",  _e1.get("depth_mm") == 175)
ck("mesh A193",  _e1.get("mesh") == "A193")
ck("mix C32/40", _e1.get("conc_mix") == "C32/40")
_e2 = extract_spec_from_text("200mm slab with two layers of A393 reinforcement C35/45")
ck("depth 200",  _e2.get("depth_mm") == 200)
ck("2 layers A393", _e2.get("mesh") == "A393" and _e2.get("layers") == 2)
_e3 = extract_spec_from_text("No specification provided")
ck("empty text -> empty spec", not any(k in _e3 for k in ("depth_mm","mesh","conc_mix")))
_e4 = extract_spec_from_text("A393 x2 250 mm C40/50")
ck("x2 notation -> 2 layers", _e4.get("layers") == 2)
_e5 = extract_spec_from_text("A252 mesh")
ck("mesh without a layer count keeps layers unknown", "layers" not in _e5, _e5)
ck("mesh-only human summary says layers not provided",
   describe_spec(_e5) == "A252 mesh (layers not provided)", describe_spec(_e5))

# Inderjit's first live-use review exposed a 150 mm false result while the sheet stated
# 190 mm.  Scale-bar print dimensions and neighbouring build-ups are not slab evidence.
_e6 = extract_spec_from_text(
    "CHECK: scale bar must measure 150 mm when printed.\n"
    "EXTERNAL SERVICE YARD — 190 mm thick reinforced concrete slab with A252 fabric.",
    source_name="Live Yard Joint Layout.pdf", page_number=1,
    context="external service yard",
)
ck("explicit 190mm Yard slab outranks a 150mm scale-bar print dimension",
   _e6.get("depth_mm") == 190, _e6)
ck("extracted thickness carries auditable file/page/text evidence",
   (_e6.get("_evidence") or {}).get("depth_mm", {}).get("file") ==
       "Live Yard Joint Layout.pdf" and
   (_e6.get("_evidence") or {}).get("depth_mm", {}).get("page") == 1 and
   "190 mm thick" in (_e6.get("_evidence") or {}).get("depth_mm", {}).get("text", ""),
   _e6.get("_evidence"))
_e6_detail = extract_spec_from_text(
    "CONCRETE YARD — Concrete Slab: 205mm thickness of PAV2. "
    "Min 300mm thickness of ground improvement. Sub Base 150mm Type 1. "
    "H12 U-bars at 225mm crs.",
    source_name="External Works Details.pdf", page_number=1,
    context="external concrete yard",
)
ck("ground-improvement, sub-base and reinforcement-spacing dimensions are not slab depths",
   _e6_detail.get("depth_mm") == 205 and
   not (_e6_detail.get("_conflicts") or {}).get("depth_mm"), _e6_detail)

_e7 = extract_spec_from_text(
    "PROPOSED EXTERNAL YARD SLAB JOINT LAYOUT. "
    "Allow for joints in external slab at maximum 4.9 by 6.5 metre centres.",
    source_name="Yard Joint Layout.pdf", page_number=2,
    context="external yard",
)
ck("joint-layout spacing is extracted instead of reported as no details",
   _e7.get("bay_sizes") == "4.9 m x 6.5 m centres" and
   (_e7.get("_evidence") or {}).get("bay_sizes", {}).get("page") == 2,
   _e7)

# Portal uploads from every project share one persistent drawings directory.  A details PDF
# from project A must never supply project B's spec merely because it sorts first.
_spec_scope_dir = Path("/tmp/_fortel_spec_scope")
shutil.rmtree(_spec_scope_dir, ignore_errors=True)
_spec_scope_dir.mkdir()
for _spec_name, _spec_text in (
    ("AAA_Other_External_Construction_Details.pdf", "150 mm thick concrete slab A193"),
    ("BBB_Target_Yard.pdf", "190 mm thick external service yard slab A252"),
):
    _spec_canvas = canvas.Canvas(str(_spec_scope_dir / _spec_name), pagesize=(500, 300))
    _spec_canvas.drawString(40, 240, _spec_text)
    _spec_canvas.save()
from takeoff_pipeline import find_engineer_spec as _find_engineer_spec
_scoped_spec = _find_engineer_spec(
    str(_spec_scope_dir / "BBB_Target_Yard.pdf"), project_ref="BBB")
ck("project-scoped lookup cannot import another project's 150mm spec",
   _scoped_spec and _scoped_spec.get("depth_mm") == 190 and
   "AAA_Other" not in str(_scoped_spec), _scoped_spec)
_same_project_detail = _spec_scope_dir / "BBB_Yard_Build-Up.pdf"
_spec_canvas = canvas.Canvas(str(_same_project_detail), pagesize=(500, 300))
_spec_canvas.drawString(40, 240, "CONCRETE YARD — Concrete slab 205mm thick with A393 mesh")
_spec_canvas.save()
_conflicted_spec = _find_engineer_spec(
    str(_spec_scope_dir / "BBB_Target_Yard.pdf"), project_ref="BBB")
ck("competing same-project slab callouts remain a visible conflict, never an auto-confirmed value",
   _conflicted_spec and "depth_mm" not in _conflicted_spec and
   {record["value"] for record in _conflicted_spec["_conflicts"]["depth_mm"]} ==
       {190, 205}, _conflicted_spec)

# Portal project membership, unlike same-directory discovery, is explicit. A later detail
# upload may therefore live in another persisted folder and must still supply cited evidence.
_cross_project_dir = _spec_scope_dir / "project_registry"
_cross_layout_dir = _cross_project_dir / "layout_upload"
_cross_detail_dir = _cross_project_dir / "later_detail_upload"
_cross_layout_dir.mkdir(parents=True)
_cross_detail_dir.mkdir(parents=True)
_cross_target = _cross_layout_dir / "CROSS_Yard_Layout.pdf"
_cross_detail = _cross_detail_dir / "CROSS_External_Works_Details.pdf"
for _cross_path, _cross_text in (
    (_cross_target, "EXTERNAL SERVICE YARD GENERAL ARRANGEMENT"),
    (_cross_detail,
     "EXTERNAL SERVICE YARD — 215 mm thick concrete slab with A393 mesh, C35/45 concrete"),
):
    _spec_canvas = canvas.Canvas(str(_cross_path), pagesize=(500, 300))
    _spec_canvas.drawString(40, 240, _cross_text)
    _spec_canvas.save()
_cross_without_registry = _find_engineer_spec(
    str(_cross_target), project_ref="CROSS")
_cross_with_registry = _find_engineer_spec(
    str(_cross_target), project_ref="CROSS",
    project_files=[str(_cross_target), str(_cross_detail)])
ck("project-wide spec lookup crosses upload folders only through the explicit job registry",
   _cross_without_registry is None and
   _cross_with_registry.get("depth_mm") == 215 and
   _cross_with_registry.get("mesh") == "A393" and
   _cross_with_registry.get("conc_mix") == "C35/45",
   {"without_registry":_cross_without_registry,
    "with_registry":_cross_with_registry})
ck("project-wide extracted fields retain exact drawing and page citations",
   all((_cross_with_registry.get("_evidence") or {}).get(field, {}).get("file") ==
       _cross_detail.name for field in ("depth_mm","mesh","conc_mix")) and
   all((_cross_with_registry.get("_evidence") or {}).get(field, {}).get("page") == 1
       for field in ("depth_mm","mesh","conc_mix")),
   _cross_with_registry.get("_evidence"))

_joint_fixture = Path(
    "drawings/inderjit_markups_31jul/2165 Tanro- Voltage Business Park/"
    "Tanro Voltage Business Park/Bid_Drawings/A)-Tender-Stage/Current/"
    "Engineer---Baynham-Meikle/"
    "13897-114-Proposed-External-Yard-Slab-Joint-Layout-Rev.T1.pdf")
try:
    _require_fixture(_joint_fixture, "Tanro joint-layout fixture not present")
    _joint_fixture_spec = extract_spec(str(_joint_fixture), context="external yard")
    ck("real joint-layout scale-bar 100mm is never confirmed as slab thickness",
       _joint_fixture_spec.get("depth_mm") != 100 and
       bool(_joint_fixture_spec.get("_joint_layout")), _joint_fixture_spec)
except _FixtureNotPresent as _e:
    print(f"  [SKIP] real joint-layout spec guard — {_e} — fixture not present")

_multi_build_fixture = Path(
    "drawings/aryan_31jul/Bidding_Documents___27_0/Tender_drawings/Tender-Stage/"
    "Current/Engineer---PRP/64426-112-External-Works-Details-Rev.T2.pdf")
try:
    _require_fixture(_multi_build_fixture, "PRP multi-build-up detail fixture not present")
    _multi_build_spec = extract_spec(
        str(_multi_build_fixture), context="external service yard concrete yard")
    _depth_conflicts = {
        record.get("value")
        for record in (_multi_build_spec.get("_conflicts") or {}).get("depth_mm", [])
    }
    ck("real multi-build-up detail refuses scope ambiguity without retaining non-slab dimensions",
       "depth_mm" not in _multi_build_spec and "mesh" not in _multi_build_spec and
       {150, 190, 205}.issubset(_depth_conflicts) and
       not ({155, 225, 300} & _depth_conflicts), _multi_build_spec)
except _FixtureNotPresent as _e:
    print(f"  [SKIP] real multi-build-up spec guard — {_e} — fixture not present")

print("Fortel Brief_Spec schema + field provenance")
from slab_spec import (COMMON_FIELDS as _SPEC_COMMON_FIELDS, SLAB_SPEC_SCHEMA as _SPEC_SCHEMA,
                       brief_spec_signature as _brief_signature,
                       build_brief_spec as _build_brief_spec,
                       empty_brief_spec as _empty_brief_spec)
_expected_spec_fields = {
    "external_yard": ("depth_mm", "conc_mix", "mesh", "layers", "bay_sizes", "joint_details"),
    "dock": ("depth_mm", "conc_mix", "mesh", "layers", "bay_sizes", "joint_details"),
    "ground_floor": ("depth_mm", "conc_mix", "mesh", "layers", "joint_details"),
    "upper_floor": ("depth_mm", "conc_mix", "mesh", "layers"),
}
ck("Brief_Spec schema has the exact four slab types and applicable fields",
   set(_SPEC_SCHEMA) == set(_expected_spec_fields) and
   all(tuple(_SPEC_SCHEMA[key]["fields"]) == fields
       for key, fields in _expected_spec_fields.items()))
_empty_yard_spec = _empty_brief_spec("external_yard")
ck("blank Brief_Spec has no invented values and every field is provisional",
   all(field["value"] is None and field["provisional"]
       for field in _empty_yard_spec["fields"].values()), _empty_yard_spec)
_effective_spec = {"depth_mm": 190, "conc_mix": "C32/40", "mesh": "A252", "layers": 1}
_assumed_brief = _build_brief_spec("external_yard", effective_spec=_effective_spec)
_confirmed_brief = _build_brief_spec(
    "external_yard", effective_spec=_effective_spec, confirmed=_effective_spec,
    source="engineer_drawing")
ck("effective costing values remain field-by-field assumed until confirmed",
   all(_assumed_brief["fields"][key]["provisional"] for key in _SPEC_COMMON_FIELDS) and
   all(not _confirmed_brief["fields"][key]["provisional"] for key in _SPEC_COMMON_FIELDS))
ck("assumed and confirmed copies of the same effective spec have different aggregation identity",
   _brief_signature(_assumed_brief) != _brief_signature(_confirmed_brief))
_mesh_only_brief = _build_brief_spec(
    "external_yard", confirmed={"mesh": "A252"}, source="engineer_drawing",
    evidence={"mesh": {"file": "Yard Detail D-112.pdf", "page": 3,
                       "text": "A252 fabric reinforcement"}},
)
ck("a fabric callout confirms mesh but never silently confirms one reinforcement layer",
   not _mesh_only_brief["fields"]["mesh"]["provisional"] and
   _mesh_only_brief["fields"]["layers"]["provisional"] and
   _mesh_only_brief["fields"]["layers"]["value"] is None,
   _mesh_only_brief)

print("quotation generator")
from quotation import (generate_quotation, quotation_text, quotation_html, quotation_json,
                       quotation_xlsx, SECTION_ORDER, PROVISIONAL_LABEL, PROVISIONAL_COL,
                       FORTEL_MH_ROW, FORTEL_CHANNEL_ROW, FORTEL_TRANSITION_ROW)
from geometry import polygon_perimeter_lm
from openpyxl import load_workbook as _load_workbook
from io import BytesIO as _BytesIO
import copy as _copy
_demo_result = {
    "file": "D77.pdf", "type": "UNMARKED vector", "confidence": "medium",
    "source_discipline": "architect",
    "costing": {
        "area_m2": 3172, "rate": 44.89, "total_gbp": 142391.08, "assumed": True,
        "spec": {"depth_mm": 190, "mesh": "A252", "conc_mix": "C32/40", "layers": 1, "conc_rate": 128},
        "breakdown": {"concrete": 25.05, "steel": 4.30, "dpm": 0.46,
                      "curing": 0.23, "labour": 10.00, "trim": 0.40, "nett": 40.44, "margin%": 11},
    },
    "flags": ["BUILD-UP ASSUMED: 190mm / A252 / C32/40"],
}
_q = generate_quotation(_demo_result, project="Test", client="Client", ref="TST-001")
ck("quotation total > 0",     _q["total_gbp"] > 0)
ck("quotation assumed=True",  _q["assumed"] is True)
ck("has declaration",         len(_q["declarations"]) >= 1)
ck("slab line item present",  any("slab" in li["description"].lower() for li in _q["line_items"]))
ck("text contains total",     "TOTAL NETT" in quotation_text(_q))
ck("html contains total",     "TOTAL NETT" in quotation_html(_q) or "Total" in quotation_html(_q))
ck("html is valid-ish",       quotation_html(_q).startswith("<!DOCTYPE html>"))

# Client-call quotation requirements: one editable tab, canonical section order, aggregate
# identical unit specs, retain different specs, mark assumptions provisional, and expose an
# informational perimeter without pricing it. Synthetic geometry only — no drawings fixture.
def _quotation_unit(filename, section, area, *, mesh="A252", assumed=False):
    unit = _copy.deepcopy(_demo_result)
    unit.update({"file": filename, "quotation_section": section, "area_m2": area,
                 "source_discipline": "engineer", "flags": []})
    unit["costing"].update({"area_m2": area, "assumed": assumed})
    unit["costing"]["spec"] = dict(unit["costing"]["spec"], mesh=mesh)
    return unit

_quote_units = [
    _quotation_unit("Upper.pdf", "Upper floor slabs", 40),
    _quotation_unit("Yard-A.pdf", "External yard slabs", 100, assumed=True),
    _quotation_unit("Footpath.pdf", "Footpath slabs", 15),
    _quotation_unit("Dock.pdf", "Dock slabs", 30),
    _quotation_unit("Ground.pdf", "Ground floor slabs", 20),
    _quotation_unit("Yard-B.pdf", "External yard slabs", 150, assumed=True),
]
_q_multi = generate_quotation(
    _quote_units, project="Multi-unit", client="Fortel", ref="TST-MULTI",
    extras=[{"section": "Prelims", "description": "Existing prelim item",
             "qty": 1, "unit": "Item", "rate": _demo_result["costing"]["rate"]}],
)
_actual_sections = list(dict.fromkeys(li["section"] for li in _q_multi["line_items"]))
ck("quotation sections follow client order", _actual_sections == list(SECTION_ORDER), _actual_sections)
_yard_slabs = [li for li in _q_multi["line_items"]
               if li["section"] == "External yard slabs"
               and li.get("line_role") == "concrete_slab"]
ck("matching-spec units aggregate into one slab row", len(_yard_slabs) == 1 and
   _yard_slabs[0]["qty"] == 250, _yard_slabs)
_q_diff_spec = generate_quotation([
    _quotation_unit("Yard-A.pdf", "External yard slabs", 100),
    _quotation_unit("Yard-C.pdf", "External yard slabs", 50, mesh="A393"),
], ref="TST-DIFF-SPEC")
ck("different unit specs remain separate on one quotation",
   len([li for li in _q_diff_spec["line_items"]
        if li.get("line_role") == "concrete_slab"]) == 2)
_unit1_ground = _quotation_unit(
    "Unit-1 Ground Floor Core.pdf", "Ground floor slabs", 100)
_unit5_ground = _quotation_unit(
    "Unit-5 Ground Floor Core.pdf", "Ground floor slabs", 80)
for _unit, _depth in ((_unit1_ground, 193), (_unit5_ground, 150)):
    _unit["costing"]["spec"].update({"depth_mm":_depth, "mesh":"A252", "layers":1})
    _unit["brief_spec"] = _build_brief_spec(
        "ground_floor", effective_spec=_unit["costing"]["spec"],
        confirmed={"depth_mm":_depth, "mesh":"A252", "layers":1},
        source="engineer_drawing")
_q_ground_spec_split = generate_quotation(
    [_unit1_ground, _unit5_ground], ref="TST-GROUND-SPEC-SPLIT")
_ground_spec_rows = [
    item for item in _q_ground_spec_split["line_items"]
    if item["section"] == "Ground floor slabs"
    and item.get("line_role") == "concrete_slab"
]
ck("193mm A252 and 150mm A252 ground slabs in one case are never merged",
   len(_ground_spec_rows) == 2 and
   {item["qty"] for item in _ground_spec_rows} == {100.0, 80.0} and
   {spec["fields"]["depth_mm"]["value"]
    for spec in _q_ground_spec_split["specifications"]} == {193, 150},
   {"items":_ground_spec_rows, "specs":_q_ground_spec_split["specifications"]})
_confirmed_unit_a = _quotation_unit("Confirmed-A.pdf", "External yard slabs", 60)
_confirmed_unit_b = _quotation_unit("Confirmed-B.pdf", "External yard slabs", 40)
_confirmed_unit_a["brief_spec"] = _confirmed_brief
_confirmed_unit_b["brief_spec"] = _copy.deepcopy(_confirmed_brief)
_q_confirmed_aggregate = generate_quotation([_confirmed_unit_a, _confirmed_unit_b])
ck("equally confirmed Brief_Spec units aggregate",
   len([li for li in _q_confirmed_aggregate["line_items"]
        if li.get("line_role") == "concrete_slab"]) == 1)
_assumed_unit = _quotation_unit("Assumed.pdf", "External yard slabs", 40, assumed=True)
_assumed_unit["brief_spec"] = _assumed_brief
_q_provenance_split = generate_quotation([_confirmed_unit_a, _assumed_unit])
ck("assumed and confirmed equal-value Brief_Spec units remain separate",
   len([li for li in _q_provenance_split["line_items"]
        if li.get("line_role") == "concrete_slab"]) == 2)

_rect = [[0, 0], [40, 0], [40, 20], [0, 20]]
ck("perimeter_lm rectangle: 20m × 10m -> 60.0m",
   polygon_perimeter_lm(_rect, 0.5) == 60.0)
_perimeter_result = _quotation_unit("Dock-Perimeter.pdf", "Dock slabs", 200)
_perimeter_result.update({"polygon_pts": _rect, "scale_k": 0.5})
_q_perimeter = generate_quotation(_perimeter_result, ref="TST-PERIMETER")
ck("quotation surfaces perimeter as informational, unpriced quantity",
   _q_perimeter["perimeter_lm"] == 60.0 and
   any(m["description"] == "Slab perimeter" and m["qty"] == 60.0
       for m in _q_perimeter["measurements"]) and
   all("perimeter" not in li["description"].lower() for li in _q_perimeter["line_items"]))

# A real case can mix pipeline-measured siblings (top-level perimeter_lm) with an
# assessor-traced sibling whose perimeter lives on a zone.  Both sources intentionally
# share the same measurement key and must aggregate without either provenance path assuming
# it was the one that created the record.
_mixed_top_perimeter = _quotation_unit(
    "Yard Unit-1.pdf", "External yard slabs", 100)
_mixed_top_perimeter["perimeter_lm"] = 40.0
_mixed_zone_perimeter = _quotation_unit(
    "Yard Unit-2.pdf", "External yard slabs", 120)
_mixed_zone_perimeter["zones"] = [{
    "category": "external_yard", "area_m2": 120.0, "perimeter_lm": 50.0,
}]
_q_mixed_perimeter = generate_quotation(
    [_mixed_top_perimeter, _mixed_zone_perimeter],
    project="Mixed perimeter case", ref="TST-MIXED-PERIMETER")
_mixed_outputs = (
    quotation_text(_q_mixed_perimeter), quotation_html(_q_mixed_perimeter),
    quotation_json(_q_mixed_perimeter), quotation_xlsx(_q_mixed_perimeter),
)
_mixed_ws = _load_workbook(
    _BytesIO(_mixed_outputs[3]), data_only=False)["REV_01"]
_mixed_measurement = next(
    measurement for measurement in _q_mixed_perimeter["measurements"]
    if measurement["section"] == "External yard slabs"
    and measurement["description"] == "Slab perimeter")
ck("case quotation renders txt/html/json/xlsx with mixed perimeter provenance",
   _mixed_measurement["qty"] == 90.0 and
   len(_mixed_measurement["quantity_rows"]) == 2 and
   all(output for output in _mixed_outputs[:3]) and
   _mixed_ws.max_row > 7,
   _mixed_measurement)

_q_text = quotation_text(_q)
_q_html = quotation_html(_q)
_q_json = quotation_json(_q)
ck("assumed build-up is provisional in text/html/json",
   all(PROVISIONAL_LABEL in output for output in (_q_text, _q_html, _q_json)))
ck("unknown client spec fields are visible in text/html/json",
   all("ASSUMED / no details provided" in output for output in (_q_text, _q_html, _q_json)))
_xss_brief = _build_brief_spec(
    "external_yard", effective_spec=_effective_spec,
    confirmed=dict(_effective_spec, joint_details="<script>alert(1)</script>"),
)
_xss_unit = _quotation_unit("Safe.pdf", "External yard slabs", 10)
_xss_unit["brief_spec"] = _xss_brief
_xss_html = quotation_html(generate_quotation(_xss_unit))
ck("assessor-entered Brief_Spec text is HTML-escaped in served quotation",
   "<script>alert(1)</script>" not in _xss_html and
   "&lt;script&gt;alert(1)&lt;/script&gt;" in _xss_html)

_cited_brief = _build_brief_spec(
    "external_yard", effective_spec=_effective_spec,
    confirmed={"depth_mm": 190, "mesh": "A252"}, source="engineer_drawing",
    evidence={
        "depth_mm": {"file": "64426-112-External-Works-Details-Rev.T2.pdf", "page": 2,
                     "text": "minimum 190mm developer specification"},
        "mesh": {"file": "64426-112-External-Works-Details-Rev.T2.pdf", "page": 2,
                 "text": "A252 mesh fabric reinforcement"},
    },
)
_cited_unit = _quotation_unit("64426-111-Yard.pdf", "External yard slabs", 10)
_cited_unit["brief_spec"] = _cited_brief
_cited_quote = generate_quotation(_cited_unit, ref="TST-SPEC-CITATION")
_cited_outputs = (
    quotation_text(_cited_quote), quotation_html(_cited_quote),
    quotation_json(_cited_quote), quotation_xlsx(_cited_quote),
)
_cited_wb = _load_workbook(_BytesIO(_cited_outputs[3]), data_only=False)
_cited_xlsx_text = "\n".join(
    str(cell.value or "") for row in _cited_wb.active.iter_rows() for cell in row)
_source_citation = "64426-112-External-Works-Details-Rev.T2.pdf, page 2"
ck("every quotation format cites the drawing file and page for extracted specification fields",
   all(_source_citation in output for output in _cited_outputs[:3]) and
   _source_citation in _cited_xlsx_text,
   {"text": _cited_outputs[0], "xlsx": _cited_xlsx_text})

_xlsx_bytes = quotation_xlsx(_q_multi)
_xlsx_wb = _load_workbook(_BytesIO(_xlsx_bytes), data_only=False)
_xlsx_ws = _xlsx_wb["REV_01"]
ck("xlsx export reopens as exactly one editable quotation tab", _xlsx_wb.sheetnames == ["REV_01"])
ck("xlsx uses client BOQ column labels and order",
   tuple(_xlsx_ws.cell(7, col).value for col in range(1, 6)) ==
   ("DESCRIPTION", "QTY", "UNIT", "RATE", "VALUE"))
ck("xlsx header matches real BOQ project/client/date/rev/drawing layout",
   _xlsx_ws["A1"].value == "Project: Multi-unit" and
   _xlsx_ws["A2"].value == "Client: Fortel" and
   str(_xlsx_ws["A3"].value).startswith("Date: ") and
   _xlsx_ws["A4"].value == "Rev: TST-MULTI" and
   {str(cell_range) for cell_range in _xlsx_ws.merged_cells.ranges} == {"D4:E4"})
ck("xlsx drawing register is multiline in the real BOQ's A5 cell",
   _xlsx_ws["A5"].value.startswith("Drawing ref available at tender:\n") and
   all(name in _xlsx_ws["A5"].value for name in ("Yard-A.pdf", "Yard-B.pdf")))
_expected_xlsx_sections = (
    "External Yard Slabs- Provisional Cost (No Details)",
    "Footpath Slabs- Provisional Cost (No Details)",
    "Dock Slabs- Provisional Cost (No Details)",
    "Ground Floor Slabs- Provisional Cost (No Details)",
    "Upper Floors- Provisional Cost (No Details)",
    "Prelims",
)
_xlsx_section_rows = {
    _xlsx_ws.cell(row, 1).value: row for row in range(1, _xlsx_ws.max_row + 1)
    if _xlsx_ws.cell(row, 1).value in _expected_xlsx_sections
}
ck("xlsx section headers follow client order",
   tuple(_xlsx_section_rows) == _expected_xlsx_sections, _xlsx_section_rows)
_xlsx_item_row = next(row for row in range(1, _xlsx_ws.max_row + 1)
                      if "Concrete Slabs" in str(_xlsx_ws.cell(row, 1).value or ""))
_xlsx_source_rows = [row for row in range(1, _xlsx_ws.max_row + 1)
                     if _xlsx_ws.cell(row, 1).value in ("Yard-A.pdf", "Yard-B.pdf")]
_xlsx_area_total_row = next(row for row in range(1, _xlsx_ws.max_row + 1)
                            if _xlsx_ws.cell(row, 1).value == "Total Area Take Off:")
ck("xlsx keeps editable numeric per-unit source quantities and formula aggregate",
   [float(_xlsx_ws.cell(row, 2).value) for row in _xlsx_source_rows] == [100.0, 150.0] and
   _xlsx_ws.cell(_xlsx_area_total_row, 2).value ==
   f"=SUM(B{_xlsx_source_rows[0]}:B{_xlsx_source_rows[-1]})")
ck("xlsx priced qty references aggregate, rate is numeric, and value is rounded qty*rate",
   _xlsx_ws.cell(_xlsx_item_row, 2).value == f"=B{_xlsx_area_total_row}" and
   isinstance(_xlsx_ws.cell(_xlsx_item_row, 4).value, (int, float)) and
   _xlsx_ws.cell(_xlsx_item_row, 5).data_type == "f" and
   _xlsx_ws.cell(_xlsx_item_row, 5).value ==
   f"=ROUND(B{_xlsx_item_row}*D{_xlsx_item_row},2)")
_yard_section_start = _xlsx_section_rows[_expected_xlsx_sections[0]]
_footpath_section_start = _xlsx_section_rows[_expected_xlsx_sections[1]]
_yard_rows_by_label = {
    str(_xlsx_ws.cell(row, 1).value or "").splitlines()[0]: row
    for row in range(_yard_section_start + 1, _footpath_section_start)
    if str(_xlsx_ws.cell(row, 1).value or "").splitlines()
}
_included_labels = (
    "A252 Mesh Fabric x Single Layer", "Curing Agent",
    "DPM 1200G (Excl. tapes and seals to laps)", "Brush Finish",
)
ck("xlsx uses Fortel's Incl. convention without splitting or inventing component rates",
   all(label in _yard_rows_by_label and
       _xlsx_ws.cell(_yard_rows_by_label[label], 2).value == f"=B{_xlsx_area_total_row}" and
       _xlsx_ws.cell(_yard_rows_by_label[label], 4).value is None and
       _xlsx_ws.cell(_yard_rows_by_label[label], 5).value == "Incl."
       for label in _included_labels),
   {label:(_xlsx_ws.cell(_yard_rows_by_label.get(label, 1), 4).value,
           _xlsx_ws.cell(_yard_rows_by_label.get(label, 1), 5).value)
    for label in _included_labels})
ck("xlsx quantity/unit display matches client template without forced .00",
   _xlsx_ws.cell(_xlsx_item_row, 2).number_format == "#,##0.##" and
   _xlsx_ws.cell(_xlsx_item_row, 3).value == "m2")
_xlsx_total_row = next(row for row in range(1, _xlsx_ws.max_row + 1)
                       if _xlsx_ws.cell(row, 1).value == "TOTAL NETT")
ck("xlsx follows real BOQ: no section subtotals and one nett formula",
   not any(str(_xlsx_ws.cell(row, 1).value or "").startswith("Subtotal —")
           for row in range(1, _xlsx_ws.max_row + 1)) and
   _xlsx_ws.cell(_xlsx_total_row, 5).value == f"=SUM(E7:E{_xlsx_total_row - 1})")
ck("xlsx matches real BOQ widths, accounting display, and portrait layout",
   abs(_xlsx_ws.column_dimensions["A"].width - 82.43) < .01 and
   "£" in _xlsx_ws.cell(_xlsx_item_row, 5).number_format and
   _xlsx_ws.page_setup.orientation == "portrait" and _xlsx_ws.freeze_panes is None and
   _xlsx_ws.auto_filter.ref is None)
ck("xlsx visibly marks assumed quantity provisional",
   # Marker relocated out of DESCRIPTION into its own column; check either so the assertion is
   # about the marker being VISIBLE, not about which cell holds it.
   any(PROVISIONAL_LABEL in str(_xlsx_ws.cell(row, col).value or "")
       for row in range(1, _xlsx_ws.max_row + 1) for col in (1, PROVISIONAL_COL)))
ck("xlsx visibly carries every unknown client checklist field",
   any("Bay sizes if joint layout available: ASSUMED / no details provided" in
       str(_xlsx_ws.cell(row, 1).value or "") for row in range(1, _xlsx_ws.max_row + 1)))
_dock_foundation_row = next(
    row for row in range(1, _xlsx_ws.max_row + 1)
    if _xlsx_ws.cell(row, 1).value ==
       "Foundation thickenings directly underneath Dock Slab region")
ck("xlsx carries the Dock foundation-thickening subgroup without inventing a quantity or rate",
   all(_xlsx_ws.cell(_dock_foundation_row, col).value is None for col in range(2, 6)))

_bay_unit = _quotation_unit("Unit-1 Yard Joint Layout.pdf", "External yard slabs", 100)
_bay_unit["brief_spec"] = _build_brief_spec(
    "external_yard", effective_spec=_bay_unit["costing"]["spec"],
    confirmed={"bay_sizes":_e7["bay_sizes"]}, source="engineer_drawing",
    evidence={"bay_sizes":(_e7.get("_evidence") or {}).get("bay_sizes")},
)
_bay_quote = generate_quotation(_bay_unit, ref="TST-JOINT-BAYS")
_bay_joint_row = next(
    item for item in _bay_quote["line_items"]
    if item["description"].startswith("Joints (Excl. Mastic)"))
ck("extracted bay dimensions flow into the Fortel Joints row wording",
   _bay_joint_row["description"] ==
   "Joints (Excl. Mastic) - Based on 4.9m wide x 6.5m long bays",
   _bay_joint_row)
_q_status = generate_quotation(
    _quotation_unit("Status.pdf", "External yard slabs", 10),
    extras=[{"section": "Prelims", "description": "Commercial option", "qty": 1,
             "unit": "Item", "rate": 12.34, "value_status": "RATE ONLY"}],
)
_status_ws = _load_workbook(_BytesIO(quotation_xlsx(_q_status)), data_only=False)["REV_01"]
_status_row = next(row for row in range(1, _status_ws.max_row + 1)
                   if _status_ws.cell(row, 1).value == "Commercial option")
ck("xlsx supports the real BOQ's explicit RATE ONLY value token without inferring it",
   _status_ws.cell(_status_row, 4).value == 12.34 and
   _status_ws.cell(_status_row, 5).value == "RATE ONLY")

_channel_quote_unit = _quotation_unit("External Unit-1.pdf", "External yard slabs", 100)
_channel_quote_unit["channel_proposals"] = [{
    "proposal_id":"channel-dock-loading-face", "component":"dock_retaining_wall",
    "proposed_length_lm":90.0,
    "basis":"two straight runs where channels are not drawn",
}]
_channel_quote_unit["channel_proposal_decisions"] = {
    "channel-dock-loading-face": {
        "decision":"accepted", "length_lm":96.7, "edited":True,
    }
}
_q_channel = generate_quotation(_channel_quote_unit, ref="CHANNEL-QUOTE-001")
_channel_rows = [item for item in _q_channel["line_items"]
                 if item["description"] == FORTEL_CHANNEL_ROW]
ck("accepted/edited channel proposal becomes a provisional blank-rate Lm quote line",
   len(_channel_rows) == 1 and _channel_rows[0]["qty"] == 96.7 and
   _channel_rows[0]["unit"] == "Lm" and _channel_rows[0]["rate"] is None and
   _channel_rows[0]["value"] is None and _channel_rows[0]["assessor_rate_required"] and
   _channel_rows[0]["provisional"], _channel_rows)
_channel_ws = _load_workbook(_BytesIO(quotation_xlsx(_q_channel)), data_only=False)["REV_01"]
_channel_row = next(row for row in range(1, _channel_ws.max_row + 1)
                    if str(_channel_ws.cell(row, 1).value or "").startswith(
                        FORTEL_CHANNEL_ROW))
ck("accepted channel quantity is exact in text/HTML/XLSX and is never auto-priced",
   "96.7" in quotation_text(_q_channel) and "96.7 Lm" in quotation_html(_q_channel) and
   _channel_ws.cell(_channel_row, 2).value == 96.7 and
   _channel_ws.cell(_channel_row, 4).value is None and
   _channel_ws.cell(_channel_row, 5).data_type == "f" and
   # The provisional marker moved out of DESCRIPTION into its own column (right of VALUE,
   # mirroring Fortel's REMEASURE caveat column) so rows stop wrapping onto two lines. The row
   # must still be MARKED provisional — only where the marker sits changed.
   PROVISIONAL_LABEL in str(_channel_ws.cell(_channel_row, PROVISIONAL_COL).value))
_pending_channel_unit = _copy.deepcopy(_channel_quote_unit)
_pending_channel_unit["channel_proposal_decisions"] = {}
_q_pending_channel = generate_quotation(_pending_channel_unit, ref="CHANNEL-PENDING-001")
ck("unactioned channel proposal is declared explicitly and never becomes a quote quantity",
   not any(item["description"] == FORTEL_CHANNEL_ROW
           for item in _q_pending_channel["line_items"]) and
   any("has not been actioned" in note and "no channel quantity or price is included" in note
       for note in _q_pending_channel["declarations"]))

_transition_quote_unit = _quotation_unit("External Unit-2.pdf", "External yard slabs", 100)
_transition_quote_unit["transition_candidates"] = [{
    "candidate_id":"transition-yard-region-1", "region_id":"yard-region-1",
    "proposed_length_lm":15.0,
    "basis":"macadam-to-concrete boundary at the Yard entrance",
}]
_transition_quote_unit["transition_candidate_decisions"] = {
    "transition-yard-region-1": {
        "decision":"accepted", "length_lm":17.5, "edited":True,
    }
}
_q_transition_candidate = generate_quotation(
    _transition_quote_unit, ref="TRANSITION-QUOTE-001")
_transition_candidate_rows = [
    item for item in _q_transition_candidate["line_items"]
    if item["description"] == FORTEL_TRANSITION_ROW
]
ck("accepted/edited Transition candidate becomes provisional blank-rate Lm quote line",
   len(_transition_candidate_rows) == 1 and
   _transition_candidate_rows[0]["qty"] == 17.5 and
   _transition_candidate_rows[0]["unit"] == "Lm" and
   _transition_candidate_rows[0]["rate"] is None and
   _transition_candidate_rows[0]["value"] is None and
   _transition_candidate_rows[0]["assessor_rate_required"] and
   _transition_candidate_rows[0]["provisional"] and
   "macadam-to-concrete" in _transition_candidate_rows[0]["assumption_basis"],
   _transition_candidate_rows)
_pending_transition_unit = _copy.deepcopy(_transition_quote_unit)
_pending_transition_unit["transition_candidate_decisions"] = {}
_q_pending_transition = generate_quotation(
    _pending_transition_unit, ref="TRANSITION-PENDING-001")
ck("unactioned Transition candidate stays outside totals and is declared explicitly",
   not any(item["description"] == FORTEL_TRANSITION_ROW
           for item in _q_pending_transition["line_items"]) and
   any("has not been actioned" in note and
       "no Transition quantity or price is included" in note
       for note in _q_pending_transition["declarations"]),
   _q_pending_transition["declarations"])

# Fortel's standard costing sheet has one visible extra-over row for each quantity class.
# Exercise all three together, including two accepted channel runs, so the regression test
# proves the final workbook shape rather than three isolated generator branches.
_fortel_eo_unit = _quotation_unit("External Unit-3.pdf", "External yard slabs", 100)
_fortel_eo_unit["perimeter_lm"] = 80.0
_fortel_eo_unit["channel_proposals"] = [
    {"proposal_id":"channel-dock", "component":"dock_retaining_wall",
     "basis":"assessor-reviewed dock run"},
    {"proposal_id":"channel-yard", "component":"yard_wall_adjacent_run",
     "basis":"assessor-reviewed Yard run"},
]
_fortel_eo_unit["channel_proposal_decisions"] = {
    "channel-dock":{"decision":"accepted", "length_lm":12.5},
    "channel-yard":{"decision":"accepted", "length_lm":20.0},
}
_fortel_eo_unit["transition_candidates"] = [{
    "candidate_id":"transition-yard-region-1", "region_id":"yard-region-1",
    "basis":"assessor-reviewed Yard entrance",
}]
_fortel_eo_unit["transition_candidate_decisions"] = {
    "transition-yard-region-1":{"decision":"accepted", "length_lm":8.25},
}
_q_fortel_eo = generate_quotation(
    _fortel_eo_unit, ref="FORTEL-EO-ROWS",
    extras=[{"section":"External yard slabs", "description":FORTEL_MH_ROW,
             "qty":2, "unit":"Nr", "rate":None}],
)
_fortel_eo_lines = {
    item["description"]:item for item in _q_fortel_eo["line_items"]
    if item["description"] in {FORTEL_MH_ROW, FORTEL_CHANNEL_ROW, FORTEL_TRANSITION_ROW}
}
ck("channel, Transition and manhole quantities use exactly Fortel's three row labels",
   set(_fortel_eo_lines) == {FORTEL_MH_ROW, FORTEL_CHANNEL_ROW, FORTEL_TRANSITION_ROW} and
   _fortel_eo_lines[FORTEL_MH_ROW]["unit"] == "Nr" and
   _fortel_eo_lines[FORTEL_CHANNEL_ROW]["unit"] == "Lm" and
   _fortel_eo_lines[FORTEL_TRANSITION_ROW]["unit"] == "Lm" and
   _fortel_eo_lines[FORTEL_CHANNEL_ROW]["qty"] == 32.5,
   _fortel_eo_lines)
_fortel_eo_ws = _load_workbook(
    _BytesIO(quotation_xlsx(_q_fortel_eo)), data_only=False)["REV_01"]
_fortel_eo_row_numbers = {
    str(_fortel_eo_ws.cell(row, 1).value or "").splitlines()[0]:row
    for row in range(1, _fortel_eo_ws.max_row + 1)
    if str(_fortel_eo_ws.cell(row, 1).value or "").splitlines() and
       str(_fortel_eo_ws.cell(row, 1).value or "").splitlines()[0] in
       {"Slab perimeter", FORTEL_MH_ROW, FORTEL_CHANNEL_ROW, FORTEL_TRANSITION_ROW}
}
ck("xlsx writes perimeter then MH/channel/Transition with editable blank rates and formulas",
   list(_fortel_eo_row_numbers) == [
       "Slab perimeter", FORTEL_MH_ROW, FORTEL_CHANNEL_ROW, FORTEL_TRANSITION_ROW] and
   [_fortel_eo_ws.cell(_fortel_eo_row_numbers[label], 3).value
    for label in (FORTEL_MH_ROW, FORTEL_CHANNEL_ROW, FORTEL_TRANSITION_ROW)] ==
       ["Nr", "Lm", "Lm"] and
   all(_fortel_eo_ws.cell(_fortel_eo_row_numbers[label], 4).value is None and
       _fortel_eo_ws.cell(_fortel_eo_row_numbers[label], 5).value ==
       f'=IF(D{_fortel_eo_row_numbers[label]}="","",ROUND('
       f'B{_fortel_eo_row_numbers[label]}*D{_fortel_eo_row_numbers[label]},2))'
       for label in (FORTEL_MH_ROW, FORTEL_CHANNEL_ROW, FORTEL_TRANSITION_ROW)),
   _fortel_eo_row_numbers)

print("marked zone-aware measurement + multi-unit BOQ allocation")
import fitz as _fitz_zones
from robust_takeoff import read_marked as _read_marked_legacy, read_marked_zones as _read_marked_zones
from markup import parse_area_m2 as _parse_area_m2
ck("prefix-free Bluebeam measurement lines parse without accepting incidental area prose",
   _parse_area_m2("Unit-1&2\r235.37 sq m") == 235.37 and
   _parse_area_m2("Unit-3\r133.79 sq m") == 133.79 and
   _parse_area_m2("Rate note: allow 12.50 sq m at this location") is None and
   _parse_area_m2("Pricing note A = 12.50 sq m at this rate") is None and
   _parse_area_m2("This note contains 12.50 sq m for context only") is None)
ck("area truth parser still rejects linear, thickness and scale labels",
   all(_parse_area_m2(label) is None for label in ("12.30 m", "150 mm", "1:200")))
_zone_pdf = "/tmp/ci_marked_zones.pdf"
_zone_doc = _fitz_zones.open()
_zone_page = _zone_doc.new_page(width=600, height=600)
for _subject, _value, _rect in (
        ("Yard", 100, (20, 20, 220, 220)),
        ("Dock ", 30, (250, 20, 400, 120)),
        ("Mystery slab", 5, (250, 150, 350, 250))):
    _annot = _zone_page.add_polygon_annot([
        (_rect[0], _rect[1]), (_rect[2], _rect[1]),
        (_rect[2], _rect[3]), (_rect[0], _rect[3]),
    ])
    _annot.set_info(title="Fortel QA", subject=_subject, content=f"Area\n{_value:.2f} sq m")
    _annot.update()
_channel_annot = _zone_page.add_polyline_annot([(20, 300), (120, 300), (160, 330)])
_channel_annot.set_info(title="Fortel QA", subject="Channel", content="Channel\n12.50 m")
_channel_annot.update()
_zone_doc.save(_zone_pdf)
_zone_doc.close()
_zone_read = _read_marked_zones(_zone_pdf)
_zone_by_category = {zone["category"]: zone for zone in _zone_read["zones"]}
ck("marked zone reader preserves legacy aggregate across every labelled polygon",
   _zone_read["area_m2"] == 135.0 and _zone_read["regions"] == 3 and
   _read_marked_legacy(_zone_pdf) == (135.0, 3), _zone_read)
ck("Bluebeam subjects separate Yard, Dock and Channel without colour fallback",
   _zone_by_category["external_yard"]["area_m2"] == 100.0 and
   _zone_by_category["dock"]["area_m2"] == 30.0 and
   _zone_by_category["channel"]["length_lm"] == 12.5, _zone_by_category)
ck("unknown measurable subject is unclassified and visibly flagged",
   _zone_by_category["unclassified"]["area_m2"] == 5.0 and
   any("assessor: classify zone 'Mystery slab'" in flag for flag in _zone_read["flags"]),
   _zone_read["flags"])
ck("zone reader retains per-annotation subject/author/colour evidence",
   len(_zone_read["markup_annotations"]) == 4 and
   all("subject" in record and "author" in record and "stroke_color" in record
       for record in _zone_read["markup_annotations"]))

_exclusion_pdf = "/tmp/ci_marked_slab_exclusions.pdf"
_exclusion_doc = _fitz_zones.open()
_exclusion_page = _exclusion_doc.new_page(width=600, height=600)
for _subject, _value, _x in (
        ("Ground floor core", 500.0, 20),
        ("Lift shaft", 12.0, 260),
        ("Data riser", 4.0, 380),
        ("Precast staircase foundation", 8.0, 480)):
    _annot = _exclusion_page.add_polygon_annot([
        (_x, 20), (_x + 80, 20), (_x + 80, 120), (_x, 120),
    ])
    _annot.set_info(title="Fortel QA", subject=_subject,
                    content=f"Area\n{_value:.2f} sq m")
    _annot.update()
_exclusion_doc.save(_exclusion_pdf)
_exclusion_doc.close()
_exclusion_read = _read_marked_zones(_exclusion_pdf)
ck("explicit lift/riser/stair-foundation markups are excluded, recorded and never summed",
   _exclusion_read["area_m2"] == 500.0 and
   _exclusion_read["regions"] == 1 and
   len(_exclusion_read["zones"]) == 1 and
   {item["exclusion_id"] for item in _exclusion_read["exclusions"]} == {
       "lift_void", "service_data_riser", "precast_stair_foundation"} and
   all(record["excluded_from_slab"]
       for record in _exclusion_read["markup_annotations"][1:]),
   _exclusion_read)
from measurement_rules import (
    classify_exclusion as _classify_client_exclusion,
    exclusion_review_prompts as _exclusion_review_prompts,
)
_yard_exclusion_prompts = _exclusion_review_prompts(
    ["external_yard"], "Proposed Gatehouse and Hub Office")
ck("raw labels create visible unresolved Gatehouse/Hub-office prompts, never fake geometry",
   {prompt["exclusion_id"] for prompt in _yard_exclusion_prompts
    if prompt["status"] == "outline_unresolved"} == {"gatehouse", "hub_office"} and
   all(prompt.get("requires_assessor_confirmation")
       for prompt in _yard_exclusion_prompts), _yard_exclusion_prompts)
ck("a bare pit is never guessed to be a lift void or precast stair foundation",
   _classify_client_exclusion("Pit", "300 345 600 mm") is None)
ck("a lift lobby is slab circulation space, not guessed to be a lift void",
   _classify_client_exclusion("Lift lobby", "") is None)
_explicit_client_exclusions = {
    subject: _classify_client_exclusion(subject, content)
    for subject, content in (
        ("Gatehouse", ""),
        ("Hub office", ""),
        ("Area Measurement", "Lift bit"),
        ("Riser data", ""),
        ("Pre-cast concrete staircase foundation", ""),
    )
}
ck("5-Aug exclusion wording is recognised from explicit subject/content evidence",
   {subject: record and record.get("exclusion_id")
    for subject, record in _explicit_client_exclusions.items()} == {
       "Gatehouse":"gatehouse", "Hub office":"hub_office",
       "Area Measurement":"lift_void", "Riser data":"service_data_riser",
       "Pre-cast concrete staircase foundation":"precast_stair_foundation",
   }, _explicit_client_exclusions)
_ground_exclusion_checklist = _exclusion_review_prompts(["ground_floor"], "")
ck("every ground-floor trace keeps the lift/riser/stair checklist visible without a text hit",
   {prompt["exclusion_id"] for prompt in _ground_exclusion_checklist} == {
       "lift_void", "service_data_riser", "precast_stair_foundation"} and
   all(prompt.get("status") == "assessor_check" and
       prompt.get("requires_assessor_confirmation") and prompt.get("assumed") and
       prompt.get("basis") for prompt in _ground_exclusion_checklist),
   _ground_exclusion_checklist)
_labelled_ground_exclusions = _exclusion_review_prompts(
    ["ground_floor"], "Lift pit and data riser")
ck("drawing-text evidence gates unresolved lift/data outlines instead of missing slash labels",
   {prompt["exclusion_id"] for prompt in _labelled_ground_exclusions
    if prompt["status"] == "outline_unresolved"} == {
       "lift_void", "service_data_riser"}, _labelled_ground_exclusions)

_scope_pdf = "/tmp/ci_internal_warehouse_scope.pdf"
_scope_doc = _fitz_zones.open()
_scope_page = _scope_doc.new_page(width=600, height=300)
for _subject, _value, _x in (("Ground floor core", 120.0, 20),
                             ("Internal warehouse slab", 900.0, 260)):
    _annot = _scope_page.add_polygon_annot([
        (_x, 20), (_x + 180, 20), (_x + 180, 180), (_x, 180),
    ])
    _annot.set_info(title="Fortel QA", subject=_subject,
                    content=f"Area\n{_value:.2f} sq m")
    _annot.update()
_scope_doc.save(_scope_pdf)
_scope_doc.close()
_scope_read = _read_marked_zones(_scope_pdf)
ck("explicit internal warehouse slabs stay visible but never enter our area or BOQ zones",
   _scope_read["area_m2"] == 120.0 and _scope_read["regions"] == 1 and
   len(_scope_read["zones"]) == 1 and
   _scope_read["zones"][0]["category"] == "ground_floor" and
   any(item["exclusion_id"] == "internal_warehouse_scope" and
       item["area_m2"] == 900.0 for item in _scope_read["exclusions"]) and
   any("Internal warehouse slab" in flag for flag in _scope_read["flags"]),
   _scope_read)

_transition_pdf = "/tmp/ci_unit_yard_transitions.pdf"
_transition_doc = _fitz_zones.open()
_transition_page = _transition_doc.new_page(width=600, height=300)
for _subject, _value, _y in (("Unit 1 Transition", 12.5, 50),
                             ("Unit 2 Transition", 8.25, 150)):
    _annot = _transition_page.add_polyline_annot([(20, _y), (220, _y)])
    _annot.set_info(title="Fortel QA", subject=_subject,
                    content=f"{_value:.2f} m")
    _annot.update()
_transition_doc.save(_transition_pdf)
_transition_doc.close()
_transition_read = _read_marked_zones(_transition_pdf)
_transition_zones = [zone for zone in _transition_read["zones"]
                     if zone["category"] == "transition"]
ck("yard-entrance transitions remain one measured Lm zone per unit, never one anonymous lump",
   len(_transition_zones) == 2 and
   [zone["unit_label"] for zone in _transition_zones] == ["unit 1", "unit 2"] and
   [zone["length_lm"] for zone in _transition_zones] == [12.5, 8.25] and
   all("tarmac-to-concrete" in zone["basis"] for zone in _transition_zones),
   _transition_zones)
_transition_quote = generate_quotation({
    "file":"Site yard transitions.pdf", "source_discipline":"engineer",
    "zones":_transition_zones, "flags":[], "area_m2":None,
}, project="Transition QA")
_transition_measurement = next(
    item for item in _transition_quote["measurements"]
    if item["description"] == FORTEL_TRANSITION_ROW)
ck("measured Yard transitions reach the quotation as blank-rate Lm source rows",
   _transition_measurement["qty"] == 20.75 and
   _transition_measurement["assessor_rate_required"] and
   [row["description"] for row in _transition_measurement["quantity_rows"]] ==
   ["unit 1", "unit 2"], _transition_measurement)

from takeoff_unmarked import detect_raw_construction_joint as _detect_raw_cj
_raw_cj_pdf = "/tmp/ci_raw_construction_joint.pdf"
_raw_cj_doc = _fitz_zones.open()
_raw_cj_page = _raw_cj_doc.new_page(width=600, height=500)
_raw_cj_page.insert_text((180, 80), "CJ indicates internal construction joint")
_raw_cj_legend = _raw_cj_page.new_shape()
_raw_cj_legend.draw_line((100, 76), (160, 76))
_raw_cj_legend.finish(color=(0.0, 0.75, 0.0)); _raw_cj_legend.commit()
_raw_cj_body = _raw_cj_page.new_shape()
_raw_cj_body.draw_line((50, 200), (150, 200))
_raw_cj_body.draw_line((200, 300), (300, 300))
_raw_cj_body.finish(color=(0.0, 0.75, 0.0)); _raw_cj_body.commit()
_raw_cj_other = _raw_cj_page.new_shape()
_raw_cj_other.draw_line((50, 400), (350, 400))
_raw_cj_other.finish(color=(0.0, 0.0, 0.0)); _raw_cj_other.commit()
_raw_cj_doc.save(_raw_cj_pdf); _raw_cj_doc.close()
_raw_cj_detected = _detect_raw_cj(_raw_cj_pdf, 0.1)
ck("raw CJ reads its green legend swatch and measures only matching body vectors",
   _raw_cj_detected.get("zone", {}).get("category") == "construction_joint" and
   _raw_cj_detected["zone"]["length_lm"] == 20.0 and
   len(_raw_cj_detected["zone"]["polyline_segments"]) == 2 and
   "600 centres" in _raw_cj_detected["zone"]["joint_detail"],
   _raw_cj_detected)

_marked_cj_pdf = "/tmp/Ground Floor CJ QA.pdf"
_marked_cj_doc = _fitz_zones.open()
_marked_cj_page = _marked_cj_doc.new_page(width=500, height=300)
_marked_cj = _marked_cj_page.add_polyline_annot([(20, 50), (145, 50)])
_marked_cj.set_info(title="Fortel QA", subject="CJ internal", content="12.50 m")
_marked_cj.update()
_roller = _marked_cj_page.add_polyline_annot([(20, 150), (300, 150)])
_roller.set_info(title="Fortel QA", subject="Roller shutter detail", content="99.00 m")
_roller.update()
_marked_cj_doc.save(_marked_cj_pdf); _marked_cj_doc.close()
_marked_cj_read = _read_marked_zones(_marked_cj_pdf)
_marked_cj_zone = next(zone for zone in _marked_cj_read["zones"]
                       if zone["category"] == "construction_joint")
ck("explicit CJ is a measured Lm zone while roller-shutter detail stays out of scope",
   _marked_cj_zone["length_lm"] == 12.5 and
   _marked_cj_zone.get("slab_category") == "ground_floor" and
   any(zone["category"] == "other" for zone in _marked_cj_read["zones"]),
   _marked_cj_read["zones"])
_marked_cj_quote = generate_quotation({
    "file":"Ground Floor CJ QA.pdf", "source_discipline":"engineer",
    "zones":_marked_cj_read["zones"], "flags":[], "area_m2":None,
}, project="CJ QA")
_cj_measurement = next(item for item in _marked_cj_quote["measurements"]
                       if item["description"] == "Internal construction joint (CJ)")
ck("CJ reaches Ground-floor quotation as quantity-only Lm with blank assessor rate",
   _cj_measurement["section"] == "Ground floor slabs" and
   _cj_measurement["qty"] == 12.5 and _cj_measurement["assessor_rate_required"] and
   any("150x150x8 mm" in note for note in _marked_cj_quote["declarations"]),
   {"measurement":_cj_measurement,
    "declarations":_marked_cj_quote["declarations"]})

_upper_scope_pdf = "/tmp/ci_upper_floor_scopes.pdf"
_upper_scope_doc = _fitz_zones.open()
_upper_scope_page = _upper_scope_doc.new_page(width=700, height=300)
for _subject, _value, _x in (
        ("Unit 5 First Floor", 100.0, 20),
        ("Unit 5 Plant Deck", 30.0, 260),
        ("Unit 5 POD First Floor", 40.0, 500)):
    _annot = _upper_scope_page.add_polygon_annot([
        (_x, 20), (_x + 120, 20), (_x + 120, 150), (_x, 150),
    ])
    _annot.set_info(title="Fortel QA", subject=_subject,
                    content=f"Area\n{_value:.2f} sq m")
    _annot.update()
_upper_scope_doc.save(_upper_scope_pdf)
_upper_scope_doc.close()
_upper_scope_read = _read_marked_zones(_upper_scope_pdf)
ck("main upper floor, Plant deck and Unit-5 POD stay three explicit measured scopes",
   len(_upper_scope_read["zones"]) == 3 and
   {zone["boq_scope"] for zone in _upper_scope_read["zones"]} == {
       "main_upper_floor", "plant_deck", "pod_first_floor"} and
   all(zone["category"] == "upper_floor" and
       zone["boundary_rule"] == "measure to the edge of the metal decking"
       for zone in _upper_scope_read["zones"]), _upper_scope_read["zones"])
_upper_scope_quote = generate_quotation({
    "file":"Unit-5 Upper Floors.pdf", "source_discipline":"engineer",
    "zones":_upper_scope_read["zones"], "flags":[], "area_m2":170.0,
    "costing":{"area_m2":170.0, "rate":None, "total_gbp":None,
               "assumed":True, "spec":{}, "breakdown":{}, "extras":[]},
}, project="Upper Scope QA")
_upper_scope_items = [item for item in _upper_scope_quote["line_items"]
                      if item["section"] == "Upper floor slabs" and
                      item.get("line_role") == "concrete_slab"]
ck("Plant deck and POD are separate Upper-floor BOQ rows, never merged with main floor",
   len(_upper_scope_items) == 3 and
   any(item["description"].startswith("Plant deck —") for item in _upper_scope_items) and
   any(item["description"].startswith("POD first floor —") for item in _upper_scope_items),
   _upper_scope_items)

_unit4_pdf = "/tmp/Ground Floor Unit 4 Subunits.pdf"
_unit4_doc = _fitz_zones.open()
_unit4_page = _unit4_doc.new_page(width=700, height=300)
for _index, _subject in enumerate(("Unit 4A", "Unit 4B", "Unit 4C", "Unit 4D")):
    _x = 20 + _index * 160
    _annot = _unit4_page.add_polygon_annot([
        (_x, 20), (_x + 100, 20), (_x + 100, 120), (_x, 120),
    ])
    _annot.set_info(title="Fortel QA", subject=_subject, content="Area\n25.00 sq m")
    _annot.update()
_unit4_doc.save(_unit4_pdf)
_unit4_doc.close()
_unit4_read = _read_marked_zones(_unit4_pdf)
ck("complete Unit-4A/4B/4C/4D markup is one combined ground-floor slab zone",
   len(_unit4_read["zones"]) == 1 and
   _unit4_read["zones"][0]["category"] == "ground_floor" and
   _unit4_read["zones"][0]["area_m2"] == 100.0 and
   _unit4_read["zones"][0]["annotation_count"] == 4 and
   _unit4_read["zones"][0]["unit_label"] == "Unit 4 (4A-4D combined)" and
   not _unit4_read["unit_group_review_required"], _unit4_read)

_unit_zone_pdf = "/tmp/Yard Markup Unit Labels.pdf"
_unit_zone_doc = _fitz_zones.open()
_unit_zone_page = _unit_zone_doc.new_page(width=600, height=600)
for _subject, _value, _x in (("Unit-1&2", 235.37, 20), ("Unit-3", 133.79, 300)):
    _annot = _unit_zone_page.add_polygon_annot([
        (_x, 20), (_x + 200, 20), (_x + 200, 220), (_x, 220),
    ])
    _annot.set_info(title="Fortel QA", subject=_subject,
                    content=f"Area\n{_value:.2f} sq m")
    _annot.update()
_unit_zone_doc.save(_unit_zone_pdf)
_unit_zone_doc.close()
_unit_zone_read = _read_marked_zones(_unit_zone_pdf)
_unit_yard_zones = [zone for zone in _unit_zone_read["zones"]
                    if zone["category"] == "external_yard"]
ck("unit-labelled regions inherit a strong Yard filename context and preserve each area",
   len(_unit_yard_zones) == 2 and
   sorted(zone["area_m2"] for zone in _unit_yard_zones) == [133.79, 235.37] and
   sorted(zone["unit_label"] for zone in _unit_yard_zones) == ["Unit-1&2", "Unit-3"] and
   _unit_zone_read["area_m2"] == 369.2,
   _unit_zone_read)
_weak_unit_zone_pdf = "/tmp/General Markup Unit Labels.pdf"
shutil.copyfile(_unit_zone_pdf, _weak_unit_zone_pdf)
_weak_unit_read = _read_marked_zones(_weak_unit_zone_pdf)
ck("unit labels with weak context stay unclassified while every measured area is preserved",
   all(zone["category"] == "unclassified" and zone["needs_assessor"]
       for zone in _weak_unit_read["zones"]) and
   abs(sum(zone["area_m2"] for zone in _weak_unit_read["zones"]) - 369.16) < 0.001 and
   _weak_unit_read["area_m2"] == 369.2 and
   len(_weak_unit_read["flags"]) == 2,
   _weak_unit_read)

_zone_quote_results = []
for _unit_n in range(1, 5):
    _unit = _quotation_unit(f"Castle Unit-{_unit_n}.pdf", "External yard slabs", 1)
    _unit["zones"] = [
        {"category":"external_yard", "area_m2":100 * _unit_n, "perimeter_lm":10 * _unit_n},
        {"category":"dock", "area_m2":10 * _unit_n, "perimeter_lm":5 * _unit_n},
        {"category":"ground_floor", "area_m2":5 * _unit_n, "perimeter_lm":3 * _unit_n},
        {"category":"upper_floor", "area_m2":20 * _unit_n, "perimeter_lm":4 * _unit_n},
        {"category":"channel", "length_lm":7 * _unit_n},
        {"category":"transition", "length_lm":2 * _unit_n},
    ]
    _unit["brief_specs"] = {
        category: _empty_brief_spec(category)
        for category in ("external_yard", "dock", "ground_floor", "upper_floor")
    }
    _zone_quote_results.append(_unit)
_q_zones = generate_quotation(_zone_quote_results, project="Castle", client="Winvic",
                              ref="ZONE-001")
_zone_section_order = [
    "External yard slabs", "Dock slabs", "Ground floor slabs", "Upper floor slabs",
]
ck("mixed marked files allocate into all four BOQ sections",
   [spec["section"] for spec in _q_zones["specifications"]] == _zone_section_order)
ck("each BOQ section keeps four numeric Unit-N source rows",
   all([row["description"] for row in spec["area_rows"]] ==
       ["Unit-1", "Unit-2", "Unit-3", "Unit-4"]
       for spec in _q_zones["specifications"]), _q_zones["specifications"])
ck("aggregate job rate is never copied onto mixed zones",
   all(item.get("rate") is None and item.get("value") is None
       for item in _q_zones["line_items"]
       if item.get("line_role") == "concrete_slab"))
ck("channel, transition and zone perimeters remain unpriced Lm source quantities",
   any(m["description"] == FORTEL_CHANNEL_ROW and m["qty"] == 70
       for m in _q_zones["measurements"]) and
   any(m["description"] == FORTEL_TRANSITION_ROW and m["qty"] == 20
       for m in _q_zones["measurements"]) and
   all(m.get("assessor_rate_required") for m in _q_zones["measurements"]))
_zone_ws = _load_workbook(_BytesIO(quotation_xlsx(_q_zones)), data_only=False)["REV_01"]
_zone_xlsx_sections = [
    _expected_xlsx_sections[0], _expected_xlsx_sections[2],
    _expected_xlsx_sections[3], _expected_xlsx_sections[4],
]
_zone_section_labels = [
    _zone_ws.cell(row, 1).value for row in range(1, _zone_ws.max_row + 1)
    if _zone_ws.cell(row, 1).value in _zone_xlsx_sections
]
_channel_row = next(row for row in range(1, _zone_ws.max_row + 1)
                    if _zone_ws.cell(row, 1).value == FORTEL_CHANNEL_ROW)
ck("zone XLSX preserves section order with editable blank assessor rates",
   _zone_section_labels == _zone_xlsx_sections and
   _zone_ws.cell(_channel_row, 4).value is None and
   _zone_ws.cell(_channel_row, 5).value ==
   f'=IF(D{_channel_row}="","",ROUND(B{_channel_row}*D{_channel_row},2))')

_castle_dir = Path("drawings/castle_donington")
_castle_names = [
    *(f"External Markup Unit-{number}.pdf" for number in range(1, 5)),
    *(f"Office Floors Unit-{number}.pdf" for number in range(1, 5)),
]
try:
    for _castle_name in _castle_names:
        _require_fixture(_castle_dir / _castle_name, "Castle Donington client zone-gold checks")
    import json as _json_zone_gold
    from takeoff_pipeline import _zone_reference_flags as _zone_reference_flags_test
    _zone_gold = _json_zone_gold.loads(Path("gold.json").read_text())
    _castle_reads = {}
    for _castle_name in _castle_names:
        _castle_path = _castle_dir / _castle_name
        _castle_marked = _read_marked_zones(str(_castle_path))
        _castle_reads[_castle_name] = _castle_marked
        _entry = _zone_gold[str(_castle_path)]
        _actual = {z["category"]: z.get("area_m2") for z in _castle_marked["zones"]
                   if z.get("area_m2") is not None}
        _aggregate_delta = abs(_castle_marked["area_m2"] - _entry["net_m2"]) / _entry["net_m2"] * 100
        _zones_pass = all(
            category in _actual and abs(_actual[category] - expected) / expected * 100 <=
            _entry["zone_tol_pct"]
            for category, expected in _entry["zones_m2"].items()
        )
        ck(f"Castle zone gold: {_castle_name} aggregate + BOQ sections",
           _aggregate_delta <= _entry["tol_pct"] and _zones_pass,
           {"actual": _actual, "gold": _entry["zones_m2"]})
    _dock_perimeter = sum(
        next(z["perimeter_lm"] for z in _castle_reads[f"External Markup Unit-{n}.pdf"]["zones"]
             if z["category"] == "dock") for n in range(1, 5)
    )
    ck("Castle Dock polygon perimeter reproduces client BOQ 967 Lm",
       abs(_dock_perimeter - 967) / 967 * 100 <= 1, _dock_perimeter)
    _unit3_path = _castle_dir / "External Markup Unit-3.pdf"
    ck("Castle Unit-3 channel-vs-BOQ mismatch emits assessor flag",
       any("channel measured 545.36 Lm" in flag
           for flag in _zone_reference_flags_test(str(_unit3_path),
                                                  _castle_reads[_unit3_path.name]["zones"])))
except _FixtureNotPresent as _e:
    print(f"  [SKIP] {_e} — fixture not present")

print("raw external zone measurement — Yard/Dock split + low-confidence state cap")
import numpy as _np_multi_yard
from takeoff_unmarked import segment_hatch as _segment_hatch_multi_yard
_multi_yard_image = _np_multi_yard.full((100, 100, 3), 255, dtype=_np_multi_yard.uint8)
_multi_yard_image[10:25, 10:30] = (180, 180, 180)   # 300 m2 primary
_multi_yard_image[45:55, 10:22] = (180, 180, 180)   # 120 m2 second unit (<200)
_multi_yard_image[70:72, 80:84] = (180, 180, 180)   # 8 m2 legend chip
_multi_yard_diag = {}
_multi_yard_mask = _segment_hatch_multi_yard(
    _multi_yard_image, (180, 180, 180), tol=0, close=1, k=1.0, S=1.0,
    exclude_border=False, legend_exclusion_bbox=[79, 69, 85, 73],
    full_rgb=True, _diag=_multi_yard_diag)
ck("same-tint segmentation retains a real second unit below the single-yard 200 m2 floor",
   int(_multi_yard_mask.sum()) == 420 and
   len(_multi_yard_diag.get("_retained_component_masks", [])) == 2,
   {"pixels": int(_multi_yard_mask.sum()),
    "components": _multi_yard_diag.get("component_candidates")})
ck("matched legend chip is geometrically excluded, never promoted as another Yard",
   not _multi_yard_mask[70:72, 80:84].any(),
    {"excluded_legend_m2": _multi_yard_diag.get("excluded_legend_m2"),
     "components": _multi_yard_diag.get("component_candidates")})

# Raster boundaries can be serrated by internal linework.  Perimeter is permitted only when
# an independently encoded closed CAD path strongly overlaps the segmented tint.  This fixture
# has two disjoint regions and deliberately nicks raster pixels along one edge; neither expected
# perimeter is inferred from an area or a client value.
import fitz as _fitz_native_boundary
from takeoff_unmarked import _native_boundary_for_mask as _native_boundary_for_mask_test

def _closed_line_drawing(_points):
    _closed = _points + [_points[0]]
    return {
        "rect": _fitz_native_boundary.Rect(
            min(p.x for p in _points), min(p.y for p in _points),
            max(p.x for p in _points), max(p.y for p in _points)),
        "items": [("l", start, end) for start, end in zip(_closed, _closed[1:])],
        "closePath": True,
    }

class _BoundaryPage:
    rotation_matrix = _fitz_native_boundary.Matrix(1, 1)
    def __init__(self, drawings):
        self._drawings = drawings
    def get_drawings(self):
        return self._drawings

_boundary_a = [_fitz_native_boundary.Point(10, 10), _fitz_native_boundary.Point(60, 10),
               _fitz_native_boundary.Point(60, 40), _fitz_native_boundary.Point(10, 40)]
_boundary_b = [_fitz_native_boundary.Point(90, 55), _fitz_native_boundary.Point(140, 55),
               _fitz_native_boundary.Point(140, 90), _fitz_native_boundary.Point(120, 90),
               _fitz_native_boundary.Point(120, 75), _fitz_native_boundary.Point(90, 75)]
_boundary_page = _BoundaryPage([
    _closed_line_drawing(_boundary_a), _closed_line_drawing(_boundary_b)])
_boundary_mask_a = _np_multi_yard.zeros((120, 170), dtype=bool)
_boundary_mask_a[10:41, 10:61] = True
_boundary_mask_a[10:12, 20:25] = False  # raster-only nick: must not become extra perimeter
_boundary_mask_b = _np_multi_yard.zeros((120, 170), dtype=_np_multi_yard.uint8)
import cv2 as _cv2_native_boundary
_cv2_native_boundary.fillPoly(
    _boundary_mask_b, [_np_multi_yard.array(_boundary_b, dtype=_np_multi_yard.int32)], 1)
_native_a, _native_a_reason = _native_boundary_for_mask_test(
    _boundary_page, _boundary_mask_a, S=1.0, k=0.1)
_native_b, _native_b_reason = _native_boundary_for_mask_test(
    _boundary_page, _boundary_mask_b.astype(bool), S=1.0, k=0.1)
ck("native CAD perimeter ignores raster/internal-edge serration on each disjoint region",
   _native_a is not None and _native_b is not None and
   abs(_native_a["perimeter_lm"] - 16.0) < 0.01 and
   abs(_native_b["perimeter_lm"] - 17.0) < 0.01,
   {"region_a": _native_a or _native_a_reason,
    "region_b": _native_b or _native_b_reason})
_unresolved_boundary, _unresolved_reason = _native_boundary_for_mask_test(
    _BoundaryPage([]), _boundary_mask_a, S=1.0, k=0.1)
ck("perimeter refuses when no corroborating native boundary exists",
   _unresolved_boundary is None and "no explicit closed" in _unresolved_reason,
   _unresolved_reason)

from takeoff_unmarked import (
    _transition_candidates_from_surface_mask as _transition_candidates_from_mask_test,
    _native_boundary_stream_budget as _native_boundary_stream_budget_test,
    MAX_NATIVE_BOUNDARY_STREAM_BYTES as _native_boundary_stream_limit_test,
)
ck("dense native-vector pages refuse optional perimeter parsing before the robustness timeout",
   _native_boundary_stream_budget_test(_native_boundary_stream_limit_test) and
   not _native_boundary_stream_budget_test(_native_boundary_stream_limit_test + 1))
_transition_surface = _np_multi_yard.zeros((100, 160), dtype=bool)
_transition_surface[41:46, 10:61] = True
_transition_yard = [{
    "region_id": "yard-region-1",
    "polygon_pts": [[10, 10], [60, 10], [60, 40], [10, 40]],
    "perimeter_confidence": "high",
}]
_transition_prefills, _transition_prefill_reasons = \
    _transition_candidates_from_mask_test(
        _transition_yard, _transition_surface, k=0.1, S=1.0)
ck("legend-surface adjacency creates one assisted Transition prefill outside measured zones",
   len(_transition_prefills) == 1 and
   _transition_prefills[0]["proposed_length_lm"] == 5.0 and
   _transition_prefills[0]["assumed"] is True and
   "length_lm" not in _transition_prefills[0] and
   _transition_prefills[0]["category"] == "transition",
   _transition_prefills or _transition_prefill_reasons)
_transition_ambiguous_surface = _transition_surface.copy()
_transition_ambiguous_surface[4:10, 10:61] = True
_transition_ambiguous, _transition_ambiguous_reasons = \
    _transition_candidates_from_mask_test(
        _transition_yard, _transition_ambiguous_surface, k=0.1, S=1.0)
ck("two disjoint adjacent surface runs refuse instead of guessing a Yard entrance",
   not _transition_ambiguous and
   any("disjoint" in reason for reason in _transition_ambiguous_reasons),
   _transition_ambiguous_reasons)

# Source: Fortel's Yard Markup.pdf supplied for 2165 Tanro Voltage Business Park:
# Unit-1&2 = 235.37 m² and Unit-3 = 133.79 m².  Production sees only a temporary
# annotation-stripped copy; these client answers stay here in the validation assertion.
try:
    import glob as _glob_tanro_regions
    import tempfile as _tempfile_tanro_regions
    from accuracy_report import strip_annotations as _strip_tanro_regions
    from pathlib import Path as _Path_tanro_regions
    import takeoff_unmarked as _takeoff_unmarked_tanro_regions
    _tanro_matches = _glob_tanro_regions.glob(
        "drawings/aryan_drive/**/2165 Tanro- Voltage Business Park/Markup/Yard Markup.pdf",
        recursive=True,
    )
    if not _tanro_matches:
        raise _FixtureNotPresent("Tanro Yard Markup.pdf")
    with _tempfile_tanro_regions.TemporaryDirectory() as _tanro_tmp:
        _tanro_raw = _Path_tanro_regions(_tanro_tmp) / "Yard raw.pdf"
        _strip_tanro_regions(_Path_tanro_regions(_tanro_matches[0]), _tanro_raw)
        _tanro_result = _takeoff_unmarked_tanro_regions.takeoff(str(_tanro_raw))
        _tanro_regions = _tanro_result.get("yard_regions") or []
        _tanro_areas = sorted(region.get("area_m2") for region in _tanro_regions)
        _tanro_truth = sorted([235.37, 133.79])
        ck("Tanro raw Yard separates both client unit regions within 5%",
           len(_tanro_areas) == 2 and all(
               abs(actual - expected) / expected * 100 <= 5
               for actual, expected in zip(_tanro_areas, _tanro_truth)
           ), {"actual": _tanro_areas, "truth": _tanro_truth})
        ck("every retained raw Yard region surfaces its own area, bbox and perimeter",
           all(region.get("area_m2") and region.get("bbox_pdf_pts")
               and region.get("perimeter_lm") for region in _tanro_regions),
           _tanro_regions)
        _tanro_actual = sorted(
            (region["area_m2"], region["perimeter_lm"]) for region in _tanro_regions)
        _tanro_perimeter_truth = sorted([(235.37, 61.54), (133.79, 49.80)])
        ck("Tanro raw Yard native per-region perimeters are within 5%",
           len(_tanro_actual) == 2 and all(
               abs(actual_perimeter - truth_perimeter) / truth_perimeter * 100 <= 5
               for (_, actual_perimeter), (_, truth_perimeter)
               in zip(_tanro_actual, _tanro_perimeter_truth)
           ), {"actual": _tanro_actual, "truth": _tanro_perimeter_truth})
        _tanro_transition_truth = sorted([14.25, 12.30])
        _tanro_transition_candidates = _tanro_result.get("transition_candidates") or []
        _tanro_transition_actual = sorted(
            candidate.get("proposed_length_lm")
            for candidate in _tanro_transition_candidates)
        ck("Tanro raw macadam/Yard adjacency prefills both client Transition runs within 5%",
           len(_tanro_transition_actual) == 2 and all(
               abs(actual - expected) / expected * 100 <= 5
               for actual, expected in zip(
                   _tanro_transition_actual, _tanro_transition_truth)
           ), {"actual": _tanro_transition_actual,
               "truth": _tanro_transition_truth})
        ck("raw Transition prefills never leak into measured zones or the measured total",
           all(candidate.get("assumed") is True and
               "length_lm" not in candidate
               for candidate in _tanro_transition_candidates) and
           not any(zone.get("category") == "transition"
                   for zone in (_tanro_result.get("zones") or [])),
           {"candidates": _tanro_transition_candidates,
            "zones": _tanro_result.get("zones")})
except _FixtureNotPresent as _e:
    print(f"  [SKIP] {_e} — fixture not present")
try:
    _external_marked_paths = [
        _castle_dir / f"External Markup Unit-{number}.pdf" for number in range(1, 5)
    ]
    for _path in _external_marked_paths:
        _require_fixture(_path, "Castle Donington raw Yard/Dock validation")
    from accuracy_report import strip_annotations as _strip_annotations_external
    from takeoff_pipeline import takeoff as _pipeline_takeoff_external
    import takeoff_unmarked as _takeoff_unmarked_external
    import os as _os_external

    _raw_external_results = {}
    try:
        for _unit_number, _marked_path in enumerate(_external_marked_paths, start=1):
            _raw_path = Path(f"/tmp/ci_external_unit_{_unit_number}_stripped.pdf")
            _strip_annotations_external(_marked_path, _raw_path)
            _truth = _read_marked_zones(str(_marked_path))
            _truth_by_zone = {
                zone["category"]: zone
                for zone in _truth["zones"]
            }
            _measured = _pipeline_takeoff_external(
                str(_raw_path), send_approval=False, auto_extract_spec=False)
            _raw_external_results[_unit_number] = _measured
            _measured_by_zone = {
                zone["category"]: zone
                for zone in _measured.get("zones", [])
            }
            _yard_ok = (
                "external_yard" in _measured_by_zone
                and abs(
                    _measured_by_zone["external_yard"]["area_m2"]
                    - _truth_by_zone["external_yard"]["area_m2"]
                ) / _truth_by_zone["external_yard"]["area_m2"] * 100 <= 5
            )
            _dock_ok = (
                "dock" in _measured_by_zone
                and abs(
                    _measured_by_zone["dock"]["area_m2"]
                    - _truth_by_zone["dock"]["area_m2"]
                ) / _truth_by_zone["dock"]["area_m2"] * 100 <= 5
            )
            _zone_sum = sum(
                zone.get("area_m2") or 0
                for zone in _measured.get("zones", [])
                if zone.get("category") in ("external_yard", "dock")
            )
            ck(f"raw External Unit-{_unit_number}: one Yard + one Dock, each within 5%",
               _yard_ok and _dock_ok and
               [zone["category"] for zone in _measured.get("zones", [])].count(
                   "external_yard") == 1 and
               [zone["category"] for zone in _measured.get("zones", [])].count("dock") == 1,
               {"truth": {key: value.get("area_m2")
                          for key, value in _truth_by_zone.items()},
                "measured": {key: value.get("area_m2")
                             for key, value in _measured_by_zone.items()}})
            ck(f"raw External Unit-{_unit_number}: zone total reconciles exactly, no double count",
               abs(_zone_sum - _measured["zones_total_area_m2"]) < 0.01
               and _measured["area_m2"] == _measured_by_zone["external_yard"]["area_m2"],
               {"zones": _zone_sum,
                "zone_total": _measured["zones_total_area_m2"],
                "legacy_yard_area": _measured["area_m2"]})
    finally:
        for _unit_number in range(1, 5):
            try:
                _os_external.remove(f"/tmp/ci_external_unit_{_unit_number}_stripped.pdf")
            except FileNotFoundError:
                pass

    ck("raw channel assumptions stay outside measured zones and accuracy totals",
       all(
           not any(zone.get("category") in ("channel", "transition")
                   for zone in result.get("zones", []))
           and len(result.get("channel_proposals", [])) == 2
           and {proposal.get("component") for proposal in result["channel_proposals"]} ==
               {"dock_retaining_wall", "yard_longest_contained_run"}
           and all(proposal.get("assumed") is True
                   and proposal.get("requires_assessor_confirmation") is True
                   and len(proposal.get("polyline_pts", [])) == 2
                   and (abs(proposal["polyline_pts"][0][0] -
                            proposal["polyline_pts"][1][0]) <= 0.01
                        or abs(proposal["polyline_pts"][0][1] -
                               proposal["polyline_pts"][1][1]) <= 0.01)
                   for proposal in result["channel_proposals"])
           and any("Channel MEASUREMENT not attempted" in flag
                   for flag in result.get("flags", []))
           for result in _raw_external_results.values()
       ))

    _marked_with_real_channel = _pipeline_takeoff_external(
        str(_external_marked_paths[0]), send_approval=False, auto_extract_spec=False)
    ck("real marked Channel linework wins; no assumed proposal is created",
       any(zone.get("category") == "channel" and zone.get("length_lm")
           for zone in _marked_with_real_channel.get("zones", []))
       and not _marked_with_real_channel.get("channel_proposals"),
       {"zones": _marked_with_real_channel.get("zones"),
        "proposals": _marked_with_real_channel.get("channel_proposals")})

    _dock_only_proposals, _dock_only_flags = _takeoff_unmarked_external.propose_channels(
        [], {
            "loading_face_lm": 42.0,
            "loading_face_pts": [[10.0, 20.0], [10.0, 62.0]],
        }, 1.0, scale_verified=True)
    ck("unconfident Yard geometry refuses its run instead of guessing",
       len(_dock_only_proposals) == 1
       and _dock_only_proposals[0]["component"] == "dock_retaining_wall"
       and any("Yard run refused" in flag for flag in _dock_only_flags),
       {"proposals": _dock_only_proposals, "flags": _dock_only_flags})

    _channel_enabled_before = _takeoff_unmarked_external.CHANNEL_PROPOSALS_ENABLED
    try:
        _takeoff_unmarked_external.CHANNEL_PROPOSALS_ENABLED = False
        _disabled_proposals, _disabled_flags = _takeoff_unmarked_external.propose_channels(
            [[0,0],[100,0],[100,100],[0,100]], {
                "loading_face_lm": 42.0,
                "loading_face_pts": [[10.0,20.0],[10.0,62.0]],
            }, 1.0, scale_verified=True)
    finally:
        _takeoff_unmarked_external.CHANNEL_PROPOSALS_ENABLED = _channel_enabled_before
    ck("one gate disables every channel proposal without touching measurement",
       not _disabled_proposals
       and _disabled_flags == [
           "CHANNEL PROPOSALS disabled by CHANNEL_PROPOSALS_ENABLED"],
       {"proposals": _disabled_proposals, "flags": _disabled_flags})

    # Direct takeoff proves the state cap is caused by the non-grey swatch fallback,
    # independently of Unit 3's downstream portal/job handling.
    _unit3_raw = Path("/tmp/ci_external_unit_3_state_cap.pdf")
    try:
        _strip_annotations_external(_external_marked_paths[2], _unit3_raw)
        _unit3_direct = _takeoff_unmarked_external.takeoff(str(_unit3_raw))
        ck("non-grey/white swatch fallback can never reach MEASURED_VERIFIED",
           _unit3_direct.get("region_confidence") == "low" and
           _unit3_direct.get("measurement_state") == "MEASURED_UNVERIFIED" and
           _unit3_direct.get("needs_assessor") is True,
           {"confidence": _unit3_direct.get("region_confidence"),
            "state": _unit3_direct.get("measurement_state")})
    finally:
        try:
            _os_external.remove(_unit3_raw)
        except FileNotFoundError:
            pass

    # An ambiguous native loading-face signal must remain visible in the zone
    # contract and must cap approval; it cannot be silently folded into Yard.
    _ambiguous_raw = Path("/tmp/ci_external_ambiguous_dock.pdf")
    _real_dock_detector = _takeoff_unmarked_external.detect_raw_dock_zone
    try:
        _strip_annotations_external(_external_marked_paths[0], _ambiguous_raw)
        _takeoff_unmarked_external.detect_raw_dock_zone = lambda *args, **kwargs: {
            "zone": None,
            "reason": "multiple plausible loading faces",
            "evidence_seen": True,
        }
        _ambiguous_result = _takeoff_unmarked_external.takeoff(str(_ambiguous_raw))
        ck("ambiguous raw Dock stays unclassified + flagged and cannot be VERIFIED",
           any(zone.get("category") == "unclassified"
               and zone.get("needs_assessor") is True
               for zone in _ambiguous_result.get("zones", []))
           and _ambiguous_result.get("measurement_state") == "MEASURED_UNVERIFIED"
           and any("classify/trace Dock zone" in flag
                   for flag in _ambiguous_result.get("flags", [])),
           {"zones": _ambiguous_result.get("zones"),
            "state": _ambiguous_result.get("measurement_state")})
    finally:
        _takeoff_unmarked_external.detect_raw_dock_zone = _real_dock_detector
        try:
            _os_external.remove(_ambiguous_raw)
        except FileNotFoundError:
            pass
except _FixtureNotPresent as _e:
    print(f"  [SKIP] {_e} — fixture not present")

print("raw multi-region Yard fixture + assessor keep/exclude loop")
try:
    _tanro_marked = Path(
        "drawings/inderjit_markups_31jul/2165 Tanro- Voltage Business Park/Markup/Yard Markup.pdf")
    _require_fixture(_tanro_marked, "Tanro multi-region Yard validation")
    from accuracy_report import strip_annotations as _strip_tanro_annotations
    from takeoff_pipeline import takeoff as _takeoff_tanro_multi
    _tanro_raw = Path("/tmp/ci_tanro_multi_region_stripped.pdf")
    try:
        _strip_tanro_annotations(_tanro_marked, _tanro_raw)
        _tanro_result = _takeoff_tanro_multi(
            str(_tanro_raw), send_approval=False, auto_extract_spec=False)
        _tanro_regions = _tanro_result.get("yard_regions", [])
        ck("Tanro raw drawing retains both distant unit Yards as separate visible regions",
           len(_tanro_regions) == 2 and all(region.get("bbox_pdf_pts")
                                            for region in _tanro_regions),
           [{"area_m2":region.get("area_m2"), "bbox":region.get("bbox_pdf_pts")}
            for region in _tanro_regions])
        ck("Tanro candidate total sums both retained regions within 5% of client truth",
           abs(_tanro_result.get("area_m2", 0) - 369.2) / 369.2 * 100 <= 5 and
           _tanro_result.get("yard_region_review_required") is True,
           {"measured": _tanro_result.get("area_m2"), "truth": 369.2})
    finally:
        try:
            _tanro_raw.unlink()
        except FileNotFoundError:
            pass
except _FixtureNotPresent as _e:
    print(f"  [SKIP] {_e} — fixture not present")

print("assessor Yard-region exclusion endpoint")
import tempfile as _tempfile_yard_review
import approval_server as _AS_yard_review
_yard_review_tmp = Path(_tempfile_yard_review.mkdtemp(prefix="ci_yard_regions_"))
_yard_review_jobs_before = _AS_yard_review.JOBS_FILE
_yard_review_rates_before = _AS_yard_review.CLIENT_RATES_FILE
try:
    _AS_yard_review.JOBS_FILE = _yard_review_tmp / "jobs.json"
    _AS_yard_review.CLIENT_RATES_FILE = _yard_review_tmp / "client_rates.json"
    _yard_review_id = "bbbbbbbb-bbbb-4bbb-8bbb-bbbbbbbbbbbb"
    _yard_review_regions = [
        {"region_id":"yard-region-1", "area_m2":234.2, "included":True,
         "bbox_pdf_pts":[10,10,30,25], "polygon_pts":[[10,10],[30,10],[30,25],[10,25]]},
        {"region_id":"yard-region-2", "area_m2":132.0, "included":True,
         "bbox_pdf_pts":[60,45,72,55], "polygon_pts":[[60,45],[72,45],[72,55],[60,55]]},
    ]
    _yard_review_zones = [{
        "zone_key":"external_yard", "category":"external_yard", "area_m2":366.0,
        "measurement_kind":"area", "needs_assessor":True,
    }]
    _AS_yard_review.save_jobs({_yard_review_id: {
        "id":_yard_review_id, "status":"pending", "decision":None,
        "area_m2":366.0, "measurement_state":"MEASURED_UNVERIFIED",
        "scale_k":0.1, "yard_regions":_yard_review_regions,
        "yard_region_review_required":True, "zones":_yard_review_zones,
        "flags":["YARD REGION REVIEW REQUIRED: synthetic"],
        "result":{"file":"Multi Yard.pdf", "area_m2":366.0, "scale_k":0.1,
                  "measurement_state":"MEASURED_UNVERIFIED",
                  "yard_regions":_yard_review_regions,
                  "yard_region_review_required":True, "zones":_yard_review_zones,
                  "flags":["YARD REGION REVIEW REQUIRED: synthetic"]},
    }})
    _yard_review_client = _AS_yard_review.app.test_client()
    _partial_yard_review = _yard_review_client.post(
        f"/yard-regions/{_yard_review_id}",
        json={"decisions":[{"region_id":"yard-region-1", "action":"keep"}]})
    ck("Yard-region review must cover every retained component, never silently omit one",
       _partial_yard_review.status_code == 409, _partial_yard_review.get_json())
    _complete_yard_review = _yard_review_client.post(
        f"/yard-regions/{_yard_review_id}", json={"decisions":[
            {"region_id":"yard-region-1", "action":"keep"},
            {"region_id":"yard-region-2", "action":"exclude"},
        ]})
    _reviewed_yard_job = _AS_yard_review.load_jobs()[_yard_review_id]
    ck("assessor exclusion removes exactly that component from Yard and its BOQ zone",
       _complete_yard_review.status_code == 200 and
       _complete_yard_review.get_json().get("area_m2") == 234.2 and
       _reviewed_yard_job["area_m2"] == 234.2 and
       _reviewed_yard_job["zones"][0]["area_m2"] == 234.2 and
       _reviewed_yard_job["yard_regions"][1]["included"] is False,
       {"response":_complete_yard_review.get_json(),
        "regions":_reviewed_yard_job.get("yard_regions")})
    ck("completed Yard-region review clears only its gate and preserves four-state review",
       not _reviewed_yard_job.get("yard_region_review_required") and
       "same-tint Yard regions" not in
       (_AS_yard_review._approve_block_reason(_reviewed_yard_job) or "") and
       _reviewed_yard_job.get("measurement_state") == "MEASURED_UNVERIFIED",
       _AS_yard_review._approve_block_reason(_reviewed_yard_job))
finally:
    _AS_yard_review.JOBS_FILE = _yard_review_jobs_before
    _AS_yard_review.CLIENT_RATES_FILE = _yard_review_rates_before
    shutil.rmtree(_yard_review_tmp, ignore_errors=True)

print("channel proposal geometry — retaining-wall adjacent and never diagonal")
from takeoff_unmarked import (
    _longest_contained_yard_run as _channel_yard_run,
    propose_channels as _propose_channels_axis,
)
_channel_rect = [[0,0],[1000,0],[1000,600],[0,600]]
_horizontal_wall = [
    {"polyline_pts":[[0,1],[1000,1]], "grey_wall_evidence":False},
    {"polyline_pts":[[0,1],[1000,1]], "grey_wall_evidence":False},
]
_horizontal_run, _horizontal_refusal = _channel_yard_run(
    _channel_rect, 0.1, wall_segments=_horizontal_wall)
ck("corroborated horizontal retaining-wall run is emitted exactly axis-aligned",
   _horizontal_run is not None and _horizontal_refusal is None and
   _horizontal_run["length_lm"] == 100.0 and
   _horizontal_run["polyline_pts"][0][1] == _horizontal_run["polyline_pts"][1][1],
   (_horizontal_run, _horizontal_refusal))
_vertical_wall = [
    {"polyline_pts":[[1,0],[1,600]], "grey_wall_evidence":True},
]
_vertical_run, _vertical_refusal = _channel_yard_run(
    _channel_rect, 0.1, wall_segments=_vertical_wall)
ck("corroborated vertical retaining-wall run is emitted exactly axis-aligned",
   _vertical_run is not None and _vertical_refusal is None and
   _vertical_run["length_lm"] == 60.0 and
   _vertical_run["polyline_pts"][0][0] == _vertical_run["polyline_pts"][1][0],
   (_vertical_run, _vertical_refusal))
_diagonal_run, _diagonal_refusal = _channel_yard_run(
    _channel_rect, 0.1, wall_segments=[
        {"polyline_pts":[[0,0],[1000,600]], "grey_wall_evidence":True},
        {"polyline_pts":[[0,0],[1000,600]], "grey_wall_evidence":True},
    ])
ck("diagonal-only evidence refuses the Yard component instead of falling back",
   _diagonal_run is None and "non-diagonal" in _diagonal_refusal,
   _diagonal_refusal)
_diagonal_dock_proposals, _diagonal_dock_flags = _propose_channels_axis(
    _channel_rect,
    {"loading_face_lm":100.0, "loading_face_pts":[[0,0],[1000,600]]},
    0.1, wall_segments=_horizontal_wall)
ck("diagonal Dock loading-face evidence refuses the whole assumption, not just the Yard run",
   not _diagonal_dock_proposals and
   any("not a straight non-diagonal" in flag for flag in _diagonal_dock_flags),
   _diagonal_dock_flags)
_ambiguous_wall_run, _ambiguous_wall_refusal = _channel_yard_run(
    _channel_rect, 0.1, wall_segments=[
        {"polyline_pts":[[0,1],[1000,1]], "grey_wall_evidence":False},
        {"polyline_pts":[[0,1],[1000,1]], "grey_wall_evidence":False},
        {"polyline_pts":[[0,599],[970,599]], "grey_wall_evidence":False},
        {"polyline_pts":[[0,599],[970,599]], "grey_wall_evidence":False},
    ])
ck("near-equal competing retaining-wall runs refuse rather than guessing an edge",
   _ambiguous_wall_run is None and "within 5%" in _ambiguous_wall_refusal,
   _ambiguous_wall_refusal)
_no_dock_proposals, _no_dock_flags = _propose_channels_axis(
    _channel_rect, None, 0.1, scale_verified=True,
    dock_presence="absent", wall_segments=_horizontal_wall)
ck("unit with no Dock level gets exactly one full-Yard-width assumed channel",
   len(_no_dock_proposals) == 1 and
   _no_dock_proposals[0]["component"] == "yard_longest_contained_run" and
   _no_dock_proposals[0]["channel_case"] == "no_dock_level_one_run" and
   any("one full-Yard-width" in flag for flag in _no_dock_flags),
   (_no_dock_proposals, _no_dock_flags))
_with_dock_proposals, _with_dock_flags = _propose_channels_axis(
    _channel_rect,
    {"loading_face_lm":60.0, "loading_face_pts":[[1,0],[1,600]]},
    0.1, scale_verified=True, dock_presence="present",
    wall_segments=_horizontal_wall, access_road_interruption=True)
ck("unit with a Dock level gets dock-level + full-width runs and access-road assumption",
   len(_with_dock_proposals) == 2 and
   all(proposal["channel_case"] == "with_dock_level_two_runs"
       for proposal in _with_dock_proposals) and
   _with_dock_proposals[1]["access_road_interruption"] is True and
   any("split/edit" in reason
       for reason in _with_dock_proposals[1]["confidence_reasons"]),
   (_with_dock_proposals, _with_dock_flags))
_ambiguous_dock_proposals, _ambiguous_dock_flags = _propose_channels_axis(
    _channel_rect, None, 0.1, dock_presence="ambiguous",
    wall_segments=_horizontal_wall)
ck("ambiguous Dock presence refuses instead of silently choosing one- or two-run rule",
   not _ambiguous_dock_proposals and
   any("one-channel versus two-channel" in flag for flag in _ambiguous_dock_flags),
   _ambiguous_dock_flags)

try:
    _office_marked_paths = [
        _castle_dir / f"Office Floors Unit-{number}.pdf" for number in range(1, 5)
    ]
    _office_stripped_paths = [
        _castle_dir / "_stripped" / path.name for path in _office_marked_paths
    ]
    for _path in _office_marked_paths + _office_stripped_paths:
        _require_fixture(_path, "Castle Donington stripped-office validation")
    import json as _json_office_gold
    from takeoff_pipeline import takeoff as _pipeline_takeoff_office
    _office_gold = _json_office_gold.loads(Path("gold.json").read_text())
    _outside_gate = []
    for _marked_path, _stripped_path in zip(_office_marked_paths, _office_stripped_paths):
        with _fitz_zones.open(_marked_path) as _marked_doc, _fitz_zones.open(_stripped_path) as _stripped_doc:
            _marked_annots = sum(1 for page in _marked_doc for _ in (page.annots() or []))
            _stripped_annots = sum(1 for page in _stripped_doc for _ in (page.annots() or []))
        ck(f"stripped Office fixture removes every annotation: {_marked_path.name}",
           _marked_annots > 0 and _stripped_annots == 0,
           {"marked": _marked_annots, "stripped": _stripped_annots})

        _assisted = _pipeline_takeoff_office(
            str(_stripped_path), send_approval=False, auto_extract_spec=False)
        _candidate_levels = [
            candidate["level"] for candidate in _assisted.get("candidate_polygons", [])
        ]
        with _fitz_zones.open(_stripped_path) as _office_title_doc:
            from office_candidates import _level_titles as _office_level_titles
            _expected_levels = sorted(
                title["level"] for title in _office_level_titles(_office_title_doc[0]))
        ck(f"every Office level appears exactly once (no duplicate rows): {_marked_path.name}",
           sorted(_candidate_levels) == _expected_levels and
           len(_candidate_levels) == len(set(_candidate_levels)),
           {"expected": _expected_levels, "actual": _candidate_levels})

        _level_areas = {}
        for _candidate in _assisted.get("candidate_polygons", []):
            _candidate_area = 0.0
            _regions = _candidate.get("regions") or [_candidate.get("polygon_pts", [])]
            _region_holes = _candidate.get("region_holes") or [[] for _ in _regions]
            for _region_index, _region in enumerate(_regions):
                if len(_region) < 3:
                    continue
                _region_area, _ = measure_regions(
                    [_region], _assisted["scale_k"],
                    holes={0: _region_holes[_region_index]})
                _candidate_area += _region_area
            _level = _candidate["level"]
            _level_areas[_level] = _candidate_area
        _detected_total = round(sum(_level_areas.values()), 2)
        _markup_total = _read_marked_zones(str(_marked_path))["area_m2"]
        _boq_total = _office_gold[str(_marked_path)]["net_m2"]
        _delta_pct = (_detected_total - _markup_total) / _markup_total * 100
        _outside_gate.append(abs(_delta_pct) > 5)
        ck(f"Office candidates remain assisted geometry only: {_marked_path.name}",
           _assisted.get("area_m2") is None and
           _assisted.get("measurement_state") == "UNMEASURED" and
           _assisted.get("needs_assessor") is True and
           _assisted.get("costing") is None and _assisted.get("polygon_pts") is None and
           bool(_assisted.get("candidate_polygons")) and
           all("area_m2" not in candidate
               for candidate in _assisted.get("candidate_polygons", [])),
           {"diagnostic_candidate_total": _detected_total, "markup": _markup_total,
            "boq": _boq_total, "delta_pct": round(_delta_pct, 2)})
    ck("Office auto-measure bar remains closed unless every unit is within 5%",
       len(_outside_gate) == 4 and any(_outside_gate), _outside_gate)
except _FixtureNotPresent as _e:
    print(f"  [SKIP] {_e} — fixture not present")

print("pipeline price_with_defaults")
import contextlib, io as _io
with contextlib.redirect_stdout(_io.StringIO()):
    from takeoff_pipeline import price_with_defaults, _needs_approval
_c = price_with_defaults(26080)
ck("26,080 m² at defaults -> £1,175,425.60", _c["total_gbp"] == 1175425.60)
ck("price_with_defaults assumed=True (no spec)", _c["assumed"] is True)
_c2 = price_with_defaults(3172, {"depth_mm": 200, "mesh": "A393", "layers": 1,
                                  "conc_mix": "C32/40", "conc_rate": 128})
ck("all four client construction fields supplied -> assumed=False", _c2["assumed"] is False)
_c3 = price_with_defaults(3172, {"depth_mm": 200})
ck("partial client construction spec stays assumed/provisional", _c3["assumed"] is True)
ck("approval trigger on assessor flag",
   _needs_approval({"type":"UNMARKED vector","confidence":"medium",
                    "flags":["assessor: confirm extent + scale"]}))
ck("no approval trigger on clean marked",
   not _needs_approval({"type":"MARKED vector","confidence":"high","flags":[]}))

print("client-editable rate override layer — defaults untouched, versioned/audited, quotation-stamped")
try:
    import hashlib as _hashlib_rates
    import tempfile as _tempfile_rates
    import json as _json_rates
    from client_rates import (apply_client_rates as _apply_client_rates_test,
                              load_rate_store as _load_rate_store_test,
                              save_client_rates as _save_client_rates_test)
    from defaults import DEFAULT_SPEC as _DEFAULT_SPEC_RATES
    from takeoff_pipeline import MANHOLE_EO_RATE as _MANHOLE_RATE_DEFAULT

    _rates_tmpdir = Path(_tempfile_rates.mkdtemp(prefix="ci_client_rates_"))
    _rates_path = _rates_tmpdir / "client_rates.json"
    _rate_defaults = {
        key: _DEFAULT_SPEC_RATES[key]
        for key in ("conc_rate", "steel_rate_t", "margin", "labour", "dpm", "curing",
                    "trim", "conc_wastage", "steel_wastage", "lap_acc")
    }
    _rate_defaults["manhole_eo_rate"] = _MANHOLE_RATE_DEFAULT
    try:
        _legacy_costing = price_with_defaults(26080, client_rates_path=_rates_path)
        _legacy_bytes = _json_rates.dumps(
            _legacy_costing, sort_keys=True, separators=(",", ":"), ensure_ascii=False
        ).encode()
        ck("absent client_rates.json keeps the complete legacy costing byte-identical",
           _hashlib_rates.sha256(_legacy_bytes).hexdigest() ==
           "38194af48023162689f095a4372d6293e7d4d0db87a71415b703ed576d0bab50" and
           "rates_version" not in _legacy_costing and
           "client_rates_applied" not in _legacy_costing)

        _edited_concrete = _rate_defaults["conc_rate"] * 1.01
        _edited_manhole = _rate_defaults["manhole_eo_rate"] * 1.01
        _saved_rates, _rate_changes = _save_client_rates_test(
            {"conc_rate": _edited_concrete, "manhole_eo_rate": _edited_manhole},
            _rate_defaults, path=_rates_path, who="assessor-token-authenticated",
            when="2026-07-18T10:00:00+00:00",
        )
        ck("first client-rates save creates version 1 and one audit row per changed field",
           _saved_rates["version"] == 1 and len(_saved_rates["audit"]) == 2 and
           {entry["field"] for entry in _saved_rates["audit"]} ==
           {"conc_rate", "manhole_eo_rate"})
        _audit_concrete = next(entry for entry in _saved_rates["audit"]
                               if entry["field"] == "conc_rate")
        ck("client-rates audit records authenticated assessor + exact old -> new",
           _audit_concrete["who"] == "assessor-token-authenticated" and
           _audit_concrete["old"] == _rate_defaults["conc_rate"] and
           _audit_concrete["new"] == _edited_concrete)

        _overridden_costing = price_with_defaults(
            26080, manhole_count=2, client_rates_path=_rates_path)
        ck("override changes only fresh pricing and carries rates provenance",
           _overridden_costing["rate"] != _legacy_costing["rate"] and
           _overridden_costing["total_gbp"] != _legacy_costing["total_gbp"] and
           _overridden_costing["rates_version"] == 1 and
           _overridden_costing["client_rates_applied"] is True)
        ck("manhole E/O uses the client override without changing its built-in rate",
           _overridden_costing["extras"][0]["rate"] == _edited_manhole and
           _MANHOLE_RATE_DEFAULT == _rate_defaults["manhole_eo_rate"])

        _rates_quote_result = {
            "file": "Fresh-Rates.pdf", "area_m2": 26080,
            "quotation_section": "External yard slabs",
            "costing": _overridden_costing, "flags": [],
        }
        _rates_quote = generate_quotation(
            _rates_quote_result, project="Rates Test", client="Fortel", ref="RATE-001")
        ck("fresh quotation records rates_version + CLIENT-EDITED provenance declaration",
           _rates_quote.get("rates_version") == 1 and
           _rates_quote.get("client_rates_applied") is True and
           any("CLIENT-EDITED RATES" in note and "version 1" in note
               for note in _rates_quote["declarations"]))
        ck("client-rate provenance is visible in text, HTML and JSON quotation outputs",
           all("CLIENT-EDITED RATES" in output for output in (
               quotation_text(_rates_quote), quotation_html(_rates_quote),
               quotation_json(_rates_quote))))
        _rates_book = _load_workbook(
            _BytesIO(quotation_xlsx(_rates_quote)), data_only=False)["REV_01"]
        ck("xlsx header records the applied client-rates version",
           _rates_book["D1"].value == "Client-edited rates version: 1")

        _rates_path.unlink()
        _restored_costing = price_with_defaults(26080, client_rates_path=_rates_path)
        _restored_bytes = _json_rates.dumps(
            _restored_costing, sort_keys=True, separators=(",", ":"), ensure_ascii=False
        ).encode()
        ck("removing client_rates.json restores the exact default output",
           _restored_bytes == _legacy_bytes and not _rates_path.exists())
    finally:
        shutil.rmtree(_rates_tmpdir, ignore_errors=True)
except (ImportError, OSError, ValueError) as _e:
    ck("client-editable rate override layer imports and runs", False, _e)

print("spec extractor — supplier fields")
from spec_extractor import extract_spec_from_text
_s5 = extract_spec_from_text("20mm crushed aggregate, 0.45 w/c ratio, S3 slump, air-entrained")
ck("aggregate 20mm extracted",  _s5.get("aggregate_mm") == 20)
ck("wc_ratio 0.45 extracted",   abs(_s5.get("wc_ratio", 0) - 0.45) < 0.001)
ck("slump S3 extracted",        _s5.get("slump_class") == "S3")
ck("air_entrained extracted",   _s5.get("air_entrained") is True)
_s6 = extract_spec_from_text("12mm aggregate 0.50 w/c S4 slump class CEM I")
ck("aggregate 12mm extracted",  _s6.get("aggregate_mm") == 12)
ck("slump S4 extracted",        _s6.get("slump_class") == "S4")

print("supplier inquiry generator")
from supplier_inquiry import generate_inquiry, format_cubes
ck("cubes calc 26080m² 190mm 3%",
   format_cubes(26080, 190, 0.03) == round(26080 * 0.190 * 1.03, 1))
_demo = {
    "area_m2": 3172, "project_name": "Test Project", "project_ref": "2132",
    "costing": {"spec": {"depth_mm": 190, "conc_mix": "C32/40", "cement_type": "CEM I",
                          "air_entrained": True, "aggregate_mm": 20, "wc_ratio": 0.45,
                          "slump_class": "S3", "conc_wastage": 0.03}}
}
_inq = generate_inquiry(_demo)
ck("inquiry has subject",        bool(_inq["subject"]))
ck("inquiry subject has mix",    "C32/40" in _inq["subject"])
ck("inquiry subject has cubes",  "m³" in _inq["subject"])
ck("inquiry text has slump",     "S3" in _inq["text"])
ck("inquiry text has aggregate", "20 mm" in _inq["text"])
ck("inquiry text has wc",        "0.45" in _inq["text"])
ck("inquiry html starts DOCTYPE","<!DOCTYPE" in _inq["html"])
ck("inquiry cubes > 0",          _inq["cubes_m3"] > 0)
_inq_no_proj = generate_inquiry({"area_m2": 1000, "costing": {}})
ck("inquiry works with no project info", bool(_inq_no_proj["subject"]))

print("measurement state machine (sanity.measurement_state)")
from sanity import measurement_state, MEASURED_VERIFIED, MEASURED_UNVERIFIED, UNMEASURED, REJECTED
ck("verified + high conf -> MEASURED_VERIFIED",
   measurement_state(26080, scale_verified=True, confidence="high")[0] == MEASURED_VERIFIED)
ck("unverified scale -> MEASURED_UNVERIFIED",
   measurement_state(26080, scale_verified=False)[0] == MEASURED_UNVERIFIED)
ck("low confidence -> MEASURED_UNVERIFIED",
   measurement_state(100, scale_verified=True, confidence="low")[0] == MEASURED_UNVERIFIED)
ck("implausible area -> MEASURED_UNVERIFIED (blocks pricing AND approval)",
   measurement_state(95463, site_m2=34329)[0] == MEASURED_UNVERIFIED)
ck("over single-zone bound -> MEASURED_UNVERIFIED",
   measurement_state(70000, scale_verified=True, confidence="high")[0] == MEASURED_UNVERIFIED)
ck("no area -> UNMEASURED", measurement_state(None)[0] == UNMEASURED)
ck("rejected_reason short-circuits -> REJECTED",
   measurement_state(26080, scale_verified=True, rejected_reason="encrypted PDF")[0] == REJECTED)
ck("plausible + verified -> no flags", measurement_state(26080, scale_verified=True, confidence="high")[1] == [])

print("multi-page routing (router.rank_pages / classify_page)")
try:
    from reportlab.pdfgen import canvas as _canvas
    import fitz as _fitz
    from router import rank_pages, classify_page, drawing_priority
    _d = _fitz.open()
    _p0 = _d.new_page(width=1000, height=1000)
    _p0.insert_text((50, 50), "Proposed Site Plan")
    for _i in range(5):
        _p0.draw_line((100 + _i, 100), (100 + _i, 200))     # < 50 vector paths -> low priority / raster-ish
    _p1 = _d.new_page(width=1000, height=1000)
    _p1.insert_text((50, 50), "External Construction Thickness Layout")
    for _i in range(60):
        _p1.draw_line((100 + _i, 300), (100 + _i, 400))     # >= 50 vector paths -> UNMARKED vector
    _d.save("/tmp/_ci_rank_test.pdf")

    _ranked = rank_pages("/tmp/_ci_rank_test.pdf")
    ck("rank_pages classifies every page", len(_ranked) == 2)
    ck("rank_pages best candidate is page 1 (construction-thickness beats site plan)",
       _ranked[0]["page"] == 1)
    ck("rank_pages best candidate score > runner-up", _ranked[0]["score"] > _ranked[1]["score"])
    ck("classify_page(1) matches rank_pages page-1 type",
       classify_page("/tmp/_ci_rank_test.pdf", 1)[0] == _ranked[0]["type"])
    ck("classify_page(0) is page-0-only (never assumes the whole doc)",
       classify_page("/tmp/_ci_rank_test.pdf", 0)[0] == "RASTER / scanned")
except ImportError as _e:
    print(f"  [SKIP] router multi-page tests — missing dependency: {_e}")

print("pipeline multi-page + raster UNMEASURED (takeoff_pipeline.takeoff)")
try:
    import os as _os
    _os.environ["SKIP_APPROVAL_LOG"] = "1"
    with contextlib.redirect_stdout(_io.StringIO()):
        from takeoff_pipeline import takeoff as _pipeline_takeoff
    import fitz as _fitz2

    # Multi-page MARKED pack: page 0 is a decoy site plan, page 1 has the priced markup.
    # takeoff() must measure page 1, not silently default to page 0.
    _mp = _fitz2.open()
    _mp0 = _mp.new_page(width=1400, height=2200)
    _mp0.insert_text((100, 100), "Proposed Site Plan")
    for _i in range(5):
        _mp0.draw_line((100 + _i, 300), (100 + _i, 400))
    _mp1 = _mp.new_page(width=1400, height=2200)
    _mp1.insert_text((100, 100), "External Construction Thickness Layout")
    for _i in range(60):
        _mp1.draw_line((100 + _i, 300), (100 + _i, 400))
    _annot = _mp1.add_polygon_annot([(100, 100), (600, 100), (600, 500), (100, 500)])
    _annot.set_info(content="Area = 3000.0 sq m")
    _annot.update()
    _mp.save("/tmp/_ci_pipeline_multipage.pdf")

    with contextlib.redirect_stdout(_io.StringIO()):
        _rmp = _pipeline_takeoff("/tmp/_ci_pipeline_multipage.pdf")
    ck("multi-page pipeline measures the ranked page, not page 0", _rmp.get("page") == 1)
    ck("multi-page pipeline area comes from page 1's markup", _rmp.get("area_m2") == 3000.0)
    ck("multi-page pipeline flags which page was chosen",
       any("MULTI-PAGE" in f and "page 1 of 2" in f for f in _rmp.get("flags", [])))
    ck("multi-page pipeline lists the other candidate page",
       any("other candidates" in f for f in _rmp.get("flags", [])))

    # Single-page MARKED vector -> MEASURED_VERIFIED, matches the four-state contract.
    _sp = _fitz2.open()
    _spp = _sp.new_page(width=1400, height=2200)
    for _i in range(60):
        _spp.draw_line((100 + _i, 100), (100 + _i, 200))
    _sannot = _spp.add_polygon_annot([(100, 100), (600, 100), (600, 500), (100, 500)])
    _sannot.set_info(content="Area = 2000.0 sq m")
    _sannot.update()
    _sp.save("/tmp/_ci_pipeline_marked.pdf")
    with contextlib.redirect_stdout(_io.StringIO()):
        _rsp = _pipeline_takeoff("/tmp/_ci_pipeline_marked.pdf")
    ck("MARKED vector -> MEASURED_VERIFIED", _rsp.get("measurement_state") == MEASURED_VERIFIED)
    ck("MARKED vector -> status mirrors measurement_state", _rsp.get("status") == MEASURED_VERIFIED)
    ck("MARKED vector -> needs_assessor False", _rsp.get("needs_assessor") is False)

    # RASTER/scanned (few vector paths, e.g. a scanned/flattened sheet) -> proper UNMEASURED
    # job, never a bare flag-only stub. area_m2 stays None; needs_assessor True.
    _rast = _fitz2.open()
    _rp = _rast.new_page(width=1400, height=2200)
    _rp.insert_text((100, 100), "Scanned Site Photo")   # < 50 vector paths -> RASTER / scanned
    _rast.save("/tmp/_ci_pipeline_raster.pdf")
    with contextlib.redirect_stdout(_io.StringIO()):
        _rr = _pipeline_takeoff("/tmp/_ci_pipeline_raster.pdf")
    ck("raster drawing -> area_m2 stays None", _rr.get("area_m2") is None)
    ck("raster drawing -> UNMEASURED (not a crash, not a bare flag)",
       _rr.get("measurement_state") == UNMEASURED)
    ck("raster drawing -> needs_assessor True", _rr.get("needs_assessor") is True)
    ck("raster drawing -> flag explains mandatory assessor trace",
       any("mandatory assessor trace" in f.lower() or "UNMEASURED" in f for f in _rr.get("flags", [])))
except ImportError as _e:
    print(f"  [SKIP] pipeline multi-page/raster tests — missing dependency: {_e}")

print("D77 accuracy invariant (measurement math unchanged)")
try:
    _require_fixture("drawings/_int_d77.pdf", "D77 accuracy test")
    from takeoff_unmarked import takeoff as _tu_takeoff
    _d77 = _tu_takeoff("drawings/_int_d77.pdf")
    ck("D77 area unchanged at 3,159 m² (Smita gold 3,156)", _d77.get("area_m2") == 3159.0)
    ck("D77 scale verified True (bar agrees with title via scale_consensus)",
       _d77.get("scale_verified") is True)
    ck("D77 measurement_state MEASURED_VERIFIED", _d77.get("measurement_state") == MEASURED_VERIFIED)
    ck("D77 needs_assessor False", _d77.get("needs_assessor") is False)
    ck("D77 VERIFIED state has independent native-boundary extent corroboration",
       _d77.get("extent_corroborated") is True and
       all(region.get("perimeter_confidence") == "high"
           for region in _d77.get("yard_regions", []) if region.get("included")),
       {"extent_corroborated": _d77.get("extent_corroborated"),
        "regions": _d77.get("yard_regions")})

    # Corpus-backed confidence regression for the live 575 vs 1,212 failure class. Keep the
    # real D77 pixels, legend and verified scale, but remove only the independent native-boundary
    # corroboration. The candidate area must remain unchanged while VERIFIED is withheld.
    import takeoff_unmarked as _tu_extent_guard
    _real_native_drawings_extent = _tu_extent_guard._native_boundary_drawings
    try:
        _tu_extent_guard._native_boundary_drawings = lambda page: (
            [], "regression fixture: native extent unavailable")
        _d77_uncorroborated = _tu_extent_guard.takeoff("drawings/_int_d77.pdf")
        from takeoff_pipeline import takeoff as _pipeline_extent_guard
        _d77_uncorroborated_pipeline = _pipeline_extent_guard(
            "drawings/_int_d77.pdf", send_approval=False, auto_extract_spec=False)
    finally:
        _tu_extent_guard._native_boundary_drawings = _real_native_drawings_extent
    ck("uncorroborated raw extent preserves the measured candidate number",
       _d77_uncorroborated.get("area_m2") == _d77.get("area_m2") == 3159.0,
       {"corroborated": _d77.get("area_m2"),
        "uncorroborated": _d77_uncorroborated.get("area_m2")})
    ck("verified scale + legend alone cannot promote a partial raw extent to VERIFIED",
       _d77_uncorroborated.get("scale_verified") is True and
       _d77_uncorroborated.get("extent_corroborated") is False and
       _d77_uncorroborated.get("measurement_state") == MEASURED_UNVERIFIED and
       _d77_uncorroborated.get("needs_assessor") is True and
       any("YARD EXTENT UNCORROBORATED" in flag
           for flag in _d77_uncorroborated.get("flags", [])),
       {"state": _d77_uncorroborated.get("measurement_state"),
        "flags": _d77_uncorroborated.get("flags")})
    ck("pipeline preserves the raw extent confidence cap",
       _d77_uncorroborated_pipeline.get("extent_corroborated") is False and
       _d77_uncorroborated_pipeline.get("measurement_state") == MEASURED_UNVERIFIED and
       _d77_uncorroborated_pipeline.get("needs_assessor") is True,
       {"state": _d77_uncorroborated_pipeline.get("measurement_state"),
        "extent": _d77_uncorroborated_pipeline.get("extent_corroborated")})
except _FixtureNotPresent as _e:
    print(f"  [SKIP] {_e} — fixture not present")
except (ImportError, FileNotFoundError) as _e:
    print(f"  [SKIP] D77 accuracy test — missing dependency or file: {_e}")

print("D77 border/legend exclusion (Aryan field report: real SGP sheet over-measures by "
      "border strips + legend swatch that share the yard's grey)")
try:
    _require_fixture("drawings", "D77 border/legend exclusion test")
    import fitz as _fitz_b
    from takeoff_unmarked import takeoff as _tu_takeoff2, segment_hatch as _seg_b

    def _gen_d77_borders(out_path, with_borders):
        """Rebuild _int_d77.pdf's exact yard rect + scale bar (same geometry, so the
        measured area is directly comparable), optionally adding:
          - a grey sheet-frame border strip running around the full page edge
            (same fill colour as the yard hatch — this is what a real SGP sheet's
            outer frame line looks like when rendered to raster and colour-segmented)
          - a small grey legend swatch rectangle near the title block (isolated,
            far from the yard, same grey) — mimics a legend colour chip.
        WITHOUT the fix these must inflate the measured area; WITH the fix
        (segment_hatch exclude_border=True, default) the result must match
        plain _int_d77.pdf (3,159 m²) within 0.5%.
        """
        d = _fitz_b.open()
        W, H = 1067.7659912109375, 824.853515625
        pg = d.new_page(width=W, height=H)
        GREY = (0.84, 0.84, 0.84)
        # Same yard rect + scale bar + title text as drawings/_int_d77.pdf
        pg.draw_rect(_fitz_b.Rect(130.0, 120.0, 937.765625, 624.853515625),
                     color=(0, 0, 0), fill=GREY, width=1.0)
        pg.draw_line(_fitz_b.Point(130.0, 714.853515625), _fitz_b.Point(696.9290771484375, 714.853515625),
                     color=(0, 0, 0), width=2.0)
        pg.insert_text((130.0, 80), "PROPOSED HARD LANDSCAPING - CONCRETE SERVICE YARD    Scale 1:250",
                       fontsize=13)
        pg.insert_text((126.0, 731), "0", fontsize=11)
        pg.insert_text((678.9, 731), "50 m", fontsize=11)
        if with_borders:
            # Sheet-frame border strip: four thin grey rects running along the outer page
            # edge (inside the outer ~1% margin), same grey as the yard hatch — drawn as
            # separate strips (not a filled rect + white hole) so they don't cover other
            # content or perturb the solid-fill drawing-style heuristic.
            bw = 6  # strip thickness in pt
            m = 4   # inset from the physical page edge
            for r in (
                _fitz_b.Rect(m, m, W - m, m + bw),                 # top
                _fitz_b.Rect(m, H - m - bw, W - m, H - m),          # bottom
                _fitz_b.Rect(m, m, m + bw, H - m),                  # left
                _fitz_b.Rect(W - m - bw, m, W - m, H - m),          # right
            ):
                pg.draw_rect(r, color=None, fill=GREY, width=0)
            # Thin grey bridging tail connecting the left border strip to the yard rect's
            # own left edge — this reproduces the real failure mode Aryan found: a border
            # line that runs close enough to the yard boundary that binary_closing (kernel
            # size 9) fuses it into the SAME connected component as the yard hatch, directly
            # inflating the measured area rather than appearing as an isolated, easily-
            # skipped satellite blob. A frame that stays fully isolated out in the margin is
            # already handled by the pre-existing best-plausible-component selection, so it
            # alone would not exercise this fix.
            pg.draw_rect(_fitz_b.Rect(m + bw, 300, 130, 306), color=None, fill=GREY, width=0)
            # Legend colour swatch: small isolated grey chip near the title block, far
            # from the yard polygon (same grey, small — a real legend colour key patch).
            pg.draw_rect(_fitz_b.Rect(950, 760, 966, 776), color=(0, 0, 0), fill=GREY, width=0.5)
        d.save(out_path)
        d.close()

    _p_plain = "/tmp/_ci_d77_borders_plain.pdf"
    _p_bord = "drawings/_int_d77_borders.pdf"
    _gen_d77_borders(_p_plain, with_borders=False)
    _gen_d77_borders(_p_bord, with_borders=True)

    # Sanity: the regenerated plain fixture reproduces the real _int_d77.pdf's area.
    _r_plain = _tu_takeoff2(_p_plain)
    ck("regenerated D77 fixture matches real _int_d77.pdf area (3,159 m²)",
       _r_plain.get("area_m2") == 3159.0, f"got {_r_plain.get('area_m2')}")

    # WITHOUT the exclusion: border pixels (frame touches the mask + legend swatch)
    # must inflate the measured area if segmented with exclude_border=False.
    import numpy as _np_b
    from PIL import Image as _Image_b
    _pgb = _fitz_b.open(_p_bord)[0]
    _pixb = _pgb.get_pixmap(matrix=_fitz_b.Matrix(2.0, 2.0))
    _imb = _np_b.frombuffer(_pixb.samples, _np_b.uint8).reshape(_pixb.height, _pixb.width, _pixb.n)[..., :3]
    _GREY_RGB = (214, 214, 214)
    _k77 = 0.08819   # same k as D77 (1:250, verified)
    _comp_noex = _seg_b(_imb, _GREY_RGB, k=_k77, S=2.0, exclude_border=False)
    _area_noex = round(int(_comp_noex.sum()) * (1 / 2.0) ** 2 * _k77 * _k77, 0)
    ck("WITHOUT exclusion: borders+legend over-measure vs plain 3,159 m²",
       _area_noex > 3159.0 + 15, f"got {_area_noex}")

    # WITH the exclusion (default path, via full takeoff()): must land back on 3,159 ± 0.5%.
    _r_bord = _tu_takeoff2(_p_bord)
    _area_bord = _r_bord.get("area_m2")
    ck("WITH exclusion: _int_d77_borders.pdf area back to 3,159 m² (±0.5%)",
       _area_bord is not None and abs(_area_bord - 3159.0) / 3159.0 <= 0.005,
       f"got {_area_bord}")
    ck("WITH exclusion: flag lists excluded border/legend components",
       any("excluded" in f and "border/legend" in f for f in _r_bord.get("flags", [])),
       _r_bord.get("flags"))
except _FixtureNotPresent as _e:
    print(f"  [SKIP] {_e} — fixture not present")
except (ImportError, FileNotFoundError) as _e:
    print(f"  [SKIP] D77 border/legend exclusion test — missing dependency or file: {_e}")

print("manhole counting — MARKED path (robust_takeoff.count_manholes_marked)")
try:
    _require_fixture("drawings", "manhole counting (marked path) test")
    import fitz as _fitz_mh
    from robust_takeoff import read_marked as _read_marked_mh, count_manholes_marked

    def _gen_synthetic_yard(out_path, n_manholes=26):
        """drawings/synthetic_yard.pdf: the gold.json 'synthetic_yard.pdf' fixture —
        a yard boundary labelled with its NET area (25,920 sq m — gross 26,080 minus a
        160 m² void, mirroring how gold.json tracks gross_m2/void_m2/net_m2 for this
        fixture and how a real Bluebeam net-area markup states the final net figure
        directly on the polygon, not gross+void as two separate summed entries) plus
        n_manholes Circle annots scattered inside (Fortel's manhole-marker convention
        on the MARKED path). read_marked() sums Polygon-labelled areas, so a single
        polygon labelled with the net figure reproduces net_m2 exactly.
        """
        d = _fitz_mh.open()
        pg = d.new_page(width=1800, height=1800)
        ox, oy = 50, 50
        W, H = 1630, 1600
        # router.classify() gates MARKED-vs-RASTER on vector path count (vec >= 50); a plain
        # annot-only PDF has 0 page-content vector paths and would misclassify as RASTER.
        # Draw the actual yard boundary + a filler grid as real vector lines (matching how
        # ci_tests.py's own multi-page/marked fixtures push vec >= 50) so this fixture
        # classifies as MARKED vector like a real Bluebeam-marked drawing does.
        pg.draw_rect(_fitz_mh.Rect(ox, oy, ox + W, oy + H), color=(0, 0, 0), width=1.5)
        for i in range(60):
            pg.draw_line(_fitz_mh.Point(ox + 10 + i, oy + H + 40), _fitz_mh.Point(ox + 10 + i, oy + H + 140))
        # Yard boundary polygon, labelled with the NET area (gross 26,080 - void 160).
        outer = [(ox, oy), (ox + W, oy), (ox + W, oy + H), (ox, oy + H)]
        a = pg.add_polygon_annot(outer)
        a.set_info(content="L = 6,460.0 m\rA = 25,920.0 sq m")
        a.update()
        # A drawn (non-annotated) void rectangle purely for visual/context completeness —
        # NOT a separate Polygon annot, so read_marked (which sums every Polygon annot's
        # labelled area) doesn't double count it against the net figure above.
        pg.draw_rect(_fitz_mh.Rect(ox + 700, oy + 700, ox + 800, oy + 860), color=(0.5, 0.5, 0.5), width=1)
        # 26 manhole markers: small Circle annots scattered on a grid inside the yard,
        # avoiding the void rectangle.
        placed = 0
        gx, gy = 0, 0
        cols = 6
        while placed < n_manholes:
            cx = ox + 120 + (gx % cols) * 260
            cy = oy + 120 + gy * 260
            if not (ox + 680 <= cx <= ox + 820 and oy + 680 <= cy <= oy + 880):
                c = pg.add_circle_annot(_fitz_mh.Rect(cx - 6, cy - 6, cx + 6, cy + 6))
                c.set_info(content="MH")
                c.update()
                placed += 1
            gx += 1
            if gx % cols == 0:
                gy += 1
        d.save(out_path)
        d.close()

    _p_synth = "drawings/synthetic_yard.pdf"
    _gen_synthetic_yard(_p_synth, n_manholes=26)

    _area_synth, _n_regions = _read_marked_mh(_p_synth)
    ck("synthetic_yard net area == gold net_m2 (25,920 = 26,080 gross - 160 void)",
       _area_synth == 25920.0, f"got {_area_synth}")
    _mh_count = count_manholes_marked(_p_synth)
    ck("synthetic_yard manhole_count (Circle annots) == 26 (gold marker_count/manhole_count)",
       _mh_count == 26, f"got {_mh_count}")

    # A drawing with no Circle annots at all -> 0, not a crash.
    _d_nomh = _fitz_mh.open(); _p_nomh = _d_nomh.new_page()
    _a_nomh = _p_nomh.add_polygon_annot([(10, 10), (100, 10), (100, 100), (10, 100)])
    _a_nomh.set_info(content="A = 100 sq m"); _a_nomh.update()
    _d_nomh.save("/tmp/_ci_no_manholes.pdf"); _d_nomh.close()
    ck("no Circle annots -> manhole_count 0 (not a crash)",
       count_manholes_marked("/tmp/_ci_no_manholes.pdf") == 0)

    # Real Winvic marked yard PDF: as shipped in this repo it carries NO Circle annots
    # (Fortel has not yet placed manhole markers on it — confirmed by direct inspection;
    # its 18 Square annots are AutoCAD SHX Text bounding boxes for street names/numbers,
    # not manhole markers). count_manholes_marked must report that honestly (0), never
    # fabricate the Winvic costing sheet's "26 Nr" figure from thin air.
    _winvic_yard = "drawings/winvic/Yard_Area_Proposed_Site_Plan.pdf"
    if Path(_winvic_yard).is_file():
        ck("real Winvic yard PDF has 0 Circle annots today (no markers placed yet -> honest 0, "
           "not a fabricated 26)", count_manholes_marked(_winvic_yard) == 0)
    else:
        print("  [SKIP] real Winvic yard manhole regression — fixture not present")
except _FixtureNotPresent as _e:
    print(f"  [SKIP] {_e} — fixture not present")
except (ImportError, FileNotFoundError) as _e:
    print(f"  [SKIP] manhole counting (marked path) test — missing dependency or file: {_e}")

print("manhole counting — UNMARKED path (takeoff_unmarked.detect_manholes, conservative ESTIMATE)")
try:
    _require_fixture("drawings/_int_d77.pdf", "manhole counting (unmarked path) D77 test")
    import numpy as _np_mh, cv2 as _cv2_mh
    from takeoff_unmarked import detect_manholes, takeoff as _tu_takeoff3

    def _gen_yard_with_circles(n_circles, diam_m, k=0.05, S=2.0):
        """A grey yard rect rendered directly as a numpy image (no PDF round-trip needed —
        detect_manholes takes the rendered array + mask + k directly), with n_circles dark
        rings drawn inside at diam_m real-world diameter, converted to px via k/S."""
        H_px, W_px = 900, 1200
        im = _np_mh.full((H_px, W_px, 3), 255, _np_mh.uint8)
        im[100:800, 100:1100] = (214, 214, 214)   # yard hatch
        r_px = int(round((diam_m / 2) * (S / k)))
        centres = []
        cols = 6
        for i in range(n_circles):
            cx = 200 + (i % cols) * 150
            cy = 200 + (i // cols) * 150
            _cv2_mh.circle(im, (cx, cy), r_px, (60, 60, 60), thickness=2)
            centres.append((cx, cy))
        comp = _np_mh.zeros((H_px, W_px), bool)
        comp[100:800, 100:1100] = True
        return im, comp, centres

    # 6 manhole-sized circles (0.9 m diameter, mid-band) inside the yard -> detector finds them.
    _im_mh, _comp_mh, _true_centres = _gen_yard_with_circles(6, diam_m=0.9, k=0.05, S=2.0)
    _n_mh, _found_centres = detect_manholes(_im_mh, _comp_mh, k=0.05, S=2.0)
    ck("detect_manholes finds manhole-sized circles inside the yard (>=4 of 6)", _n_mh >= 4,
       f"found {_n_mh}")

    # No circles at all -> 0, not a crash (D77-style plain rect).
    _im_none = _np_mh.full((400, 400, 3), 255, _np_mh.uint8); _im_none[50:350, 50:350] = (214, 214, 214)
    _comp_none = _np_mh.zeros((400, 400), bool); _comp_none[50:350, 50:350] = True
    _n_none, _ = detect_manholes(_im_none, _comp_none, k=0.05, S=2.0)
    ck("no circular features -> manhole_count_estimate 0 (not a crash)", _n_none == 0)

    # Oversized circles (e.g. 6 m diameter — a roundabout/planter, not a manhole) must NOT
    # be counted: the radius band excludes anything outside MANHOLE_DIAM_M_MIN..MAX.
    _im_big, _comp_big, _ = _gen_yard_with_circles(2, diam_m=6.0, k=0.05, S=2.0)
    _n_big, _ = detect_manholes(_im_big, _comp_big, k=0.05, S=2.0)
    ck("oversized circles (6 m dia, not manhole-sized) excluded by radius band", _n_big == 0,
       f"found {_n_big}")

    # End-to-end: D77 (plain rect, no circular features) -> manhole_count_estimate present,
    # zero, and no false "confirm" flag fired when there's nothing to confirm.
    _d77_mh = _tu_takeoff3("drawings/_int_d77.pdf")
    ck("D77 takeoff() carries manhole_count_estimate field", "manhole_count_estimate" in _d77_mh)
    ck("D77 manhole_count_estimate is 0 (plain rect, no circular features)",
       _d77_mh.get("manhole_count_estimate") == 0)
    # Inderjit's rule (last Fortel call): no drainage layout / no drawn symbols -> ASSUME 1 per
    # 1,000 m². D77 measures ~3,159 m² with a legend label, so the assumed count = round(3159/1000)
    # = 3. It's a SEPARATE field (never manhole_count_estimate, which auto-prices) so it never
    # feeds the £75/Nr E/O line automatically — the assessor confirms first.
    ck("D77 takeoff() carries manhole_count_assumed field", "manhole_count_assumed" in _d77_mh)
    ck("D77 manhole_count_assumed == round(area/1000), floor 1 (Inderjit's 1-per-1,000 rule)",
       _d77_mh.get("manhole_count_assumed") == max(1, round((_d77_mh.get("area_m2") or 0) / 1000.0)),
       f"assumed={_d77_mh.get('manhole_count_assumed')} area={_d77_mh.get('area_m2')}")
    ck("D77 manhole_count_assumed is 3 for the ~3,159 m² fixture",
       _d77_mh.get("manhole_count_assumed") == 3, f"got {_d77_mh.get('manhole_count_assumed')}")
except _FixtureNotPresent as _e:
    print(f"  [SKIP] {_e} — fixture not present")
except (ImportError, FileNotFoundError) as _e:
    print(f"  [SKIP] manhole counting (unmarked path) test — missing dependency or file: {_e}")

print("refuse-instead-of-guess guard — non-slab sheets must REFUSE, not emit a garbage area")
try:
    import os as _os_rg
    import takeoff_pipeline as _tp_rg
    # Four real tender-pack sheets that are NOT concrete slabs. Before the guard they emitted
    # confident 5,000-6,000 m² areas (no legend label + unverified scale). They must now REFUSE
    # cleanly. Files are gitignored client drawings, so this block skips in CI (drawings/ absent);
    # it runs locally as the regression that pins the fix.
    _fp_files = [
        "drawings/tender_pack/2-Enquiry/01-Tender/Drawings/Proposed_GA_Elevations.pdf",
        "drawings/tender_pack/2-Enquiry/01-Tender/Drawings/Proposed_GA_Office_Elevations.pdf",
        "drawings/tender_pack/2-Enquiry/01-Tender/Drawings/Proposed_Gatehouse.pdf",
        "drawings/tender_pack/2-Enquiry/01-Tender/Planning-Documentation/Site_Location_Plan.pdf",
    ]
    _fp_present = [f for f in _fp_files if _os_rg.path.exists(f)]
    for _f in _fp_files:
        if not _os_rg.path.exists(_f):
            print(f"  [SKIP] refuse-guard regression for {_os_rg.path.basename(_f)} — fixture not present")
    for _f in _fp_present:
        _r = _tp_rg.takeoff(_f, send_approval=False)
        _b = _os_rg.path.basename(_f)
        ck(f"non-slab '{_b}' refuses -> area_m2 is None", _r.get("area_m2") is None,
           f"got area={_r.get('area_m2')}")
        ck(f"non-slab '{_b}' -> UNMEASURED", _r.get("measurement_state") == "UNMEASURED",
           f"got {_r.get('measurement_state')}")
        ck(f"non-slab '{_b}' carries a REFUSED flag",
           any("REFUSED" in _fl for _fl in _r.get("flags", [])))
    # Positive control: real D77 gold has a legend label, so the guard must NOT fire even though
    # its scale bar is unverified — it must still measure the slab (~3,156 m²).
    _d77_positive_control_ran = False
    for _d77f in ("drawings/real_sgp/D77_Hard_Landscaping.pdf", "drawings/_int_d77.pdf"):
        if _os_rg.path.exists(_d77f):
            _rd = _tp_rg.takeoff(_d77f, send_approval=False)
            ck(f"legend'd D77 '{_os_rg.path.basename(_d77f)}' NOT refused by guard (area still emitted)",
               _rd.get("area_m2") is not None and _rd.get("area_m2") > 2500,
               f"got area={_rd.get('area_m2')}")
            _d77_positive_control_ran = True
            break
    if not _d77_positive_control_ran:
        print("  [SKIP] refuse-guard D77 positive control — fixture not present")
except (ImportError, FileNotFoundError) as _e:
    print(f"  [SKIP] refuse-guard regression — missing dependency or file: {_e}")

print("hatch-drawn surfaces (MJM 9000 class: slanted legend hatching, not a solid fill)")
try:
    import os as _os_hx
    import json as _json_hx
    import cv2 as _cv_hx
    import fitz as _fitz_hx
    import numpy as _np_hx
    import takeoff_unmarked as _tu_hx

    # ── Router guards. No fixture needed, so these run everywhere and pin the behaviour
    # that protects the gold sheets: a solid fill must NEVER be routed to a wide kernel.
    _hx_solid = _np_hx.zeros((700, 700), bool)
    _hx_solid[50:650, 50:650] = True
    _k_s, _i_s = _tu_hx._hatch_closing_kernel(_hx_solid, 6, 0.1, 2.0)
    ck("hatch router: a solid filled block is NOT routed to the wide kernel",
       _k_s is None, _i_s.get("reason"))

    _hx_hatch = _np_hx.zeros((700, 700), bool)
    _hx_hatch[50:650, 50:650:26] = True          # strokes 26 px apart, MJM's own spacing
    _k_h, _i_h = _tu_hx._hatch_closing_kernel(_hx_hatch, 6, 0.1, 2.0)
    ck("hatch router: 26 px-spaced strokes ARE routed to a wider kernel",
       _k_h is not None and _k_h > 6, _i_h.get("reason"))

    # The cap is in METRES, not pixels: the same 26 px spacing that is a legitimate hatch on a
    # 1:500 sheet would bridge 13 m on a coarse one, which is fusing the drawing rather than
    # reading it. Same mask as above, coarser scale, and it must refuse.
    _k_w, _i_w = _tu_hx._hatch_closing_kernel(_hx_hatch, 6, 1.0, 2.0)
    ck("hatch router: a kernel bridging too much REAL distance REFUSES, however it looks in px",
       _k_w is None and "cap" in (_i_w.get("reason") or ""), _i_w.get("reason"))

    _hx_sparse = _np_hx.zeros((700, 700), bool)
    _hx_sparse[10:30, 10:30] = True              # far too little tint to classify
    _k_sp, _i_sp = _tu_hx._hatch_closing_kernel(_hx_sparse, 6, 0.1, 2.0)
    ck("hatch router: too little matching tint REFUSES to classify",
       _k_sp is None, _i_sp.get("reason"))

    # ── Real-sheet gold. Client drawings are gitignored, so skip VISIBLY when absent.
    _hx_pdf = ("drawings/inderjit_p7/"
               "7_25195-MJM-00-00-DR-C-9000-D2-P04-External_Works_Layout.pdf")
    if not (_os_hx.path.exists(_hx_pdf) and _os_hx.path.exists("ground_truth_polygons.json")):
        print(f"  [SKIP] MJM hatch gold — client fixture not present ({_hx_pdf})")
    else:
        _hx_gt = _json_hx.loads(Path("ground_truth_polygons.json").read_text())
        _hx_entry = _hx_gt[_hx_pdf]
        _hx_res = _tu_hx.takeoff(_hx_pdf, source="architect")
        _hx_area = _hx_res.get("area_m2")
        _hx_truth = _hx_entry["area_m2"]

        # Area alone is not enough (CLAUDE.md rule 4: an agent once matched a gold area to
        # 0.1% with the WRONG region), so the shape is scored too.
        ck("MJM hatch gold: measures the area the hatch ENCLOSES, not its ink",
           _hx_area is not None and abs(_hx_area - _hx_truth) / _hx_truth * 100 <= 2.0,
           f"got {_hx_area} vs truth {_hx_truth}")
        ck("MJM hatch gold: the widened-kernel path is FLAGGED, never silent",
           any("HATCH-DRAWN SURFACE" in _f for _f in _hx_res.get("flags") or []))

        _hx_doc = _fitz_hx.open(_hx_pdf)
        _hx_pg = _hx_doc[0]
        _hx_S = 2.0
        _hx_pix = _hx_pg.get_pixmap(matrix=_fitz_hx.Matrix(_hx_S, _hx_S))
        _hx_im = _np_hx.frombuffer(_hx_pix.samples, _np_hx.uint8).reshape(
            _hx_pix.height, _hx_pix.width, _hx_pix.n)[:, :, :3].copy()
        _hx_H, _hx_W = _hx_im.shape[:2]
        _hx_R = _hx_pg.rotation_matrix
        _hx_poly = _np_hx.array(
            [[(_fitz_hx.Point(_x, _y) * _hx_R).x * _hx_S,
              (_fitz_hx.Point(_x, _y) * _hx_R).y * _hx_S]
             for _x, _y in _hx_entry["polygon_pts"]], _np_hx.int32)
        _hx_truth_mask = _np_hx.zeros((_hx_H, _hx_W), _np_hx.uint8)
        _cv_hx.fillPoly(_hx_truth_mask, [_hx_poly], 1)
        _hx_truth_mask = _hx_truth_mask.astype(bool)

        _hx_mask = _np_hx.all(
            _np_hx.abs(_hx_im.astype(_np_hx.int16) - _np_hx.array([254, 0, 0], _np_hx.int16))
            <= 14, axis=2).astype(_np_hx.uint8)
        _hx_my = max(1, int(round(_hx_H * _tu_hx.MARGIN_FRAC)))
        _hx_mx = max(1, int(round(_hx_W * _tu_hx.MARGIN_FRAC)))
        _hx_mask[:_hx_my, :] = 0
        _hx_mask[-_hx_my:, :] = 0
        _hx_mask[:, :_hx_mx] = 0
        _hx_mask[:, -_hx_mx:] = 0
        _hx_lb = _tu_hx._legend_sample_bbox_for(_hx_pdf, _tu_hx.CONCRETE_LABELS)
        if _hx_lb:
            _a, _b, _c, _d = [int(round(_v * _hx_S)) for _v in _hx_lb]
            _hx_mask[max(0, _b):min(_hx_H, _d), max(0, _a):min(_hx_W, _c)] = 0
        _hx_closed = _cv_hx.morphologyEx(
            _hx_mask, _cv_hx.MORPH_CLOSE, _np_hx.ones((51, 51), _np_hx.uint8))
        _hx_n, _hx_lab, _hx_st, _ = _cv_hx.connectedComponentsWithStats(_hx_closed, 8)
        _hx_i = 1 + int(_np_hx.argmax(_hx_st[1:, _cv_hx.CC_STAT_AREA]))
        _hx_region = (_hx_lab == _hx_i)
        _hx_iou = ((_hx_region & _hx_truth_mask).sum()
                   / max((_hx_region | _hx_truth_mask).sum(), 1))
        ck("MJM hatch gold: the closed region is the RIGHT region (IoU vs client markup)",
           _hx_iou >= 0.90, f"IoU {_hx_iou:.3f}")

        # The runtime acceptance gate, not a dev-time one: production sheets have no truth
        # polygon, so a hatch is only measured when its outline corroborates a native CAD path.
        _hx_native, _hx_reason = _tu_hx._native_boundary_for_mask(
            _hx_pg, _hx_region, _hx_S, _hx_entry["k_m_per_pt"])
        ck("MJM hatch gold: closed outline corroborates a native CAD boundary at IoU 0.90",
           _hx_native is not None, _hx_reason)
        _hx_doc.close()
except (ImportError, FileNotFoundError, KeyError) as _e:
    print(f"  [SKIP] hatch-drawn surface regression — missing dependency or file: {_e}")

print("manhole E/O costing line (costing.py Winvic rate: £75.00/Nr)")
try:
    from quotation import generate_quotation as _gen_q_mh

    MANHOLE_EO_RATE = 75.00   # £/Nr — "E/O for MH details" from the real Winvic costing sheet

    _demo_confirmed = {
        "file": "Yard.pdf", "type": "MARKED vector", "confidence": "high",
        "source_discipline": "engineer",
        "costing": {"area_m2": 26080, "rate": 44.89, "total_gbp": 1170731.20, "assumed": False,
                    "spec": {"depth_mm": 190, "mesh": "A252", "conc_mix": "C32/40", "layers": 1, "conc_rate": 128}},
        "flags": [], "manhole_count": 26,
    }
    _extras_confirmed = [("E/O for MH details", 26, "Nr", MANHOLE_EO_RATE)]
    _q_confirmed = _gen_q_mh(_demo_confirmed, project="Winvic Yard", client="Winvic",
                             ref="TST-MH-CONFIRMED", extras=_extras_confirmed)
    _mh_line = next((li for li in _q_confirmed["line_items"] if "MH details" in li["description"]), None)
    ck("confirmed manhole_count -> E/O line present", _mh_line is not None)
    ck("confirmed E/O line value = 26 x £75.00 = £1,950.00",
       _mh_line is not None and _mh_line["value"] == 1950.00, _mh_line)
    ck("confirmed E/O line NOT marked ESTIMATE", _mh_line is not None and "ESTIMATE" not in _mh_line["description"])

    _demo_estimate = dict(_demo_confirmed)
    _demo_estimate["manhole_count_estimate"] = 3
    _demo_estimate.pop("manhole_count", None)
    _extras_estimate = [("E/O for MH details (ESTIMATE — assessor confirm)", 3, "Nr", MANHOLE_EO_RATE)]
    _q_estimate = _gen_q_mh(_demo_estimate, project="D77", client="Fortel",
                            ref="TST-MH-ESTIMATE", extras=_extras_estimate)
    _mh_line_est = next((li for li in _q_estimate["line_items"] if "MH details" in li["description"]), None)
    ck("estimated manhole_count_estimate -> E/O line present and marked ESTIMATE",
       _mh_line_est is not None and "ESTIMATE" in _mh_line_est["description"])
    ck("estimated E/O line value = 3 x £75.00 = £225.00",
       _mh_line_est is not None and _mh_line_est["value"] == 225.00, _mh_line_est)
except ImportError as _e:
    print(f"  [SKIP] manhole E/O costing test — missing dependency: {_e}")

print("marked-PDF export: permanent Bluebeam-ready markup + recoverable manifest")
try:
    import json as _json_marked
    import fitz as _fitz_marked
    from marked_pdf import (MANIFEST_NAME as _MARKED_MANIFEST_NAME,
                            MarkedPdfError as _MarkedPdfError,
                            build_marked_pdf as _build_marked_pdf,
                            marked_pdf_filename as _marked_pdf_filename)

    _marked_source = Path("/tmp/ci_marked_pdf_source.pdf")
    _marked_doc = _fitz_marked.open()
    _marked_page = _marked_doc.new_page(width=300, height=200)
    _marked_page.insert_text((20, 30), "FORTEL MARKED PDF ROTATION QA")
    _source_annot = _marked_page.add_rect_annot(_fitz_marked.Rect(20, 45, 80, 85))
    _source_annot.set_info(content="Existing Bluebeam-style source annotation")
    _source_annot.update()
    _marked_page.set_rotation(90)
    _marked_doc.save(_marked_source)
    _marked_doc.close()

    # Assessor geometry is saved in 2 px/PDF-point snapshot space. Region gross=400m2,
    # contained cut-out=16m2, therefore the exported/priced net region label must be 384m2.
    _marked_job = {
        "id": "marked-pdf-ci-job", "project_ref": "MARKED/QA 01",
        "project_name": "Marked PDF QA", "decision": "adjusted",
        "quotation_revision": 3, "pdf_path": str(_marked_source),
        "result": {"file": "Rotated Drawing.pdf", "pdf_path": str(_marked_source),
                   "page": 0, "area_m2": 384.0, "scale_k": 0.2},
        "adjusted": {
            "regions": [[[40,40],[240,40],[240,240],[40,240]]],
            "region_categories": ["external_yard"], "region_scopes": ["main"],
            "scale_k": 0.1, "snapshot_scale": 2.0, "area_m2": 384.0,
            "cutout_regions": [[[80,80],[120,80],[120,120],[80,120]]],
            "user_channels": [[[40,300],[140,300],[140,380]]],
        },
        "channel_proposals": [{
            "proposal_id": "channel-assumed", "component": "Dock channel",
            "polyline_pts": [[30,230],[130,230]], "basis": "client rule",
        }],
        "channel_proposal_decisions": {"channel-assumed": {
            "decision": "accepted", "length_lm": 10.0,
            "polyline_pts": [[30,230],[130,230]],
        }},
        "transition_candidates": [{
            "candidate_id": "transition-1", "polyline_pts": [[30,200],[130,200]],
            "basis": "yard entrance",
        }],
        "transition_candidate_decisions": {"transition-1": {
            "decision": "accepted", "length_lm": 10.0,
        }},
    }
    _marked_bytes, _marked_manifest = _build_marked_pdf(_marked_job)
    ck("marked-PDF filename carries project, drawing and commercial revision",
       _marked_pdf_filename(_marked_job) ==
       "MARKED_QA_01_Rotated_Drawing_REV_03_MARKED.pdf",
       _marked_pdf_filename(_marked_job))
    _prefixed_marked_job = _copy.deepcopy(_marked_job)
    _prefixed_marked_job["result"]["file"] = "MARKED_QA_01_Rotated_Drawing.pdf"
    ck("marked-PDF client filename does not repeat the storage collision prefix",
       _marked_pdf_filename(_prefixed_marked_job) ==
       "MARKED_QA_01_Rotated_Drawing_REV_03_MARKED.pdf",
       _marked_pdf_filename(_prefixed_marked_job))
    with _fitz_marked.open(stream=_marked_bytes, filetype="pdf") as _marked_roundtrip:
        _marked_annots = sum(1 for _page in _marked_roundtrip
                             for _annot in (_page.annots() or []))
        _marked_embedded = _json_marked.loads(
            _marked_roundtrip.embfile_get(_MARKED_MANIFEST_NAME))
        _marked_text = "\n".join(page.get_text() for page in _marked_roundtrip)
        ck("marked-PDF reopens with original page rotation and all annotations burned in",
           _marked_roundtrip.page_count == 1 and
           _marked_roundtrip[0].rotation == 90 and _marked_annots == 0 and
           _marked_embedded["source_annotations_burned_in"] == 1,
           {"pages":_marked_roundtrip.page_count,"rotation":_marked_roundtrip[0].rotation,
            "annots":_marked_annots})
        ck("marked-PDF embeds versioned job/page geometry for future re-import",
           _marked_roundtrip.embfile_names() == [_MARKED_MANIFEST_NAME] and
           _marked_embedded["schema"] == "fortel.markup.v1" and
           _marked_embedded["job_id"] == "marked-pdf-ci-job" and
           _marked_embedded["source_page_index"] == 0 and
           _marked_embedded["geometry"]["regions"][0]["points"][0] == [20.0,20.0] and
           _marked_embedded["geometry"]["regions"][0]["area_m2"] == 384.0,
           _marked_embedded)
        ck("marked-PDF permanently labels net area, cut-out, channel and transition quantities",
           all(marker in _marked_text for marker in (
               "Service yard - 384.00 m2", "Cut-out 1 - 16.00 m2",
               "Channel 1 - 18.00 Lm", "Dock channel - 10.00 Lm - PROVISIONAL",
               "Transition - 10.00 Lm - PROVISIONAL")), _marked_text[-1000:])
        ck("marked-PDF metadata identifies Fortel schema/job/revision",
           "job_id=marked-pdf-ci-job" in _marked_roundtrip.metadata.get("keywords", "") and
           _marked_roundtrip.metadata.get("subject") ==
               "Fortel assessor marked-up takeoff drawing")

    _marked_blank_source = Path("/tmp/ci_marked_pdf_blank_source.pdf")
    _marked_blank_doc = _fitz_marked.open()
    _marked_blank_doc.new_page(width=300, height=200)
    _marked_blank_doc.save(_marked_blank_source)
    _marked_blank_doc.close()
    try:
        _build_marked_pdf({
            "id":"empty-markup", "decision":"adjusted", "pdf_path":str(_marked_blank_source),
            "result":{"pdf_path":str(_marked_blank_source), "page":0},
        })
        ck("marked-PDF refuses a job with no exportable geometry", False)
    except _MarkedPdfError as _marked_error:
        ck("marked-PDF refuses a job with no exportable geometry",
           "no measured/assessor markup geometry" in str(_marked_error), str(_marked_error))
except (ImportError, FileNotFoundError) as _e:
    print(f"  [SKIP] marked-PDF export tests — missing dependency or file: {_e}")

print("approval_server: upload format handling + approve hard-block")
try:
    import approval_server as _AS
    import fitz as _fitz3, zipfile as _zipfile, tempfile as _tempfile, io as _io3

    _tmpdir = Path(_tempfile.mkdtemp(prefix="ci_upload_"))

    # .zip with two PDFs -> both extracted, ranked by drawing_priority, no zip-slip
    _pdf_a = _tmpdir / "Proposed Site Plan.pdf"
    _pdf_b = _tmpdir / "External Construction Thickness Layout.pdf"
    for _pp in (_pdf_a, _pdf_b):
        _dd = _fitz3.open(); _dd.new_page(); _dd.save(str(_pp))
    _zip_path = _tmpdir / "pack.zip"
    with _zipfile.ZipFile(_zip_path, "w") as _zf:
        _zf.write(_pdf_a, _pdf_a.name)
        _zf.write(_pdf_b, _pdf_b.name)
    _extracted, _zflags = _AS._safe_extract_zip(_zip_path, _tmpdir)
    ck("zip extraction pulls both PDFs", len(_extracted) == 2)
    _ranked_zip = _AS._rank_pdfs_by_priority(_extracted)
    ck("zip PDFs ranked — construction-thickness beats site plan",
       "Construction_Thickness" in _ranked_zip[0].name or "Thickness" in _ranked_zip[0].name)

    # zip-slip guard: a malicious entry name must never escape dest_dir
    _evil_zip = _tmpdir / "evil.zip"
    with _zipfile.ZipFile(_evil_zip, "w") as _zf:
        _zf.writestr("../../etc/evil.pdf", b"%PDF-1.4 fake")
    _esc_before = set(_tmpdir.parent.glob("evil.pdf"))
    _extracted_evil, _eflags = _AS._safe_extract_zip(_evil_zip, _tmpdir)
    ck("zip-slip entry sanitised to a safe basename (stays inside dest_dir)",
       all(str(p).startswith(str(_tmpdir.resolve())) for p in _extracted_evil))

    # encrypted / zero-byte PDF -> rejected reason, not a crash
    _zero = _tmpdir / "zero.pdf"; _zero.write_bytes(b"")
    _doc, _reason = _AS._open_pdf_safely(_zero)
    ck("zero-byte PDF -> rejected with reason (not a crash)", _doc is None and "zero-byte" in _reason)

    _enc = _tmpdir / "enc.pdf"
    _ed = _fitz3.open(); _ed.new_page()
    _ed.save(str(_enc), encryption=_fitz3.PDF_ENCRYPT_AES_256, owner_pw="x", user_pw="y")
    _doc2, _reason2 = _AS._open_pdf_safely(_enc)
    ck("encrypted PDF -> rejected with reason (not a crash)",
       _doc2 is None and "encrypted" in _reason2.lower())

    _orig_jobs_file_up = _AS.JOBS_FILE
    _orig_jobs_archive_file_up = _AS.JOBS_ARCHIVE_FILE
    _orig_backup_dir_up = _AS.BACKUP_DIR
    _orig_drawings_dir_up = _AS.DRAWINGS_DIR
    _orig_quotations_dir_up = _AS.QUOTATIONS_DIR
    _orig_server_file_up = _AS.__file__
    _orig_dispatcher_up = _AS._TAKEOFF_DISPATCHER
    _started_up = []

    class _RecordingDispatcher:
        def submit(self, *args):
            _started_up.append(args)

    try:
        _AS.JOBS_FILE = _tmpdir / "multi_upload_jobs.json"
        _AS.JOBS_ARCHIVE_FILE = _tmpdir / "multi_upload_jobs_archive.json"
        _AS.BACKUP_DIR = _tmpdir / "multi_upload_backups"
        _AS.DRAWINGS_DIR = _tmpdir / "drawings"
        _AS.QUOTATIONS_DIR = _tmpdir / "quotations"
        _AS.__file__ = str(_tmpdir / "approval_server.py")
        _AS._TAKEOFF_DISPATCHER = _RecordingDispatcher()
        _client_up = _AS.app.test_client()
        _pdf_a_bytes = _pdf_a.read_bytes()
        _pdf_b_bytes = _pdf_b.read_bytes()

        # Full HTTP lifecycle: a pending AI result cannot be issued as assessor markup;
        # /adjust persists exact snapshot coordinates, then the real download route returns
        # a permanent PDF with the same recoverable geometry and a client-safe filename.
        _marked_route_job = _copy.deepcopy(_marked_job)
        _marked_route_job.update({
            "status":"pending", "decision":None,
            "measurement_state":"MEASURED_UNVERIFIED", "adjusted":None,
            "flags":["assessor: confirm extent + scale"],
        })
        _marked_route_job["result"].update({
            "measurement_state":"MEASURED_UNVERIFIED",
            "flags":["assessor: confirm extent + scale"],
        })
        _AS.save_jobs({_marked_route_job["id"]:_marked_route_job})
        _premature_marked = _client_up.get(
            f"/marked-pdf/{_marked_route_job['id']}.pdf")
        ck("marked-PDF route refuses unreviewed AI geometry",
           _premature_marked.status_code == 409 and
           "after assessor approval or adjustment" in
               (_premature_marked.get_json().get("error") or ""),
           _premature_marked.get_json())
        import gzip as _gzip_marked
        _vector_view = _client_up.get(
            f"/snapshot-vector/{_marked_route_job['id']}.svg",
            headers={"Accept-Encoding":"gzip"})
        _vector_svg = _gzip_marked.decompress(_vector_view.data).decode("utf-8")
        _base_snapshot = _client_up.get(f"/snapshot/{_marked_route_job['id']}")
        ck("true-resolution viewport serves the rotated measured page as scalable SVG",
           _vector_view.status_code == 200 and
           _vector_view.mimetype == "image/svg+xml" and
           _vector_view.headers.get("Content-Encoding") == "gzip" and
           _vector_view.headers.get("X-Vector-Coordinate-Space") ==
               "rotated_pdf_points" and
           _vector_view.headers.get("X-Vector-Page-Width") == "200" and
           _vector_view.headers.get("X-Vector-Page-Height") == "300" and
           '<svg ' in _vector_svg and 'viewBox="0 0 200 300"' in _vector_svg,
           dict(_vector_view.headers))
        ck("vector viewport leaves the exact PNG snapshot coordinate mapping unchanged",
           _base_snapshot.status_code == 200 and
           float(_base_snapshot.headers["X-Snapshot-Scale"]) == 4.0,
           _base_snapshot.headers.get("X-Snapshot-Scale"))
        _marked_adjust = _client_up.post(f"/adjust/{_marked_route_job['id']}", json={
            "regions":[[[40,40],[240,40],[240,240],[40,240]]],
            "region_categories":["external_yard"], "region_scopes":["main"],
            "scale_k":0.1, "snapshot_scale":2.0,
            "cutout_regions":[[[80,80],[120,80],[120,120],[80,120]]],
            "user_channels":[[[40,300],[140,300],[140,380]]],
        })
        _marked_saved = _AS.load_jobs()[_marked_route_job["id"]]
        ck("adjust route persists the exact snapshot transform beside assessor geometry",
           _marked_adjust.status_code == 200 and
           _marked_saved["adjusted"]["snapshot_scale"] == 2.0 and
           _marked_saved["adjusted"]["area_m2"] == 384.0,
           _marked_adjust.get_json())
        _marked_route_response = _client_up.get(
            f"/marked-pdf/{_marked_route_job['id']}.pdf")
        with _fitz_marked.open(
                stream=_marked_route_response.data, filetype="pdf") as _route_marked_doc:
            _route_manifest = _json_marked.loads(
                _route_marked_doc.embfile_get(_MARKED_MANIFEST_NAME))
            _route_annots = sum(1 for page in _route_marked_doc
                                for _annot in (page.annots() or []))
        ck("real marked-PDF endpoint returns Bluebeam-openable permanent PDF + manifest",
           _marked_route_response.status_code == 200 and
           _marked_route_response.mimetype == "application/pdf" and
           _marked_route_response.headers.get("X-Fortel-Markup-Schema") ==
               "fortel.markup.v1" and
           "MARKED_QA_01_Rotated_Drawing_REV_03_MARKED.pdf" in
               _marked_route_response.headers.get("Content-Disposition", "") and
           _route_manifest["geometry"]["regions"][0]["area_m2"] == 384.0 and
           _route_annots == 0,
           {"status":_marked_route_response.status_code,
            "content_disposition":_marked_route_response.headers.get("Content-Disposition"),
            "manifest":_route_manifest})

        # P2 lifecycle: adding two independently named areas to an already-adjusted main
        # slab must preserve 384m2 (not turn it into 454m2), and each name must become its
        # own line in the same case workbook. This uses the real HTTP routes.
        _named_area_adjust = _client_up.post(f"/adjust/{_marked_route_job['id']}", json={
            "scale_k":0.1, "snapshot_scale":2.0,
            "area_elements":[
                {"element_id":"footpath-1", "name":"Footpath",
                 "category":"external_yard", "boq_scope":"main",
                 "polygon_pts":[[300,40],[400,40],[400,90],[300,90]]},
                {"element_id":"duct-slab-1", "name":"Duct slab",
                 "category":"dock", "boq_scope":"main",
                 "polygon_pts":[[300,120],[340,120],[340,170],[300,170]]},
            ],
        })
        _named_area_saved = _AS.load_jobs()[_marked_route_job["id"]]
        ck("+Area preserves the main measured total and measures each polygon independently",
           _named_area_adjust.status_code == 200 and
           _named_area_adjust.get_json()["main_area_m2"] == 384.0 and
           _named_area_adjust.get_json()["area_elements_total_m2"] == 70.0 and
           _named_area_saved["area_m2"] == 384.0 and
           _named_area_saved["cutout_regions"] ==
               [[[80,80],[120,80],[120,120],[80,120]]] and
           _named_area_saved["user_channels"] ==
               [[[40,300],[140,300],[140,380]]] and
           [(element["name"], element["area_m2"])
            for element in _named_area_saved["area_elements"]] ==
               [("Footpath",50.0),("Duct slab",20.0)],
           _named_area_adjust.get_json())
        ck("unclassified +Area is approval-blocked instead of silently filename-bucketed",
           "explicit BOQ section" in (_AS._area_element_block_reason({
               "area_elements":[{"name":"Unknown separate slab",
                                 "category":"unclassified"}]
           }) or ""))
        _named_area_approve = _client_up.post(
            f"/approve/{_marked_route_job['id']}", json={"note":"named areas confirmed"})
        ck("classified +Area elements complete the real approval flow",
           _named_area_approve.status_code == 200,
           _named_area_approve.get_json())

        _named_q_response = _client_up.get(
            f"/quotation/{_marked_route_job['id']}.json")
        _named_q = _json_marked.loads(_named_q_response.data)
        _named_rows = [item for item in _named_q["line_items"]
                       if item.get("assessor_named_area")]
        _main_slab_rows = [item for item in _named_q["line_items"]
                           if "slab" in item.get("description", "").lower()
                           and not item.get("assessor_named_area")]
        ck("each +Area name is one distinct quotation row in its assessor-selected section",
           [(item["section"],item["description"],item["qty"],item["rate"])
            for item in _named_rows] == [
                ("External yard slabs","Footpath",50.0,45.07),
                ("Dock slabs","Duct slab",20.0,None),
            ], _named_rows)
        ck("+Area rows do not merge into or inflate the 384m2 main slab row",
           any(item.get("qty") == 384.0 for item in _main_slab_rows) and
           not any(item.get("qty") == 454.0 for item in _named_q["line_items"]),
           _main_slab_rows)

        from openpyxl import load_workbook as _load_named_workbook
        _named_xlsx_response = _client_up.get(
            f"/quotation/{_marked_route_job['id']}.xlsx")
        _named_wb = _load_named_workbook(
            _io3.BytesIO(_named_xlsx_response.data), data_only=False)
        _named_ws = _named_wb.active
        _named_xlsx_rows = [
            (cell.value.split("\n",1)[0], _named_ws.cell(cell.row,2).value,
             _named_ws.cell(cell.row,3).value, _named_ws.cell(cell.row,4).value,
             _named_ws.cell(cell.row,5).value)
            for row in _named_ws.iter_rows() for cell in row if cell.column == 1
            and isinstance(cell.value, str)
            and cell.value.split("\n",1)[0] in {"Footpath","Duct slab"}
        ]
        ck("same-workbook xlsx keeps named quantities numeric and values as live formulas",
           len(_named_wb.sheetnames) == 1 and
           [(row[0],row[1],row[2],row[3]) for row in _named_xlsx_rows] == [
               ("Footpath",50,"m2",45.07),
               ("Duct slab",20,"m2",None),
           ] and isinstance(_named_xlsx_rows[0][4], str) and
           _named_xlsx_rows[0][4].startswith("=ROUND(B") and
           "*D" in _named_xlsx_rows[0][4] and
           isinstance(_named_xlsx_rows[1][4], str) and
           _named_xlsx_rows[1][4].startswith('=IF(D') and
           "ROUND(B" in _named_xlsx_rows[1][4] and
           "*D" in _named_xlsx_rows[1][4], _named_xlsx_rows)

        _named_marked_response = _client_up.get(
            f"/marked-pdf/{_marked_route_job['id']}.pdf")
        with _fitz_marked.open(
                stream=_named_marked_response.data, filetype="pdf") as _named_marked_doc:
            _named_marked_text = "\n".join(
                page.get_text() for page in _named_marked_doc)
            _named_manifest = _json_marked.loads(
                _named_marked_doc.embfile_get(_MARKED_MANIFEST_NAME))
        ck("marked-PDF burns separately named areas and preserves their independent identity",
           "Footpath - 50.00 m2" in _named_marked_text and
           "Duct slab - 20.00 m2" in _named_marked_text and
           sum(bool(region.get("independent_area_element"))
               for region in _named_manifest["geometry"]["regions"]) == 2,
           _named_manifest["geometry"]["regions"])

        _AS.save_jobs({})
        _started_up.clear()
        _multi_resp = _client_up.post("/upload", data={
            "project_ref": "MULTI-001",
            "project_name": "Four Slab Project",
            "client_name": "Fortel QA",
            "pdf": [(_io3.BytesIO(_pdf_a_bytes), "Yard.pdf"),
                    (_io3.BytesIO(_pdf_b_bytes), "Dock.pdf")],
        }, content_type="multipart/form-data")
        _multi_json = _multi_resp.get_json()
        _multi_jobs = _AS.load_jobs()
        ck("multi-file upload returns two job_ids", _multi_resp.status_code == 202 and
           len(_multi_json.get("job_ids", [])) == 2, _multi_json)
        ck("multi-file upload creates one job per drawing under one project",
           len(_multi_jobs) == 2 and
           {j.get("project_ref") for j in _multi_jobs.values()} == {"MULTI-001"} and
           {j.get("project_name") for j in _multi_jobs.values()} == {"Four Slab Project"})
        ck("multi-file upload preserves prefixed, non-overwriting source paths",
           len({j.get("pdf_path") for j in _multi_jobs.values()}) == 2 and
           all(Path(j["pdf_path"]).name.startswith("MULTI-001_") for j in _multi_jobs.values()))
        ck("multi-file upload queues every drawing for bounded takeoff",
           len(_started_up) == 2 and all(len(args) == 4 for args in _started_up))

        # Adding a later drawing through an existing job anchor must use the persisted project
        # identity, even if a stale/tampered browser submits different visible form text.
        _existing_anchor_id = _multi_json["job_ids"][0]
        _existing_add_resp = _client_up.post("/upload", data={
            "existing_project_job_id": _existing_anchor_id,
            "project_ref": "WRONG-REF", "project_name": "Wrong look-alike project",
            "client_name": "Wrong client",
            "pdf": (_io3.BytesIO(_pdf_a_bytes), "Later folder drawing.pdf"),
        }, content_type="multipart/form-data")
        _existing_add_json = _existing_add_resp.get_json()
        _existing_add_jobs = _AS.load_jobs()
        _existing_new_job = _existing_add_jobs.get(_existing_add_json.get("job_id"), {})
        ck("existing-project upload adds one job without replacing prior project drawings",
           _existing_add_resp.status_code == 202 and
           _existing_add_json.get("added_to_project") is True and
           len(_existing_add_jobs) == 3 and
           all(job_id in _existing_add_jobs for job_id in _multi_json["job_ids"]),
           _existing_add_json)
        ck("existing-project anchor is authoritative for ref/name/client grouping",
           _existing_new_job.get("project_ref") == "MULTI-001" and
           _existing_new_job.get("project_name") == "Four Slab Project" and
           _existing_new_job.get("client_name") == "Fortel QA",
           _existing_new_job)
        _missing_existing_add = _client_up.post("/upload", data={
            "existing_project_job_id":"missing-anchor",
            "pdf":(_io3.BytesIO(_pdf_a_bytes),"orphan.pdf"),
        }, content_type="multipart/form-data")
        ck("existing-project upload refuses an unknown anchor without creating an orphan",
           _missing_existing_add.status_code == 404 and
           len(_AS.load_jobs()) == 3,
           _missing_existing_add.get_json())

        _registered_project_paths = _AS._project_pdf_paths(
            "MULTI-001", _existing_new_job["pdf_path"])
        ck("server resolves every readable same-project PDF from persisted job membership",
           len(_registered_project_paths) == 3 and
           set(_registered_project_paths) ==
               {job["pdf_path"] for job in _existing_add_jobs.values()},
           _registered_project_paths)

        # The background worker must hand that registry to the real pipeline. A narrow legacy
        # callable remains compatible because _run_takeoff introspects the optional parameter.
        import sys as _sys_project_files
        _real_project_pipeline = _sys_project_files.modules.get("takeoff_pipeline")
        _project_file_capture = {}
        class _ProjectFilePipeline:
            @staticmethod
            def takeoff(pdf_path, project_name=None, project_ref=None,
                        client_rates_path=None, approval_job_id=None, project_files=None):
                _project_file_capture["paths"] = list(project_files or [])
                return {
                    "file":Path(pdf_path).name, "pdf_path":pdf_path,
                    "project_name":project_name, "project_ref":project_ref,
                    "area_m2":None, "measurement_state":"UNMEASURED",
                    "needs_assessor":True, "flags":["project registry handoff test"],
                }
        try:
            _sys_project_files.modules["takeoff_pipeline"] = _ProjectFilePipeline
            _AS._run_takeoff(
                _existing_anchor_id,
                _existing_add_jobs[_existing_anchor_id]["pdf_path"],
                "Four Slab Project", "MULTI-001")
        finally:
            if _real_project_pipeline is None:
                _sys_project_files.modules.pop("takeoff_pipeline", None)
            else:
                _sys_project_files.modules["takeoff_pipeline"] = _real_project_pipeline
        ck("takeoff worker passes the complete project registry to spec extraction",
           set(_project_file_capture.get("paths") or []) ==
               set(_registered_project_paths), _project_file_capture)

        _AS.save_jobs({})
        _started_up.clear()
        _single_resp = _client_up.post("/upload", data={
            "project_ref": "SINGLE-001", "project_name": "Single Drawing Project",
            "pdf": (_io3.BytesIO(_pdf_a_bytes), "Yard.pdf"),
        }, content_type="multipart/form-data")
        _single_json = _single_resp.get_json()
        ck("single-file upload keeps legacy one-job response shape",
           _single_resp.status_code == 202 and "job_id" in _single_json and
           "job_ids" not in _single_json and len(_AS.load_jobs()) == 1, _single_json)

        _AS.save_jobs({})
        _started_up.clear()
        _zip_resp = _client_up.post("/upload", data={
            "project_ref": "ZIP-001", "project_name": "ZIP Slab Project",
            "pdf": (_io3.BytesIO(_zip_path.read_bytes()), "slabs.zip"),
        }, content_type="multipart/form-data")
        _zip_json = _zip_resp.get_json()
        _zip_jobs = _AS.load_jobs()
        ck("ZIP upload creates a job for every contained PDF",
           _zip_resp.status_code == 202 and len(_zip_json.get("job_ids", [])) == 2 and
           len(_zip_jobs) == 2 and len(_started_up) == 2, _zip_json)
        ck("ZIP jobs share the project ref and record all-drawings provenance",
           {j.get("project_ref") for j in _zip_jobs.values()} == {"ZIP-001"} and
           all(any("every PDF queued" in f for f in j.get("flags", []))
               for j in _zip_jobs.values()))

        _candidate_job_id = "99999999-9999-4999-8999-999999999999"
        _candidate_records = [
            {"candidate_id":"office-p0-level-00-1", "page":0, "level":0,
             "category":"ground_floor", "boq_scope":"ground_floor_core",
             "polygon_pts":[[0,0],[100,0],[100,100],[0,100]]},
            {"candidate_id":"office-p0-level-01-1", "page":0, "level":1,
             "category":"upper_floor", "boq_scope":"main_upper_floor",
             "polygon_pts":[[200,0],[300,0],[300,100],[200,100]]},
        ]
        _AS.save_jobs({_candidate_job_id: {
            "id":_candidate_job_id, "status":"pending", "decision":None,
            "measurement_state":"UNMEASURED", "scale_confirmed":False,
            "candidate_polygons":_candidate_records,
            "result":{"file":"Office-GA.pdf", "area_m2":None,
                      "measurement_state":"UNMEASURED",
                      "candidate_polygons":_candidate_records},
        }})
        _stale_candidate_resp = _client_up.post(f"/adjust/{_candidate_job_id}", json={
            "regions":[[[0,0],[100,0],[100,100],[0,100]]], "scale_k":0.1,
            "candidate_ids":["office-p0-level-99-1"],
        })
        ck("assisted adjustment rejects stale/unknown candidate IDs",
           _stale_candidate_resp.status_code == 409, _stale_candidate_resp.get_json())
        _multi_region_resp = _client_up.post(f"/adjust/{_candidate_job_id}", json={
            "regions":[[[0,0],[100,0],[100,100],[0,100]],
                       [[200,0],[300,0],[300,100],[200,100]]],
            "scale_k":0.1,
            "candidate_ids":["office-p0-level-00-1", "office-p0-level-01-1"],
            "region_categories":["ground_floor", "upper_floor"],
            "region_scopes":["ground_floor_core", "main_upper_floor"],
            "note":"assessor accepted two Office GA regions",
        })
        _multi_region_job = _AS.load_jobs()[_candidate_job_id]
        ck("assessor adjustment measures several Office regions in one atomic decision",
           _multi_region_resp.status_code == 200 and
           _multi_region_resp.get_json()["area_m2"] == 200.0 and
           _multi_region_resp.get_json()["region_count"] == 2 and
           len(_multi_region_job["adjusted"]["regions"]) == 2 and
           _multi_region_job["adjusted"]["candidate_ids"] ==
           ["office-p0-level-00-1", "office-p0-level-01-1"],
           _multi_region_resp.get_json())
        ck("candidate selection alone stayed UNMEASURED; assessor POST performs confirmation",
           _multi_region_job["scale_confirmed"] is True and
           _multi_region_job["measurement_state"] == "MEASURED_VERIFIED")
        ck("assisted Office adjustment persists one real BOQ zone per supplied region category",
           {zone["category"]:zone["area_m2"] for zone in _multi_region_job["zones"]} ==
           {"ground_floor":100.0, "upper_floor":100.0} and
           not _multi_region_job.get("zone_allocation_stale") and
           _multi_region_job["adjusted"]["region_categories"] ==
           ["ground_floor", "upper_floor"] and
           _multi_region_job["adjusted"]["region_scopes"] ==
           ["ground_floor_core", "main_upper_floor"], _multi_region_job.get("zones"))
        _office_adjust_quote = _AS._quotation_for_job(_candidate_job_id)
        ck("real Office region categories override the filename and reach separate BOQ sections",
           [spec["section"] for spec in _office_adjust_quote["specifications"]] ==
           ["Ground floor slabs", "Upper floor slabs"],
           _office_adjust_quote["specifications"])
        _manual_unknown_resp = _client_up.post(f"/adjust/{_candidate_job_id}", json={
            "regions":[[[0,0],[100,0],[100,100],[0,100]]], "scale_k":0.1,
            "region_categories":["unclassified"],
            "note":"manual outline; floor level not resolved",
        })
        _manual_unknown_job = _AS.load_jobs()[_candidate_job_id]
        ck("unresolved manual region is preserved as unclassified and visibly blocks approval",
           _manual_unknown_resp.status_code == 200 and
           _manual_unknown_job["zones"][0]["area_m2"] == 100.0 and
           _manual_unknown_job["zones"][0]["category"] == "unclassified" and
           _manual_unknown_job["zone_classification_required"] and
           "unclassified" in (_AS._approve_block_reason(_manual_unknown_job) or ""),
           _manual_unknown_job.get("zones"))

        _channel_job_id = "aaaaaaaa-aaaa-4aaa-8aaa-aaaaaaaaaaaa"
        _channel_zones = [
            {"zone_key":"external_yard", "category":"external_yard", "area_m2":100.0},
            {"zone_key":"dock", "category":"dock", "area_m2":20.0,
             "loading_face_lm":30.0, "loading_face_pts":[[10,10],[10,310]]},
        ]
        _channel_proposals = [
            {"proposal_id":"channel-dock-loading-face", "component":"dock_retaining_wall",
             "proposed_length_lm":30.0, "polyline_pts":[[10,10],[10,310]],
             "assumed":True, "requires_assessor_confirmation":True},
            {"proposal_id":"channel-yard-longest-contained-run",
             "component":"yard_longest_contained_run", "proposed_length_lm":90.0,
             "polyline_pts":[[20,20],[920,20]], "assumed":True,
             "requires_assessor_confirmation":True},
        ]
        _channel_costing = _copy.deepcopy(_demo_result["costing"])
        _AS.save_jobs({_channel_job_id: {
            "id":_channel_job_id, "status":"pending", "decision":None,
            "measurement_state":"MEASURED_VERIFIED", "scale_confirmed":False,
            "zones":_copy.deepcopy(_channel_zones),
            "channel_proposals":_copy.deepcopy(_channel_proposals),
            "channel_proposal_decisions":{}, "costing":_copy.deepcopy(_channel_costing),
            "result":{"file":"Raw External.pdf", "area_m2":120.0,
                      "scale_k":0.1,
                      "measurement_state":"MEASURED_VERIFIED",
                      "zones":_copy.deepcopy(_channel_zones),
                      "channel_proposals":_copy.deepcopy(_channel_proposals),
                      "channel_proposal_decisions":{},
                      "costing":_copy.deepcopy(_channel_costing)},
        }})
        ck("unreviewed assumed channel proposals block approval",
           "accept/edit/remove" in (_AS._approve_block_reason(
               _AS.load_jobs()[_channel_job_id]) or ""))
        _channel_first_resp = _client_up.post(
            f"/channel-proposals/{_channel_job_id}", json={"decisions":[{
                "proposal_id":"channel-dock-loading-face", "action":"accept",
                "length_lm":31.25,
            }]})
        _channel_first_job = _AS.load_jobs()[_channel_job_id]
        ck("assessor can edit and accept a proposed channel length without measuring it",
           _channel_first_resp.status_code == 200 and
           not _channel_first_resp.get_json()["review_complete"] and
           _channel_first_job["channel_proposal_decisions"][
               "channel-dock-loading-face"]["length_lm"] == 31.25 and
           _channel_first_job["channel_proposal_decisions"][
               "channel-dock-loading-face"]["edited"] is True)
        ck("one pending channel proposal continues to block approval",
           _AS._approve_block_reason(_channel_first_job) is not None)
        _channel_geometry_resp = _client_up.post(
            f"/channel-proposals/{_channel_job_id}", json={"decisions":[{
                "proposal_id":"channel-dock-loading-face", "action":"accept",
                # Deliberately contradictory browser length: server must derive 35m from
                # the two PDF-point endpoints and stored 0.1m/pt scale.
                "length_lm":1.0, "polyline_pts":[[10,10],[360,10]],
            }]})
        _channel_geometry_job = _AS.load_jobs()[_channel_job_id]
        _channel_geometry_decision = _channel_geometry_job["channel_proposal_decisions"][
            "channel-dock-loading-face"]
        ck("assessor can drag a channel endpoint; server persists axis-aligned geometry",
           _channel_geometry_resp.status_code == 200 and
           _channel_geometry_decision["polyline_pts"] == [[10.0,10.0],[360.0,10.0]] and
           _channel_geometry_decision["geometry_edited"] is True,
           _channel_geometry_decision)
        ck("edited channel length is derived from geometry and scale, not browser arithmetic",
           _channel_geometry_decision["length_lm"] == 35.0,
           _channel_geometry_decision)
        _channel_diagonal_resp = _client_up.post(
            f"/channel-proposals/{_channel_job_id}", json={"decisions":[{
                "proposal_id":"channel-yard-longest-contained-run", "action":"accept",
                "length_lm":90.0, "polyline_pts":[[20,20],[920,120]],
            }]})
        ck("channel edit API refuses diagonal geometry rather than storing it",
           _channel_diagonal_resp.status_code == 400 and
           "non-diagonal" in _channel_diagonal_resp.get_json()["error"] and
           "channel-yard-longest-contained-run" not in
               _AS.load_jobs()[_channel_job_id]["channel_proposal_decisions"],
           _channel_diagonal_resp.get_json())
        _channel_second_resp = _client_up.post(
            f"/channel-proposals/{_channel_job_id}", json={"decisions":[{
                "proposal_id":"channel-yard-longest-contained-run", "action":"remove",
            }]})
        _channel_reviewed_job = _AS.load_jobs()[_channel_job_id]
        ck("assessor remove completes channel-proposal review and releases its gate",
           _channel_second_resp.status_code == 200 and
           _channel_second_resp.get_json()["review_complete"] and
           _AS._approve_block_reason(_channel_reviewed_job) is None)
        ck("channel decisions never enter measured zones or an approvable costing total",
           _channel_reviewed_job["zones"] == _channel_zones and
           _channel_reviewed_job["result"]["zones"] == _channel_zones and
           not any(z.get("category") == "channel" for z in _channel_reviewed_job["zones"]) and
           _channel_reviewed_job["costing"] == _channel_costing and
           _channel_reviewed_job["result"]["costing"] == _channel_costing)
        _channel_approve = _client_up.post(f"/approve/{_channel_job_id}", json={
            "note":"channel assumptions reviewed",
        })
        _channel_approved_job = _AS.load_jobs()[_channel_job_id]
        _channel_quote_paths = _channel_approved_job.get("quotation_paths") or {}
        _channel_saved_quote = (__import__("json").loads(
            Path(_channel_quote_paths["json"]).read_text())
            if _channel_quote_paths.get("json") and
            Path(_channel_quote_paths["json"]).exists() else {})
        _channel_saved_rows = [item for item in _channel_saved_quote.get("line_items", [])
                               if item.get("description") == FORTEL_CHANNEL_ROW]
        ck("approved server quotation carries accepted channel Lm with blank assessor rate",
           _channel_approve.status_code == 200 and len(_channel_saved_rows) == 1 and
           _channel_saved_rows[0]["qty"] == 35.0 and
           _channel_saved_rows[0]["rate"] is None and
           _channel_saved_rows[0]["value"] is None and
           _channel_saved_rows[0]["provisional"],
           _channel_saved_rows)

        _transition_job_id = "bbbbbbbb-bbbb-4bbb-8bbb-bbbbbbbbbbbb"
        _transition_candidates = [
            {"candidate_id":"transition-yard-region-1", "region_id":"yard-region-1",
             "category":"transition", "proposed_length_lm":14.0,
             "polyline_pts":[[10,10],[150,10]], "assumed":True,
             "basis":"macadam-to-concrete Yard entrance boundary"},
            {"candidate_id":"transition-yard-region-2", "region_id":"yard-region-2",
             "category":"transition", "proposed_length_lm":12.0,
             "polyline_pts":[[20,20],[140,20]], "assumed":True,
             "basis":"macadam-to-concrete Yard entrance boundary"},
        ]
        _AS.save_jobs({_transition_job_id: {
            "id":_transition_job_id, "status":"pending", "decision":None,
            "measurement_state":"MEASURED_VERIFIED", "scale_confirmed":False,
            "zones":_copy.deepcopy(_channel_zones),
            "transition_candidates":_copy.deepcopy(_transition_candidates),
            "transition_candidate_decisions":{},
            "accepted_transition_quantities":[],
            "costing":_copy.deepcopy(_channel_costing),
            "result":{"file":"Raw External transitions.pdf", "area_m2":120.0,
                      "scale_k":0.1, "measurement_state":"MEASURED_VERIFIED",
                      "zones":_copy.deepcopy(_channel_zones),
                      "transition_candidates":_copy.deepcopy(_transition_candidates),
                      "transition_candidate_decisions":{},
                      "accepted_transition_quantities":[],
                      "costing":_copy.deepcopy(_channel_costing)},
        }})
        ck("unreviewed assumed Transition candidates block approval",
           "accept/edit/remove" in (_AS._approve_block_reason(
               _AS.load_jobs()[_transition_job_id]) or ""))
        _transition_accept_resp = _client_up.post(
            f"/transition-candidates/{_transition_job_id}", json={"decisions":[{
                "candidate_id":"transition-yard-region-1", "action":"accept",
                "length_lm":15.25,
            }]})
        _transition_accept_job = _AS.load_jobs()[_transition_job_id]
        ck("Transition endpoint persists assessor-edited accepted quantity",
           _transition_accept_resp.status_code == 200 and
           not _transition_accept_resp.get_json()["review_complete"] and
           _transition_accept_job["transition_candidate_decisions"][
               "transition-yard-region-1"]["length_lm"] == 15.25 and
           _transition_accept_job["transition_candidate_decisions"][
               "transition-yard-region-1"]["edited"] is True and
           _transition_accept_job["accepted_transition_quantities"] == [{
               "candidate_id":"transition-yard-region-1",
               "region_id":"yard-region-1", "category":"transition",
               "measurement_kind":"length", "length_lm":15.25, "unit":"Lm",
               "assumed":True, "provisional":True,
               "basis":"macadam-to-concrete Yard entrance boundary",
               "source":None, "assessor_edited":True,
           }], _transition_accept_resp.get_json())
        ck("one pending Transition candidate continues to block approval",
           _AS._approve_block_reason(_transition_accept_job) is not None)
        _transition_remove_resp = _client_up.post(
            f"/transition-candidates/{_transition_job_id}", json={"decisions":[{
                "candidate_id":"transition-yard-region-2", "action":"remove",
            }]})
        _transition_reviewed_job = _AS.load_jobs()[_transition_job_id]
        ck("Transition remove completes review and releases the approval gate",
           _transition_remove_resp.status_code == 200 and
           _transition_remove_resp.get_json()["review_complete"] and
           _AS._approve_block_reason(_transition_reviewed_job) is None)
        ck("accepted Transition remains provisional state, never measured zone or costing",
           not any(zone.get("category") == "transition"
                   for zone in _transition_reviewed_job["zones"]) and
           _transition_reviewed_job["costing"] == _channel_costing and
           _transition_reviewed_job["accepted_transition_quantities"][0][
               "provisional"] is True)
        _transition_quote = _AS._quotation_for_job(_transition_job_id)
        _transition_rows = [item for item in _transition_quote["line_items"]
                            if item.get("description") == FORTEL_TRANSITION_ROW]
        ck("persisted accepted Transition reaches server quotation with blank rate",
           len(_transition_rows) == 1 and _transition_rows[0]["qty"] == 15.25 and
           _transition_rows[0]["rate"] is None and
           _transition_rows[0]["value"] is None and
           _transition_rows[0]["provisional"], _transition_rows)

        # Critical assisted-loop regression: begin with a real /upload-created record, then
        # install the completed zoned MEASURED_UNVERIFIED takeoff that the background worker
        # would save. Confirming the existing scale/extent must preserve both zones, release
        # the gate, approve, and write a real quotation — no destructive /adjust in between.
        _AS.save_jobs({})
        _started_up.clear()
        _loop_upload = _client_up.post("/upload", data={
            "project_ref":"E2E-ZONE-001", "project_name":"Zoned confirmation loop",
            "client_name":"Fortel QA",
            "pdf":(_io3.BytesIO(_pdf_a_bytes), "External Unit-1.pdf"),
        }, content_type="multipart/form-data")
        _loop_job_id = _loop_upload.get_json()["job_id"]
        _loop_zones = [
            {"zone_key":"external_yard", "category":"external_yard", "area_m2":100.0},
            {"zone_key":"dock", "category":"dock", "area_m2":20.0},
        ]
        _loop_jobs = _AS.load_jobs()
        _loop_jobs[_loop_job_id].update({
            "status":"pending", "measurement_state":"MEASURED_UNVERIFIED",
            "scale_confirmed":False, "zone_allocation_stale":True,
            "zones":_copy.deepcopy(_loop_zones),
            "flags":["ZONE ALLOCATION STALE: retained zones await assessor confirmation"],
            "result":{
                "file":"E2E-ZONE-001_External Unit-1.pdf", "area_m2":120.0,
                "scale_k":0.1, "measurement_state":"MEASURED_UNVERIFIED",
                "zones":_copy.deepcopy(_loop_zones), "zone_allocation_stale":True,
                "flags":["ZONE ALLOCATION STALE: retained zones await assessor confirmation"],
            },
        })
        _AS.save_jobs(_loop_jobs)
        _loop_confirm = _client_up.post(f"/confirm-measurement/{_loop_job_id}", json={
            "confirm_scale_extent":True, "note":"scale and coloured extents checked",
        })
        _loop_confirmed_job = _AS.load_jobs()[_loop_job_id]
        ck("existing scale+extent confirmation preserves zoned measurement and clears stale gate",
           _loop_confirm.status_code == 200 and
           _loop_confirm.get_json()["zone_count"] == 2 and
           _loop_confirmed_job["zones"] == _loop_zones and
           _loop_confirmed_job["result"]["zones"] == _loop_zones and
           _loop_confirmed_job["scale_confirmed"] is True and
           not _loop_confirmed_job["zone_allocation_stale"] and
           _AS._approve_block_reason(_loop_confirmed_job) is None,
           _loop_confirm.get_json())
        _loop_approve = _client_up.post(f"/approve/{_loop_job_id}", json={
            "note":"approved after non-destructive confirmation",
        })
        _loop_approved_job = _AS.load_jobs()[_loop_job_id]
        _loop_paths = _loop_approved_job.get("quotation_paths") or {}
        _loop_quote = (__import__("json").loads(Path(_loop_paths["json"]).read_text())
                       if _loop_paths.get("json") and Path(_loop_paths["json"]).exists() else {})
        ck("E2E upload -> confirm -> approve terminates in an APPROVED zoned quotation",
           _loop_approve.status_code == 200 and
           _loop_approved_job["status"] == "approved" and
           Path(_loop_paths.get("xlsx", "missing")).exists() and
           [spec["section"] for spec in _loop_quote.get("specifications", [])] ==
           ["External yard slabs", "Dock slabs"],
            {"confirm":_loop_confirm.get_json(), "approve":_loop_approve.get_json(),
             "status":_loop_approved_job.get("status"), "quotation_paths":_loop_paths})

        # Money-path regression: assessor geometry, cut-outs, and scale must replace the AI
        # inputs everywhere that can be costed or quoted. These exercise the real Flask routes
        # and then inspect the generated quotation, rather than trusting only /adjust's reply.
        def _assessor_truth_job(job_id, project_ref, original_area=6816.0):
            original_costing = _copy.deepcopy(_demo_result["costing"])
            original_costing["area_m2"] = original_area
            zones = [{"zone_key":"external_yard", "category":"external_yard",
                      "area_m2":original_area}]
            result = {
                "file":f"{project_ref}_External.pdf", "type":"UNMARKED vector",
                "source_discipline":"architect", "area_m2":original_area,
                "scale_k":0.17639, "measurement_state":"MEASURED_UNVERIFIED",
                "zones":_copy.deepcopy(zones), "flags":[],
                "costing":_copy.deepcopy(original_costing),
            }
            return {
                "id":job_id, "status":"pending", "decision":None,
                "project_ref":project_ref, "project_name":"Assessor truth regression",
                "client_name":"Fortel QA", "created_at":"2026-08-12T00:00:00",
                "measurement_state":"MEASURED_UNVERIFIED", "scale_confirmed":False,
                "zones":_copy.deepcopy(zones), "costing":original_costing,
                "result":result,
            }

        def _confirm_and_adjust(job_id, *, width=200, cutouts=None, channels=None):
            blocked_before = _AS._approve_block_reason(_AS.load_jobs()[job_id]) is not None
            confirmed_response = _client_up.post(
                f"/confirm-measurement/{job_id}", json={
                    "confirm_scale_extent":True, "note":"assessor confirmed extent",
                })
            adjusted_response = _client_up.post(f"/adjust/{job_id}", json={
                "regions":[[[100,100],[100 + width,100],[100 + width,200],[100,200]]],
                "region_categories":["external_yard"], "region_scopes":["main"],
                "scale_k":0.1, "cutout_regions":cutouts or [],
                "user_channels":channels or [], "note":"categorized assessor trace",
            })
            return blocked_before, confirmed_response, adjusted_response

        _money_cutout_id = "91000000-0000-4000-8000-000000000001"
        _AS.save_jobs({_money_cutout_id:_assessor_truth_job(
            _money_cutout_id, "MONEY-CUTOUT")})
        _cutout_gate, _cutout_confirm, _cutout_adjust = _confirm_and_adjust(
            _money_cutout_id,
            cutouts=[[[150,120],[200,120],[200,140],[150,140]]])
        _cutout_job = _AS.load_jobs()[_money_cutout_id]
        _cutout_quote = _AS._quotation_for_job(_money_cutout_id)
        _cutout_slab = next(item for item in _cutout_quote["line_items"]
                            if item.get("line_role") == "concrete_slab")
        _cutout_json = __import__("json").loads(quotation_json(_cutout_quote))
        _cutout_json_slab = next(item for item in _cutout_json["line_items"]
                                 if item.get("line_role") == "concrete_slab")
        _cutout_text = quotation_text(_cutout_quote)
        _cutout_html = quotation_html(_cutout_quote)
        _cutout_text_slab = next(line for line in _cutout_text.splitlines()
                                  if "Concrete Slabs" in line)
        _cutout_wb = _load_workbook(_BytesIO(quotation_xlsx(_cutout_quote)), data_only=False)
        _cutout_ws = _cutout_wb["REV_01"]
        _cutout_xlsx_source_values = [
            _cutout_ws.cell(row, 2).value for row in range(1, _cutout_ws.max_row + 1)
            if "External.pdf" in str(_cutout_ws.cell(row, 1).value or "")
        ]
        ck("assessor cut-out net area reaches zones and every quotation format",
           _cutout_gate and _cutout_confirm.status_code == 200 and
           _cutout_adjust.get_json()["area_m2"] == 190.0 and
           _cutout_job["zones"][0]["area_m2"] == 190.0 and
           _cutout_slab["qty"] == 190.0 and _cutout_json_slab["qty"] == 190.0 and
           "190" in _cutout_text_slab and ">190 m²</td>" in _cutout_html and
           190.0 in _cutout_xlsx_source_values and
           200.0 not in _cutout_xlsx_source_values,
           {"adjust":_cutout_adjust.get_json(), "zones":_cutout_job["zones"],
            "slab":_cutout_slab, "xlsx_sources":_cutout_xlsx_source_values})
        ck("assessor adjustment preserves four-state gate and explicit verification",
           _cutout_job["measurement_state"] == "MEASURED_VERIFIED" and
           _cutout_job["scale_confirmed"] is True and
           _AS._approve_block_reason(_cutout_job) is None)

        _money_approve_id = "91000000-0000-4000-8000-000000000002"
        _AS.save_jobs({_money_approve_id:_assessor_truth_job(
            _money_approve_id, "MONEY-APPROVE")})
        _confirm_and_adjust(_money_approve_id, width=190)
        _approve_response = _client_up.post(f"/approve/{_money_approve_id}", json={
            "note":"approve assessor correction",
        })
        _approved_truth_job = _AS.load_jobs()[_money_approve_id]
        _approved_truth_quote = _AS._quotation_for_job(_money_approve_id)
        _approved_truth_slab = next(item for item in _approved_truth_quote["line_items"]
                                    if item.get("line_role") == "concrete_slab")
        ck("approval costs and quotes assessor-adjusted area instead of original AI area",
           _approve_response.status_code == 200 and
           _approve_response.get_json()["costing"]["area_m2"] == 190.0 and
           _approved_truth_job["costing"]["area_m2"] == 190.0 and
           _approved_truth_slab["qty"] == 190.0,
           {"approve":_approve_response.get_json(), "slab":_approved_truth_slab})

        _money_channel_id = "91000000-0000-4000-8000-000000000003"
        _AS.save_jobs({_money_channel_id:_assessor_truth_job(
            _money_channel_id, "MONEY-CHANNEL")})
        _confirm_and_adjust(
            _money_channel_id, channels=[[[100,250],[200,250]]])
        _channel_truth_quote = _AS._quotation_for_job(_money_channel_id)
        _channel_truth_row = next(item for item in _channel_truth_quote["line_items"]
                                  if item["description"] == FORTEL_CHANNEL_ROW)
        ck("assessor channel quotation length uses assessor scale travelling with geometry",
           _channel_truth_row["qty"] == 10.0 and
           _AS._quotation_result_for_job(
               _AS.load_jobs()[_money_channel_id])["scale_k"] == 0.1,
           _channel_truth_row)

        _money_bend_id = "91000000-0000-4000-8000-000000000005"
        _AS.save_jobs({_money_bend_id:_assessor_truth_job(
            _money_bend_id, "MONEY-BENDING-CHANNEL")})
        _, _bend_confirm, _bend_adjust = _confirm_and_adjust(
            _money_bend_id,
            # Two segments: 30 px + 40 px at assessor k=0.1 m/px => 7.00 Lm.
            channels=[[[100,250],[130,250],[130,290]]],
        )
        _bend_job = _AS.load_jobs()[_money_bend_id]
        _bend_quote = _AS._quotation_for_job(_money_bend_id)
        _bend_row = next(item for item in _bend_quote["line_items"]
                         if item["description"] == FORTEL_CHANNEL_ROW)
        ck("assessor can persist a bending channel polyline instead of a forced straight chord",
           _bend_confirm.status_code == 200 and _bend_adjust.status_code == 200 and
           _bend_job["adjusted"]["user_channels"] ==
               [[[100,250],[130,250],[130,290]]],
           _bend_adjust.get_json())
        ck("quotation sums every segment of the assessor channel polyline",
           _bend_row["qty"] == 7.0, _bend_row)

        _money_plain_id = "91000000-0000-4000-8000-000000000004"
        _plain_job = _assessor_truth_job(_money_plain_id, "MONEY-PLAIN", original_area=125.0)
        _plain_job["measurement_state"] = "MEASURED_VERIFIED"
        _plain_job["scale_confirmed"] = True
        _plain_job["result"]["measurement_state"] = "MEASURED_VERIFIED"
        _AS.save_jobs({_money_plain_id:_plain_job})
        _plain_approve = _client_up.post(f"/approve/{_money_plain_id}", json={
            "note":"unchanged normal path",
        })
        _plain_saved = _AS.load_jobs()[_money_plain_id]
        _plain_quote = _AS._quotation_for_job(_money_plain_id)
        _plain_slab = next(item for item in _plain_quote["line_items"]
                           if item.get("line_role") == "concrete_slab")
        ck("unadjusted verified job keeps its original pricing path unchanged",
           _plain_approve.status_code == 200 and
           _plain_approve.get_json()["costing"]["area_m2"] == 125.0 and
           _plain_saved["costing"]["area_m2"] == 125.0 and
           _plain_slab["qty"] == 125.0,
           {"approve":_plain_approve.get_json(), "slab":_plain_slab})

        _route_costing_a = _copy.deepcopy(_demo_result["costing"])
        _route_costing_a.update({"area_m2": 100, "assumed": True})
        _route_costing_b = _copy.deepcopy(_demo_result["costing"])
        _route_costing_b.update({"area_m2": 150, "assumed": True})
        _route_jobs = {
            "11111111-1111-4111-8111-111111111111": {
                "id": "11111111-1111-4111-8111-111111111111", "decision": "approved",
                "status": "approved", "project_ref": "QUOTE-MULTI-001",
                "project_name": "Two Yard Units", "client_name": "Fortel QA",
                "created_at": "2026-07-15T10:00:00",
                "costing": _route_costing_a,
                "result": {"file": "Yard-A.pdf", "quotation_section": "External yard slabs",
                           "area_m2": 100, "costing": _route_costing_a, "flags": []},
            },
            "22222222-2222-4222-8222-222222222222": {
                "id": "22222222-2222-4222-8222-222222222222", "decision": "adjusted",
                "status": "adjusted", "project_ref": "QUOTE-MULTI-001",
                "project_name": "Two Yard Units", "client_name": "Fortel QA",
                "created_at": "2026-07-15T10:01:00",
                "costing": _route_costing_b,
                "result": {"file": "Yard-B.pdf", "quotation_section": "External yard slabs",
                           "area_m2": 150, "costing": _route_costing_b, "flags": []},
            },
        }
        _AS.save_jobs(_route_jobs)
        _xlsx_route_resp = _client_up.get(
            "/quotation/11111111-1111-4111-8111-111111111111.xlsx")
        _xlsx_route_wb = _load_workbook(_BytesIO(_xlsx_route_resp.data), data_only=False)
        _xlsx_route_ws = _xlsx_route_wb["REV_01"]
        _xlsx_route_slab_row = next(
            row for row in range(1, _xlsx_route_ws.max_row + 1)
            if "Concrete Slabs" in str(_xlsx_route_ws.cell(row, 1).value or ""))
        ck("xlsx download route returns a valid attachment",
           _xlsx_route_resp.status_code == 200 and
           _xlsx_route_resp.mimetype ==
           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" and
           "QUOTE-MULTI-001.xlsx" in _xlsx_route_resp.headers.get("Content-Disposition", ""))
        ck("xlsx route aggregates approved sibling units sharing project_ref",
           _xlsx_route_ws.cell(_xlsx_route_slab_row, 2).data_type == "f" and
           any(_xlsx_route_ws.cell(row, 1).value == "Total Area Take Off:" and
               _xlsx_route_ws.cell(row, 2).data_type == "f"
               for row in range(1, _xlsx_route_ws.max_row + 1)) and
           sorted(float(_xlsx_route_ws.cell(row, 2).value)
                  for row in range(1, _xlsx_route_ws.max_row + 1)
                  if _xlsx_route_ws.cell(row, 1).value in ("Yard-A.pdf", "Yard-B.pdf"))
           == [100.0, 150.0])

        # Aryan field report 17 Jul: uploading a fresh case produced a SEPARATE xlsx per
        # document (pending siblings were excluded from aggregation), and unmeasured office
        # GA plans (line/hatch -> assessor trace) vanished from the output entirely. The
        # case quotation must be ONE workbook: pending-but-measured siblings included
        # (marked provisional), unmeasured documents listed as awaiting trace — never absent.
        _pend_costing_a = _copy.deepcopy(_demo_result["costing"]); _pend_costing_a.update({"area_m2": 100, "assumed": True})
        _pend_costing_b = _copy.deepcopy(_demo_result["costing"]); _pend_costing_b.update({"area_m2": 150, "assumed": True})
        _pend_jobs = {
            "aaaaaaaa-1111-4111-8111-111111111111": {
                "id": "aaaaaaaa-1111-4111-8111-111111111111", "decision": None,
                "status": "pending", "project_ref": "CASE-PEND-001",
                "project_name": "Fresh Case", "client_name": "Fortel QA",
                "created_at": "2026-07-17T10:00:00", "costing": _pend_costing_a,
                "result": {"file": "Yard-A.pdf", "quotation_section": "External yard slabs",
                           "area_m2": 100, "costing": _pend_costing_a, "flags": []},
            },
            "bbbbbbbb-2222-4222-8222-222222222222": {
                "id": "bbbbbbbb-2222-4222-8222-222222222222", "decision": None,
                "status": "pending", "project_ref": "CASE-PEND-001",
                "project_name": "Fresh Case", "client_name": "Fortel QA",
                "created_at": "2026-07-17T10:01:00", "costing": _pend_costing_b,
                "result": {"file": "Yard-B.pdf", "quotation_section": "External yard slabs",
                           "area_m2": 150, "costing": _pend_costing_b, "flags": []},
            },
            "cccccccc-3333-4333-8333-333333333333": {
                "id": "cccccccc-3333-4333-8333-333333333333", "decision": None,
                "status": "pending", "project_ref": "CASE-PEND-001",
                "project_name": "Fresh Case", "client_name": "Fortel QA",
                "created_at": "2026-07-17T10:02:00",
                "result": {"file": "Office-Floors-U1.pdf", "area_m2": None,
                           "measurement_state": "UNMEASURED",
                           "flags": ["NON-COLOUR-CODED (line/hatch) drawing — assessor trace"]},
            },
        }
        _AS.save_jobs(_pend_jobs)
        _pend_resp = _client_up.get("/quotation/aaaaaaaa-1111-4111-8111-111111111111.xlsx")
        ck("case xlsx succeeds for a fresh (all-pending) case", _pend_resp.status_code == 200,
           _pend_resp.status_code)
        _pend_ws = _load_workbook(_BytesIO(_pend_resp.data), data_only=False)["REV_01"]
        _pend_cells = [str(_pend_ws.cell(r, 1).value or "") for r in range(1, _pend_ws.max_row + 1)]
        ck("pending-but-measured siblings aggregate into ONE case workbook",
           any("Yard-A.pdf" in c for c in _pend_cells) and any("Yard-B.pdf" in c for c in _pend_cells))
        ck("unmeasured document is LISTED in the case workbook (never silently absent)",
           any("Office-Floors-U1.pdf" in c and "NOT YET MEASURED" in c for c in _pend_cells),
           [c for c in _pend_cells if "Office" in c])
        ck("pending quantities are marked provisional pending approval",
           any("not yet" in c.lower() and "approved" in c.lower() for c in _pend_cells))
        _pend_json = _client_up.get("/quotation/aaaaaaaa-1111-4111-8111-111111111111.json").get_json()
        ck("case quotation JSON carries the unmeasured document list",
           any(d.get("file") == "Office-Floors-U1.pdf" for d in _pend_json.get("unmeasured", [])))

        # Aryan follow-up 18 Jul: a case containing ONLY unmeasurable docs (e.g. office GA
        # plans, all awaiting assessor trace) previously 400'd on download — the office
        # drawings "still skipped". An office-only case must still yield the case workbook.
        _office_only = {
            "dddddddd-4444-4444-8444-444444444444": {
                "id": "dddddddd-4444-4444-8444-444444444444", "decision": None,
                "status": "pending", "project_ref": "CASE-OFFICE-ONLY",
                "project_name": "Office Only", "client_name": "Fortel QA",
                "created_at": "2026-07-18T09:00:00",
                "result": {"file": "Office-GA-L00.pdf", "area_m2": None,
                           "measurement_state": "UNMEASURED", "flags": ["line/hatch"]},
            },
            "eeeeeeee-5555-4555-8555-555555555555": {
                "id": "eeeeeeee-5555-4555-8555-555555555555", "decision": None,
                "status": "pending", "project_ref": "CASE-OFFICE-ONLY",
                "project_name": "Office Only", "client_name": "Fortel QA",
                "created_at": "2026-07-18T09:01:00",
                "result": {"file": "Office-GA-L01.pdf", "area_m2": None,
                           "measurement_state": "UNMEASURED", "flags": ["line/hatch"]},
            },
        }
        _AS.save_jobs(_office_only)
        _oo_resp = _client_up.get("/quotation/dddddddd-4444-4444-8444-444444444444.xlsx")
        ck("office-only case still yields the case workbook (no 400)",
           _oo_resp.status_code == 200, _oo_resp.status_code)
        _oo_ws = _load_workbook(_BytesIO(_oo_resp.data), data_only=False)["REV_01"]
        _oo_cells = [str(_oo_ws.cell(r, 1).value or "") for r in range(1, _oo_ws.max_row + 1)]
        ck("BOTH office docs listed as NOT YET MEASURED in the office-only workbook",
           all(any(f in c and "NOT YET MEASURED" in c for c in _oo_cells)
               for f in ("Office-GA-L00.pdf", "Office-GA-L01.pdf")), _oo_cells[-6:])
        # Unit labels derived from Fortel filename conventions (per-unit BOQ rows)
        from quotation import _unit_label_from_filename as _ulf
        ck("unit label: 'External Markup Unit-1.pdf' + D-ref context -> 'Unit 1'",
           _ulf("External Markup Unit-1.pdf") == "Unit 1")
        ck("unit label: D-ref included when present",
           _ulf("Unit_3 D410 Hard Landscaping.pdf") == "Unit 3 (D410)")
        ck("unit label: no unit pattern -> None (caller keeps filename)",
           _ulf("Proposed_Site_Plan.pdf") is None)

        # /upload stores enquiry identification (subject/body) on every job in the batch
        _AS.save_jobs({})
        _id_resp = _client_up.post("/upload", data={
            "project_ref": "IDENT-1", "project_name": "Ident Case",
            "email_subject": "RE: Winwick tender enquiry",
            "email_body": "Please price the attached drawings.",
            "pdf": (_io3.BytesIO(_pdf_a_bytes), "yard.pdf"),
        }, content_type="multipart/form-data")
        _id_jobs = _AS.load_jobs()
        ck("upload stores email_subject/email_body for enquiry identification",
           _id_resp.status_code in (201, 202) and
           all(j.get("email_subject") == "RE: Winwick tender enquiry" and
               "attached drawings" in j.get("email_body", "") for j in _id_jobs.values()),
           list(_id_jobs.values())[:1])

        # count_manholes_marked: unreadable file -> None (couldn't check), never a silent 0
        from robust_takeoff import count_manholes_marked as _cmm
        ck("count_manholes_marked returns None (not 0) for an unreadable file",
           _cmm("/nonexistent/nope.pdf") is None)

        _AS.save_jobs(_route_jobs)   # restore the store for the spec-capture tests below

        # Fortel's supplied Brief_Spec is a blank checklist. Capture applicable fields
        # atomically without touching the job's four-state/measurement record; a partial
        # pricing spec remains assumed even though the unchanged calculation can re-price.
        _spec_job_id = "11111111-1111-4111-8111-111111111111"
        _spec_jobs_before = _AS.load_jobs()
        _spec_jobs_before[_spec_job_id]["measurement_state"] = "MEASURED_VERIFIED"
        _spec_jobs_before[_spec_job_id]["adjusted"] = {
            "area_m2": 100, "scale_k": 0.2,
            "polygon_pts": [[0, 0], [1, 0], [1, 1]],
        }
        _initial_spec_costing = _copy.deepcopy(_spec_jobs_before[_spec_job_id]["costing"])
        _initial_spec_paths = _AS._save_quotation(
            _spec_job_id,
            _AS._quotation_result_for_job(_spec_jobs_before[_spec_job_id]),
            _initial_spec_costing,
            file_stem="QUOTE-MULTI-001-REV_01",
        )
        _spec_jobs_before[_spec_job_id].update({
            "quotation_paths": _initial_spec_paths,
            "quotation_revision": 1,
            "quotation_status": "ready",
            "quotation_history": [{
                "revision": 1, "label": "REV_01",
                "issued_at": "2026-07-15T10:02:00",
                "reason": "initial approval", "paths": dict(_initial_spec_paths),
            }],
        })
        _initial_spec_bytes = {
            fmt: Path(path).read_bytes() for fmt, path in _initial_spec_paths.items()
        }
        _AS.save_jobs(_spec_jobs_before)
        _spec_resp = _client_up.post(f"/spec-override/{_spec_job_id}", json={
            "slab_type": "external_yard",
            "fields": {"depth_mm": 200, "conc_mix": None, "mesh": None, "layers": None,
                       "bay_sizes": "5m x 5m", "joint_details": None},
        })
        _spec_json = _spec_resp.get_json()
        _spec_saved_job = _AS.load_jobs()[_spec_job_id]
        ck("partial Brief_Spec capture succeeds but keeps costing provisional",
           _spec_resp.status_code == 200 and _spec_json["repriced"] is True and
           _spec_json["costing"]["assumed"] is True and
           not _spec_json["brief_spec"]["fields"]["depth_mm"]["provisional"] and
           _spec_json["brief_spec"]["fields"]["mesh"]["provisional"], _spec_json)
        ck("approved specification correction preserves four-state, geometry and assessor approval",
           _spec_saved_job["brief_spec"]["fields"]["bay_sizes"]["value"] == "5m x 5m" and
           _spec_saved_job["measurement_state"] == "MEASURED_VERIFIED" and
           _spec_saved_job["status"] == "approved" and
           _spec_saved_job["decision"] == "approved" and
           _spec_saved_job["adjusted"] == _spec_jobs_before[_spec_job_id]["adjusted"])
        ck("approved specification correction re-costs the assessor-approved area",
           _spec_json["post_approval_correction"] is True and
           _spec_saved_job["costing"]["area_m2"] == 100 and
           _spec_saved_job["costing"]["spec"]["depth_mm"] == 200 and
           _spec_saved_job["costing"]["total_gbp"] !=
               _initial_spec_costing["total_gbp"],
           {"before": _initial_spec_costing, "after": _spec_saved_job["costing"]})
        _revised_spec_paths = _spec_saved_job.get("quotation_paths") or {}
        ck("post-approval correction creates visible REV_02 files and immutable revision history",
           _spec_json["quotation_revision"] == 2 and
           _spec_saved_job["quotation_revision"] == 2 and
           [entry["label"] for entry in _spec_saved_job["quotation_history"]] ==
               ["REV_01", "REV_02"] and
           all("REV_02" in Path(path).stem for path in _revised_spec_paths.values()) and
           all(Path(_initial_spec_paths[fmt]).read_bytes() == original
               for fmt, original in _initial_spec_bytes.items()),
           _spec_saved_job.get("quotation_history"))
        _revised_spec_text = Path(_revised_spec_paths["txt"]).read_text()
        ck("corrected quotation visibly declares correction-after-approval and revision",
           "CORRECTION AFTER APPROVAL" in _revised_spec_text and
           "REV_02" in _revised_spec_text, _revised_spec_text[:500])
        _sibling_revised_text = quotation_text(_AS._quotation_for_job(
            "22222222-2222-4222-8222-222222222222"))
        ck("case correction remains REV_02 with a caveat when downloaded from a sibling drawing",
           "CORRECTION AFTER APPROVAL" in _sibling_revised_text and
           "REV_02" in _sibling_revised_text, _sibling_revised_text[:500])

        # Normal pipeline jobs carry per-zone checklists. Their approved correction route must
        # re-cost the corrected zone (and only that zone), otherwise the portal says REV_02 while
        # the mixed-zone quotation still contains the pre-correction blank/stale rate.
        from slab_spec import empty_brief_spec as _empty_zone_brief
        _zone_spec_job_id = "11111111-1111-4111-8111-111111111119"
        _zone_spec_zones = [
            {"category": "external_yard", "area_m2": 100.0, "measurement_kind": "area"},
            {"category": "dock", "area_m2": 20.0, "measurement_kind": "area"},
        ]
        _zone_briefs = {
            "external_yard": _empty_zone_brief("external_yard"),
            "dock": _empty_zone_brief("dock"),
        }
        _zone_result = {
            "file": "Approved Mixed Zones.pdf", "area_m2": 120.0,
            "measurement_state": "MEASURED_VERIFIED", "scale_k": 0.1,
            "zones": _zone_spec_zones, "brief_specs": _zone_briefs,
            "costing": _copy.deepcopy(_initial_spec_costing),
        }
        _zone_job = {
            "id": _zone_spec_job_id, "project_ref": "ZONE-SPEC-CORR-1",
            "project_name": "Zone correction QA", "status": "approved",
            "decision": "approved", "decided_at": "2026-07-15T11:00:00",
            "measurement_state": "MEASURED_VERIFIED", "scale_confirmed": True,
            "result": _zone_result, "zones": _zone_spec_zones,
            "brief_specs": _zone_briefs, "costing": _copy.deepcopy(_initial_spec_costing),
        }
        _zone_jobs = _AS.load_jobs()
        _zone_jobs[_zone_spec_job_id] = _zone_job
        _AS.save_jobs(_zone_jobs)
        _zone_rev1_paths = _AS._save_quotation(
            _zone_spec_job_id, _AS._quotation_result_for_job(_zone_job),
            _zone_job["costing"], file_stem="ZONE-SPEC-CORR-1-REV_01")
        _zone_jobs = _AS.load_jobs()
        _zone_jobs[_zone_spec_job_id].update({
            "quotation_paths": _zone_rev1_paths, "quotation_revision": 1,
            "quotation_status": "ready", "quotation_history": [{
                "revision": 1, "label": "REV_01", "issued_at": "2026-07-15T11:00:00",
                "reason": "initial approval", "paths": dict(_zone_rev1_paths),
            }],
        })
        _AS.save_jobs(_zone_jobs)
        _zone_spec_resp = _client_up.post(f"/spec-override/{_zone_spec_job_id}", json={
            "zone_category": "external_yard", "fields": {
                "depth_mm": 200, "conc_mix": "C32/40", "mesh": "A252", "layers": 1,
                "bay_sizes": None, "joint_details": None,
            },
        })
        _zone_spec_json = _zone_spec_resp.get_json()
        _zone_spec_saved = _AS.load_jobs()[_zone_spec_job_id]
        _zone_costing = (_zone_spec_saved.get("zone_costings") or {}).get("external_yard")
        ck("approved per-zone specification correction re-costs the exact assessor-scoped zone",
           _zone_spec_resp.status_code == 200 and _zone_spec_json["repriced"] is True and
           _zone_costing and _zone_costing["area_m2"] == 100.0 and
           _zone_costing["spec"]["depth_mm"] == 200 and
           "dock" not in (_zone_spec_saved.get("zone_costings") or {}),
           {"response": _zone_spec_json, "zone_costings": _zone_spec_saved.get("zone_costings")})
        _zone_quote = _AS._quotation_for_job(_zone_spec_job_id)
        _zone_yard_lines = [item for item in _zone_quote["line_items"]
                            if item["section"] == "External yard slabs"]
        _zone_dock_lines = [item for item in _zone_quote["line_items"]
                            if item["section"] == "Dock slabs"]
        ck("zone correction prices Yard in REV_02 while unsupplied Dock stays blank",
           _zone_spec_json["quotation_revision"] == 2 and
           _zone_yard_lines and isinstance(_zone_yard_lines[0].get("rate"), (int, float)) and
           _zone_dock_lines and _zone_dock_lines[0].get("rate") is None,
           {"yard": _zone_yard_lines, "dock": _zone_dock_lines})
        ck("per-zone correction never overwrites the legacy aggregate costing",
           _zone_spec_saved["costing"] == _zone_job["costing"],
           _zone_spec_saved["costing"])

        _bad_spec_resp = _client_up.post(f"/spec-override/{_spec_job_id}", json={
            "slab_type": "upper_floor", "fields": {"bay_sizes": "5m x 5m"},
        })
        ck("non-applicable Brief_Spec field is rejected cleanly",
           _bad_spec_resp.status_code == 400 and "does not apply" in
           (_bad_spec_resp.get_json().get("error") or ""), _bad_spec_resp.get_json())
        _unsupported_resp = _client_up.post(f"/spec-override/{_spec_job_id}", json={
            "slab_type": "external_yard",
            "fields": {"depth_mm": 200, "conc_mix": "Client mix", "mesh": "CLIENT-MESH",
                       "layers": 1, "bay_sizes": None, "joint_details": None},
        })
        _unsupported_json = _unsupported_resp.get_json()
        _unsupported_saved = _AS.load_jobs()[_spec_job_id]
        ck("unsupported open-text client spec is saved without inventing a rate",
           _unsupported_resp.status_code == 200 and
           _unsupported_json["repriced"] is False and
           bool(_unsupported_json["pricing_warning"]) and
           _unsupported_saved["brief_spec"]["fields"]["mesh"]["value"] == "CLIENT-MESH")
        ck("unsupported client spec hard-blocks approval for human pricing review",
           _AS._approve_block_reason(_unsupported_saved) is not None)
        _blocked_quote_resp = _client_up.get(f"/quotation/{_spec_job_id}.json")
        ck("unsupported client spec blocks stale quotation downloads even after prior decision",
           _blocked_quote_resp.status_code == 409 and "human pricing review" in
           (_blocked_quote_resp.get_json().get("error") or ""),
           _blocked_quote_resp.get_json())

        _unmeasured_spec_id = "33333333-3333-4333-8333-333333333333"
        _unmeasured_jobs = _AS.load_jobs()
        _unmeasured_jobs[_unmeasured_spec_id] = {
            "id": _unmeasured_spec_id, "status": "error", "decision": None,
            "measurement_state": "UNMEASURED",
            "result": {"file": "Dock-Unmeasured.pdf", "measurement_state": "UNMEASURED"},
        }
        _AS.save_jobs(_unmeasured_jobs)
        _unmeasured_spec_resp = _client_up.post(
            f"/spec-override/{_unmeasured_spec_id}", json={
                "slab_type": "dock", "fields": {"depth_mm": 225, "conc_mix": None,
                                                    "mesh": None, "layers": None,
                                                    "bay_sizes": None, "joint_details": None},
            })
        ck("Brief_Spec capture works before a drawing has a measurable area",
           _unmeasured_spec_resp.status_code == 200 and
           _unmeasured_spec_resp.get_json()["repriced"] is False and
           _AS.load_jobs()[_unmeasured_spec_id]["brief_spec"]["fields"]["depth_mm"]["value"] == 225,
           _unmeasured_spec_resp.get_json())

        _zone_job_id = "44444444-4444-4444-8444-444444444444"
        _zone_jobs = _AS.load_jobs()
        _zone_jobs[_zone_job_id] = {
            "id": _zone_job_id, "status": "pending", "decision": None,
            "measurement_state": "MEASURED_VERIFIED", "zone_classification_required": True,
            "zones": [{"zone_key":"unclassified:fdns", "category":"unclassified",
                       "subjects":["FDNS"], "measurement_kind":"unparsed",
                       "area_m2":None, "length_lm":None, "annotation_count":4}],
            "markup_annotations": [{"subject":"FDNS", "type":"Polygon"}],
            "flags": ["assessor: classify zone 'FDNS'"],
            "result": {"file":"External Markup Unit-1.pdf", "area_m2":3185.8,
                       "measurement_state":"MEASURED_VERIFIED",
                       "zone_classification_required":True,
                       "zones":[{"zone_key":"unclassified:fdns", "category":"unclassified",
                                 "subjects":["FDNS"], "measurement_kind":"unparsed",
                                 "area_m2":None, "length_lm":None, "annotation_count":4}],
                       "markup_annotations":[{"subject":"FDNS", "type":"Polygon"}],
                       "flags":["assessor: classify zone 'FDNS'"]},
        }
        _AS.save_jobs(_zone_jobs)
        ck("unclassified marked zone hard-blocks approval",
           _AS._approve_block_reason(_zone_jobs[_zone_job_id]) is not None)
        _classify_resp = _client_up.post(f"/zones/{_zone_job_id}", json={
            "classifications":[{"zone_key":"unclassified:fdns", "category":"other"}],
        })
        _classified_job = _AS.load_jobs()[_zone_job_id]
        ck("assessor can classify out-of-scope FDNS without changing its measurement",
           _classify_resp.status_code == 200 and
           _classified_job["zones"][0]["category"] == "other" and
           not _classified_job["zone_classification_required"] and
           _classified_job["markup_annotations"] == [{"subject":"FDNS", "type":"Polygon"}],
           _classify_resp.get_json())
        _ack_jobs = _AS.load_jobs()
        _ack_jobs[_zone_job_id]["zone_reference_mismatch"] = True
        _ack_jobs[_zone_job_id]["result"]["zone_reference_mismatch"] = True
        _AS.save_jobs(_ack_jobs)
        _ack_resp = _client_up.post(f"/zones/{_zone_job_id}", json={
            "acknowledge_reference_mismatch": True,
        })
        _ack_job = _AS.load_jobs()[_zone_job_id]
        ck("assessor can explicitly acknowledge a BOQ mismatch before approval",
           _ack_resp.status_code == 200 and not _ack_job["zone_reference_mismatch"] and
           _ack_job["result"].get("zone_reference_reviewed_at") and
           _AS._approve_block_reason(_ack_job) is None, _ack_resp.get_json())

        _mixed_zone_id = "55555555-5555-4555-8555-555555555555"
        _mixed_jobs = _AS.load_jobs()
        _mixed_costing = _copy.deepcopy(_demo_result["costing"])
        _mixed_jobs[_mixed_zone_id] = {
            "id": _mixed_zone_id, "status":"pending", "decision":None,
            "measurement_state":"MEASURED_VERIFIED", "costing":_mixed_costing,
            "zones":[{"zone_key":"external_yard", "category":"external_yard", "area_m2":100},
                     {"zone_key":"dock", "category":"dock", "area_m2":20}],
            "brief_specs":{"external_yard":_empty_brief_spec("external_yard"),
                           "dock":_empty_brief_spec("dock")},
            "result":{"file":"External Markup Unit-9.pdf", "area_m2":120,
                      "measurement_state":"MEASURED_VERIFIED", "costing":_mixed_costing,
                      "zones":[{"zone_key":"external_yard", "category":"external_yard", "area_m2":100},
                               {"zone_key":"dock", "category":"dock", "area_m2":20}],
                      "brief_specs":{"external_yard":_empty_brief_spec("external_yard"),
                                     "dock":_empty_brief_spec("dock")}},
        }
        _AS.save_jobs(_mixed_jobs)
        _zone_spec_resp = _client_up.post(f"/spec-override/{_mixed_zone_id}", json={
            "zone_category":"dock", "slab_type":"dock",
            "fields":{"depth_mm":250, "conc_mix":None, "mesh":None, "layers":None,
                      "bay_sizes":None, "joint_details":None},
        })
        _zone_spec_job = _AS.load_jobs()[_mixed_zone_id]
        ck("per-zone slab checklist re-prices only its zone without overwriting aggregate rate",
           _zone_spec_resp.status_code == 200 and _zone_spec_resp.get_json()["repriced"] and
           _zone_spec_job["brief_specs"]["dock"]["fields"]["depth_mm"]["value"] == 250 and
           _zone_spec_job["zone_costings"]["dock"]["area_m2"] == 20 and
           _zone_spec_job["zone_costings"]["dock"]["spec"]["depth_mm"] == 250 and
           _zone_spec_job["costing"] == _mixed_costing,
           _zone_spec_resp.get_json())
        _zone_adjust_resp = _client_up.post(f"/adjust/{_mixed_zone_id}", json={
            "assessed_area_m2":125, "note":"aggregate correction",
        })
        _zone_adjusted_job = _AS.load_jobs()[_mixed_zone_id]
        ck("aggregate adjustment clears stale split and re-blocks zone approval",
           _zone_adjust_resp.status_code == 200 and _zone_adjusted_job["zones"] == [] and
           _zone_adjusted_job["zone_allocation_stale"] and
           _AS._approve_block_reason(_zone_adjusted_job) is not None)
        _empty_stale_confirm = _client_up.post(
            f"/confirm-measurement/{_mixed_zone_id}", json={"confirm_scale_extent":True})
        ck("confirmation cannot resurrect zones erased by a true aggregate replacement",
           _empty_stale_confirm.status_code == 409 and
           _AS.load_jobs()[_mixed_zone_id]["zone_allocation_stale"],
           _empty_stale_confirm.get_json())

        _portal_html_up = (Path(_orig_server_file_up).parent / "assessor_portal.html").read_text()
        ck("portal file input allows multiple PDFs and ZIPs",
           'accept=".pdf,.zip" multiple' in _portal_html_up)
        ck("portal submits every selected file under the backward-compatible pdf field",
           "files.forEach(file => fd.append('pdf', file))" in _portal_html_up)
        ck("portal groups repeated project refs under collapsible project headers",
           "projectCounts.get(ref)" in _portal_html_up and
           'class="project-group-header' in _portal_html_up and
           "toggleProjectGroup(this)" in _portal_html_up)
        ck("portal exposes editable xlsx quotation download",
           'id="linkXlsx"' in _portal_html_up and
           "quotation/${job.id}.xlsx" in _portal_html_up)
        ck("portal exposes marked-up PDF independently of quotation pricing state",
           all(marker in _portal_html_up for marker in (
               'id="markedPdfLinks"', 'id="linkMarkedPdf"',
               "marked-pdf/${job.id}.pdf", "Permanent Bluebeam-ready markup",
               "recoverable Fortel geometry embedded")))
        ck("portal exposes renameable +Area geometry separately from main regions",
           all(marker in _portal_html_up for marker in (
               'id="btnNewArea"', "startNewAreaElement",
               "Separate area name", "area_elements: namedAreaEntries.map",
               "main area:")))
        ck("portal accumulates folder selections and can target an existing project anchor",
           all(marker in _portal_html_up for marker in (
               "selectedUploadFiles.push(...Array.from(fileList || []))",
               "document.getElementById('upFile').value = ''",
               'id="upFileList"', 'id="btnAddDrawings"',
               "beginAddDrawingsById", "existing_project_job_id")))
        ck("portal switches to a scalable PDF visual layer on zoom without changing snapScale",
           all(marker in _portal_html_up for marker in (
               "ensureVectorSurface", "snapshot-vector/${encodeURIComponent(requestedJobId)}.svg",
               "if (!vectorSurfaceReady && img)",
               "if (vectorSurface) vectorSurface.style.transform = transform",
               "res.scale_k ? res.scale_k / snapScale : null")))
        ck("portal exposes exact Brief_Spec fields without silent fallback form values",
           all(label in _portal_html_up for label in (
               "External/Service Yard Slabs", "Dock Slabs", "Ground Floor Slabs(Core Areas)",
               "Upper Floors", "Bay sizes if joint layout available", "Nr of mesh layers")) and
           "${spec.depth_mm||190}" not in _portal_html_up and
           "${esc(spec.mesh||'A252')}" not in _portal_html_up and
           "ASSUMED / no details provided" in _portal_html_up and
           "projectPricingBlocked" in _portal_html_up)
        ck("portal displays drawing-file and page citations only when extraction evidence exists",
           all(marker in _portal_html_up for marker in (
               "function specSourceCitation", "evidence.file", "evidence.page",
               "Source: ${esc(citation)}")))
        ck("portal permits an approved spec correction and labels its immutable quote revision",
           all(marker in _portal_html_up for marker in (
               "Correct approved specification", "Save as new quotation revision",
               "post_approval_correction", "CORRECTED AFTER APPROVAL", "quotation_revision")))
        ck("portal presents the measurement scale as both internal m/px and conventional 1:N",
           all(marker in _portal_html_up for marker in (
               'id="scaleDisplay"', 'id="scaleRatioDisplay"', "Drawing scale ratio",
               "(mpp * snapScale) * 72 / 0.0254")))
        ck("portal renders and captures per-zone quantities/classifications/specs",
           all(marker in _portal_html_up for marker in (
               "Measured zones", "ZONE REVIEW REQUIRED", "classifyZone(",
               "acknowledgeZoneReferenceMismatch", "zone_category", "effectiveBriefSpecs")))
        ck("portal shows correct banner for yard-region review vs zone classification",
           all(marker in _portal_html_up for marker in (
               "yardRegionReview", "YARD REGION REVIEW REQUIRED",
               "Multiple same-tint Yard regions detected",
               "yard_region_review_required")))
        ck("portal exposes assisted Office candidates without auto-submitting them",
           all(marker in _portal_html_up for marker in (
               "ASSISTED TRACE CANDIDATES", "candidate_polygons", "loadTraceCandidate(",
               "Add to trace", "btnNewRegion", "traceRegions", 'id="traceScope"',
               "ground_floor_core", "main_upper_floor", "plant_deck",
               "pod_first_floor", "region_scopes: regionEntries.map")) and
           "function loadTraceCandidate" in _portal_html_up and
           "const proposed = candidatePolygons.find" in _portal_html_up and
           "loadTraceCandidate(proposed.candidate_id)" in _portal_html_up and
           "regions: regionPayload" in _portal_html_up and
           "region_categories: regionEntries.map" in _portal_html_up)
        ck("portal offers non-destructive existing scale+extent confirmation",
           all(marker in _portal_html_up for marker in (
               "btnConfirmExisting", "confirmExistingMeasurement",
               "/confirm-measurement/", "Confirm scale + extent")))
        ck("portal explains candidate confidence and keeps unresolved levels visible",
           all(marker in _portal_html_up for marker in (
               "confidence_reasons", "confidence_score", "outline_status",
               "Trace manually", "candidate.regions")))
        ck("portal labels channel proposals as assumptions and provides review controls",
           all(marker in _portal_html_up for marker in (
               "ASSUMED CHANNEL PROPOSALS - NOT MEASURED OR PRICED",
               "channel_proposals", "reviewChannelProposal(",
               "/channel-proposals/", "Accept / save edit", "Remove",
               "Dock-level retaining-wall/loading face", "Full Yard width (no Dock level)")))
        ck("portal gives Transition candidates the full accept/edit/remove lifecycle",
           all(marker in _portal_html_up for marker in (
               "ASSUMED TRANSITION CANDIDATES - NOT PRICED UNTIL REVIEWED",
               "transition_candidate_decisions", "editTransitionLength(",
               "reviewTransitionCandidate(", "/transition-candidates/",
               "Accept / save edit", "Remove", "blank assessor rate")))
        ck("portal visibly carries exclusion checks and construction-joint classification",
           all(marker in _portal_html_up for marker in (
               "SLAB EXCLUSIONS", "CHECK EXCLUSION", "EXCLUDED ·",
               "Construction joint (CJ)", "construction_joint")))
        ck("portal supports axis-locked endpoint drag plus numeric channel geometry edits",
           all(marker in _portal_html_up for marker in (
               "channelDrag", "function editChannelLength", "syncChannelLengthInput",
               "line remains straight/non-diagonal", "polyline_pts:polylinePts")))
        ck("portal separately supports assessor-drawn bending channel polylines",
           all(marker in _portal_html_up for marker in (
               "click each bend", "active.points.push(p)", "active.points.length",
               "ch.points.length >= 2", "all bends reach the server/quotation")))
        ck("portal surfaces every retained Yard region and posts explicit keep/exclude decisions",
           all(marker in _portal_html_up for marker in (
               "effectiveYardRegions", "yard-region-toggle", "saveYardRegionReview",
               "/yard-regions/", "Save kept/excluded regions", "bbox_pdf_pts")))
        ck("portal gives assumed/proposed/provisional values one unmissable provenance style",
           all(marker in _portal_html_up for marker in (
               'class="assumption-badge" data-provenance="assumed"',
               "assumption-item", "assumption-legend", "assumption-basis",
               "assumptionBadge('PROPOSED'", "assumptionBadge('PROVISIONAL'",
               "assumptionBadge('ASSUMED'", "assumptionBadge('ESTIMATED'")))
        _zone_assumption_fn = _portal_html_up.split("function zoneAssumption", 1)[1].split(
            "function provenanceFlag", 1)[0]
        ck("measured zone quantities never gain an assumption badge without explicit provenance",
           all(marker in _zone_assumption_fn for marker in (
               "zone.assumed", "zone.proposed", "zone.provisional", "zone.estimate")) and
           all(marker not in _zone_assumption_fn for marker in (
               "measurement_state", "measurement_kind", "area_m2", "length_lm")))
        _candidate_fn = _portal_html_up.split("function loadTraceCandidate", 1)[1].split(
            "function calcArea", 1)[0]
        ck("one-click candidate load is non-mutating until Submit Adjustment",
           "fetch(" not in _candidate_fn and "poly = regions[0]" in _candidate_fn)
    finally:
        _AS._TAKEOFF_DISPATCHER = _orig_dispatcher_up
        _AS.__file__ = _orig_server_file_up
        _AS.JOBS_FILE = _orig_jobs_file_up
        _AS.JOBS_ARCHIVE_FILE = _orig_jobs_archive_file_up
        _AS.BACKUP_DIR = _orig_backup_dir_up
        _AS.DRAWINGS_DIR = _orig_drawings_dir_up
        _AS.QUOTATIONS_DIR = _orig_quotations_dir_up

    # approve hard-block mirrors the >£200k escalation guard mechanism (fb5b92b)
    ck("UNMEASURED job blocks approve",
       _AS._approve_block_reason({"measurement_state": "UNMEASURED", "scale_confirmed": False}) is not None)
    ck("MEASURED_UNVERIFIED job blocks approve",
       _AS._approve_block_reason({"measurement_state": "MEASURED_UNVERIFIED", "scale_confirmed": False}) is not None)
    ck("MEASURED_VERIFIED job does not block approve",
       _AS._approve_block_reason({"measurement_state": "MEASURED_VERIFIED", "scale_confirmed": False}) is None)
    ck("assessor-confirmed UNMEASURED job no longer blocks approve",
       _AS._approve_block_reason({"measurement_state": "UNMEASURED", "scale_confirmed": True}) is None)
    ck("REJECTED job blocks approve", _AS._approve_block_reason({"measurement_state": "REJECTED"}) is not None)

    shutil.rmtree(_tmpdir, ignore_errors=True)
except ImportError as _e:
    print(f"  [SKIP] approval_server upload/approve tests — missing dependency: {_e}")

print("approval_server: bounded 26-PDF queue excludes wait time from watchdog budget")
try:
    import approval_server as _AS_queue
    import fitz as _fitz_queue, io as _io_queue, tempfile as _tempfile_queue
    import sys as _sys_queue, time as _time_queue, threading as _threading_queue

    _queue_tmp = Path(_tempfile_queue.mkdtemp(prefix="ci_takeoff_queue_"))
    _queue_originals = {
        "jobs": _AS_queue.JOBS_FILE,
        "archive": _AS_queue.JOBS_ARCHIVE_FILE,
        "backup": _AS_queue.BACKUP_DIR,
        "drawings": _AS_queue.DRAWINGS_DIR,
        "dispatcher": _AS_queue._TAKEOFF_DISPATCHER,
        "timeout": _AS_queue.TAKEOFF_TIMEOUT_S,
    }
    _real_pipeline_queue = _sys_queue.modules.get("takeoff_pipeline")
    _active_queue = 0
    _max_active_queue = 0
    _active_lock_queue = _threading_queue.Lock()

    ck("takeoff worker default is CPU-sized and deliberately capped at two",
       _AS_queue._takeoff_worker_count(raw_value="", cpu_count=8) == 2 and
       _AS_queue._takeoff_worker_count(raw_value="", cpu_count=1) == 1)
    ck("TAKEOFF_WORKERS explicitly overrides the CPU-sized default",
       _AS_queue._takeoff_worker_count(raw_value="3", cpu_count=1) == 3)

    class _QueuePipeline:
        @staticmethod
        def takeoff(pdf_path, project_name=None, project_ref=None,
                    client_rates_path=None, approval_job_id=None):
            global _active_queue, _max_active_queue
            with _active_lock_queue:
                _active_queue += 1
                _max_active_queue = max(_max_active_queue, _active_queue)
            try:
                _time_queue.sleep(0.03)
                return {
                    "file": Path(pdf_path).name,
                    "project_name": project_name,
                    "project_ref": project_ref,
                    "area_m2": 250.0,
                    "measurement_state": "MEASURED_VERIFIED",
                    "needs_assessor": False,
                    "scale_verified": True,
                    "flags": ["queue regression stub completed"],
                }
            finally:
                with _active_lock_queue:
                    _active_queue -= 1

    try:
        _AS_queue.JOBS_FILE = _queue_tmp / "approval_jobs.json"
        _AS_queue.JOBS_ARCHIVE_FILE = _queue_tmp / "approval_jobs_archive.json"
        _AS_queue.BACKUP_DIR = _queue_tmp / "backups"
        _AS_queue.DRAWINGS_DIR = _queue_tmp / "drawings"
        _AS_queue.TAKEOFF_TIMEOUT_S = 0.15
        _AS_queue._TAKEOFF_DISPATCHER = _AS_queue._TakeoffDispatcher(2)
        _sys_queue.modules["takeoff_pipeline"] = _QueuePipeline
        _AS_queue.save_jobs({})

        _queue_doc = _fitz_queue.open()
        _queue_doc.new_page(width=300, height=200)
        _queue_bytes = _queue_doc.tobytes()
        _queue_doc.close()
        _queue_files = [
            (_io_queue.BytesIO(_queue_bytes), f"Tender_Drawing_{index:02d}.pdf")
            for index in range(1, 27)
        ]
        _queue_started = _time_queue.monotonic()
        _queue_response = _AS_queue.app.test_client().post("/upload", data={
            "project_ref": "QUEUE-26",
            "project_name": "26 Drawing Tender Pack",
            "pdf": _queue_files,
        }, content_type="multipart/form-data")
        _AS_queue._TAKEOFF_DISPATCHER.wait_for_idle()
        _queue_elapsed = _time_queue.monotonic() - _queue_started
        _queue_jobs = _AS_queue.load_jobs()
        _queue_states = [job.get("measurement_state") for job in _queue_jobs.values()]
        _queue_timeout_flags = [
            flag for job in _queue_jobs.values() for flag in (job.get("flags") or [])
            if "PIPELINE TIMEOUT" in flag
        ]

        ck("26-PDF upload returns one queued job per drawing",
           _queue_response.status_code == 202 and len(_queue_jobs) == 26 and
           len(_queue_response.get_json().get("job_ids", [])) == 26,
           {"http": _queue_response.status_code,
            "jobs": len(_queue_jobs), "body": _queue_response.get_json()})
        ck("bounded takeoff queue never exceeds configured worker concurrency",
           _max_active_queue == 2, _max_active_queue)
        ck("every queued drawing reaches a legitimate four-state outcome",
           all(state in {"MEASURED_VERIFIED", "MEASURED_UNVERIFIED", "UNMEASURED", "REJECTED"}
               for state in _queue_states) and
           all(job.get("takeoff_phase") == "completed" for job in _queue_jobs.values()),
           {"states": _queue_states,
            "phases": [job.get("takeoff_phase") for job in _queue_jobs.values()]})
        ck("queue wait longer than one watchdog budget causes zero watchdog failures",
           _queue_elapsed > _AS_queue.TAKEOFF_TIMEOUT_S and not _queue_timeout_flags and
           not any(job.get("status") == "error" for job in _queue_jobs.values()),
           {"elapsed_s": round(_queue_elapsed, 3),
            "watchdog_s": _AS_queue.TAKEOFF_TIMEOUT_S,
            "timeout_flags": _queue_timeout_flags})
        ck("26-file bounded batch finishes promptly",
           _queue_elapsed < 5.0, round(_queue_elapsed, 3))
        print(f"  [EVIDENCE] 26 files, workers=2, watchdog=0.15s, "
              f"batch={_queue_elapsed:.3f}s, max_active={_max_active_queue}, "
              f"watchdog_kills={len(_queue_timeout_flags)}")
    finally:
        _AS_queue.JOBS_FILE = _queue_originals["jobs"]
        _AS_queue.JOBS_ARCHIVE_FILE = _queue_originals["archive"]
        _AS_queue.BACKUP_DIR = _queue_originals["backup"]
        _AS_queue.DRAWINGS_DIR = _queue_originals["drawings"]
        _AS_queue._TAKEOFF_DISPATCHER = _queue_originals["dispatcher"]
        _AS_queue.TAKEOFF_TIMEOUT_S = _queue_originals["timeout"]
        if _real_pipeline_queue is None:
            _sys_queue.modules.pop("takeoff_pipeline", None)
        else:
            _sys_queue.modules["takeoff_pipeline"] = _real_pipeline_queue
        shutil.rmtree(_queue_tmp, ignore_errors=True)
except ImportError as _e:
    print(f"  [SKIP] bounded takeoff queue tests — missing dependency: {_e}")

print("approval_server: /snapshot status codes for all four measurement states "
      "(Aryan field report — 'session which renders screenshots is not working properly')")
try:
    import approval_server as _AS2
    import fitz as _fitz4, uuid as _uuid2, tempfile as _tempfile2

    _client = _AS2.app.test_client()
    _tmpdir2 = Path(_tempfile2.mkdtemp(prefix="ci_snapshot_"))

    # Save/restore the real jobs file around this block — snapshot() reads via load_jobs()
    # which is a real file read, not mockable without a live Flask app context.
    _jobs_backup = _AS2.JOBS_FILE.read_text() if _AS2.JOBS_FILE.exists() else None

    def _mk_pdf(path, w=600, h=400, n_pages=1):
        d = _fitz4.open()
        for _ in range(n_pages):
            d.new_page(width=w, height=h)
        d.save(str(path))
        return path

    try:
        _jobs = _AS2.load_jobs()

        # 1. REJECTED job (no pdf_path at all) -> 404, not 500
        _jid_rej = str(_uuid2.uuid4())
        _jobs[_jid_rej] = {"id": _jid_rej, "status": "rejected", "measurement_state": "REJECTED",
                           "pdf_path": None, "result": {"measurement_state": "REJECTED"}}

        # 2. UNMEASURED job with a real PDF on disk -> 200 (assessor still needs to see it to trace)
        _pdf_unm = _mk_pdf(_tmpdir2 / "unmeasured.pdf")
        _jid_unm = str(_uuid2.uuid4())
        _jobs[_jid_unm] = {"id": _jid_unm, "status": "error", "measurement_state": "UNMEASURED",
                           "pdf_path": str(_pdf_unm),
                           "result": {"pdf_path": str(_pdf_unm), "page": 0, "measurement_state": "UNMEASURED"}}

        # 3. UNMEASURED job whose PDF is missing from disk (temp dir cleaned up) -> 404, not 500
        _jid_gone = str(_uuid2.uuid4())
        _jobs[_jid_gone] = {"id": _jid_gone, "status": "error", "measurement_state": "UNMEASURED",
                            "pdf_path": str(_tmpdir2 / "does_not_exist.pdf"),
                            "result": {"pdf_path": str(_tmpdir2 / "does_not_exist.pdf"),
                                      "measurement_state": "UNMEASURED"}}

        # 4. MEASURED_VERIFIED multi-page job whose result["page"] != 0 -> snapshot must render
        # THAT page (this was the root cause of "AI polygon not shown": /snapshot always
        # rendered page 0 regardless of which page the pipeline actually measured).
        _pdf_multi = _mk_pdf(_tmpdir2 / "multi.pdf", n_pages=3)
        _jid_page = str(_uuid2.uuid4())
        _jobs[_jid_page] = {"id": _jid_page, "status": "pending", "measurement_state": "MEASURED_VERIFIED",
                            "pdf_path": str(_pdf_multi),
                            "result": {"pdf_path": str(_pdf_multi), "page": 2,
                                      "polygon_pts": [[10, 10], [100, 10], [100, 100], [10, 100]],
                                      "measurement_state": "MEASURED_VERIFIED"}}

        # 5. Out-of-range page index (stale data) -> must fall back to page 0, never 500
        _jid_badpage = str(_uuid2.uuid4())
        _jobs[_jid_badpage] = {"id": _jid_badpage, "status": "pending", "measurement_state": "MEASURED_VERIFIED",
                               "pdf_path": str(_pdf_multi),
                               "result": {"pdf_path": str(_pdf_multi), "page": 99,
                                         "measurement_state": "MEASURED_VERIFIED"}}

        _AS2.save_jobs(_jobs)

        _r_rej = _client.get(f"/snapshot/{_jid_rej}")
        ck("REJECTED job snapshot -> 404 (not 500)", _r_rej.status_code == 404, _r_rej.status_code)

        _r_unm = _client.get(f"/snapshot/{_jid_unm}")
        ck("UNMEASURED job with PDF on disk -> 200 (assessor can still trace)",
           _r_unm.status_code == 200, _r_unm.status_code)

        _r_gone = _client.get(f"/snapshot/{_jid_gone}")
        ck("UNMEASURED job with missing PDF -> 404 (not 500)", _r_gone.status_code == 404, _r_gone.status_code)

        _r_page = _client.get(f"/snapshot/{_jid_page}")
        ck("multi-page job snapshot -> 200", _r_page.status_code == 200, _r_page.status_code)
        # Verify it actually rendered page 2's dimensions, not page 0's (both pages here are
        # the same size so we check indirectly: render page 2 directly and diff against the
        # response bytes' pixel dimensions via the PNG header — same width guaranteed by
        # construction, so the meaningful assertion is the X-Snapshot-Scale header matches
        # snapshot_scale() computed for page 2 specifically.
        from approval_email import snapshot_scale as _snap_scale_fn
        _expected_scale = f"{_snap_scale_fn(str(_pdf_multi), page=2):.6f}"
        ck("multi-page snapshot X-Snapshot-Scale computed for the MEASURED page (not page 0)",
           _r_page.headers.get("X-Snapshot-Scale") == _expected_scale,
           (_r_page.headers.get("X-Snapshot-Scale"), _expected_scale))

        _r_badpage = _client.get(f"/snapshot/{_jid_badpage}")
        ck("out-of-range page index falls back to page 0 (not 500)",
           _r_badpage.status_code == 200, _r_badpage.status_code)

        _r_404job = _client.get(f"/snapshot/{_uuid2.uuid4()}")
        ck("nonexistent job -> 404", _r_404job.status_code == 404, _r_404job.status_code)

    finally:
        if _jobs_backup is not None:
            _AS2.JOBS_FILE.write_text(_jobs_backup)
        shutil.rmtree(_tmpdir2, ignore_errors=True)
except ImportError as _e:
    print(f"  [SKIP] approval_server snapshot tests — missing dependency: {_e}")

print("approval_server: watchdog-vs-completion race (Aryan field report — 'server is unstable')")
try:
    import approval_server as _AS3
    import sys as _sys3, time as _time3, uuid as _uuid3
    from unittest import mock as _mock3

    _jobs_backup3 = _AS3.JOBS_FILE.read_text() if _AS3.JOBS_FILE.exists() else None
    try:
        _jid_wd = str(_uuid3.uuid4())
        _jobs3 = _AS3.load_jobs()
        _jobs3[_jid_wd] = {"id": _jid_wd, "status": "processing", "flags": []}
        _AS3.save_jobs(_jobs3)

        # _mark_job_unmeasured with watchdog_fired=True sets the sentinel used to detect the race
        _AS3._mark_job_unmeasured(_jid_wd, "PIPELINE TIMEOUT: took too long", watchdog_fired=True)
        _j_after_wd = _AS3.load_jobs()[_jid_wd]
        ck("watchdog fire sets _watchdog_fired sentinel", _j_after_wd.get("_watchdog_fired") is True)
        ck("watchdog fire flips job to UNMEASURED", _j_after_wd.get("measurement_state") == "UNMEASURED")
        ck("watchdog fire records a PIPELINE TIMEOUT flag",
           any("PIPELINE TIMEOUT" in f for f in _j_after_wd.get("flags", [])))

        # Now simulate the pipeline finishing LATE (after the watchdog already fired) by
        # driving the real _run_takeoff() with a stubbed takeoff_pipeline module whose
        # takeoff() sleeps past a 1s watchdog timeout — exercises the actual production
        # code path, not a re-implementation of its logic.
        _orig_timeout = _AS3.TAKEOFF_TIMEOUT_S
        _AS3.TAKEOFF_TIMEOUT_S = 1
        _jid_wd2 = str(_uuid3.uuid4())
        _jobs3 = _AS3.load_jobs()
        _jobs3[_jid_wd2] = {"id": _jid_wd2, "status": "processing", "flags": []}
        _AS3.save_jobs(_jobs3)

        _fake_pipeline = _mock3.MagicMock()
        def _slow_takeoff(pdf_path, project_name=None, project_ref=None):
            _time3.sleep(2.2)
            return {"measurement_state": "MEASURED_VERIFIED", "area_m2": 3159.0,
                    "flags": ["completed ok"], "project_name": project_name, "project_ref": project_ref,
                    "candidate_polygons":[{"candidate_id":"office-p0-level-01-1",
                                           "polygon_pts":[[0,0],[1,0],[1,1]]}]}
        _fake_pipeline.takeoff = _slow_takeoff
        _real_module = _sys3.modules.get("takeoff_pipeline")
        _sys3.modules["takeoff_pipeline"] = _fake_pipeline
        try:
            _AS3._run_takeoff(_jid_wd2, "drawings/_int_d77.pdf", "QA WD race", "QA-PORTAL-CI-WDRACE")
        finally:
            if _real_module is not None:
                _sys3.modules["takeoff_pipeline"] = _real_module
            else:
                _sys3.modules.pop("takeoff_pipeline", None)
            _AS3.TAKEOFF_TIMEOUT_S = _orig_timeout

        _j_final = _AS3.load_jobs()[_jid_wd2]
        ck("late pipeline completion overwrites watchdog UNMEASURED with the real result",
           _j_final.get("measurement_state") == "MEASURED_VERIFIED", _j_final.get("measurement_state"))
        ck("stale 'PIPELINE TIMEOUT' flag stripped once the pipeline actually completes",
           not any("PIPELINE TIMEOUT" in f for f in _j_final.get("flags", [])), _j_final.get("flags"))
        ck("_watchdog_fired sentinel cleared after the race resolves",
           "_watchdog_fired" not in _j_final)
        ck("background takeoff mirrors assisted candidates at job and result level",
           _j_final.get("candidate_polygons") ==
           _j_final.get("result", {}).get("candidate_polygons") and
           _j_final.get("candidate_polygons", [])[0]["candidate_id"] ==
           "office-p0-level-01-1")

        _jobs3 = _AS3.load_jobs()
        _jobs3.pop(_jid_wd, None); _jobs3.pop(_jid_wd2, None)
        _AS3.save_jobs(_jobs3)
    finally:
        if _jobs_backup3 is not None:
            _AS3.JOBS_FILE.write_text(_jobs_backup3)
except ImportError as _e:
    print(f"  [SKIP] approval_server watchdog-race tests — missing dependency: {_e}")

print("approval_server: approval_jobs.json concurrent read/write does not raise "
      "(Aryan field report — 'the server is unstable')")
try:
    import approval_server as _AS4
    import threading as _threading4, tempfile as _tempfile4

    _tmp_jobs_file = Path(_tempfile4.mkdtemp(prefix="ci_atomic_")) / "jobs.json"
    _orig_jobs_file = _AS4.JOBS_FILE
    _AS4.JOBS_FILE = _tmp_jobs_file
    try:
        _big = {str(_i): {"x": "y" * 500} for _i in range(500)}
        _AS4.save_jobs(_big)

        _errors4 = []
        def _reader4():
            for _ in range(150):
                try:
                    _d = _AS4.load_jobs()
                    if not isinstance(_d, dict):
                        _errors4.append("load_jobs did not return a dict")
                except Exception as _e:
                    _errors4.append(str(_e))

        def _writer4():
            for _ in range(150):
                _AS4.save_jobs(_big)

        _t1 = _threading4.Thread(target=_reader4)
        _t2 = _threading4.Thread(target=_writer4)
        _t1.start(); _t2.start(); _t1.join(); _t2.join()

        ck("concurrent load_jobs()/save_jobs() never raises or returns a torn read",
           len(_errors4) == 0, _errors4[:3])
        ck("no leftover .tmp files after concurrent saves",
           list(_tmp_jobs_file.parent.glob("*.tmp*")) == [])
    finally:
        _AS4.JOBS_FILE = _orig_jobs_file
        shutil.rmtree(_tmp_jobs_file.parent, ignore_errors=True)
except ImportError as _e:
    print(f"  [SKIP] approval_server atomic-write tests — missing dependency: {_e}")

print("D77 swatch-locked grey band vs 'Footpaths (ancillary): Concrete' annexation "
      "(Aryan field report: real SGP sheet measured 3,172 vs Smita gold 3,156 — root cause was "
      "the generic 214±14 grey band admitting a darker, adjacent ancillary-concrete legend "
      "colour and binary_closing fusing it into the yard's own connected component)")
try:
    _require_fixture("drawings/_int_d77.pdf", "D77 swatch-locked grey band test")
    import fitz as _fitz_fp
    from takeoff_unmarked import (takeoff as _tu_takeoff_fp, segment_hatch as _seg_fp,
                                   PLAUSIBLE_MIN_M2 as _PMIN_FP, PLAUSIBLE_MAX_M2 as _PMAX_FP)

    def _gen_d77_footpath(out_path, chip_grey=0.878, yard_grey=0.878):
        """Same D77 yard geometry (page 1067.766x824.854pt, scale bar, 1:250 title) plus:
          - a darker 'Footpaths (ancillary): Concrete' strip (204 grey) sitting 0.65pt below
            the yard's bottom edge — close enough for binary_closing (any close>=2) to bridge,
            reproducing the real-sheet CONNECTED over-measure (not a satellite blob).
          - a legend swatch chip + label 'Concrete Service Yard construction' (readable by
            find_concrete_swatch_rgb) so the swatch-lock path engages.
          - a second, non-matching legend line 'Footpaths (ancillary): Concrete' with its own
            (darker) swatch chip — must NOT be picked up as the concrete-yard label anchor.
        Title text deliberately avoids CONCRETE_LABELS substrings (unlike drawings/_int_d77.pdf,
        whose title text IS the label match and has no nearby swatch chip -> unreadable swatch,
        which is why that fixture stays on the generic-band fallback path untouched by this fix).
        """
        d = _fitz_fp.open()
        W, H = 1067.7659912109375, 824.853515625
        pg = d.new_page(width=W, height=H)
        pg.insert_text((130.0, 80), "PROPOSED HARD LANDSCAPING - UNIT 1 SITE PLAN    Scale 1:250",
                       fontsize=13)
        pg.draw_line(_fitz_fp.Point(130.0, 714.853515625), _fitz_fp.Point(696.9290771484375, 714.853515625),
                     color=(0, 0, 0), width=2.0)
        pg.insert_text((126.0, 731), "0", fontsize=11)
        pg.insert_text((678.9, 731), "50 m", fontsize=11)

        yg = (yard_grey, yard_grey, yard_grey)
        pg.draw_rect(_fitz_fp.Rect(130.0, 120.0, 937.765625, 624.853515625),
                     color=(0, 0, 0), fill=yg, width=1.0)

        # Ancillary footpath strip: 230x9pt = 16.1 m² at k=0.08819, darker grey (204), 0.65pt
        # gap below the yard's own bottom edge (bridged by binary_closing regardless of the
        # exact close value, same mechanism as the real sheet's kerb-line gap).
        strip_grey = (0.80, 0.80, 0.80)
        pg.draw_rect(_fitz_fp.Rect(350.0, 625.503515625, 580.0, 634.503515625),
                     color=None, fill=strip_grey, width=0)

        # Legend: matching swatch chip + label (concrete-yard anchor).
        cg = (chip_grey, chip_grey, chip_grey)
        pg.draw_rect(_fitz_fp.Rect(330.0, 762.0, 360.0, 776.0), color=(0, 0, 0), fill=cg, width=0.5)
        pg.insert_text((400.0, 772.0), "Concrete Service Yard construction", fontsize=9)

        # Second legend line: non-matching label + its own (darker) swatch chip — must not be
        # mistaken for the concrete-yard anchor, and is small/isolated -> satellite-dropped.
        pg.draw_rect(_fitz_fp.Rect(330.0, 784.0, 360.0, 796.0), color=(0, 0, 0), fill=strip_grey, width=0.5)
        pg.insert_text((400.0, 792.0), "Footpaths (ancillary): Concrete", fontsize=9)

        d.save(out_path)
        d.close()

    _p_fp = "drawings/_int_d77_footpath.pdf"
    _gen_d77_footpath(_p_fp)

    # BEFORE: old generic-band segmentation (direct segment_hatch call, mirroring the borders
    # test's "WITHOUT exclusion" pattern) — proves the annexation is real and CONNECTED (not
    # something the satellite-fraction filter would already have dropped).
    _pgfp = _fitz_fp.open(_p_fp)[0]
    _pixfp = _pgfp.get_pixmap(matrix=_fitz_fp.Matrix(2.0, 2.0))
    import numpy as _np_fp
    _imfp = _np_fp.frombuffer(_pixfp.samples, _np_fp.uint8).reshape(_pixfp.height, _pixfp.width, _pixfp.n)[..., :3]
    _k_fp = 0.08819
    _comp_old_fp = _seg_fp(_imfp, (214, 214, 214), k=_k_fp, S=2.0, exclude_border=True)
    _area_old_fp = round(int(_comp_old_fp.sum()) * (1 / 2.0) ** 2 * _k_fp * _k_fp, 0)
    ck("BEFORE fix (generic 214 band): footpath strip annexed, area > 3,159 + 10 m² "
       "(connected over-measure, not a dropped satellite)",
       _area_old_fp > 3159.0 + 10, f"got {_area_old_fp}")

    # AFTER: full takeoff() with the swatch-lock fix — flags show the lock, area back to gold.
    _r_fp = _tu_takeoff_fp(_p_fp)
    _area_fp = _r_fp.get("area_m2")
    ck("AFTER fix: swatch (224ish) LOCKED — footpath strip excluded, area within 0.5% of 3,159 m²",
       _area_fp is not None and abs(_area_fp - 3159.0) / 3159.0 <= 0.005, f"got {_area_fp}")
    ck("AFTER fix: flags show the swatch-locked band",
       any("LOCKED" in f for f in _r_fp.get("flags", [])), _r_fp.get("flags"))
    ck("AFTER fix: measurement_state MEASURED_VERIFIED",
       _r_fp.get("measurement_state") == MEASURED_VERIFIED, _r_fp.get("measurement_state"))

    # DEMO-4 REGRESSION GUARD: swatch reads far enough from the yard's own fill (232 vs 214)
    # that the locked band [218,246] misses the 214 yard entirely -> must FALL BACK, never
    # silently return area=None on a perfectly measurable sheet.
    _p_fp_d4 = "/tmp/_ci_d77_footpath_demo4.pdf"
    _gen_d77_footpath(_p_fp_d4, chip_grey=0.910, yard_grey=0.84)
    _r_fp_d4 = _tu_takeoff_fp(_p_fp_d4)
    ck("DEMO-4 GUARD: swatch-locked band misses the yard fill -> FELL BACK (flag present)",
       any("FELL BACK" in f for f in _r_fp_d4.get("flags", [])), _r_fp_d4.get("flags"))
    ck("DEMO-4 GUARD: fallback still produces a measurable area (never area=None)",
       _r_fp_d4.get("area_m2") is not None, _r_fp_d4.get("area_m2"))
    ck("DEMO-4 GUARD: low-confidence fallback is MEASURED_UNVERIFIED (measurable, approve-blocked)",
       _r_fp_d4.get("measurement_state") == MEASURED_UNVERIFIED
       and _r_fp_d4.get("needs_assessor") is True,
       f"state={_r_fp_d4.get('measurement_state')} needs_assessor={_r_fp_d4.get('needs_assessor')}")

    # GOLD GUARDS unchanged: both pre-existing synthetic fixtures have unreadable swatches
    # (title text IS the label match, no nearby swatch chip) -> always take the fallback path,
    # golds untouched by this change.
    _d77_regress = _tu_takeoff_fp("drawings/_int_d77.pdf")
    ck("GOLD GUARD: _int_d77.pdf still exactly 3,159 m² (swatch-lock did not touch it)",
       _d77_regress.get("area_m2") == 3159.0, _d77_regress.get("area_m2"))
except _FixtureNotPresent as _e:
    print(f"  [SKIP] {_e} — fixture not present")
except (ImportError, FileNotFoundError) as _e:
    print(f"  [SKIP] D77 swatch-locked grey band test — missing dependency or file: {_e}")

print("approval_server: soft-delete (archive/unarchive) — Aryan's portal delete-estimation request")
try:
    import approval_server as _AS5
    import tempfile as _tempfile5

    _tmpdir5 = Path(_tempfile5.mkdtemp(prefix="ci_archive_"))
    _orig_jobs_file5 = _AS5.JOBS_FILE
    _orig_archive_file5 = _AS5.JOBS_ARCHIVE_FILE
    _AS5.JOBS_FILE = _tmpdir5 / "jobs.json"
    _AS5.JOBS_ARCHIVE_FILE = _tmpdir5 / "jobs_archive.json"
    try:
        _app5 = _AS5.app
        _app5.testing = True
        _client5 = _app5.test_client()

        # Ordinary pending job -> archivable
        _jid5 = "job-pending-1"
        _AS5.save_jobs({_jid5: {"id": _jid5, "status": "pending", "decision": None,
                                 "project_name": "Test Yard", "flags": []}})
        _r5 = _client5.post(f"/archive/{_jid5}", json={"note": "duplicate upload"})
        ck("archive: pending job archives with 200", _r5.status_code == 200, _r5.status_code)
        _jobs_after5 = _AS5.load_jobs()
        ck("archive: job removed from hot jobs file", _jid5 not in _jobs_after5)
        _archive5 = _AS5._load_archive()
        ck("archive: job present in archive file", _jid5 in _archive5)
        ck("archive: archived record carries archived=True + archived_at",
           _archive5.get(_jid5, {}).get("archived") is True and _archive5.get(_jid5, {}).get("archived_at"),
           _archive5.get(_jid5))
        ck("archive: status set to 'deleted' in the archive record",
           _archive5.get(_jid5, {}).get("status") == "deleted", _archive5.get(_jid5, {}).get("status"))
        ck("archive: no data lost — project_name preserved",
           _archive5.get(_jid5, {}).get("project_name") == "Test Yard")

        # /jobs/archived surfaces it, default /jobs does not
        _r_list5 = _client5.get("/jobs/archived")
        ck("GET /jobs/archived includes the archived job", _jid5 in _r_list5.get_json())
        _r_hot5 = _client5.get("/jobs")
        ck("GET /jobs (default) excludes the archived job", _jid5 not in _r_hot5.get_json())

        # Unarchive restores it
        _r_un5 = _client5.post(f"/unarchive/{_jid5}")
        ck("unarchive: restores with 200", _r_un5.status_code == 200, _r_un5.status_code)
        _jobs_restored5 = _AS5.load_jobs()
        ck("unarchive: job back in hot jobs file", _jid5 in _jobs_restored5)
        ck("unarchive: archived flag cleared", _jobs_restored5.get(_jid5, {}).get("archived") is False)
        _archive_after_un5 = _AS5._load_archive()
        ck("unarchive: removed from archive file", _jid5 not in _archive_after_un5)

        # Approved job -> BLOCKED (needs Jas, not a portal button)
        _jid5b = "job-approved-1"
        _AS5.save_jobs({_jid5b: {"id": _jid5b, "status": "approved", "decision": "approved",
                                  "project_name": "Approved Yard", "flags": []}})
        _r5b = _client5.post(f"/archive/{_jid5b}")
        ck("archive: approved job is BLOCKED (409)", _r5b.status_code == 409, _r5b.status_code)
        ck("archive: blocked-job error mentions Jas / manual",
           "jas" in _r5b.get_json().get("error", "").lower(), _r5b.get_json())
        ck("archive: approved job NOT removed from hot jobs file after a blocked attempt",
           _jid5b in _AS5.load_jobs())

        # Processing job -> BLOCKED (409), same pattern as approve/reject/adjust
        _jid5c = "job-processing-1"
        _AS5.save_jobs({_jid5c: {"id": _jid5c, "status": "processing", "decision": None, "flags": []}})
        _r5c = _client5.post(f"/archive/{_jid5c}")
        ck("archive: processing job is BLOCKED (409)", _r5c.status_code == 409, _r5c.status_code)

        # Unknown job -> 404, never a crash
        _r5d = _client5.post("/archive/does-not-exist")
        ck("archive: unknown job -> 404 (not a crash)", _r5d.status_code == 404, _r5d.status_code)
        _r5e = _client5.post("/unarchive/does-not-exist")
        ck("unarchive: unknown archived job -> 404 (not a crash)", _r5e.status_code == 404, _r5e.status_code)
    finally:
        _AS5.JOBS_FILE = _orig_jobs_file5
        _AS5.JOBS_ARCHIVE_FILE = _orig_archive_file5
        shutil.rmtree(_tmpdir5, ignore_errors=True)
except ImportError as _e:
    print(f"  [SKIP] approval_server soft-delete tests — missing dependency: {_e}")

print("approval_server: PORTAL_TOKEN auth gate (prod-audit MUST — unauthenticated approve/reject/adjust)")
try:
    import approval_server as _AS6
    import tempfile as _tempfile6

    _tmpdir6 = Path(_tempfile6.mkdtemp(prefix="ci_auth_"))
    _orig_jobs_file6 = _AS6.JOBS_FILE
    _AS6.JOBS_FILE = _tmpdir6 / "jobs.json"
    _orig_token6 = _AS6.APPROVAL_TOKEN
    _AS6.APPROVAL_TOKEN = "test-secret-token-123"
    try:
        _app6 = _AS6.app
        _app6.testing = True
        _client6 = _app6.test_client()

        _jid6 = "job-auth-1"
        _AS6.save_jobs({_jid6: {"id": _jid6, "status": "pending", "decision": None, "flags": []}})

        # No token at all -> 401, never a silent pass-through
        _r6 = _client6.get("/jobs")
        ck("no token -> /jobs is 401 when APPROVAL_TOKEN is set", _r6.status_code == 401, _r6.status_code)

        # Wrong token -> 401
        _r6b = _client6.get("/jobs", headers={"Authorization": "Bearer wrong-token"})
        ck("wrong Bearer token -> 401", _r6b.status_code == 401, _r6b.status_code)

        # Correct Bearer token -> 200
        _r6c = _client6.get("/jobs", headers={"Authorization": "Bearer test-secret-token-123"})
        ck("correct Bearer token -> 200", _r6c.status_code == 200, _r6c.status_code)

        # /status always exempt (health-check must work for deploy monitoring pre-auth)
        _r6d = _client6.get("/status")
        ck("/status is exempt from the token gate", _r6d.status_code == 200, _r6d.status_code)

        # / stays reachable so it can redirect a browser into the portal login flow.
        _r6d0 = _client6.get("/")
        ck("/ is exempt from the token gate so the landing redirect works",
           _r6d0.status_code in (301, 302), _r6d0.status_code)

        # /portal?token=<correct> sets a cookie and redirects
        _r6e = _client6.get(f"/portal?token=test-secret-token-123")
        ck("/portal?token=<correct> redirects (sets cookie)", _r6e.status_code in (301, 302), _r6e.status_code)
        _set_cookie6 = _r6e.headers.get("Set-Cookie", "")
        ck("/portal?token=<correct> Set-Cookie contains the token cookie name",
           "approval_token" in _set_cookie6, _set_cookie6)

        # /portal?token=<wrong> does not authorise — use a FRESH client (no cookie carried
        # over from the earlier correct-token request on _client6, which would mask this).
        _client6fresh = _app6.test_client()
        _r6f = _client6fresh.get("/portal?token=nope")
        ck("/portal?token=<wrong> -> login form, not silently served",
           _r6f.status_code == 200 and b"Access code" in _r6f.data
           and b"Fortel Approval Portal" in _r6f.data, _r6f.status_code)

        # Cookie-based auth works for a mutating route (mirrors what the portal's own fetch()
        # calls will do once the browser holds the cookie from the bootstrap redirect above)
        _client6.set_cookie("approval_token", "test-secret-token-123")
        _r6g = _client6.get(f"/job/{_jid6}")
        ck("cookie auth authorises a normal route", _r6g.status_code == 200, _r6g.status_code)

        # With APPROVAL_TOKEN unset, auth is fully disabled (back-compat / local dev)
        _AS6.APPROVAL_TOKEN = ""
        _client6b = _app6.test_client()
        _r6h = _client6b.get("/jobs")
        ck("no APPROVAL_TOKEN configured -> auth disabled, /jobs open", _r6h.status_code == 200, _r6h.status_code)
    finally:
        _AS6.JOBS_FILE = _orig_jobs_file6
        _AS6.APPROVAL_TOKEN = _orig_token6
        shutil.rmtree(_tmpdir6, ignore_errors=True)
except ImportError as _e:
    print(f"  [SKIP] approval_server auth-gate tests — missing dependency: {_e}")

print("approval_server: token-gated rates API + build visibility")
try:
    import approval_server as _AS_rates
    import tempfile as _tempfile_rates_api
    import inspect as _inspect_rates_api
    import os as _os_rates_api
    import client_rates as _client_rates_api
    from unittest import mock as _mock_rates_api

    _rates_api_tmpdir = Path(_tempfile_rates_api.mkdtemp(prefix="ci_rates_api_"))
    _orig_rates_file_api = _AS_rates.CLIENT_RATES_FILE
    _orig_token_rates_api = _AS_rates.APPROVAL_TOKEN
    _AS_rates.CLIENT_RATES_FILE = _rates_api_tmpdir / "client_rates.json"
    _AS_rates.APPROVAL_TOKEN = "test-rates-token"
    try:
        _app_rates = _AS_rates.app
        _app_rates.testing = True
        _client_rates_http = _app_rates.test_client()

        _unauth_rates = _client_rates_http.get("/rates")
        ck("rates endpoint is protected by the existing portal token gate",
           _unauth_rates.status_code == 401, _unauth_rates.status_code)
        _headers_rates = {"Authorization": "Bearer test-rates-token"}
        _get_rates = _client_rates_http.get("/rates", headers=_headers_rates)
        _get_rates_json = _get_rates.get_json() or {}
        ck("GET /rates returns every effective field as DEFAULT at version 0",
           _get_rates.status_code == 200 and _get_rates_json.get("version") == 0 and
           len(_get_rates_json.get("fields") or []) == len(_client_rates_api.RATE_FIELDS) and
           all(field.get("provenance") == "DEFAULT"
               for field in _get_rates_json.get("fields") or []))

        _api_defaults = _AS_rates._client_rate_defaults()
        _new_labour = _api_defaults["labour"] * 1.01
        _post_rates_1 = _client_rates_http.post(
            "/rates", headers=_headers_rates, json={"rates": {"labour": _new_labour}})
        ck("POST /rates saves version 1 with CLIENT-EDITED provenance",
           _post_rates_1.status_code == 200 and
           _post_rates_1.get_json().get("version") == 1 and
           next(field for field in _post_rates_1.get_json()["fields"]
                if field["key"] == "labour")["provenance"] == "CLIENT-EDITED")
        _server_fresh_costing = _AS_rates._run_costing(100, {})
        ck("approval/adjust fresh-costing path applies the same saved rates version",
           _server_fresh_costing.get("client_rates_applied") is True and
           _server_fresh_costing.get("rates_version") == 1 and
           _server_fresh_costing["rate"] != price_with_defaults(
               100, client_rates_path=_rates_api_tmpdir / "absent.json")["rate"])

        _post_rates_2 = _client_rates_http.post(
            "/rates", headers=_headers_rates,
            json={"rates": {"labour": _api_defaults["labour"]}})
        _stored_rates_api = _client_rates_api.load_rate_store(_AS_rates.CLIENT_RATES_FILE)
        ck("each actual save bumps version and restoring a default removes its override",
           _post_rates_2.status_code == 200 and
           _post_rates_2.get_json().get("version") == 2 and
           "labour" not in _stored_rates_api["overrides"] and
           len(_stored_rates_api["audit"]) == 2)
        ck("rates writes use same-filesystem os.replace atomic replacement",
           "os.replace(tmp, path)" in _inspect_rates_api.getsource(
               _client_rates_api.save_client_rates))

        _status_build = _client_rates_http.get("/status")
        _status_build_json = _status_build.get_json() or {}
        ck("/status carries build sha + date and remains health-check accessible",
           _status_build.status_code == 200 and
           set((_status_build_json.get("build") or {})) == {"sha", "date"} and
           bool(_status_build_json["build"]["sha"]) and bool(_status_build_json["build"]["date"]))
        with _mock_rates_api.patch.dict(
                _os_rates_api.environ,
                {"RAILWAY_GIT_COMMIT_SHA": "", "RAILWAY_GIT_COMMIT_DATE": ""}), \
                _mock_rates_api.patch("subprocess.run", side_effect=FileNotFoundError("no git")):
            _no_git_build = _AS_rates._detect_build_info()
        ck("build detection never crashes when Railway env and git metadata are absent",
           _no_git_build == {"sha": "unknown", "date": "unknown"}, _no_git_build)

        _portal_source_rates = Path("assessor_portal.html").read_text()
        ck("portal contains Rates panel, new-pricing warning, and fixed build footer",
           all(marker in _portal_source_rates for marker in (
               'id="ratesModal"', "field.provenance", ".rate-tag.client-edited",
               "new pricing only",
               'id="buildFooter"', "loadBuild()", "fetch(`${BASE}/rates`)")))
        _login_build_html = _AS_rates._portal_login_page()
        ck("login page also shows the build footer without exposing the token",
           'id="buildFooter"' in _login_build_html and "Build " in _login_build_html and
           "test-rates-token" not in _login_build_html)
    finally:
        _AS_rates.CLIENT_RATES_FILE = _orig_rates_file_api
        _AS_rates.APPROVAL_TOKEN = _orig_token_rates_api
        shutil.rmtree(_rates_api_tmpdir, ignore_errors=True)
except (ImportError, OSError, ValueError, StopIteration) as _e:
    ck("rates API and build visibility tests import and run", False, _e)

print("approval_server: /portal login form (no-token case posts a code instead of a bare 401)")
try:
    import approval_server as _AS6b

    _orig_token6b = _AS6b.APPROVAL_TOKEN
    _AS6b.APPROVAL_TOKEN = "test-login-code"
    try:
        _app6b = _AS6b.app
        _app6b.testing = True

        _client6b1 = _app6b.test_client()
        _r6b1 = _client6b1.get("/portal")
        ck("GET /portal with no cookie/token -> login form",
           _r6b1.status_code == 200 and b"Access code" in _r6b1.data
           and b"Review Portal" not in _r6b1.data, _r6b1.status_code)

        _r6b2 = _client6b1.post("/portal", data={"code": "wrong-code"})
        ck("POST /portal wrong code -> re-shown with error, 200",
           _r6b2.status_code == 200 and b"Incorrect code" in _r6b2.data, _r6b2.status_code)

        _r6b3 = _client6b1.post("/portal", data={"code": "test-login-code"})
        ck("POST /portal correct code -> redirect", _r6b3.status_code in (301, 302), _r6b3.status_code)
        ck("POST /portal correct code -> Set-Cookie contains the token cookie name",
           "approval_token" in _r6b3.headers.get("Set-Cookie", ""),
           _r6b3.headers.get("Set-Cookie", ""))

        _r6b4 = _client6b1.get("/portal")
        ck("GET /portal with cookie from login -> real portal, not the login form",
           _r6b4.status_code == 200 and b"Access code" not in _r6b4.data, _r6b4.status_code)
    finally:
        _AS6b.APPROVAL_TOKEN = _orig_token6b
except ImportError as _e:
    print(f"  [SKIP] approval_server /portal login-form tests — missing dependency: {_e}")

print("persistent storage: one volume-aware resolver keeps jobs, drawings and quotations together")
try:
    import tempfile as _tempfile_storage
    import shutil as _shutil_storage
    import json as _json_storage
    import fitz as _fitz_storage
    import approval_server as _AS_storage
    from storage_paths import resolve_storage_paths as _resolve_storage_paths
    from quotation import save_quotation as _save_quote_storage

    _storage_root = Path(_tempfile_storage.mkdtemp(prefix="ci_volume_"))
    _app_root = _storage_root / "ephemeral-app"
    _volume_root = _storage_root / "railway-volume"
    _local_paths = _resolve_storage_paths({}, app_dir=_app_root)
    _volume_paths = _resolve_storage_paths(
        {"RAILWAY_VOLUME_MOUNT_PATH": str(_volume_root)}, app_dir=_app_root)
    ck("storage paths stay repository-local without a Railway volume",
       _local_paths.jobs_file == _app_root / "approval_jobs.json" and
       _local_paths.drawings_dir == _app_root / "drawings" and
       _local_paths.quotations_dir == _app_root / "quotations", _local_paths)
    ck("Railway volume places job store, drawings and quotations under one persistent base",
       _volume_paths.jobs_file == _volume_root / "approval_jobs.json" and
       _volume_paths.drawings_dir == _volume_root / "drawings" and
       _volume_paths.quotations_dir == _volume_root / "quotations" and
       _volume_paths.training_log_file == _volume_root / "training_log.jsonl" and
       _volume_paths.learned_patterns_file == _volume_root / "learned_patterns.json" and
       _volume_paths.jobs_archive_file.parent == _volume_root and
       _volume_paths.backup_dir.parent == _volume_root, _volume_paths)
    _override_paths = _resolve_storage_paths({
        "RAILWAY_VOLUME_MOUNT_PATH": str(_volume_root),
        "JOBS_FILE": str(_storage_root / "custom" / "jobs.qa.json"),
        "DRAWINGS_DIR": str(_storage_root / "custom-drawings"),
        "QUOTATIONS_DIR": str(_storage_root / "custom-quotes"),
    }, app_dir=_app_root)
    ck("explicit jobs/drawings/quotations path overrides still win over the volume default",
       _override_paths.jobs_file == _storage_root / "custom" / "jobs.qa.json" and
       _override_paths.drawings_dir == _storage_root / "custom-drawings" and
       _override_paths.quotations_dir == _storage_root / "custom-quotes" and
       _override_paths.training_log_file == _storage_root / "custom" / "jobs.qa_training_log.jsonl" and
       _override_paths.learned_patterns_file == _storage_root / "custom" / "jobs.qa_learned_patterns.json",
       _override_paths)

    _orig_storage_jobs = _AS_storage.JOBS_FILE
    _orig_storage_backups = _AS_storage.BACKUP_DIR
    try:
        _AS_storage.JOBS_FILE = _volume_paths.jobs_file
        _AS_storage.BACKUP_DIR = _volume_paths.backup_dir
        _volume_paths.drawings_dir.mkdir(parents=True, exist_ok=True)
        _persistent_pdf = _volume_paths.drawings_dir / "PERSIST-001_plan.pdf"
        _persistent_doc = _fitz_storage.open()
        _persistent_doc.new_page(width=100, height=100)
        _persistent_doc.save(_persistent_pdf)
        _persistent_doc.close()
        _persistent_quote_paths = _save_quote_storage(
            _q, out_dir=str(_volume_paths.quotations_dir))
        _AS_storage.save_jobs({"persist-1": {
            "id": "persist-1", "status": "approved",
            "pdf_path": str(_persistent_pdf),
            "quotation_paths": _persistent_quote_paths,
        }})

        # Re-run the pure resolver exactly as a fresh process does after a deploy.  Nothing
        # from the ephemeral application directory is consulted.
        _fresh_paths = _resolve_storage_paths(
            {"RAILWAY_VOLUME_MOUNT_PATH": str(_volume_root)}, app_dir=_app_root)
        _fresh_jobs = _json_storage.loads(_fresh_paths.jobs_file.read_text())
        _fresh_job = _fresh_jobs["persist-1"]
        ck("deploy-cycle re-resolution preserves job record, uploaded PDF and quotation",
           Path(_fresh_job["pdf_path"]).is_file() and
           Path(_fresh_job["quotation_paths"]["xlsx"]).is_file() and
           Path(_fresh_job["quotation_paths"]["json"]).is_file() and
           Path(_fresh_job["pdf_path"]).is_relative_to(_fresh_paths.drawings_dir) and
           Path(_fresh_job["quotation_paths"]["xlsx"]).is_relative_to(
               _fresh_paths.quotations_dir), _fresh_job)
    finally:
        _AS_storage.JOBS_FILE = _orig_storage_jobs
        _AS_storage.BACKUP_DIR = _orig_storage_backups
        _shutil_storage.rmtree(_storage_root, ignore_errors=True)
except (ImportError, OSError, ValueError) as _e:
    ck("persistent storage deploy-cycle test imports and runs", False, _e)

print("approved-only pattern memory: prior approval is built from approved jobs only")
try:
    import tempfile as _tempfile_memory
    import shutil as _shutil_memory
    from training_analytics import build_learned_patterns as _build_learned_patterns
    from training_analytics import prior_approval_for_job as _prior_approval_for_job
    from training_analytics import attach_prior_approval as _attach_prior_approval
    import approval_server as _AS_memory

    _memory_root = Path(_tempfile_memory.mkdtemp(prefix="ci_memory_"))
    _orig_memory_jobs = _AS_memory.JOBS_FILE
    _orig_memory_log = _AS_memory.TRAINING_LOG
    _orig_memory_patterns = _AS_memory.LEARNED_PATTERNS_FILE
    try:
        _AS_memory.JOBS_FILE = _memory_root / "jobs.json"
        _AS_memory.TRAINING_LOG = _memory_root / "training_log.jsonl"
        _AS_memory.LEARNED_PATTERNS_FILE = _memory_root / "learned_patterns.json"
        _approved_job = {
            "id": "approved-1", "status": "approved", "decision": "approved",
            "document_sha256": "a" * 64,
            "project_ref": "MEM-001", "project_name": "Memory Case",
            "result": {"file": "memory.pdf", "area_m2": 100.0, "project_ref": "MEM-001",
                        "project_name": "Memory Case", "scale_src": "bar"},
            "costing": {"assumed": False},
            "flags": ["approved-flag"],
            "decided_at": "2026-08-13T10:00:00",
        }
        _adjusted_job = {
            "id": "adjusted-1", "status": "adjusted", "decision": "adjusted",
            "project_ref": "MEM-001", "project_name": "Memory Case",
            "result": {"file": "memory.pdf", "area_m2": 110.0, "project_ref": "MEM-001",
                        "project_name": "Memory Case"},
            "adjusted": {"area_m2": 110.0},
            "flags": ["adjusted-flag"],
            "decided_at": "2026-08-13T11:00:00",
        }
        _pending_job = {
            "id": "pending-1", "status": "pending", "decision": None,
            "document_sha256": "a" * 64,
            "project_ref": "MEM-001", "project_name": "Memory Case",
            "result": {"file": "memory.pdf", "area_m2": 120.0, "project_ref": "MEM-001",
                        "project_name": "Memory Case"},
        }
        _AS_memory.save_jobs({
            "approved-1": _approved_job,
            "adjusted-1": _adjusted_job,
            "pending-1": _pending_job,
        })
        _patterns_memory = _build_learned_patterns(_AS_memory.load_jobs())
        ck("approved-only builder ignores adjusted jobs",
           "memory.pdf" in _patterns_memory["by_file"] and
           _patterns_memory["by_file"]["memory.pdf"]["latest"]["job_id"] == "approved-1" and
           len(_patterns_memory["by_file"]["memory.pdf"]["history"]) == 1,
           _patterns_memory["by_file"].get("memory.pdf"))
        _prior_memory = _prior_approval_for_job(_pending_job, _patterns_memory)
        ck("matching pending job gets prior approval from approved job only",
           _prior_memory and _prior_memory.get("job_id") == "approved-1" and
           _prior_memory.get("matched_on") == "exact document SHA-256", _prior_memory)
        _attached_memory = _attach_prior_approval("pending-1", _pending_job, _patterns_memory)
        ck("attach_prior_approval adds informational prior_approval payload",
           _attached_memory.get("prior_approval", {}).get("job_id") == "approved-1", _attached_memory)
        ck("prior approval payload contains identity only, never learned quantities or money",
           set(_attached_memory.get("prior_approval", {})) == {
               "job_id", "approved_at", "document_sha256", "matched_on"
           }, _attached_memory.get("prior_approval"))
        _different_drawing = dict(_pending_job, document_sha256="b" * 64)
        ck("same filename/project never decorates a different drawing as previously approved",
           _prior_approval_for_job(_different_drawing, _patterns_memory) is None)
        ck("an approved job is never presented as its own prior approval",
           "prior_approval" not in _attach_prior_approval(
               "approved-1", _approved_job, _patterns_memory))
        _memory_portal = (Path(__file__).parent / "assessor_portal.html").read_text()
        ck("prior-approval UI is neutral and explicitly says no measurement/price was reused",
           "This exact drawing file was approved previously" in _memory_portal and
           "no prior measurement or price was reused" in _memory_portal and
           "Previously approved as-is" not in _memory_portal)
        ck("prior_approval does not affect approve-block logic",
           _AS_memory._approve_block_reason({
               "measurement_state": "MEASURED_VERIFIED",
               "scale_confirmed": False,
               "prior_approval": _attached_memory.get("prior_approval"),
           }) is None)
    finally:
        _AS_memory.JOBS_FILE = _orig_memory_jobs
        _AS_memory.TRAINING_LOG = _orig_memory_log
        _AS_memory.LEARNED_PATTERNS_FILE = _orig_memory_patterns
        _shutil_memory.rmtree(_memory_root, ignore_errors=True)
except ImportError as _e:
    print(f"  [SKIP] approved-only pattern memory tests — missing dependency: {_e}")

print("training log cleanup: fixture contamination filter keeps only real assessor history")
try:
    import json as _json_cleanup
    from cleanup_training_log import clean_entries as _clean_training_log_entries
    from cleanup_training_log import backup_path_for as _cleanup_backup_path

    _clean_sample = [
        _json_cleanup.dumps({"event": "approve", "job_id": "job-csrf-approve", "file": "csrf_test.pdf"}),
        _json_cleanup.dumps({"event": "reject", "job_id": "real-1", "file": "12345-real.pdf"}),
        _json_cleanup.dumps({"event": "approve", "job_id": "55555555-5555-4555-8555-555555555555",
                             "file": "Mixed perimeter case.pdf"}),
        _json_cleanup.dumps({"event": "adjust", "job_id": "client-real",
                             "file": "Office-GA.pdf", "project_ref": "CASE-REAL-2026"}),
        "malformed-but-potentially-real-evidence",
    ]
    _kept_log, _removed_log = _clean_training_log_entries(_clean_sample)
    ck("cleanup removes explicit fixture lines and preserves real entries",
       len(_kept_log) == 3 and len(_removed_log) == 2 and
       any('real-1' in line for line in _kept_log) and
       any('client-real' in line for line in _kept_log) and
       "malformed-but-potentially-real-evidence" in _kept_log, _removed_log)
    ck("cleanup backup names are collision-resistant even within one second",
       _cleanup_backup_path(Path("training.jsonl")) !=
       _cleanup_backup_path(Path("training.jsonl")))
except ImportError as _e:
    print(f"  [SKIP] training log cleanup tests — missing dependency: {_e}")

print("learning episodes: atomic before/after, refusal recovery, agreement and coverage")
try:
    import copy as _copy_learning
    import json as _json_learning
    import tempfile as _tempfile_learning
    import shutil as _shutil_learning
    import approval_server as _AS_learning
    from learning_capture import ensure_learning_episode as _ensure_episode
    from learning_capture import measurement_snapshot as _learning_snapshot
    from training_analytics import prior_approval_for_job as _prior_exact
    from training_analytics import analytics_report as _learning_analytics

    _learning_root = Path(_tempfile_learning.mkdtemp(prefix="ci_learning_episode_"))
    _old_learning_paths = (
        _AS_learning.JOBS_FILE, _AS_learning.BACKUP_DIR, _AS_learning.TRAINING_LOG,
        _AS_learning.LEARNED_PATTERNS_FILE,
    )
    _old_learning_functions = (_AS_learning._save_quotation, _AS_learning._run_costing)
    try:
        _AS_learning.JOBS_FILE = _learning_root / "jobs.json"
        _AS_learning.BACKUP_DIR = _learning_root / "backups"
        # Directories deliberately make both derivative writes fail. The authoritative job
        # episode must still commit and the HTTP action must still succeed.
        _AS_learning.TRAINING_LOG = _learning_root
        _AS_learning.LEARNED_PATTERNS_FILE = _learning_root
        _AS_learning._save_quotation = lambda *_a, **_k: {"xlsx": "scratch.xlsx"}
        _AS_learning._run_costing = lambda area, _result: {
            "area_m2": area, "rate": 1.0, "total_gbp": area,
        }
        _approved_seed = {
            "id": "learn-agree", "status": "pending", "decision": None,
            "measurement_state": "MEASURED_VERIFIED", "area_m2": 100.0,
            "document_sha256": "c" * 64,
            "prior_approval": {"area_m2": 999999.0, "rate": 999999.0,
                               "total_gbp": 999999.0},
            "result": {"file": "Exact.pdf", "area_m2": 100.0,
                       "measurement_state": "MEASURED_VERIFIED", "flags": []},
            "flags": [],
        }
        _ensure_episode("learn-agree", _approved_seed, source="pipeline")
        _AS_learning.save_jobs({"learn-agree": _approved_seed})
        with _AS_learning.app.test_client() as _client_learning:
            _agree_response = _client_learning.post(
                "/approve/learn-agree", json={"note": "agrees with AI"})
        _agree_saved = _AS_learning.load_jobs()["learn-agree"]
        _agree_episode = _agree_saved.get("learning_episode") or {}
        ck("training/cache write failure cannot split an approved decision from its evidence",
           _agree_response.status_code == 200 and _agree_saved.get("status") == "approved" and
           (_agree_episode.get("terminal") or {}).get("event") == "approved_unchanged",
           {"http": _agree_response.status_code, "episode": _agree_episode})
        ck("learned numeric payload cannot influence fresh approved costing",
           (_agree_response.get_json().get("costing") or {}).get("area_m2") == 100.0 and
           (_agree_response.get_json().get("costing") or {}).get("area_m2") != 999999.0,
           _agree_response.get_json())
        _commercial_probe = _learning_snapshot({
            "zones": [{"area_m2": 10.0, "rate": 99.0, "total_gbp": 990.0}],
            "adjusted": {"area_m2": 10.0, "costing": {"rate": 99.0}},
        })
        ck("atomic learning snapshots exclude rates, totals and nested costing",
           "rate" not in _json_learning.dumps(_commercial_probe) and
           "total_gbp" not in _json_learning.dumps(_commercial_probe) and
           "costing" not in _json_learning.dumps(_commercial_probe), _commercial_probe)

        _refused_seed = {
            "id": "learn-refusal", "status": "error", "decision": None,
            "measurement_state": "UNMEASURED", "area_m2": None,
            "result": {"file": "Office raw.pdf", "area_m2": None,
                       "measurement_state": "UNMEASURED", "flags": ["trace manually"]},
            "flags": ["trace manually"],
        }
        _ensure_episode("learn-refusal", _refused_seed, source="pipeline")
        _jobs_learning = _AS_learning.load_jobs()
        _jobs_learning["learn-refusal"] = _refused_seed
        _AS_learning.save_jobs(_jobs_learning)
        with _AS_learning.app.test_client() as _client_learning:
            _refusal_response = _client_learning.post("/adjust/learn-refusal", json={
                "regions": [[[0, 0], [100, 0], [100, 100], [0, 100]]],
                "region_categories": ["ground_floor"],
                "region_scopes": ["ground_floor_core"], "scale_k": 0.1,
            })
        _refusal_saved = _AS_learning.load_jobs()["learn-refusal"]
        _refusal_events = (_refusal_saved.get("learning_episode") or {}).get("events") or []
        _refusal_event = _refusal_events[-1] if _refusal_events else {}
        ck("UNMEASURED manual trace is captured as refusal_recovered with original and final",
           _refusal_response.status_code == 200 and
           _refusal_event.get("event") == "refusal_recovered" and
           _refusal_event.get("before", {}).get("area_m2") is None and
           _refusal_event.get("after", {}).get("area_m2") == 100.0,
           {"http": _refusal_response.status_code, "event": _refusal_event})
        with _AS_learning.app.test_client() as _client_learning:
            _refusal_approve = _client_learning.post("/approve/learn-refusal", json={})
        _analytics_rows = {
            row["job_id"]: row for row in
            _learning_analytics(_AS_learning.load_jobs())["approved_jobs"]
        }
        ck("analytics uses the frozen episode pair and keeps refusal originals honest",
           _refusal_approve.status_code == 200 and
           _analytics_rows["learn-agree"]["learning_outcome"] == "approved_unchanged" and
           _analytics_rows["learn-agree"]["delta_m2"] == 0.0 and
           _analytics_rows["learn-refusal"]["learning_outcome"] ==
           "approved_after_correction" and
           _analytics_rows["learn-refusal"]["ai_area_m2"] is None and
           _analytics_rows["learn-refusal"]["assessed_area_m2"] == 100.0,
           _analytics_rows)

        _malicious = {"by_document_sha256": {"d" * 64: {"latest": {
            "job_id": "old", "approved_at": "now", "document_sha256": "d" * 64,
            "area_m2": 999999, "rate": 999999, "total_gbp": 999999,
        }}}}
        _safe_prior = _prior_exact({"document_sha256": "d" * 64}, _malicious)
        ck("even a malicious learned cache exposes no measurement, rate or total",
           set(_safe_prior or {}) == {
               "job_id", "approved_at", "document_sha256", "matched_on"
           }, _safe_prior)
        from training_analytics import update_learned_patterns as _update_safe_patterns
        _scrubbed_patterns = _update_safe_patterns(_malicious, "new-safe", {
            "status": "approved", "decision": "approved",
            "document_sha256": "e" * 64, "result": {"file": "safe.pdf"},
        })
        _scrubbed_old = (_scrubbed_patterns["by_document_sha256"]["d" * 64]
                         .get("latest") or {})
        ck("refresh scrubs legacy area/rate/total fields from the derivative cache",
           not ({"area_m2", "rate", "total_gbp"} & set(_scrubbed_old)), _scrubbed_old)

        import threading as _threading_learning
        import time as _time_learning
        import training_analytics as _TA_learning
        _AS_learning.LEARNED_PATTERNS_FILE = _learning_root / "concurrent_patterns.json"
        _original_increment = _TA_learning.update_learned_patterns
        def _slow_increment(*args, **kwargs):
            value = _original_increment(*args, **kwargs)
            _time_learning.sleep(0.03)
            return value
        _TA_learning.update_learned_patterns = _slow_increment
        try:
            _threads_learning = [
                _threading_learning.Thread(
                    target=_AS_learning._refresh_learned_patterns,
                    args=(f"parallel-{index}", {
                        "status": "approved", "decision": "approved",
                        "document_sha256": str(index) * 64,
                        "result": {"file": f"parallel-{index}.pdf"},
                    }),
                ) for index in (1, 2)
            ]
            for _thread_learning in _threads_learning:
                _thread_learning.start()
            for _thread_learning in _threads_learning:
                _thread_learning.join()
        finally:
            _TA_learning.update_learned_patterns = _original_increment
        _parallel_patterns = _json_learning.loads(
            _AS_learning.LEARNED_PATTERNS_FILE.read_text())
        ck("concurrent approvals cannot lose one another from the learned-pattern cache",
           set(_parallel_patterns.get("by_document_sha256", {})) == {
               "1" * 64, "2" * 64
           }, _parallel_patterns.get("by_document_sha256"))
        _AS_learning.LEARNED_PATTERNS_FILE = _learning_root

        _coverage_seed = {
            "id": "learn-coverage", "status": "pending", "decision": None,
            "measurement_state": "MEASURED_VERIFIED", "area_m2": 100.0,
            "scale_k": 0.1, "flags": [],
            "zones": [{"zone_key": "unknown-1", "category": "unclassified",
                       "area_m2": 100.0, "measurement_kind": "area",
                       "subjects": ["Unit 1"], "needs_assessor": True}],
            "channel_proposals": [{"proposal_id": "channel-1",
                                   "proposed_length_lm": 10.0,
                                   "polyline_pts": [[0, 0], [100, 0]]}],
            "transition_candidates": [{"candidate_id": "transition-1",
                                       "proposed_length_lm": 5.0,
                                       "polyline_pts": [[0, 0], [50, 0]],
                                       "basis": "assessor review"}],
        }
        _coverage_seed["result"] = {
            "file": "Coverage.pdf", "area_m2": 100.0,
            "measurement_state": "MEASURED_VERIFIED", "flags": [],
            "zones": _copy_learning.deepcopy(_coverage_seed["zones"]),
            "channel_proposals": _copy_learning.deepcopy(_coverage_seed["channel_proposals"]),
            "transition_candidates": _copy_learning.deepcopy(
                _coverage_seed["transition_candidates"]),
            "scale_k": 0.1,
        }
        _ensure_episode("learn-coverage", _coverage_seed, source="pipeline")
        _jobs_learning = _AS_learning.load_jobs()
        _jobs_learning["learn-coverage"] = _coverage_seed
        _AS_learning.save_jobs(_jobs_learning)
        with _AS_learning.app.test_client() as _client_learning:
            _coverage_http = [
                _client_learning.post("/zones/learn-coverage", json={
                    "classifications": [{"zone_key": "unknown-1",
                                         "category": "external_yard"}],
                }).status_code,
                _client_learning.post("/spec-override/learn-coverage", json={
                    "zone_category": "external_yard", "fields": {"depth_mm": 190},
                }).status_code,
                _client_learning.post("/channel-proposals/learn-coverage", json={
                    "decisions": [{"proposal_id": "channel-1", "action": "accept",
                                   "length_lm": 10.0}],
                }).status_code,
                _client_learning.post("/transition-candidates/learn-coverage", json={
                    "decisions": [{"candidate_id": "transition-1", "action": "accept",
                                   "length_lm": 4.5}],
                }).status_code,
            ]
        _coverage_saved = _AS_learning.load_jobs()["learn-coverage"]
        _coverage_events = (_coverage_saved.get("learning_episode") or {}).get("events") or []
        _coverage_types = [event.get("event") for event in _coverage_events]
        ck("zone/spec/channel/transition reviews each preserve atomic before+after evidence",
           _coverage_http == [200, 200, 200, 200] and _coverage_types[-4:] == [
               "zones_classified", "spec_overridden", "channel_proposals_reviewed",
               "transition_candidates_reviewed",
           ] and all(event.get("before") and event.get("after")
                     for event in _coverage_events[-4:]),
           {"http": _coverage_http, "events": _coverage_types})
    finally:
        (_AS_learning.JOBS_FILE, _AS_learning.BACKUP_DIR, _AS_learning.TRAINING_LOG,
         _AS_learning.LEARNED_PATTERNS_FILE) = _old_learning_paths
        (_AS_learning._save_quotation, _AS_learning._run_costing) = _old_learning_functions
        _shutil_learning.rmtree(_learning_root, ignore_errors=True)
except (ImportError, OSError, ValueError) as _e:
    ck("learning episode safeguards import and run", False, _e)

print("critical approval safeguards: quotation errors visible; email uses canonical saved job")
try:
    import tempfile as _tempfile_critical
    import shutil as _shutil_critical
    import os as _os_critical
    from unittest import mock as _mock_critical
    import approval_server as _AS_critical
    import approval_email as _AE_critical
    import takeoff_pipeline as _TP_critical

    _critical_root = Path(_tempfile_critical.mkdtemp(prefix="ci_critical_"))
    _critical_jobs = _critical_root / "jobs.json"
    _orig_critical_jobs = _AS_critical.JOBS_FILE
    _orig_critical_backups = _AS_critical.BACKUP_DIR
    _orig_critical_quotes = _AS_critical.QUOTATIONS_DIR
    _orig_critical_token = _AS_critical.APPROVAL_TOKEN
    _orig_smtp_pass = _AE_critical.SMTP_PASS
    _AS_critical.JOBS_FILE = _critical_jobs
    _AS_critical.BACKUP_DIR = _critical_root / "backups"
    _AS_critical.QUOTATIONS_DIR = _critical_root / "quotations"
    _AS_critical.APPROVAL_TOKEN = ""
    try:
        _app_critical = _AS_critical.app
        _app_critical.testing = True
        _client_critical = _app_critical.test_client()

        _quotation_failure_id = "quotation-failure-visible"
        _AS_critical.save_jobs({_quotation_failure_id: {
            "id": _quotation_failure_id, "status": "pending", "decision": None,
            "measurement_state": "MEASURED_VERIFIED", "scale_confirmed": False,
            "flags": [], "result": {
                "file": "Mixed perimeter case.pdf", "area_m2": 100.0,
                "measurement_state": "MEASURED_VERIFIED", "flags": [],
            },
        }})
        with _mock_critical.patch.object(
                _AS_critical, "_run_costing", return_value={"area_m2":100.0}), \
                _mock_critical.patch.object(
                    _AS_critical, "_save_quotation",
                    return_value={"error":"KeyError: quantity_rows"}):
            _quotation_failure_response = _client_critical.post(
                f"/approve/{_quotation_failure_id}", json={"note":"reviewed"})
        _quotation_failure_job = _AS_critical.load_jobs()[_quotation_failure_id]
        ck("quotation generation failure returns HTTP 500 and is persisted on the approved job",
           _quotation_failure_response.status_code == 500 and
           _quotation_failure_response.get_json()["status"] == "quotation_error" and
           _quotation_failure_job["decision"] == "approved" and
           _quotation_failure_job["quotation_status"] == "error" and
           _quotation_failure_job["quotation_error"] == "KeyError: quantity_rows" and
           any(flag.startswith("QUOTATION GENERATION ERROR:")
               for flag in _quotation_failure_job["flags"]),
           _quotation_failure_response.get_json())
        ck("portal source renders a specific quotation-error banner and hides broken links",
           "job.quotation_error" in Path("assessor_portal.html").read_text() and
           "Quotation generation failed" in Path("assessor_portal.html").read_text())

        # SEND_APPROVAL_EMAILS can no longer make the pipeline create a second job.  Direct
        # pipeline use has no authoritative persisted context, so it only emits a loud defer
        # flag; the approval-server worker owns delivery after its atomic save.
        _deferred_result = {"file":"direct.pdf", "flags":[]}
        with _mock_critical.patch.object(_AE_critical, "create_job") as _create_job_mock:
            _TP_critical._trigger_approval(
                "direct.pdf", _deferred_result, approval_job_id="real-job",
                send_requested=True)
        ck("direct pipeline approval trigger never creates a duplicate job",
           _create_job_mock.call_count == 0 and
           any(flag.startswith("APPROVAL EMAIL DEFERRED:")
               for flag in _deferred_result["flags"]), _deferred_result)

        _real_job_id = "portal-real-job-id"
        _AS_critical.save_jobs({_real_job_id: {
            "id":_real_job_id, "status":"processing", "decision":None,
            "project_name":"Email case", "project_ref":"EMAIL-001", "flags":[],
            "result":{"file":"EMAIL-001_plan.pdf"},
        }})
        _email_targets = []
        _pipeline_context = []

        def _fake_takeoff_critical(pdf, project_name=None, project_ref=None,
                                   client_rates_path=None, approval_job_id=None):
            _pipeline_context.append(approval_job_id)
            return {
                "file":"EMAIL-001_plan.pdf", "area_m2":100.0,
                "measurement_state":"MEASURED_UNVERIFIED", "needs_assessor":True,
                "project_name":project_name, "project_ref":project_ref,
                "flags":[], "zones":[],
            }

        def _fake_email_critical(job_id, pdf_path, result,
                                 project_name=None, project_ref=None, to=None):
            # This assertion happens inside the notifier: delivery must not start until the
            # completed result is visible in the canonical store.
            saved = _AS_critical.load_jobs()[job_id]
            _email_targets.append((job_id, saved["status"], saved["result"]["file"]))
            return {"job_id":job_id, "sent":True, "status":"sent", "reason":""}

        with _mock_critical.patch.object(
                _TP_critical, "takeoff", new=_fake_takeoff_critical), \
                _mock_critical.patch.object(
                    _AE_critical, "send_job_approval_email",
                    side_effect=_fake_email_critical), \
                _mock_critical.patch.dict(
                    _os_critical.environ, {"SEND_APPROVAL_EMAILS":"1"}):
            _AS_critical._run_takeoff(
                _real_job_id, str(_critical_root / "EMAIL-001_plan.pdf"),
                "Email case", "EMAIL-001")
        _email_jobs = _AS_critical.load_jobs()
        ck("worker email targets the portal's real saved job_id and creates no duplicate",
           list(_email_jobs) == [_real_job_id] and
           _pipeline_context == [_real_job_id] and
           _email_targets == [(_real_job_id, "pending", "EMAIL-001_plan.pdf")] and
           _email_jobs[_real_job_id]["approval_email_status"] == "sent",
           {"jobs":list(_email_jobs), "targets":_email_targets,
            "pipeline_context":_pipeline_context})

        # Missing credentials are checked before snapshot rendering and stored on the job;
        # there is no automatic approval_emails/*.html fallback on this production path.
        _AE_critical.SMTP_PASS = ""
        with _mock_critical.patch.dict(
                _os_critical.environ, {"SEND_APPROVAL_EMAILS":"1"}):
            _missing_smtp = _AS_critical._notify_saved_review_job(
                _real_job_id, str(_critical_root / "missing.pdf"),
                _email_jobs[_real_job_id]["result"], "Email case", "EMAIL-001")
        _missing_smtp_job = _AS_critical.load_jobs()[_real_job_id]
        ck("missing SMTP configuration is recorded visibly and never treated as sent",
           _missing_smtp["status"] == "not_configured" and
           _missing_smtp["sent"] is False and
           _missing_smtp_job["approval_email_status"] == "not_configured" and
           "SMTP_PASS" in _missing_smtp_job["approval_email_error"] and
           any(flag.startswith("APPROVAL EMAIL NOT SENT:")
               for flag in _missing_smtp_job["flags"]) and
           not (_critical_root / "approval_emails" / f"{_real_job_id}.html").exists(),
           _missing_smtp)
    finally:
        _AS_critical.JOBS_FILE = _orig_critical_jobs
        _AS_critical.BACKUP_DIR = _orig_critical_backups
        _AS_critical.QUOTATIONS_DIR = _orig_critical_quotes
        _AS_critical.APPROVAL_TOKEN = _orig_critical_token
        _AE_critical.SMTP_PASS = _orig_smtp_pass
        _shutil_critical.rmtree(_critical_root, ignore_errors=True)
except (ImportError, OSError, ValueError, KeyError) as _e:
    ck("critical approval safeguard tests import and run", False, _e)

print("approval_server: jobs-file backup rotation + corrupt-file preservation (prod-audit MUST)")
try:
    import approval_server as _AS7
    import tempfile as _tempfile7

    _tmpdir7 = Path(_tempfile7.mkdtemp(prefix="ci_backup_"))
    _orig_jobs_file7 = _AS7.JOBS_FILE
    _orig_backup_dir7 = _AS7.BACKUP_DIR
    _AS7.JOBS_FILE = _tmpdir7 / "jobs.json"
    _AS7.BACKUP_DIR = _tmpdir7 / "backups"
    try:
        # First-ever save: JOBS_FILE doesn't exist yet, so there's nothing to snapshot —
        # _rotate_backup is a correct no-op here (never backs up a file that isn't there yet).
        _AS7.save_jobs({"a": {"id": "a"}})
        # Backup filenames are keyed off JOBS_FILE.stem ("jobs" here, not "approval_jobs") —
        # see approval_server._rotate_backup's stem-based naming (item 5, QA-instance isolation).
        _backups7 = list(_AS7.BACKUP_DIR.glob("jobs.*.json"))
        ck("no backup created on the very first save (nothing existed yet to snapshot)",
           len(_backups7) == 0, [str(p) for p in _backups7])

        # Second save the same day: JOBS_FILE now exists from the first save, so THIS save's
        # rotation check snapshots it before overwriting -> exactly one dated backup appears.
        _AS7.save_jobs({"a": {"id": "a"}, "b": {"id": "b"}})
        _backups7b = list(_AS7.BACKUP_DIR.glob("jobs.*.json"))
        ck("save_jobs creates a same-day backup once a prior file exists to snapshot",
           len(_backups7b) == 1, [str(p) for p in _backups7b])

        # A third save the same day must NOT create a second backup file for today
        _AS7.save_jobs({"a": {"id": "a"}, "b": {"id": "b"}, "c": {"id": "c"}})
        _backups7b2 = list(_AS7.BACKUP_DIR.glob("jobs.*.json"))
        ck("no duplicate backup for a third save on the same day",
           len(_backups7b2) == 1, [str(p) for p in _backups7b2])

        # Pruning: force more than BACKUP_KEEP dated backup files to exist, then trigger a
        # rotation check that should prune down to the newest BACKUP_KEEP.
        import datetime as _dt7
        for _i in range(20):
            _fake_date = (_dt7.date(2020, 1, 1) + _dt7.timedelta(days=_i)).isoformat()
            (_AS7.BACKUP_DIR / f"jobs.{_fake_date}.json").write_text("{}")
        _AS7._rotate_backup()  # today's backup already exists, so this call only prunes
        _backups7c = sorted(_AS7.BACKUP_DIR.glob("jobs.*.json"))
        ck(f"backup pruning keeps at most BACKUP_KEEP={_AS7.BACKUP_KEEP} files",
           len(_backups7c) <= _AS7.BACKUP_KEEP, len(_backups7c))

        # Corrupt (non-empty, unparseable) jobs file -> preserved as .corrupt-*, load returns {}
        _AS7.JOBS_FILE.write_text("{not valid json!!")
        _loaded7 = _AS7.load_jobs()
        ck("corrupt jobs file -> load_jobs returns {} (never raises)", _loaded7 == {}, _loaded7)
        _corrupt_copies7 = list(_tmpdir7.glob("jobs.json.corrupt-*"))
        ck("corrupt jobs file -> a .corrupt-* copy is preserved for recovery",
           len(_corrupt_copies7) == 1, [str(p) for p in _corrupt_copies7])
    finally:
        _AS7.JOBS_FILE = _orig_jobs_file7
        _AS7.BACKUP_DIR = _orig_backup_dir7
        shutil.rmtree(_tmpdir7, ignore_errors=True)
except ImportError as _e:
    print(f"  [SKIP] approval_server backup-rotation tests — missing dependency: {_e}")

print("approval_server: startup sweep clears stranded 'processing' jobs on restart (prod-audit MUST)")
try:
    import approval_server as _AS8
    import tempfile as _tempfile8

    _tmpdir8 = Path(_tempfile8.mkdtemp(prefix="ci_sweep_"))
    _orig_jobs_file8 = _AS8.JOBS_FILE
    _AS8.JOBS_FILE = _tmpdir8 / "jobs.json"
    try:
        _jid8 = "job-stranded-1"
        _AS8.save_jobs({_jid8: {"id": _jid8, "status": "processing", "decision": None, "flags": []}})
        _AS8._sweep_stranded_processing_jobs()
        _swept8 = _AS8.load_jobs()[_jid8]
        ck("stranded 'processing' job flipped to UNMEASURED by the startup sweep",
           _swept8.get("measurement_state") == "UNMEASURED", _swept8.get("measurement_state"))
        ck("startup sweep flag mentions PIPELINE INTERRUPTED",
           any("PIPELINE INTERRUPTED" in f for f in _swept8.get("flags", [])), _swept8.get("flags"))
        ck("startup sweep sets needs_assessor=True", _swept8.get("needs_assessor") is True)

        # A job that is NOT processing must be left untouched
        _jid8b = "job-approved-untouched"
        _AS8.save_jobs({_jid8b: {"id": _jid8b, "status": "approved", "decision": "approved", "flags": ["ok"]}})
        _AS8._sweep_stranded_processing_jobs()
        _unswept8 = _AS8.load_jobs()[_jid8b]
        ck("non-processing job untouched by the startup sweep",
           _unswept8.get("status") == "approved" and _unswept8.get("flags") == ["ok"], _unswept8)
    finally:
        _AS8.JOBS_FILE = _orig_jobs_file8
        shutil.rmtree(_tmpdir8, ignore_errors=True)
except ImportError as _e:
    print(f"  [SKIP] approval_server startup-sweep tests — missing dependency: {_e}")

print("approval_server: /webhook/n8n pdf_path containment guard (prod-audit MUST — arbitrary file read)")
try:
    import approval_server as _AS9
    import approval_email as _AE9
    import tempfile as _tempfile9
    import json as _json9

    # /webhook/n8n -> approval_email.create_job() writes straight to approval_email.JOBS_FILE.
    # Left pointed at the real approval_jobs.json, every POST in this test (three per run)
    # permanently wrote a junk "x.pdf" / area=100 pending job into the LIVE jobs file — this is
    # exactly how the ~17 junk jobs that were polluting approval_jobs.json got there. Point
    # approval_email.JOBS_FILE at a tempfile.mkdtemp() scratch path for the duration of this
    # test (never a hardcoded /tmp path) and restore it in finally, so CI is byte-stable
    # against the live jobs file no matter how many times it runs.
    _tmpdir9 = Path(_tempfile9.mkdtemp(prefix="ci_webhook_n8n_"))
    _orig_ae_jobs_file9 = _AE9.JOBS_FILE
    _AE9.JOBS_FILE = _tmpdir9 / "jobs.json"
    try:
        _app9 = _AS9.app
        _app9.testing = True
        _client9 = _app9.test_client()

        _r9 = _client9.post("/webhook/n8n", json={
            "pdf_path": "/etc/passwd",
            "result": {"area_m2": 100, "file": "x.pdf"},
        })
        ck("pdf_path outside drawings/ is rejected with 400, not read",
           _r9.status_code == 400, _r9.status_code)

        _r9b = _client9.post("/webhook/n8n", json={
            "pdf_path": "../../etc/passwd",
            "result": {"area_m2": 100, "file": "x.pdf"},
        })
        ck("path-traversal pdf_path is rejected with 400",
           _r9b.status_code == 400, _r9b.status_code)

        # Empty pdf_path (legit use case — result created without a snapshot) still works
        _r9c = _client9.post("/webhook/n8n", json={
            "pdf_path": "",
            "result": {"area_m2": 100, "file": "x.pdf"},
        })
        ck("empty pdf_path (no snapshot) is not blocked by the containment guard",
           _r9c.status_code == 200, _r9c.status_code)

        # The job this test creates must land in the scratch JOBS_FILE, never the live one.
        ck("webhook test job landed in the scratch jobs file, not the live approval_jobs.json",
           _AE9.JOBS_FILE.exists() and len(_json9.loads(_AE9.JOBS_FILE.read_text())) >= 1,
           str(_AE9.JOBS_FILE))
    finally:
        _AE9.JOBS_FILE = _orig_ae_jobs_file9
        shutil.rmtree(_tmpdir9, ignore_errors=True)
except ImportError as _e:
    print(f"  [SKIP] approval_server webhook containment tests — missing dependency: {_e}")

print("approval_server: GET on /approve /reject does NOT mutate; POST does "
      "(top-level-navigation CSRF fix — SameSite=Lax cookies + a mutating GET meant an "
      "email client's link-preview prefetch, or any page merely linking here, could "
      "silently approve/reject a job)")
try:
    import approval_server as _AS10
    import tempfile as _tempfile10

    _tmpdir10 = Path(_tempfile10.mkdtemp(prefix="ci_csrf_"))
    _orig_jobs_file10 = _AS10.JOBS_FILE
    _AS10.JOBS_FILE = _tmpdir10 / "jobs.json"
    try:
        _app10 = _AS10.app
        _app10.testing = True
        _client10 = _app10.test_client()

        # --- /approve: GET must not mutate ---
        _jid10a = "job-csrf-approve"
        _AS10.save_jobs({_jid10a: {
            "id": _jid10a, "status": "pending", "decision": None,
            "measurement_state": "MEASURED_VERIFIED", "scale_confirmed": True,
            "result": {"area_m2": 1000, "file": "csrf_test.pdf"}, "flags": [],
        }})
        _rg10a = _client10.get(f"/approve/{_jid10a}")
        ck("GET /approve/<id> returns 200 (confirm page, not a mutation)",
           _rg10a.status_code == 200, _rg10a.status_code)
        ck("GET /approve/<id> renders an HTML confirm page (not JSON)",
           "text/html" in _rg10a.content_type, _rg10a.content_type)
        _job_after_get10a = _AS10.load_jobs()[_jid10a]
        ck("GET /approve/<id> did NOT change job status (still 'pending')",
           _job_after_get10a["status"] == "pending", _job_after_get10a["status"])
        ck("GET /approve/<id> did NOT set a decision",
           _job_after_get10a.get("decision") is None, _job_after_get10a.get("decision"))
        # The confirm page must contain a POST form targeting the real action, not a link
        # that itself mutates (otherwise it's just moved the vulnerability one click later).
        _body10a = _rg10a.get_data(as_text=True)
        ck("confirm page's form method is POST", 'method="POST"' in _body10a, _body10a[:200])
        ck(f"confirm page's form posts to /approve/{_jid10a}",
           f"/approve/{_jid10a}" in _body10a)

        # Now POST actually mutates
        _rp10a = _client10.post(f"/approve/{_jid10a}", json={})
        ck("POST /approve/<id> returns 200", _rp10a.status_code == 200, _rp10a.status_code)
        _job_after_post10a = _AS10.load_jobs()[_jid10a]
        ck("POST /approve/<id> DID change job status to 'approved'",
           _job_after_post10a["status"] == "approved", _job_after_post10a["status"])

        # --- /reject: GET must not mutate ---
        _jid10b = "job-csrf-reject"
        _AS10.save_jobs({_jid10b: {
            "id": _jid10b, "status": "pending", "decision": None,
            "result": {"file": "csrf_test2.pdf"}, "flags": [],
        }})
        _rg10b = _client10.get(f"/reject/{_jid10b}")
        ck("GET /reject/<id> returns 200 (confirm page, not a mutation)",
           _rg10b.status_code == 200, _rg10b.status_code)
        _job_after_get10b = _AS10.load_jobs()[_jid10b]
        ck("GET /reject/<id> did NOT change job status (still 'pending')",
           _job_after_get10b["status"] == "pending", _job_after_get10b["status"])

        _rp10b = _client10.post(f"/reject/{_jid10b}", json={})
        ck("POST /reject/<id> returns 200", _rp10b.status_code == 200, _rp10b.status_code)
        _job_after_post10b = _AS10.load_jobs()[_jid10b]
        ck("POST /reject/<id> DID change job status to 'rejected'",
           _job_after_post10b["status"] == "rejected", _job_after_post10b["status"])

        # --- /adjust: GET already only redirects (never mutated) — confirm that holds ---
        _jid10c = "job-csrf-adjust"
        _AS10.save_jobs({_jid10c: {
            "id": _jid10c, "status": "pending", "decision": None,
            "result": {"file": "csrf_test3.pdf"}, "flags": [],
        }})
        _rg10c = _client10.get(f"/adjust/{_jid10c}", follow_redirects=False)
        ck("GET /adjust/<id> redirects into the portal (302/301), never mutates",
           _rg10c.status_code in (301, 302), _rg10c.status_code)
        _job_after_get10c = _AS10.load_jobs()[_jid10c]
        ck("GET /adjust/<id> did NOT change job status", _job_after_get10c["status"] == "pending")

        # --- Unknown job on GET -> 404, not a 200 confirm page for a job that doesn't exist ---
        _r404_10 = _client10.get("/approve/does-not-exist-10")
        ck("GET /approve/<unknown> -> 404, not a confirm page for a nonexistent job",
           _r404_10.status_code == 404, _r404_10.status_code)
    finally:
        _AS10.JOBS_FILE = _orig_jobs_file10
        shutil.rmtree(_tmpdir10, ignore_errors=True)
except ImportError as _e:
    print(f"  [SKIP] approval_server GET-no-mutation tests — missing dependency: {_e}")

print("approval_email: emailed approve/reject/adjust links carry ?token= when the token "
      "gate is enabled (token mode previously 401'd every emailed action link)")
try:
    import approval_email as _AE11
    import importlib as _importlib11

    _orig_token11 = _AE11.APPROVAL_TOKEN
    try:
        # --- Token configured: every action link + the portal link carries ?token= ---
        _AE11.APPROVAL_TOKEN = "test-email-token-456"
        _html11 = _AE11.build_html_email(
            "job-email-1",
            {"area_m2": 500, "file": "email_test.pdf", "flags": []},
            png_b64="",
        )
        ck("approve link carries ?token= when APPROVAL_TOKEN is set",
           "/approve/job-email-1?token=test-email-token-456" in _html11)
        ck("reject link carries ?token= when APPROVAL_TOKEN is set",
           "/reject/job-email-1?token=test-email-token-456" in _html11)
        ck("adjust link carries ?token= when APPROVAL_TOKEN is set",
           "/adjust/job-email-1?token=test-email-token-456" in _html11)
        ck("portal review link carries ?token= when APPROVAL_TOKEN is set",
           "/review/job-email-1?token=test-email-token-456" in _html11)

        # --- No token configured: links are unchanged (no bare '?token=' with an empty value) ---
        _AE11.APPROVAL_TOKEN = ""
        _html11b = _AE11.build_html_email(
            "job-email-2",
            {"area_m2": 500, "file": "email_test.pdf", "flags": []},
            png_b64="",
        )
        ck("no token configured -> approve link has no ?token= param at all",
           "token=" not in _html11b.split('href="')[1].split('"')[0]
           if 'href="' in _html11b else True)
        ck("no token configured -> approve link is the plain job URL",
           "/approve/job-email-2" in _html11b)
    finally:
        _AE11.APPROVAL_TOKEN = _orig_token11
except ImportError as _e:
    print(f"  [SKIP] approval_email token-link tests — missing dependency: {_e}")

print("scale: detect_scale_bar rotation-agnostic + segmented-bar + crash-guard "
      "(Aryan field report — real SGP sheet 'title 1:250 only — no scale bar detected'; root "
      "cause: every real Fortel A0/A1 sheet is landscape content in a portrait MediaBox with "
      "page /Rotate 90/270, and PyMuPDF returns RAW pre-rotation coordinates, so a visually- "
      "horizontal bar is a stack of near-VERTICAL strokes the old horizontal-only test could "
      "never match; also the real bar is SEGMENTED [alternating-fill tick blocks] with a fused "
      "'25m' terminal label, and ms[0]-anchoring on text-extraction order crashed with "
      "'max() arg is an empty sequence' on two real Winvic sheets)")
try:
    import fitz as _fitz_sb

    def _gen_rotated_segmented_bar_sb(out_path, rotation=270):
        """Portrait-mediabox page (mimics the real 2384x3370 Winvic sheets) with a scale bar drawn
        as 4 stacked alternating-fill blocks (reportlab 're' rects) + a fused '25m' terminal tick,
        then rotated via PyMuPDF post-process — reproducing 'visually horizontal bar, raw-space
        near-vertical strokes' exactly as found on drawings/winvic/Yard_Area_Proposed_Site_Plan.pdf."""
        _c = canvas.Canvas(out_path, pagesize=(850, 1200))
        _c.setFont("Helvetica", 10)
        _c.drawString(50, 150, "Scale 1:250")
        x0, y0, block_h = 120, 400, 30
        for i in range(4):
            y = y0 + i * block_h
            _c.setFillColorRGB(0, 0, 0) if i % 2 == 0 else _c.setFillColorRGB(1, 1, 1)
            _c.rect(x0, y, 6, block_h, fill=1, stroke=1)
        for i, lab in enumerate(["0", "5", "10", "15", "20", "25m"]):
            _c.drawString(x0 + 10, y0 + i * block_h - 3, lab)
        _c.save()
        _d = _fitz_sb.open(out_path)
        _d[0].set_rotation(rotation)
        _d.saveIncr()
        _d.close()

    _sb_expected_k = 25 / 120   # 25 m over the 4x30pt stacked-block span

    _path_sb270 = "/tmp/_sb_rotated270.pdf"
    _gen_rotated_segmented_bar_sb(_path_sb270, rotation=270)
    _k_sb270, _info_sb270 = detect_scale_bar(_path_sb270)
    ck("rotation=270 segmented tick-block bar (real Winvic sheet style) now detects",
       _k_sb270 is not None and abs(_k_sb270 - _sb_expected_k) < 1e-9, (_k_sb270, _info_sb270))

    _path_sb90 = "/tmp/_sb_rotated90.pdf"
    _gen_rotated_segmented_bar_sb(_path_sb90, rotation=90)
    _k_sb90, _info_sb90 = detect_scale_bar(_path_sb90)
    ck("rotation=90 segmented tick-block bar also detects",
       _k_sb90 is not None and abs(_k_sb90 - _sb_expected_k) < 1e-9, (_k_sb90, _info_sb90))

    _path_sb0 = "/tmp/_sb_unrotated_control.pdf"
    _gen_rotated_segmented_bar_sb(_path_sb0, rotation=0)
    _k_sb0, _info_sb0 = detect_scale_bar(_path_sb0)
    ck("rotation=0 control (same fixture, no rotation) also detects — proves the fix is "
       "additive, not rotation-only", _k_sb0 is not None and abs(_k_sb0 - _sb_expected_k) < 1e-9,
       (_k_sb0, _info_sb0))

    # Unrotated segmented bar with a WIDE fused terminal tick ('50m'), several alternating blocks —
    # exercises the horizontal branch of the same clustering/merge logic.
    def _gen_segmented_bar_h_sb(out_path):
        _c = canvas.Canvas(out_path, pagesize=(1400, 900))
        _c.setFont("Helvetica", 10)
        _c.drawString(100, 800, "Scale 1:200")
        x0, y0, block_w = 200, 300, 40
        for i in range(5):
            x = x0 + i * block_w
            _c.setFillColorRGB(0, 0, 0) if i % 2 == 0 else _c.setFillColorRGB(1, 1, 1)
            _c.rect(x, y0, block_w, 8, fill=1, stroke=1)
        for i, lab in enumerate(["0", "10", "20", "30", "40", "50m"]):
            _c.drawString(x0 + i * block_w - 5, y0 - 15, lab)
        _c.save()

    _path_sb_h = "/tmp/_sb_segmented_h.pdf"
    _gen_segmented_bar_h_sb(_path_sb_h)
    _k_sbh, _info_sbh = detect_scale_bar(_path_sb_h)
    _expected_sbh = 50 / (5 * 40)
    ck("horizontal segmented alternating-fill bar with fused '50m' terminal tick",
       _k_sbh is not None and abs(_k_sbh - _expected_sbh) < 1e-9, (_k_sbh, _info_sbh))

    # Crash-guard regression: an early, text-order-first 'm' token with NO nearby bar/digits at
    # all must not raise (old code: max() on an empty generator -> ValueError, reproduced directly
    # on drawings/winvic/Yard_Area_Proposed_Site_Plan.pdf and Dock_Slab_Area_Proposed_Site_Plan.pdf).
    # The real scale bar (a plain line + bare 'm' label, further down the page) must still be found.
    def _gen_bad_anchor_sb(out_path):
        _c = canvas.Canvas(out_path, pagesize=(1400, 900))
        _c.setFont("Helvetica", 10)
        _c.drawString(700, 850, "m")                      # unrelated early 'm', no nearby digits
        _c.line(100, 150, 500, 150)
        _c.drawString(250, 160, "0          40 m")
        _c.save()

    _path_bad = "/tmp/_sb_bad_anchor.pdf"
    _gen_bad_anchor_sb(_path_bad)
    try:
        _k_bad, _info_bad = detect_scale_bar(_path_bad)
        ck("no crash when the first text-order 'm' token has zero nearby bar/digits "
           "(old ms[0] anchor -> max() on empty sequence -> ValueError)", True, (_k_bad, _info_bad))
        ck("bad-anchor fixture still finds the REAL bar via a later, valid label",
           _k_bad is not None and abs(_k_bad - 0.1) < 1e-9, (_k_bad, _info_bad))
    except Exception as _e:
        ck("no crash when the first text-order 'm' token has zero nearby bar/digits "
           "(old ms[0] anchor -> max() on empty sequence -> ValueError)", False,
           f"{type(_e).__name__}: {_e}")

    # A page with literally no scale-bar shape at all must still return cleanly, never raise.
    _path_none = "/tmp/_sb_no_bar_at_all.pdf"
    _c_none = canvas.Canvas(_path_none, pagesize=(800, 600))
    _c_none.drawString(100, 100, "no bar here, just some m words and 5 10 15 numbers")
    _c_none.save()
    try:
        _k_none, _info_none = detect_scale_bar(_path_none)
        ck("page with no real scale-bar shape returns (None, ...) cleanly, never raises",
           _k_none is None, (_k_none, _info_none))
    except Exception as _e:
        ck("page with no real scale-bar shape returns (None, ...) cleanly, never raises",
           False, f"{type(_e).__name__}: {_e}")

    # Detection improvements must never bypass verification: scale_consensus still gates a
    # disagreeing bar-vs-title pair (a rotated segmented bar detected via the fix, paired with a
    # deliberately wrong title-block scale) exactly as it does for the unrotated path.
    _k_gate, _flags_gate = scale_consensus([(_k_sb270, 1), (25 / 40, 1)], tol=0.03)
    ck("scale_consensus still REFUSES when the (correctly-detected, rotation-fixed) bar "
       "disagrees with a second reference beyond tol — detection fix does not bypass the gate",
       # Assert BEHAVIOUR (refusal) plus the surviving diagnostic, not the exact copy — the
       # assessor-facing sentence was rewritten into plain English after Inderjit could not
       # read it, and a wording change must not be able to mask a bypassed gate.
       _k_gate is None and any("MIXED-SCALE" in f for f in _flags_gate)
       and any("disagree" in f.lower() for f in _flags_gate), _flags_gate)

    _k_agree, _flags_agree = scale_consensus([(_k_sb270, 1), (_k_sb270 * 1.01, 1)], tol=0.03)
    ck("scale_consensus VERIFIES the rotation-fixed bar reading when a second reference agrees "
       "within tol", _k_agree is not None, _flags_agree)

except ImportError as _e:
    print(f"  [SKIP] scale.py rotation/segmented-bar tests — missing dependency: {_e}")

print("scale: real-sheet proof — Winvic sheets that already detected still detect after the fix, "
      "and the SGP-family real sheet that previously missed entirely now detects + VERIFIES "
      "against its title-block scale (never bypassing scale_consensus)")
try:
    import os as _os_sb

    _real_sheets_unchanged = [
        ("drawings/_int_d77.pdf", 0.08819445326652144),
        ("drawings/_int_d77_borders.pdf", 0.08819445326652144),
    ]
    for _pdf_path, _expected_k in _real_sheets_unchanged:
        if _os_sb.path.exists(_pdf_path):
            _k_chk, _info_chk = detect_scale_bar(_pdf_path)
            ck(f"unrotated gold fixture {_pdf_path} still detects the same k as before the fix",
               _k_chk is not None and abs(_k_chk - _expected_k) < 1e-6, (_k_chk, _info_chk))
        else:
            print(f"  [SKIP] real-sheet scale regression for {_pdf_path} — fixture not present")

    # The real Winvic sheets (270/90-rotated) must no longer crash, and the two with a genuine
    # readable segmented bar (Yard, Dock — same title-block template) must now agree with each
    # other (same physical bar) instead of one crashing and the other silently mis-anchoring.
    _winvic_rotated = [
        "drawings/winvic/Yard_Area_Proposed_Site_Plan.pdf",
        "drawings/winvic/Dock_Slab_Area_Proposed_Site_Plan.pdf",
    ]
    _winvic_ks = {}
    for _wp in _winvic_rotated:
        if _os_sb.path.exists(_wp):
            try:
                _k_w, _info_w = detect_scale_bar(_wp)
                _winvic_ks[_wp] = _k_w
                ck(f"{_wp} (rotation 270) no longer crashes calling detect_scale_bar",
                   True, (_k_w, _info_w))
            except Exception as _e:
                ck(f"{_wp} (rotation 270) no longer crashes calling detect_scale_bar",
                   False, f"{type(_e).__name__}: {_e}")
        else:
            print(f"  [SKIP] rotated Winvic scale regression for {_wp} — fixture not present")

    if len(_winvic_ks) == 2 and all(_v is not None for _v in _winvic_ks.values()):
        _vals = list(_winvic_ks.values())
        ck("Yard and Dock (same rotated title-block template, same '0 5 10 15 20 25m' bar) "
           "agree on k within 0.1% — both correctly read the same physical scale bar",
           abs(_vals[0] - _vals[1]) / _vals[1] < 0.001, _winvic_ks)

    # Full scale_for() (takeoff_unmarked's consensus-gated wrapper) on the real UNMARKED-vector,
    # rotated, segmented-bar sheet that most closely matches Aryan's real SGP sheet's shape
    # (same title-block family: rotated A0/A1, printed 1:N scale + graphical bar) — must now
    # VERIFY rather than fall back to 'title only — no scale bar detected'.
    _tp_site_plan = ("drawings/tender_pack/2-Enquiry/01-Tender/Drawings/Proposed_Site_Plan.pdf")
    if _os_sb.path.exists(_tp_site_plan):
        import takeoff_unmarked as _TU_sb
        _k_tp, _verified_tp, _note_tp, _sources_tp = _TU_sb.scale_for(_tp_site_plan)
        ck("real rotated tender-pack Proposed_Site_Plan.pdf: scale bar now VERIFIED against "
           "title-block (was previously undetectable/unverified pre-fix)",
           _verified_tp is True, _note_tp)
        ck("...and it went through scale_consensus (both sources present), not a bypass",
           "scale_bar" in _sources_tp and "title_block" in _sources_tp, _sources_tp)
    else:
        print(f"  [SKIP] tender-pack scale verification for {_tp_site_plan} — fixture not present")
except ImportError as _e:
    print(f"  [SKIP] scale.py real-sheet regression tests — missing dependency: {_e}")



# ── Deploy survival, ghost-job neutralisation, case-level quotation ───────────
# Railway's container filesystem is EPHEMERAL and this has already destroyed real assessor
# work once.  These tests pin the three properties that keep a case resumable across a deploy:
# artifacts resolve onto the mounted volume, the approval email never mints a duplicate job,
# and a case whose documents carry perimeters from DIFFERENT sources still exports.
print("\n[deploy survival / job identity / case export]")
import os as _os_ds
from pathlib import Path as _Path_ds
from storage_paths import resolve_storage_paths as _rsp

_vol = "/mnt/fortel-data"
_paths_vol = _rsp({"RAILWAY_VOLUME_MOUNT_PATH": _vol}, app_dir="/app")
ck("volume mounted: jobs file lands on the volume, not the ephemeral app dir",
   str(_paths_vol.jobs_file) == f"{_vol}/approval_jobs.json", _paths_vol.jobs_file)
ck("volume mounted: uploaded drawings survive a deploy (on the volume)",
   str(_paths_vol.drawings_dir) == f"{_vol}/drawings", _paths_vol.drawings_dir)
ck("volume mounted: generated quotations survive a deploy (on the volume)",
   str(_paths_vol.quotations_dir) == f"{_vol}/quotations", _paths_vol.quotations_dir)
ck("volume mounted: archive + backups survive a deploy too",
   str(_paths_vol.jobs_archive_file).startswith(_vol)
   and str(_paths_vol.backup_dir).startswith(_vol),
   (_paths_vol.jobs_archive_file, _paths_vol.backup_dir))

# The real failure mode: a deploy replaces the container (new app dir, same volume). Every
# artifact a half-finished assessment needs must resolve to the SAME location afterwards.
_before = _rsp({"RAILWAY_VOLUME_MOUNT_PATH": _vol}, app_dir="/app")
_after = _rsp({"RAILWAY_VOLUME_MOUNT_PATH": _vol}, app_dir="/app-redeploy-2")
ck("deploy cycle: a new container resolves the SAME jobs/drawings/quotations paths — an "
   "in-flight assessment is still resumable after a redeploy",
   (_before.jobs_file, _before.drawings_dir, _before.quotations_dir)
   == (_after.jobs_file, _after.drawings_dir, _after.quotations_dir),
   (_after.jobs_file, _after.drawings_dir, _after.quotations_dir))
_no_vol = _rsp({}, app_dir="/app")
_no_vol_after = _rsp({}, app_dir="/app-redeploy-2")
ck("no volume: paths follow the app dir and therefore do NOT survive a deploy — the "
   "difference between the two states must stay visible, not silently equal",
   _no_vol.jobs_file != _no_vol_after.jobs_file, (_no_vol.jobs_file, _no_vol_after.jobs_file))
ck("local dev without a volume keeps repo-local paths",
   str(_no_vol.jobs_file) == "/app/approval_jobs.json", _no_vol.jobs_file)
ck("explicit env overrides still win over the volume default",
   str(_rsp({"RAILWAY_VOLUME_MOUNT_PATH": _vol,
             "DRAWINGS_DIR": "/elsewhere/d"}, app_dir="/app").drawings_dir) == "/elsewhere/d")

# Ghost jobs: the portal creates a job at upload. If the approval email creates a SECOND one,
# the assessor's emailed link points at a record the portal never shows, and approving it
# leaves the real case pending forever.  request_approval must attach to the caller's job.
import tempfile as _tf_ds, json as _json_ds
import approval_email as _ae_ds
_ae_orig = (_ae_ds.JOBS_FILE, _ae_ds.render_snapshot, _ae_ds.send_email,
            _ae_ds._send_email_result)
try:
    _tmp_jobs = _Path_ds(_tf_ds.mkdtemp()) / "approval_jobs.json"
    _ae_ds.JOBS_FILE = _tmp_jobs
    _ae_ds.render_snapshot = lambda *a, **k: b"\x89PNG\r\n\x1a\n"
    _ae_ds.send_email = lambda *a, **k: True
    _ae_ds._send_email_result = lambda *a, **k: {
        "sent": True, "status": "sent", "reason": ""}
    _portal_job = _ae_ds.create_job("/x/site.pdf", {"file": "site.pdf", "area_m2": 1.0})
    _returned = _ae_ds.request_approval("/x/site.pdf", {"file": "site.pdf", "area_m2": 2.0},
                                        job_id=_portal_job)
    _jobs_after = _json_ds.loads(_tmp_jobs.read_text())
    ck("approval email attaches to the portal's existing job — no duplicate ghost record",
       len(_jobs_after) == 1, sorted(_jobs_after))
    ck("...and the emailed link carries the portal's OWN job id",
       _returned == _portal_job, (_returned, _portal_job))
    ck("...and the job record is refreshed with the new measurement, not left stale",
       _jobs_after[_portal_job]["result"]["area_m2"] == 2.0, _jobs_after[_portal_job]["result"])
    # A standalone/CLI caller with no job of its own must still get one created.
    _standalone = _ae_ds.request_approval("/x/other.pdf", {"file": "other.pdf", "area_m2": 3.0})
    ck("standalone caller (no job_id) still gets a job created — behaviour unchanged",
       _standalone != _portal_job and len(_json_ds.loads(_tmp_jobs.read_text())) == 2)
    # A stale/unknown id means the caller's real job is gone. Creating one here would be the
    # ghost bug by another route, so it must refuse loudly and leave the store untouched.
    _before_stale = len(_json_ds.loads(_tmp_jobs.read_text()))
    try:
        _ae_ds.request_approval("/x/z.pdf", {"file": "z.pdf", "area_m2": 4.0},
                                job_id="deadbeef")
        _refused, _why = False, "no error raised"
    except ValueError as _e_stale:
        _refused, _why = True, str(_e_stale)
    ck("unknown job_id is refused loudly rather than silently creating a ghost", _refused, _why)
    ck("...and the refusal leaves the job store untouched",
       len(_json_ds.loads(_tmp_jobs.read_text())) == _before_stale)
finally:
    (_ae_ds.JOBS_FILE, _ae_ds.render_snapshot, _ae_ds.send_email,
     _ae_ds._send_email_result) = _ae_orig

# The pipeline must expose the parameter the portal now passes; a silent signature drift here
# would reinstate the ghost-job bug without any test failing.
import inspect as _insp_ds, takeoff_pipeline as _tp_ds
ck("takeoff() accepts approval_job_id so the portal can own the job identity",
   "approval_job_id" in _insp_ds.signature(_tp_ds.takeoff).parameters)

# One case, two documents, perimeters arriving by DIFFERENT routes (a top-level perimeter_lm
# and a zone-carried one). These share a measurement key, so any future divergence in how the
# two rows are seeded breaks the whole case's export — not just one document's.
from quotation import generate_quotation as _gq_ds, quotation_xlsx as _qx_ds
_doc_top = {"file": "top.pdf", "area_m2": 100.0, "perimeter_lm": 40.0,
            "costing": {"area_m2": 100.0, "rate": 50.0,
                        "spec": {"depth_mm": 150, "mesh": "A393"}, "assumed": False}}
_doc_zone = {"file": "zone.pdf", "area_m2": 200.0,
             "costing": {"area_m2": 200.0, "rate": 50.0,
                         "spec": {"depth_mm": 150, "mesh": "A393"}, "assumed": False},
             "zones": [{"category": "external_yard", "perimeter_lm": 80.0, "area_m2": 200.0}]}
for _label, _case in (("top-level only", [_doc_top]), ("zone-carried only", [_doc_zone]),
                      ("mixed", [_doc_top, _doc_zone]),
                      ("mixed, reversed order", [_doc_zone, _doc_top])):
    try:
        _q_ds = _gq_ds(_case, project="P", client="C")
        _xl_ds = _qx_ds(_q_ds)
        _ok_ds, _g_ds = bool(_xl_ds), f"{len(_xl_ds)} bytes"
    except Exception as _e_ds:
        _ok_ds, _g_ds = False, f"{type(_e_ds).__name__}: {_e_ds}"
    ck(f"case export survives perimeters from mixed sources ({_label})", _ok_ds, _g_ds)

_q_mixed = _gq_ds([_doc_top, _doc_zone], project="P", client="C")
_perims = [m for m in _q_mixed["measurements"] if m["description"] == "Slab perimeter"]
ck("both documents' perimeters reach the take-off, aggregated not dropped",
   len(_perims) == 1 and abs(_perims[0]["qty"] - 120.0) < 0.01, _perims)
ck("...and each document keeps its own provenance row",
   len(_perims[0].get("quantity_rows") or []) == 2, _perims[0].get("quantity_rows"))




# ── Tender-pack upload size (Aryan 20 Aug: "the zip upload is still not working") ──────
# Root cause found by EXECUTION, not by reading: a real 118 MB pack returned HTTP 413 from
# Flask's HTML error page, which the portal's `r.json().catch(()=>({}))` turned into an empty
# object and a meaningless toast. Grepping for accept=".pdf,.zip" had "confirmed" zip support
# and proved nothing — the cap, not the wiring, was the blocker.
print("\n[tender-pack upload size]")
import importlib as _il_up
import approval_server as _AS_up
ck("default upload cap fits a real tender pack (packs run 100 MB - 2.4 GB)",
   _AS_up.MAX_UPLOAD_MB >= 1024, f"{_AS_up.MAX_UPLOAD_MB} MB")
ck("cap is env-configurable so a small host lowers it deliberately, not by accident",
   "MAX_UPLOAD_MB" in _AS_up.os.environ or _AS_up.MAX_UPLOAD_MB == 2048, _AS_up.MAX_UPLOAD_MB)
_c_up = _AS_up.app.test_client()
_r_up = _AS_up.app.response_class
with _AS_up.app.test_request_context():
    _resp_up = _AS_up._upload_too_large(None)
_body_up, _code_up = _resp_up[0].get_json(), _resp_up[1]
ck("oversized upload answers 413 with JSON the portal can display, not an HTML page",
   _code_up == 413 and isinstance(_body_up, dict) and "error" in _body_up, _body_up)
ck("...and the message tells the user what to do about it",
   "limit" in _body_up.get("error", "") and _body_up.get("max_upload_mb"), _body_up.get("error"))




# ── Section misfiling: the 5.2 Longwell "second area" / "wrong excel sheet" bug ─────────
# Inderjit, 20 Aug, sharing his screen: "I have approved this area only. Where did this second
# area come from" — he had classified it "dock slab". Aryan, same call: "it didn't pick up the
# correct sheet to mount the values on ... designed to use this as the fallback when the
# current sheet it's unable to pick up." Root cause: _normalise_section answered with the
# External-yard fallback for ANY spelling outside its alias table, so a dock area was priced
# as yard, with the yard rows and none of the dock formulas. Two reported bugs, one defect.
print("\n[BOQ section misfiling]")
from quotation import _normalise_section as _ns, UNCLASSIFIED_SECTION as _UNCLS
from quotation import generate_quotation as _gq_sec

for _probe in ("Dock Slab", "dock_slab", "Dock-Slab", "DOCK", "dock slabs"):
    ck(f"dock spelling {_probe!r} reaches Dock slabs, never the yard fallback",
       _ns(_probe) == "Dock slabs", _ns(_probe))
for _probe, _want in (("Upper floor", "Upper floor slabs"), ("upper_floor", "Upper floor slabs"),
                      ("ground_floor", "Ground floor slabs"), ("footpaths", "Footpath slabs")):
    ck(f"{_probe!r} maps to {_want}", _ns(_probe) == _want, _ns(_probe))
ck("an UNRECOGNISED section is surfaced for classification, not silently priced as yard",
   _ns("totally unknown thing") == _UNCLS, _ns("totally unknown thing"))
ck("a blank section still uses the caller's contextual default (a real default, not a misfile)",
   _ns("") == "External yard slabs" and _ns(None) == "External yard slabs")

# End-to-end: the exact shape of Inderjit's job — one yard area he approved, plus a second
# area he classified as dock. The dock quantity must never appear under External yard slabs.
for _label in ("Dock Slab", "dock_slab", "dock"):
    _doc_sec = {"file": "longwell.pdf", "area_m2": 500.0,
                "costing": {"area_m2": 500.0, "rate": 50.0,
                            "spec": {"depth_mm": 190, "mesh": "A393", "layers": 1},
                            "assumed": False},
                "zones": [{"category": "external_yard", "area_m2": 500.0}],
                "area_elements": [{"element_id": "ae-1", "name": "Dock apron", "category": "dock",
                                   "boq_scope": "main", "area_m2": 120.0, "section": _label}]}
    _q_sec = _gq_sec([_doc_sec], project="5.2 Longwell", client="Fortel")
    _yard_qtys = {li.get("qty") for li in _q_sec["line_items"]
                  if li.get("section") == "External yard slabs" and li.get("unit") == "m²"}
    ck(f"dock area classified {_label!r} does not surface under External yard slabs",
       120.0 not in _yard_qtys, sorted(_yard_qtys))




# ── Phantom approve block after zone classification ────────────────────────────────────
# Inderjit, 20 Aug: "I just tried on one project today and tried adjusting the AI's markup but
# it is not getting approved... Cannot approve zone classification." He hit it repeatedly; the
# known workaround was Aryan's "you have to re submit the zone classification" — which is the
# diagnosis, not a fix. /zones cleared zone_classification_required but left
# zone_allocation_stale set, so approval still demanded "reclassify/remeasure" — the thing he
# had just done. A gate with no route out is a dead end, and dead ends are the bug.
print("\n[zone classification clears its own approve gate]")
import approval_server as _AS_zb

def _zb_job(zone_specs, stale=True):
    zones = [{"zone_key": k, "area_m2": a, "category": c, "subjects": [k]}
             for k, a, c in zone_specs]
    return {
        "id": "zb", "status": "adjusted", "decision": "adjusted", "scale_confirmed": True,
        "zone_allocation_stale": stale,
        "zone_classification_required": any(z["category"] == "unclassified" for z in zones),
        "zones": zones,
        "result": {"measurement_state": "MEASURED_VERIFIED", "area_m2": 100.0,
                   "zone_allocation_stale": stale,
                   "zone_classification_required": any(
                       z["category"] == "unclassified" for z in zones),
                   "zones": zones, "flags": []},
    }

_zb_unclassified = _zb_job([("z1", 100.0, "unclassified")])
ck("stale allocation + an unclassified zone blocks approval (gate must still protect)",
   _AS_zb._zone_block_reason(_zb_unclassified) is not None,
   _AS_zb._zone_block_reason(_zb_unclassified))

# Every zone classified: the staleness the gate complains about has been resolved by the very
# act of classifying, so the block must lift without a second identical submission.
_zb_done = _zb_job([("z1", 100.0, "external_yard")], stale=False)
ck("once every zone is classified, approval is no longer blocked",
   _AS_zb._zone_block_reason(_zb_done) is None, _AS_zb._zone_block_reason(_zb_done))

# The gate must NOT be blanket-disabled: a partially classified job still blocks.
_zb_partial = _zb_job([("z1", 100.0, "external_yard"), ("z2", 50.0, "unclassified")])
ck("a PARTIALLY classified job still blocks — the gate is fixed, not removed",
   _AS_zb._zone_block_reason(_zb_partial) is not None,
   _AS_zb._zone_block_reason(_zb_partial))

# And the other genuine gates are untouched.
_zb_overlap = _zb_job([("z1", 100.0, "external_yard")], stale=False)
_zb_overlap["zone_geometry_overlap"] = True
ck("a genuine geometry overlap still blocks approval",
   _AS_zb._zone_block_reason(_zb_overlap) is not None)
_zb_mismatch = _zb_job([("z1", 100.0, "external_yard")], stale=False)
_zb_mismatch["zone_reference_mismatch"] = True
ck("a genuine zone-vs-BOQ mismatch still blocks approval",
   _AS_zb._zone_block_reason(_zb_mismatch) is not None)

# ── /adjust must not wipe a fully-categorized zone submission ──────────────────────────
# Aryan, 24 Aug: "approval sometimes stays blocked even after submitting it". Root cause:
# categorized_remeasure required `confirmed` (sanity.plausible() on the NEW area), so an
# assessor who categorized every region got their zones silently reset to [] and
# zone_allocation_stale=True whenever the resulting area tripped the plausibility guard (e.g.
# a legitimately large yard over the 60,000 m^2 single-zone bound) — a dead end identical in
# shape to the /zones phantom-block bug above, just reached through /adjust instead.
print("\n[/adjust preserves categorized zones even when the area is implausible]")
import tempfile as _zw_tempfile
_zw_orig_jobs_file = _AS_zb.JOBS_FILE
_zw_tmpdir = Path(_zw_tempfile.mkdtemp())
_AS_zb.JOBS_FILE = _zw_tmpdir / "jobs.json"
_zw_client = _AS_zb.app.test_client()
_zw_job_id = "zw1"
_zw_jobs = {_zw_job_id: {
    "id": _zw_job_id, "status": "adjusted", "decision": "adjusted",
    "scale_confirmed": False, "measurement_state": "MEASURED_UNVERIFIED",
    "zones": [{"zone_key": "external_yard", "area_m2": 5000.0, "category": "external_yard",
               "subjects": ["Yard"]},
              {"zone_key": "dock", "area_m2": 200.0, "category": "dock", "subjects": ["Dock"]}],
    "result": {"measurement_state": "MEASURED_UNVERIFIED", "area_m2": 5200.0,
               "zones": [{"zone_key": "external_yard", "area_m2": 5000.0,
                          "category": "external_yard", "subjects": ["Yard"]},
                         {"zone_key": "dock", "area_m2": 200.0, "category": "dock",
                          "subjects": ["Dock"]}], "flags": []},
}}
_AS_zb.save_jobs(_zw_jobs)
_zw_big = [[0, 0], [300000, 0], [300000, 300000], [0, 300000]]  # trips the >60,000 m^2 guard
_zw_small = [[0, 0], [10, 0], [10, 10], [0, 10]]
_zw_resp = _zw_client.post(f"/adjust/{_zw_job_id}", json={
    "regions": [_zw_big, _zw_small],
    "region_categories": ["external_yard", "dock"],
    "region_scopes": ["main", "main"],
    "scale_k": 1.0,
})
ck("categorized /adjust with an implausible area is still accepted",
   _zw_resp.status_code == 200, _zw_resp.status_code)
_zw_after = _AS_zb.load_jobs()[_zw_job_id]
_zw_categories = sorted(z.get("category") for z in (_zw_after.get("zones") or []))
ck("...its regions keep the categories the assessor actually submitted, not []",
   _zw_categories == ["dock", "external_yard"], _zw_categories)
ck("...zone_allocation_stale is NOT set — a categorized submission is not a stale aggregate one",
   _zw_after.get("zone_allocation_stale") is False, _zw_after.get("zone_allocation_stale"))
_zw_reason = _AS_zb._approve_block_reason(_zw_after)
ck("...the resulting block (if any) names the real problem (implausible/unverified area), "
   "not a false 'reclassify zones' demand for zones the assessor just classified",
   _zw_reason is not None and "reclassify/remeasure the drawing zones" not in _zw_reason,
   _zw_reason)
_AS_zb.JOBS_FILE = _zw_orig_jobs_file




# ── Assessor-readable scale flags ──────────────────────────────────────────────────────
# Inderjit read the old scale-disagreement flag aloud on 20 Aug and said "Nothing makes sense
# to me, really". The gate was RIGHT (two references 3x apart, refused to auto-pick) — the copy
# was unusable. The team's proposal was to hide yellow flags, which would turn a correct
# refusal into silence. Instead the first line is now an instruction an estimator can act on,
# with the technical detail retained behind a [detail] prefix.
print("\n[assessor-readable scale flags]")
from scale import scale_consensus as _sc_copy, calibrate_verified as _cal_copy
_k_copy, _f_copy = _sc_copy([(25.0, 120.0), (25.0, 40.0)])
ck("disagreeing references still REFUSE — copy change must not weaken the gate",
   _k_copy is None, _k_copy)
ck("the first flag tells the assessor what to DO, not what went wrong internally",
   _f_copy[0].startswith("Two scale readings") and "Confirm the scale before approving" in _f_copy[0],
   _f_copy[0][:90])
ck("...and it names the ways to set it (scale bar / parking bay / printed dimension)",
   all(t in _f_copy[0] for t in ("scale bar", "parking bay", "printed dimension")))
ck("the technical detail is kept for us, marked [detail]",
   any(f.startswith("[detail]") and "MIXED-SCALE" in f for f in _f_copy), _f_copy[-1][:80])
ck("no assessor-facing line shouts in block capitals at the estimator",
   not any(w.isupper() and len(w) > 4 for w in _f_copy[0].split()), _f_copy[0][:70])
_f_unver = _cal_copy(title_denominator=250)[1]
ck("title-block-only scale reads as an instruction, not a status code",
   _f_unver[0].startswith("Scale taken from the title block only"), _f_unver[0][:70])
ck("no-reference case tells the assessor to set it manually",
   "set the scale manually" in _cal_copy()[1][0], _cal_copy()[1][0])




# ── Provisional marker out of the DESCRIPTION cell ─────────────────────────────────────
# It was appended with a newline, so every provisional row rendered as two lines and looked
# untidy beside Fortel's own template. Their sheet keeps tender caveats in a column right of
# VALUE, so the marker lives there now — still impossible to miss, no longer wrapping the label.
print("\n[provisional marker placement]")
import io as _io_pm, openpyxl as _ox_pm
from quotation import (generate_quotation as _gq_pm, quotation_xlsx as _qx_pm,
                       PROVISIONAL_LABEL as _PL_pm, PROVISIONAL_COL as _PC_pm)
_doc_pm = {"file": "yard.pdf", "area_m2": 366.2,
           "costing": {"area_m2": 366.2, "rate": None, "spec": {}, "assumed": True},
           "zones": [{"category": "external_yard", "area_m2": 366.2}],
           "manhole_count_assumed": 12}
_ws_pm = _ox_pm.load_workbook(_io_pm.BytesIO(
    _qx_pm(_gq_pm([_doc_pm], project="P", client="C")))).active
_marked_pm = [r for r in range(1, _ws_pm.max_row + 1)
              if _ws_pm.cell(r, _PC_pm).value == _PL_pm]
ck("provisional rows still carry the marker (it was relocated, not dropped)",
   len(_marked_pm) > 0, f"{len(_marked_pm)} marked rows")
_line_rows_pm = [r for r in _marked_pm if "\n" in str(_ws_pm.cell(r, 1).value or "")]
ck("no marked row wraps its DESCRIPTION onto a second line any more",
   not _line_rows_pm, [str(_ws_pm.cell(r, 1).value)[:40] for r in _line_rows_pm])
ck("the marker never appears inside a line-item DESCRIPTION cell",
   not any(_PL_pm in str(_ws_pm.cell(r, 1).value or "") for r in _marked_pm))
ck("it sits to the right of VALUE, mirroring Fortel's REMEASURE caveat column",
   _PC_pm == 6, _PC_pm)




# ── Standards citations must never be read as a drawing scale ──────────────────────────
# Inderjit, 25 Aug call: "it took the wrong scale also it should have been like one is to five
# hundred". Cause, found by running his real project-6 sheet: the spec note "75mm sand:cement
# screed to BS 8204 Part 1: 2003" contains a literal "1: 2003", which beat the genuine 1:500 in
# the title block. k became 0.70661 instead of 0.17639 — 4x out linearly, 16x out on AREA.
# The guard is a standards-citation strip plus a four-digit scale SERIES whitelist; a year is in
# no series. It deliberately does NOT ban large scales — 1:1500 and 1:2000 are real and in live
# use on the CADIC site sheets, and banning them would silently break those instead.
print("\n[standards citations are not scales]")
from takeoff_unmarked import _title_scale_denominators as _tsd, FOUR_DIGIT_SCALE_SERIES as _series
_bs_note = "75mm sand:cement screed to BS 8204 Part 1: 2003 on Rockwool RockFloor"
_title_line = "Construction Thicknessess Plan Sheet 2 S2 1:500 A1 17.08.26"
ck("a BS standard year is not accepted as a scale", _tsd(_bs_note) == [], _tsd(_bs_note))
ck("the real title-block 1:500 wins over the BS year on the same sheet",
   _tsd(f"{_bs_note} {_title_line}") == [500], _tsd(f"{_bs_note} {_title_line}"))
ck("the hyphenated standard form is rejected too (BS 8204-1:2003)",
   _tsd("to BS 8204-1:2003 on") == [], _tsd("to BS 8204-1:2003 on"))
ck("BS EN / ISO variants are rejected",
   _tsd("to BS EN 1234-1:2004 spec") == [] and _tsd("per ISO 9001 Part 2: 1999") == [])
for _s, _want in (("Site plan 1:2000", 2000), ("Masterplan 1:1500", 1500),
                  ("Layout 1:1250", 1250), ("Plan 1:1000", 1000), ("Detail 1:2500", 2500)):
    ck(f"real large scale {_want} still parses — banning years must not ban these",
       _tsd(_s) == [_want], _tsd(_s))
for _s, _want in (("Yard 1:500", 500), ("Plan 1:200", 200), ("Detail 1:75", 75)):
    ck(f"ordinary scale {_want} unaffected", _tsd(_s) == [_want], _tsd(_s))
# 1:2000 is BOTH a real drawing scale and a plausible year, so the series whitelist alone
# cannot separate them — the standards-citation strip has to carry that case. Assert the hard
# one directly rather than assuming the whitelist covers it.
ck("a standards citation quoting the year 2000 is still rejected, even though 1:2000 is a "
   "real scale the whitelist allows",
   _tsd("screed to BS 8204 Part 1: 2000 on insulation") == [],
   _tsd("screed to BS 8204 Part 1: 2000 on insulation"))
ck("...and a genuine 1:2000 site plan on the same sheet still wins",
   _tsd("screed to BS 8204 Part 1: 2000 on insulation. Site plan 1:2000") == [2000],
   _tsd("screed to BS 8204 Part 1: 2000 on insulation. Site plan 1:2000"))

# The real client sheet, when present (drawings/ is gitignored — skip cleanly if absent).
_p6 = Path("drawings/inderjit_p6/"
           "6_31941-TTE-ZF-762-DR-C-0711-P01-Construction_Thicknessess_Plan.pdf")
if _p6.exists():
    import fitz as _fitz_p6
    _txt_p6 = _fitz_p6.open(str(_p6))[0].get_text() or ""
    ck("Inderjit's real project-6 sheet now yields 1:500, not the BS year 2003",
       _tsd(_txt_p6)[:1] == [500], _tsd(_txt_p6)[:3])
else:
    print(f"  [SKIP] real project-6 sheet not present — {_p6}")


print(f"\n==== {sum(P)}/{len(P)} PASS ====")
sys.exit(0 if all(P) else 1)
