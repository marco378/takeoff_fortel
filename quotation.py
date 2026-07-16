#!/usr/bin/env python3
"""
Fortel AI Takeoff — Quotation Generator

Turns a pipeline result dict into a formatted quotation output:
  - Plain-text quotation body (for email / Word paste)
  - JSON quotation record (for the tracker / n8n)
  - HTML quotation (for email or browser view)
  - Editable Excel quotation (numeric inputs + live formulas)

The quotation matches how Fortel actually issues quotes:
  - Lists the drawing used and its discipline (engineer / architect)
  - States the area measured
  - Gives the rate build-up (depth, mesh, mix)
  - Declares ASSUMPTIONS when build-up is not from an engineer drawing
  - States "subject to confirmation" wherever assumptions were made

Usage:
  from quotation import generate_quotation, quotation_text, quotation_html, quotation_xlsx

  result = takeoff_pipeline.takeoff("drawings/D77.pdf")
  q = generate_quotation(result, project="Hemington D77 Hard Landscaping",
                         client="Fortel", ref="FTL-2026-D77")
  print(quotation_text(q))

Run standalone:
  python3 quotation.py          # self-test with synthetic data
"""
import datetime, html, json, uuid, io, contextlib, re
from collections import OrderedDict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from slab_spec import (brief_spec_signature, build_brief_spec,
                       display_lines as brief_spec_display_lines, normalise_slab_type)

with contextlib.redirect_stdout(io.StringIO()):
    from costing import rate_buildup, MESH_KG
from defaults import spec_with_defaults, assumption_note

# Fortel company details for the quotation header
FORTEL_NAME    = "Fortel Group Limited"
FORTEL_ADDRESS = "Fortel House, Birmingham"
FORTEL_EMAIL   = "estimating@fortel.co.uk"
FORTEL_TEL     = "+44 (0)121 000 0000"

STANDARD_TERMS = (
    "This quotation is based on the drawings referenced below. "
    "Areas and build-ups are subject to confirmation upon receipt of the full construction specification. "
    "Rates are inclusive of supply, labour, plant and DPM. "
    "Excludes earthworks, drainage, kerbing and any items not listed above. "
    "Validity: 30 days from date of issue."
)

SECTION_ORDER = (
    "External yard slabs",
    "Dock slabs",
    "Ground floor slabs",
    "Upper floor slabs",
    "Prelims",
)
_SECTION_RANK = {section: index for index, section in enumerate(SECTION_ORDER)}
PROVISIONAL_LABEL = "PROVISIONAL — NO DETAILS PROVIDED"

ZONE_SECTION = {
    "external_yard": "External yard slabs",
    "dock": "Dock slabs",
    "ground_floor": "Ground floor slabs",
    "upper_floor": "Upper floor slabs",
}


def _unit_name(result: dict) -> str:
    """Return only a unit label proved by the drawing filename (never invent a plot code)."""
    filename = str(result.get("file") or Path(str(result.get("pdf_path") or "")).name)
    match = re.search(r"\bUnit[- _]?(\d+)\b", filename, re.I)
    return f"Unit-{match.group(1)}" if match else filename


def _expand_zone_results(results: list[dict]) -> list[dict]:
    """Fan a mixed marked drawing into unpriced BOQ-area inputs by proven zone category.

    A file-level rate/specification cannot be copied onto multiple zones: the client's real
    BOQ proves, for example, that Yard and Dock use different build-ups.  Explicit per-zone
    costing may be used when present; otherwise mixed-zone rate cells stay blank for the
    assessor.  Legacy/no-zone and genuinely single-zone results are unchanged.
    """
    expanded = []
    for parent in results:
        zones = [zone for zone in (parent.get("zones") or [])
                 if zone.get("category") in ZONE_SECTION and zone.get("area_m2") is not None]
        if not zones:
            expanded.append(parent)
            continue

        categories = {zone["category"] for zone in zones}
        mixed = len(categories) > 1
        parent_costing = dict(parent.get("costing") or {})
        per_zone_costing = parent.get("zone_costings") or {}
        per_zone_specs = parent.get("brief_specs") or {}
        for index, zone in enumerate(zones):
            category = zone["category"]
            virtual = dict(parent)
            virtual["_zone_expanded"] = True
            virtual["zones"] = []
            virtual.pop("perimeter_lm", None)
            virtual.pop("polygon_pts", None)
            virtual["area_m2"] = float(zone["area_m2"])
            virtual["quotation_section"] = ZONE_SECTION[category]
            virtual["unit_name"] = _unit_name(parent)
            virtual["zone_category"] = category
            virtual["brief_spec"] = per_zone_specs.get(category) or build_brief_spec(category)

            explicit_costing = per_zone_costing.get(category)
            if explicit_costing:
                costing = dict(explicit_costing)
                costing["area_m2"] = float(zone["area_m2"])
            elif mixed:
                # Preserve existing extras exactly once, but never inherit an aggregate rate,
                # build-up, or value into a different zone.
                costing = {
                    "area_m2": float(zone["area_m2"]), "rate": None, "total_gbp": None,
                    "assumed": True, "spec": {}, "breakdown": {},
                    "extras": list(parent_costing.get("extras") or []) if index == 0 else [],
                }
            else:
                costing = dict(parent_costing)
                costing["area_m2"] = float(zone["area_m2"])
            if index and costing.get("extras"):
                costing["extras"] = []
            virtual["costing"] = costing
            expanded.append(virtual)
    return expanded


def quotation_section(result: dict) -> str:
    """Return the client's canonical BOQ section for a drawing/result."""
    explicit = str(result.get("quotation_section") or "").strip()
    if explicit:
        for section in SECTION_ORDER:
            if explicit.casefold() == section.casefold():
                return section

    label = " ".join(str(result.get(key) or "") for key in (
        "file", "pdf", "pdf_path", "project_name", "name", "type"
    )).casefold().replace("_", "-")
    if "prelim" in label:
        return "Prelims"
    if "dock" in label:
        return "Dock slabs"
    if any(term in label for term in (
            "upper floor", "first floor", "mezzanine", "level 1", "level-1")):
        return "Upper floor slabs"
    if any(term in label for term in (
            "ground floor", "ground-floor", "office", "transport", "internal slab")):
        return "Ground floor slabs"
    return "External yard slabs"


def _normalise_section(section, fallback="External yard slabs"):
    probe = str(section or "").strip().casefold()
    aliases = {
        "yard": "External yard slabs", "external": "External yard slabs",
        "external yard": "External yard slabs", "external yard slabs": "External yard slabs",
        "dock": "Dock slabs", "dock slabs": "Dock slabs",
        "ground": "Ground floor slabs", "ground floor": "Ground floor slabs",
        "ground floor slabs": "Ground floor slabs", "gf ancillary": "Ground floor slabs",
        "upper": "Upper floor slabs", "upper floors": "Upper floor slabs",
        "upper floor slabs": "Upper floor slabs",
        "prelims": "Prelims", "preliminaries": "Prelims",
    }
    return aliases.get(probe, fallback)


def _spec_key(costing, brief_spec=None):
    # The client checklist's field-level provisional state is part of specification identity:
    # equal effective values cannot be collapsed when one is assumed and one is confirmed.
    if brief_spec:
        return brief_spec_signature(brief_spec)
    return json.dumps(costing.get("spec") or {}, sort_keys=True, default=str, separators=(",", ":"))


def _provisional_text(li):
    return f"{li['description']} [{PROVISIONAL_LABEL}]" if li.get("provisional") else li["description"]


def _unique_notes(notes):
    return list(dict.fromkeys(note for note in notes if note))


# ── Core generator ────────────────────────────────────────────────────────────

def generate_quotation(result: dict | list, project: str = "", client: str = "",
                       ref: str = None, extras: list = None, commercial: dict = None) -> dict:
    """Build one structured quotation from one result or a project's result list.

    Units with the same canonical section, specification, existing rate and assumption
    provenance are aggregated into one quantity.  Differing specifications remain separate
    rows on the same quotation; no rate is recalculated here.
    """
    source_results = [r for r in (result if isinstance(result, (list, tuple)) else [result]) if r]
    if not source_results:
        source_results = [{}]
    results = _expand_zone_results(source_results)
    ref = ref or f"FTL-{datetime.date.today().strftime('%Y%m%d')}-{str(uuid.uuid4())[:4].upper()}"
    today = datetime.date.today().strftime("%-d %B %Y")

    groups = OrderedDict()
    extra_rows = []
    declarations = []
    pipeline_flags = []
    measurements_by_key = OrderedDict()
    drawings = []
    commercial = dict(commercial or {})

    for unit in results:
        costing = unit.get("costing") or {}
        area = costing.get("area_m2") or unit.get("area_m2") or 0
        spec = costing.get("spec") or {}
        assumed = bool(costing.get("assumed", True))
        rate = costing.get("rate")
        section = quotation_section(unit)
        drawing = unit.get("file") or Path(str(unit.get("pdf_path") or "")).name
        if drawing and drawing not in drawings:
            drawings.append(drawing)
        flags = list(unit.get("flags") or [])
        pipeline_flags.extend(flags)

        stored_brief_spec = unit.get("brief_spec")
        if stored_brief_spec:
            brief_spec = stored_brief_spec
        else:
            # Legacy records carry only an effective pricing spec, not field provenance.
            # Show those values for context but keep every field visibly provisional.
            brief_spec = build_brief_spec(
                normalise_slab_type(section, text=drawing), effective_spec=spec,
            )

        if area:
            # Existing rate is part of the key so stale/different priced results can never be
            # silently collapsed under one arbitrary rate. Matching specs/rates aggregate.
            group_provisional = assumed or any(
                field.get("provisional", True)
                for field in (brief_spec.get("fields") or {}).values()
                if isinstance(field, dict)
            )
            key = (section, _spec_key(costing, brief_spec), rate, group_provisional)
            group = groups.setdefault(key, {
                "section": section, "spec": spec, "brief_spec": brief_spec,
                "rate": rate, "assumed": group_provisional,
                "area": 0.0, "drawings": [], "area_rows": [],
                "breakdown": costing.get("breakdown") or {},
            })
            group["area"] += float(area)
            if drawing and drawing not in group["drawings"]:
                group["drawings"].append(drawing)
            area_label = (unit.get("unit_name") or unit.get("area_label") or drawing
                          or f"Measured area {len(group['area_rows']) + 1}")
            group["area_rows"].append({
                "description": str(area_label), "qty": float(area), "unit": "m²",
                "drawing": drawing,
            })

        if assumed:
            if all(spec.get(key) is not None for key in ("depth_mm", "mesh")):
                declaration = assumption_note(spec)
            else:
                declaration = "zone specification not provided; assessor must complete the slab checklist"
            declarations.append(f"{PROVISIONAL_LABEL}: {declaration}")
        if unit.get("source_discipline") == "architect":
            declarations.append(
                "Area measured from architect's hard-landscaping drawing — ±5% tolerance applies. "
                "No engineer construction-detail drawing found in the pack."
            )
        declarations.extend(f for f in flags if (
            "ASSUMED" in f or "architect" in f.lower() or "tolerance" in f.lower()))

        perimeter = unit.get("perimeter_lm")
        if perimeter is None and unit.get("polygon_pts") and unit.get("scale_k"):
            from geometry import polygon_perimeter_lm
            perimeter = polygon_perimeter_lm(unit["polygon_pts"], unit["scale_k"])
        if perimeter is not None:
            mkey = (section, "Slab perimeter", "Lm", False)
            measurement = measurements_by_key.setdefault(mkey, {
                "section": section, "description": "Slab perimeter", "qty": 0.0,
                "unit": "Lm", "provisional": False, "drawings": [],
            })
            measurement["qty"] += float(perimeter)
            if drawing and drawing not in measurement["drawings"]:
                measurement["drawings"].append(drawing)

        assumed_manhole_count = unit.get("manhole_count_assumed")
        if assumed_manhole_count is not None:
            mkey = (section, "Manholes (assumed fallback)", "Nr", True)
            measurement = measurements_by_key.setdefault(mkey, {
                "section": section, "description": "Manholes (assumed fallback)", "qty": 0,
                "unit": "Nr", "provisional": True, "drawings": [],
                "provisional_reason": PROVISIONAL_LABEL,
            })
            measurement["qty"] += int(assumed_manhole_count)
            if drawing and drawing not in measurement["drawings"]:
                measurement["drawings"].append(drawing)
            provenance = next((f for f in flags if "manhole_count_assumed=" in f), "")
            declarations.append(f"{PROVISIONAL_LABEL}: {provenance or 'assumed manhole quantity'}")

        if extras is None:
            for ex in costing.get("extras", []):
                provisional = bool(ex.get("estimate", False))
                item_rate = ex.get("rate")
                item_value = ex.get("value")
                if item_value is None and isinstance(item_rate, (int, float)):
                    item_value = round(float(ex.get("qty") or 0) * item_rate, 2)
                extra_rows.append({
                    "section": _normalise_section(ex.get("section"), section),
                    "description": ex["description"], "qty": ex["qty"], "unit": ex["unit"],
                    "rate": item_rate, "value": item_value,
                    "value_status": ex.get("value_status") or "",
                    "provisional": provisional,
                    "provisional_reason": PROVISIONAL_LABEL if provisional else "",
                    "drawings": [drawing] if drawing else [],
                })
                if provisional:
                    declarations.append(
                        f"{PROVISIONAL_LABEL}: {ex['description']} is an estimate from existing "
                        "measurement provenance and must be confirmed before issue."
                    )

    # Marked-zone lengths are source quantities, never implicit prices.  Keep the per-unit
    # provenance so the workbook can expose editable source rows just like its area take-off.
    for unit in source_results:
        drawing = unit.get("file") or Path(str(unit.get("pdf_path") or "")).name
        unit_label = _unit_name(unit)
        for zone in unit.get("zones") or []:
            category = zone.get("category")
            quantities = []
            if category in ("channel", "transition") and zone.get("length_lm") is not None:
                quantities.append((
                    "External yard slabs",
                    "Channel length" if category == "channel" else "Transition length",
                    float(zone["length_lm"]),
                ))
            if category in ZONE_SECTION and zone.get("perimeter_lm") is not None:
                quantities.append((ZONE_SECTION[category], "Slab perimeter",
                                   float(zone["perimeter_lm"])))
            for section, description, quantity in quantities:
                mkey = (section, description, "Lm", False)
                measurement = measurements_by_key.setdefault(mkey, {
                    "section": section, "description": description, "qty": 0.0,
                    "unit": "Lm", "provisional": False, "drawings": [],
                    "quantity_rows": [], "assessor_rate_required": True,
                })
                measurement["qty"] += quantity
                measurement["quantity_rows"].append({
                    "description": unit_label, "qty": quantity, "unit": "Lm",
                    "drawing": drawing,
                })
                if drawing and drawing not in measurement["drawings"]:
                    measurement["drawings"].append(drawing)

    line_items = []
    specifications = []
    for group_number, group in enumerate(
            sorted(groups.values(), key=lambda g: _SECTION_RANK[g["section"]]), 1):
        spec = group["spec"]
        area = round(group["area"], 3)
        depth_mm = spec.get("depth_mm")
        mesh = spec.get("mesh")
        mix = spec.get("conc_mix")
        layers = spec.get("layers")
        group_id = f"spec-{group_number}"
        specifications.append({
            "id": group_id,
            "section": group["section"],
            "slab_type": group["brief_spec"].get("slab_type"),
            "slab_type_label": group["brief_spec"].get("slab_type_label"),
            "fields": group["brief_spec"].get("fields") or {},
            "display_lines": brief_spec_display_lines(group["brief_spec"]),
            "provisional": group["assumed"],
            "drawings": group["drawings"],
            "area_rows": group["area_rows"],
            "area_m2": area,
        })
        common = {
            "section": group["section"], "qty": area, "unit": "m²",
            "drawings": group["drawings"], "specification_id": group_id,
        }
        known_parts = []
        if depth_mm is not None: known_parts.append(f"{depth_mm}mm")
        if mix: known_parts.append(str(mix))
        known_parts.append("slab")
        if layers is not None and mesh: known_parts.append(f"{layers}× {mesh} mesh")
        slab_desc = " ".join(known_parts) + " (supply & lay)"
        line_items.append({
            **common, "description": slab_desc, "rate": group["rate"],
            "value": (round(area * group["rate"], 2)
                      if isinstance(group.get("rate"), (int, float)) else None),
            "assessor_rate_required": group.get("rate") is None,
            "provisional": group["assumed"],
            "provisional_reason": PROVISIONAL_LABEL if group["assumed"] else "",
        })
        # Existing quotation adders are preserved unchanged; only their section/aggregated
        # quantity changes so all units share the client's requested one-tab structure.
        adders = ([
            ("Final trim ±50mm", area, "m²", 1.40),
            ("Saw cuts & bay joints", area, "m²", 4.85),
        ] if isinstance(group.get("rate"), (int, float)) else [])
        for desc, qty, unit_name, item_rate in adders:
            line_items.append({
                **common, "description": desc, "qty": qty, "unit": unit_name,
                "rate": item_rate, "value": round(qty * item_rate, 2),
                "provisional": False,
            })

    if extras is not None:
        fallback_section = quotation_section(results[0])
        for ex in extras:
            if isinstance(ex, dict):
                desc, qty, unit_name = (ex["description"], ex["qty"], ex["unit"])
                item_rate = ex.get("rate")
                section = _normalise_section(ex.get("section"), fallback_section)
                provisional = bool(ex.get("estimate", False))
                value_status = ex.get("value_status") or ""
                item_value = ex.get("value")
            else:
                desc, qty, unit_name, item_rate = ex
                section = fallback_section
                # Use the result's existing provenance; do not infer from wording alone.
                provisional = bool(results[0].get("manhole_count_estimate") and "MH" in desc)
                value_status = ""
                item_value = None
            if item_value is None and isinstance(item_rate, (int, float)):
                item_value = round(qty * item_rate, 2)
            extra_rows.append({
                "section": section, "description": desc, "qty": qty, "unit": unit_name,
                "rate": item_rate, "value": item_value, "value_status": value_status,
                "provisional": provisional,
                "provisional_reason": PROVISIONAL_LABEL if provisional else "",
                "drawings": [],
            })
            if provisional:
                declarations.append(f"{PROVISIONAL_LABEL}: {desc} must be confirmed before issue.")

    line_items.extend(extra_rows)
    line_items.sort(key=lambda li: _SECTION_RANK.get(li["section"], len(SECTION_ORDER)))
    measurements = sorted(measurements_by_key.values(),
                          key=lambda item: _SECTION_RANK[item["section"]])
    subtotal = round(sum(float(li["value"]) for li in line_items
                         if isinstance(li.get("value"), (int, float))), 2)
    assumed = any(specification.get("provisional") for specification in specifications)
    for specification in specifications:
        provisional_labels = [
            line["label"] for line in specification["display_lines"] if line["provisional"]
        ]
        if provisional_labels:
            declarations.append(
                f"{PROVISIONAL_LABEL}: {specification['slab_type_label'] or specification['section']} "
                f"— {', '.join(provisional_labels)}."
            )

    if extras is None:
        declarations.append(
            "NOTE — Extra-over items (slot drains, transitions, etc. other than manholes) "
            "not included: quantities cannot be measured from the drawing automatically "
            "and require manual assessment by the assessor before issue."
        )

    first_costing = results[0].get("costing") or {}
    first_spec = first_costing.get("spec") or {}
    disciplines = _unique_notes([unit.get("source_discipline", "unknown") for unit in results])
    perimeter_measurements = [m for m in measurements if m["description"] == "Slab perimeter"]
    total_perimeter = (round(sum(float(m["qty"]) for m in perimeter_measurements), 1)
                       if perimeter_measurements else None)
    return {
        "ref": ref, "date": today, "project": project, "client": client,
        "drawing": ", ".join(drawings), "drawings": drawings,
        "drawing_type": results[0].get("type", ""),
        "discipline": ", ".join(disciplines),
        "area_m2": round(sum(float((unit.get("costing") or {}).get("area_m2")
                                   or unit.get("area_m2") or 0) for unit in results), 3),
        "perimeter_lm": total_perimeter,
        "measurements": measurements,
        "spec": {
            "depth_mm": first_spec.get("depth_mm"),
            "mesh": first_spec.get("mesh"),
            "conc_mix": first_spec.get("conc_mix"),
            "layers": first_spec.get("layers"),
        },
        "specifications": specifications,
        "rate": first_costing.get("rate"), "breakdown": first_costing.get("breakdown") or {},
        "line_items": line_items, "section_order": list(SECTION_ORDER),
        "subtotal_gbp": subtotal, "total_gbp": subtotal,
        "assumed": assumed, "declarations": _unique_notes(declarations),
        "pipeline_flags": _unique_notes(pipeline_flags), "terms": STANDARD_TERMS,
        # The real BOQ locates these fields in the header/post-total block.  They are
        # project-specific and therefore optional; nothing is silently copied from one quote.
        "revision": commercial.get("revision") or ref,
        "measurement_basis": commercial.get("measurement_basis") or "",
        "joint_layout_note": commercial.get("joint_layout_note") or "",
        "joint_details_note": commercial.get("joint_details_note") or "",
        "market_warning": commercial.get("market_warning") or "",
        "commercial_terms": list(commercial.get("terms") or []),
    }


# ── Formatters ────────────────────────────────────────────────────────────────

def quotation_text(q: dict) -> str:
    """Plain-text quotation — suitable for email body or Word paste."""
    lines = [
        f"{FORTEL_NAME}",
        f"Quotation Ref: {q['ref']}",
        f"Date: {q['date']}",
        "",
        f"Project : {q['project'] or '—'}",
        f"Client  : {q['client'] or '—'}",
        f"Drawing : {q['drawing']} ({q.get('drawing_type','')}, {q.get('discipline','')}'s drawing)",
        "",
    ]
    if q.get("specifications"):
        lines.append("SLAB SPECIFICATION CHECKLIST:")
        for specification in q["specifications"]:
            lines.append(
                f"  {specification['section']} — "
                f"{specification.get('slab_type_label') or 'Specification'}"
            )
            for field in specification.get("display_lines", []):
                lines.append(f"    {field['label']}: {field['value']}")
        lines.append("")
    lines += [
        "=" * 70,
        f"{'DESCRIPTION':<42}{'QTY':>8}{'UNIT':>6}{'RATE £':>9}{'VALUE £':>13}",
        "-" * 70,
    ]
    last_section = None
    for li in q["line_items"]:
        if li["section"] != last_section:
            lines.append(f"\n  {li['section'].upper()}")
            last_section = li["section"]
        description = _provisional_text(li)
        rate_text = f"{li['rate']:.2f}" if isinstance(li.get("rate"), (int, float)) else ""
        value_text = (str(li.get("value_status")) if li.get("value_status") else
                      (f"{li['value']:,.2f}" if isinstance(li.get("value"), (int, float)) else ""))
        lines.append(
            f"  {description:<40}{li['qty']:>8,.0f}{li['unit']:>6}"
            f"{rate_text:>9}{value_text:>13}"
        )
    lines += [
        "-" * 70,
        f"{'TOTAL NETT (excl. VAT)':<55}{q['total_gbp']:>15,.2f}",
        "=" * 70,
        "",
    ]
    if q.get("measurements"):
        lines.append("INFORMATIONAL MEASUREMENTS (NOT PRICED):")
        for measurement in q["measurements"]:
            description = _provisional_text(measurement)
            lines.append(
                f"  {measurement['section']} — {description}: "
                f"{measurement['qty']:,.1f} {measurement['unit']}"
            )
        lines.append("")
    if q["declarations"]:
        lines.append("NOTES / ASSUMPTIONS:")
        for d in q["declarations"]:
            lines.append(f"  • {d}")
        lines.append("")
    lines += [
        "STANDARD TERMS:",
        f"  {q['terms']}",
        "",
        f"Issued by: {FORTEL_NAME} · {FORTEL_EMAIL} · {FORTEL_TEL}",
    ]
    return "\n".join(lines)


def quotation_html(q: dict) -> str:
    """Self-contained HTML quotation — suitable for email or browser view."""
    def _h(value):
        return html.escape(str(value if value is not None else ""), quote=True)

    def _row(li):
        provisional = (f" <strong style='color:#9a6500'>{PROVISIONAL_LABEL}</strong>"
                       if li.get("provisional") else "")
        rate = (f"£{li['rate']:.2f}" if isinstance(li.get("rate"), (int, float)) else "")
        value = (_h(li.get("value_status")) if li.get("value_status") else
                 (f"£{li['value']:,.2f}" if isinstance(li.get("value"), (int, float)) else ""))
        return (f"<tr><td>{_h(li['section'])}</td><td>{_h(li['description'])}{provisional}</td>"
                f"<td style='text-align:right'>{li['qty']:,.0f} {_h(li['unit'])}</td>"
                f"<td style='text-align:right'>{rate}</td>"
                f"<td style='text-align:right'>{value}</td></tr>")

    assumed_banner = ""
    if q["assumed"]:
        assumed_banner = (
            f"<div style='background:#fef9ec;border-left:4px solid #e67e22;"
            f"padding:10px 16px;margin:16px 0;font-size:13px;color:#7a5200'>"
            f"⚠ <b>Build-up assumed</b> — {_h(q['declarations'][0]) if q['declarations'] else ''}"
            f"</div>"
        )

    decl_html = ""
    if q["declarations"]:
        items = "".join(f"<li>{_h(d)}</li>" for d in q["declarations"])
        decl_html = f"<h3>Notes / Assumptions</h3><ul style='font-size:13px'>{items}</ul>"

    rows = "\n".join(_row(li) for li in q["line_items"])
    specification_html = ""
    if q.get("specifications"):
        spec_blocks = []
        for specification in q["specifications"]:
            fields = "".join(
                f"<tr><td>{_h(field['label'])}</td><td>{_h(field['value'])}</td></tr>"
                for field in specification.get("display_lines", [])
            )
            spec_blocks.append(
                f"<h4 style='margin:12px 0 4px'>{_h(specification['section'])} — "
                f"{_h(specification.get('slab_type_label') or 'Specification')}</h4>"
                f"<table style='margin:4px 0 12px'><tbody>{fields}</tbody></table>"
            )
        specification_html = "<h3>Slab specification checklist</h3>" + "".join(spec_blocks)
    measurement_html = ""
    if q.get("measurements"):
        measurement_rows = ""
        for measurement in q["measurements"]:
            marker = (f" <strong style='color:#9a6500'>{PROVISIONAL_LABEL}</strong>"
                      if measurement.get("provisional") else "")
            measurement_rows += (
                f"<tr><td>{_h(measurement['section'])}</td>"
                f"<td>{_h(measurement['description'])}{marker}</td>"
                f"<td style='text-align:right'>{measurement['qty']:,.1f} "
                f"{_h(measurement['unit'])}</td></tr>"
            )
        measurement_html = (
            "<h3>Informational measurements <small>(not priced)</small></h3>"
            "<table><thead><tr><th>Section</th><th>Measurement</th>"
            "<th style='text-align:right'>Quantity</th></tr></thead>"
            f"<tbody>{measurement_rows}</tbody></table>"
        )

    return f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<title>Quotation {_h(q['ref'])}</title>
<style>
body{{font-family:Arial,sans-serif;font-size:14px;color:#111;max-width:760px;margin:32px auto;padding:0 16px}}
header{{background:#13294b;color:#fff;padding:20px 24px;border-radius:8px 8px 0 0}}
header h1{{margin:0;font-size:22px}}
header p{{margin:4px 0 0 0;opacity:.75;font-size:13px}}
.meta{{display:flex;gap:40px;padding:14px 0;border-bottom:1px solid #ddd;font-size:13px;color:#555}}
.meta b{{color:#111}}
table{{width:100%;border-collapse:collapse;margin:16px 0}}
th{{background:#f7f8fa;padding:7px 10px;text-align:left;font-size:12px;
    text-transform:uppercase;letter-spacing:.5px;color:#666;border-bottom:2px solid #ddd}}
td{{padding:6px 10px;border-bottom:1px solid #eee;font-size:13px}}
.total-row td{{font-weight:700;font-size:15px;border-top:2px solid #111;background:#f7f8fa}}
.terms{{font-size:12px;color:#888;margin-top:20px;line-height:1.6}}
footer{{font-size:11px;color:#bbb;margin-top:20px;padding-top:12px;border-top:1px solid #eee}}
</style></head><body>
<header>
  <h1>{_h(FORTEL_NAME)}</h1>
  <p>Quotation {_h(q['ref'])} &nbsp;·&nbsp; {_h(q['date'])}</p>
</header>
<div class="meta">
  <div><b>Project</b><br>{_h(q.get('project') or '—')}</div>
  <div><b>Client</b><br>{_h(q.get('client') or '—')}</div>
  <div><b>Drawing</b><br>{_h(q['drawing'])}</div>
  <div><b>Area</b><br><b style='font-size:18px;color:#13294b'>{q['area_m2']:,.0f} m²</b></div>
</div>
{assumed_banner}
{specification_html}
<table>
  <thead><tr>
    <th>Section</th><th>Description</th><th style='text-align:right'>Qty</th>
    <th style='text-align:right'>Rate</th><th style='text-align:right'>Value</th>
  </tr></thead>
  <tbody>{rows}</tbody>
  <tfoot>
    <tr class='total-row'>
      <td colspan='4'>TOTAL NETT (excl. VAT)</td>
      <td style='text-align:right'>£{q['total_gbp']:,.2f}</td>
    </tr>
  </tfoot>
</table>
{measurement_html}
{decl_html}
<p class='terms'>{_h(q['terms'])}</p>
<footer>{_h(FORTEL_NAME)} &nbsp;·&nbsp; {_h(FORTEL_EMAIL)} &nbsp;·&nbsp; {_h(FORTEL_TEL)}</footer>
</body></html>"""


def quotation_json(q: dict) -> str:
    """JSON record — for the tracker / n8n."""
    return json.dumps(q, indent=2, default=str, ensure_ascii=False)


def _excel_text(value):
    text = str(value or "")
    return f"'{text}" if text.startswith(("=", "+", "-", "@")) else text


def _excel_date(value):
    """Return a real Excel date when the quotation date is in a known format."""
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    for fmt in ("%d %B %Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.datetime.strptime(str(value), fmt).date()
        except (TypeError, ValueError):
            pass
    return _excel_text(value)


def _excel_unit(value):
    """Match the unit spellings used in Fortel's editable BOQ template."""
    text = str(value or "")
    return {
        "m²": "m2", "m2": "m2",
        "m³": "m3", "m3": "m3",
        "lm": "Lm", "nr": "Nr", "item": "Item", "t": "T",
    }.get(text.casefold(), text)


def _excel_section_title(section, items):
    labels = {
        "External yard slabs": "External Yard Slabs",
        "Dock slabs": "Dock Slabs",
        "Ground floor slabs": "Ground Floor Slabs (Ancillary Areas)",
        "Upper floor slabs": "Upper Floors",
        "Prelims": "Prelims",
    }
    title = labels.get(section, str(section))
    if section != "Prelims" and any(item.get("provisional") for item in items):
        title += "- Provisional Cost (No Details)"
    return title


def quotation_xlsx(q: dict) -> bytes:
    """Editable one-sheet Excel quotation in Fortel's client-facing BOQ layout."""
    wb = Workbook()
    ws = wb.active
    ws.title = "REV_01"
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 130

    pale_gold = "FFF2CC"
    section_blue = "0070C0"
    black = "000000"
    white = "FFFFFF"
    red = "FF0000"
    thin_grey = Side(style="thin", color="D9D9D9")
    currency_format = ('_-[$£-809]* #,##0.00_-;\\-[$£-809]* #,##0.00_-;'
                       '_-[$£-809]* "-"??_-;_-@_-')

    def _date_text(value):
        parsed = _excel_date(value)
        if isinstance(parsed, (datetime.date, datetime.datetime)):
            return parsed.strftime("%d/%m/%Y")
        return str(parsed or "")

    # Exact header shape in the supplied Winvic BOQ.  Project-specific notices remain blank
    # unless explicitly provided; they are never copied into unrelated quotations.
    ws["A1"] = _excel_text(f"Project: {q.get('project') or '—'}")
    ws["B1"] = _excel_text(q.get("measurement_basis") or "")
    ws["A2"] = _excel_text(f"Client: {q.get('client') or '—'}")
    ws["B2"] = _excel_text(q.get("joint_layout_note") or "")
    ws["A3"] = _excel_text(f"Date: {_date_text(q.get('date'))}")
    ws["B3"] = _excel_text(q.get("joint_details_note") or "")
    ws["A4"] = _excel_text(f"Rev: {q.get('revision') or q.get('ref') or '—'}")
    ws.merge_cells("D4:E4")
    drawings = q.get("drawings") or []
    ws["A5"] = _excel_text(
        "Drawing ref available at tender:" +
        (("\n" + "\n".join(str(item) for item in drawings)) if drawings else "")
    )
    ws["A5"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[5].height = max(30, 15 * max(2, len(drawings) + 1))
    ws["A6"] = _excel_text(q.get("market_warning") or "")
    ws["A6"].font = Font(name="Arial", size=8, color=red)
    ws["A6"].alignment = Alignment(wrap_text=True, vertical="top")
    for row_index in range(1, 6):
        ws.cell(row_index, 1).font = Font(name="Arial", size=9, bold=row_index in (1, 2, 3, 4))
        ws.cell(row_index, 2).font = Font(name="Arial", size=9)

    sections = OrderedDict()
    for item in q.get("line_items", []):
        sections.setdefault(item.get("section") or "External yard slabs", []).append(item)
    assessor_measurements = [m for m in (q.get("measurements") or [])
                             if m.get("assessor_rate_required")]
    for measurement in assessor_measurements:
        sections.setdefault(measurement.get("section") or "External yard slabs", [])
    ordered_sections = [section for section in q.get("section_order", SECTION_ORDER) if section in sections]
    ordered_sections += [section for section in sections if section not in ordered_sections]

    specifications = {specification["id"]: specification
                      for specification in q.get("specifications", [])}
    row = 7
    section_fill = PatternFill("solid", fgColor=section_blue)
    provisional_fill = PatternFill("solid", fgColor=pale_gold)
    headings = ("DESCRIPTION", "QTY", "UNIT", "RATE", "VALUE")

    for col, heading in enumerate(headings, 1):
        cell = ws.cell(row, col, heading)
        cell.font = Font(name="Arial", bold=True, color=white, size=11)
        cell.fill = PatternFill("solid", fgColor=black)
        cell.alignment = Alignment(horizontal="right" if col in (2, 4, 5) else "left")
    ws.row_dimensions[row].height = 20
    row += 2  # the supplied workbook leaves row 8 blank

    for section in ordered_sections:
        section_items = sections[section]
        ws.cell(row, 1, _excel_section_title(section, section_items))
        for col in range(1, 6):
            ws.cell(row, col).fill = section_fill
            ws.cell(row, col).font = Font(name="Arial", bold=True, color=white, size=9)
        ws.cell(row, 1).alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 22
        row += 1

        group_ids = list(dict.fromkeys(
            item.get("specification_id") for item in section_items
            if item.get("specification_id")
        ))
        ungrouped = [item for item in section_items if not item.get("specification_id")]

        def _write_item(item, qty_formula=None):
            nonlocal row
            description = item["description"]
            if item.get("provisional"):
                description += f"\n{PROVISIONAL_LABEL}"
            ws.cell(row, 1, _excel_text(description))
            ws.cell(row, 2, qty_formula or float(item["qty"]))
            ws.cell(row, 3, _excel_text(_excel_unit(item["unit"])))
            rate = item.get("rate")
            if isinstance(rate, (int, float)):
                ws.cell(row, 4, float(rate))
            value_status = item.get("value_status")
            if value_status:
                ws.cell(row, 5, _excel_text(value_status))
            elif isinstance(rate, (int, float)):
                ws.cell(row, 5, f"=B{row}*D{row}")
            elif item.get("assessor_rate_required"):
                ws.cell(row, 5, f'=IF(D{row}="","",B{row}*D{row})')
            elif isinstance(item.get("value"), (int, float)):
                ws.cell(row, 5, float(item["value"]))
            ws.cell(row, 2).number_format = '#,##0.##'
            ws.cell(row, 4).number_format = currency_format
            ws.cell(row, 5).number_format = currency_format
            for col in range(1, 6):
                ws.cell(row, col).border = Border(bottom=thin_grey)
                ws.cell(row, col).alignment = Alignment(
                    horizontal="right" if col in (2, 4, 5) else "left",
                    vertical="top", wrap_text=col == 1)
                ws.cell(row, col).font = Font(name="Arial", size=8)
                if item.get("provisional"):
                    ws.cell(row, col).fill = provisional_fill
            if item.get("provisional"):
                ws.row_dimensions[row].height = 30
            row += 1

        for group_id in group_ids:
            specification = specifications.get(group_id) or {}
            source_rows = specification.get("area_rows") or []
            first_source_row = row
            for source_row in source_rows:
                ws.cell(row, 1, _excel_text(source_row.get("description") or "Measured area"))
                ws.cell(row, 2, float(source_row.get("qty") or 0))
                ws.cell(row, 3, _excel_unit(source_row.get("unit") or "m²"))
                ws.cell(row, 2).number_format = '#,##0.##'
                for col in range(1, 6):
                    ws.cell(row, col).font = Font(name="Arial", size=8)
                row += 1
            if source_rows:
                total_area_row = row
                ws.cell(row, 1, "Total Area Take Off:")
                ws.cell(row, 1).font = Font(name="Arial", size=8, bold=True)
                ws.cell(row, 2, f"=SUM(B{first_source_row}:B{row - 1})")
                ws.cell(row, 2).number_format = '#,##0.##'
                ws.cell(row, 3, "m2")
                row += 1
            else:
                total_area_row = None

            display_lines = specification.get("display_lines") or []
            if display_lines:
                ws.cell(row, 1, _excel_text(
                    f"Specification — {specification.get('slab_type_label') or section}\n" +
                    "\n".join(f"{field['label']}: {field['value']}" for field in display_lines)
                ))
                ws.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="top")
                ws.cell(row, 1).font = Font(name="Arial", size=8,
                                            italic=bool(specification.get("provisional")),
                                            color=red if specification.get("provisional") else black)
                if specification.get("provisional"):
                    for col in range(1, 6):
                        ws.cell(row, col).fill = provisional_fill
                ws.row_dimensions[row].height = max(30, 12 * (len(display_lines) + 1))
                row += 1

            qty_formula = f"=B{total_area_row}" if total_area_row else None
            for item in section_items:
                if item.get("specification_id") == group_id:
                    _write_item(item, qty_formula=qty_formula)

        for item in ungrouped:
            _write_item(item)

        for measurement in (m for m in assessor_measurements if m["section"] == section):
            quantity_rows = measurement.get("quantity_rows") or []
            first_quantity_row = row
            for quantity_row in quantity_rows:
                ws.cell(row, 1, _excel_text(quantity_row.get("description") or "Measured length"))
                ws.cell(row, 2, float(quantity_row.get("qty") or 0))
                ws.cell(row, 3, _excel_unit(quantity_row.get("unit") or "Lm"))
                ws.cell(row, 2).number_format = '#,##0.##'
                for col in range(1, 6):
                    ws.cell(row, col).font = Font(name="Arial", size=8)
                row += 1
            quantity_formula = None
            if quantity_rows:
                total_quantity_row = row
                ws.cell(row, 1, f"Total {measurement['description']}:")
                ws.cell(row, 1).font = Font(name="Arial", size=8, bold=True)
                ws.cell(row, 2, f"=SUM(B{first_quantity_row}:B{row - 1})")
                ws.cell(row, 2).number_format = '#,##0.##'
                ws.cell(row, 3, _excel_unit(measurement.get("unit") or "Lm"))
                row += 1
                quantity_formula = f"=B{total_quantity_row}"
            _write_item({
                "description": measurement["description"], "qty": measurement["qty"],
                "unit": measurement["unit"], "rate": None, "value": None,
                "provisional": measurement.get("provisional", False),
                "assessor_rate_required": True,
            }, qty_formula=quantity_formula)

        row += 2

    total_row = row
    ws.cell(total_row, 1, "TOTAL NETT")
    ws.cell(total_row, 1).font = Font(name="Arial", bold=True, size=9)
    ws.cell(total_row, 5, f"=SUM(E7:E{total_row - 1})")
    ws.cell(total_row, 5).number_format = currency_format
    ws.cell(total_row, 5).font = Font(name="Arial", bold=True, size=9)

    row = total_row + 3
    measurements = [m for m in (q.get("measurements") or [])
                    if not m.get("assessor_rate_required")]
    if measurements:
        ws.cell(row, 1, "INFORMATIONAL MEASUREMENTS — NOT PRICED")
        for col in range(1, 6):
            ws.cell(row, col).font = Font(name="Arial", bold=True, color=white, size=9)
            ws.cell(row, col).fill = section_fill
        row += 1
        for measurement in measurements:
            description = f"{_excel_section_title(measurement['section'], [])} — {measurement['description']}"
            if measurement.get("provisional"):
                description += f"\n{PROVISIONAL_LABEL}"
            ws.cell(row, 1, description)
            ws.cell(row, 2, float(measurement["qty"]))
            ws.cell(row, 3, _excel_unit(measurement["unit"]))
            ws.cell(row, 2).number_format = '#,##0.##'
            if measurement.get("provisional"):
                for col in range(1, 6):
                    ws.cell(row, col).fill = provisional_fill
                ws.row_dimensions[row].height = 30
            row += 1
        row += 1

    if q.get("declarations"):
        ws.cell(row, 1, "NOTES / ASSUMPTIONS")
        for col in range(1, 6):
            ws.cell(row, col).font = Font(name="Arial", bold=True, color=white, size=9)
            ws.cell(row, col).fill = section_fill
        row += 1
        for note in q["declarations"]:
            ws.cell(row, 1, f"• {note}")
            ws.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row, 1).font = Font(name="Arial", size=8)
            ws.row_dimensions[row].height = 30
            row += 1

    ws.cell(row + 1, 1, f"STANDARD TERMS: {q['terms']}")
    ws.cell(row + 1, 1).font = Font(name="Arial", size=8)
    ws.cell(row + 1, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row + 1].height = 45
    row += 2
    for term in q.get("commercial_terms") or []:
        ws.cell(row, 1, _excel_text(term))
        ws.cell(row, 1).font = Font(name="Arial", size=8)
        ws.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    widths = {"A": 82.43, "B": 19.43, "C": 10.29, "D": 12.71, "E": 21.14}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.print_title_rows = "1:7"
    ws.print_area = f"A1:E{row}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.oddFooter.center.text = f"{FORTEL_NAME} · {FORTEL_EMAIL} · {FORTEL_TEL}"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def save_quotation(q: dict, out_dir: str = ".") -> dict:
    """Save text, HTML, JSON and editable Excel versions to disk. Returns paths dict."""
    base = Path(out_dir) / q["ref"]
    paths = {}
    base.parent.mkdir(parents=True, exist_ok=True)
    (p := Path(f"{base}.txt")).write_text(quotation_text(q));  paths["txt"]  = str(p)
    (p := Path(f"{base}.html")).write_text(quotation_html(q)); paths["html"] = str(p)
    (p := Path(f"{base}.json")).write_text(quotation_json(q)); paths["json"] = str(p)
    (p := Path(f"{base}.xlsx")).write_bytes(quotation_xlsx(q)); paths["xlsx"] = str(p)
    return paths


# ── Standalone demo ───────────────────────────────────────────────────────────
if __name__ == "__main__":
    import sys, io, contextlib

    # Simulate a pipeline result (as returned by takeoff_pipeline.takeoff)
    demo_result = {
        "file":             "SGP-D77-Hard-Landscaping.pdf",
        "type":             "UNMARKED vector",
        "confidence":       "medium",
        "source_discipline":"architect",
        "area_m2":          3172,
        "costing": {
            "area_m2":   3172,
            "rate":      44.89,
            "total_gbp": 142_391.08,
            "assumed":   True,
            "spec": {
                "depth_mm": 190, "mesh": "A252", "conc_mix": "C32/40",
                "layers": 1, "conc_rate": 128,
            },
            "breakdown": {
                "concrete": 25.05, "steel": 4.30, "dpm": 0.46,
                "curing": 0.23, "labour": 10.00, "trim": 0.40,
                "nett": 40.44, "margin%": 11,
            },
        },
        "flags": [
            "BUILD-UP ASSUMED: 190mm / A252 / C32/40 — no engineer construction-detail found",
            "ARCHITECT drawing — build-up ASSUMED; no construction-detail sheet found.",
            "assessor: confirm extent + scale",
        ],
    }

    extras = [
        ("Manholes (inside concrete boundary)", 3, "Nr",  75.00),
        ("Slot drain channels (linear metres)", 48, "Lm", 17.50),
    ]

    q = generate_quotation(demo_result, project="Hemington D77 Hard Landscaping",
                           client="Winvic Construction", ref="FTL-2026-D77A",
                           extras=extras)

    print("── Plain text quotation ─────────────────────────────────────────────")
    print(quotation_text(q))
    print()

    paths = save_quotation(q, out_dir="quotations")
    print(f"Saved: {paths}")

    # Quick validation
    assert q["total_gbp"] > 0, "total must be positive"
    assert q["assumed"] is True, "should flag assumed build-up"
    assert any("ASSUMED" in d for d in q["declarations"]), "must declare assumption"
    assert len(q["line_items"]) >= 3, "at least slab + trim + joints"
    print("\nAll assertions passed ✅")
